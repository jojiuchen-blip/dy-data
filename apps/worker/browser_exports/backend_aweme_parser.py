from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from apps.worker.collectors.normalizers import text


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "douyin_nickname": ("\u6296\u97f3\u6635\u79f0", "\u6296\u97f3\u53f7\u6635\u79f0", "nickname", "douyin_nickname"),
    "douyin_id": ("\u6296\u97f3id", "\u6296\u97f3ID", "\u6296\u97f3\u53f7", "aweme_id", "aweme_short_id", "douyin_id"),
    "account_id": (
        "\u6240\u5c5e\u8d26\u6237id",
        "\u6240\u5c5e\u8d26\u6237ID",
        "\u8d26\u6237id",
        "\u804c\u4ebaUID",
        "account_id",
        "craftsman_uid",
    ),
    "account_name": (
        "\u6240\u5c5e\u8d26\u6237\u540d\u79f0",
        "\u6240\u5c5e\u8d26\u6237",
        "\u5546\u5bb6\u4e3b\u4f53",
        "account_name",
    ),
    "poi_id": (
        "\u6240\u5c5e\u8d26\u6237\u5173\u8054poi_id",
        "\u6240\u5c5e\u8d26\u6237\u5173\u8054POI_ID",
        "\u5173\u8054poi_id",
        "\u7ed1\u5b9a\u95e8\u5e97ID",
        "poi_id",
    ),
    "certified_subject_name": (
        "\u8ba4\u8bc1\u4e3b\u4f53",
        "certified_subject_name",
        "certified_subject",
        "subject_name",
    ),
    "binding_status": ("\u6296\u97f3\u53f7\u7ed1\u5b9a\u72b6\u6001", "\u7ed1\u5b9a\u72b6\u6001", "bind_status", "binding_status"),
}


def parse_backend_aweme_workbook(path: str | Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet.max_row == 1 and sheet.max_column == 1 and hasattr(sheet, "reset_dimensions"):
        sheet.reset_dimensions()
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_cell_text(value) for value in rows[0]]
    records: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        raw_payload = {
            header: _cell_text(raw_row[index]) if index < len(raw_row) else ""
            for index, header in enumerate(headers)
            if header
        }
        if not any(raw_payload.values()):
            continue
        record = {
            "douyin_nickname": _alias_value(raw_payload, "douyin_nickname"),
            "douyin_id": _alias_value(raw_payload, "douyin_id"),
            "account_id": _alias_value(raw_payload, "account_id"),
            "account_name": _alias_value(raw_payload, "account_name"),
            "poi_id": _alias_value(raw_payload, "poi_id"),
            "certified_subject_name": _alias_value(raw_payload, "certified_subject_name"),
            "binding_status": _alias_value(raw_payload, "binding_status"),
            "raw_payload": raw_payload,
        }
        if any(record[key] for key in ("douyin_nickname", "douyin_id", "account_id", "poi_id")):
            records.append(record)
    return records


def _alias_value(raw_payload: dict[str, str], field: str) -> str | None:
    for alias in HEADER_ALIASES[field]:
        value = text(raw_payload.get(alias))
        if value:
            return value
    return None


def _cell_text(value: Any) -> str:
    return text(value) or ""
