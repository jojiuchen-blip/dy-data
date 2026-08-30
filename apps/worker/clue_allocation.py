from __future__ import annotations

import json
import re
from collections import defaultdict
from collections.abc import Callable, Iterable, Mapping
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from math import asin, cos, radians, sin, sqrt
from pathlib import Path
from uuid import uuid4
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from sqlalchemy import and_, delete, exists, func, insert, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from apps.api.dy_api.models import (
    ClueAllocationRuleVersion,
    ClueAssignmentRound,
    ClueCenterOrder,
    ClueFollowUpRecord,
    ClueHeadquartersPoolEntry,
    ClueLeadRuleVersionBinding,
    ClueMaterializationTarget,
    ClueMaterializationWorkItem,
    ClueMasterLead,
    ClueOrderStatusEvent,
    ClueSourceRecordLink,
    ClueSourceIdentifierHistory,
    DataQualityIssue,
    DimStore,
    DimStorePoiMapping,
    JobImpact,
    JobImpactWatermark,
    RawDouyinClue,
    RawDouyinOrder,
    RawDouyinOrderCoupon,
    SettlementOrderDetail,
    SettlementProjectionActive,
    SettlementProjectionGeneration,
    SettlementProjectionPartitionManifest,
    StoreScoreSnapshot,
    StoreScoreSnapshotGeneration,
    StoreScoreSnapshotRun,
    utcnow,
)
from apps.worker.clue_headquarters_pool import (
    close_current_headquarters_pool_entry,
    ensure_active_headquarters_pool_entry,
)
from apps.worker.order_status import normalize_coupon_status, resolve_clue_order_status
from apps.worker.clue_center import refresh_clue_center_projection
from apps.worker.repositories import (
    begin_clue_materialization_cycle,
    claim_clue_materialization_batch,
    complete_clue_materialization_batch,
    renew_clue_materialization_batch,
    retry_clue_materialization_batch,
    upsert_data_quality_issue,
)
from apps.worker.projection_lineage import canonical_score_partition_key


SHANGHAI = ZoneInfo("Asia/Shanghai")
DEFAULT_STORE_WEIGHT = Decimal("1")
SCHEDULED_SCORE_REFRESH_TIME = time(hour=3)
MASTER_MATERIALIZATION_LOCK = "clue-allocation-master-materialization"
MASTER_STATUS_REPAIR_LOCK = "clue-allocation-master-status-repair"
MASTER_TERMINAL_SYNC_LOCK = "clue-allocation-master-terminal-sync"
SCHEDULED_SCORE_REFRESH_LOCK = "clue-allocation-scheduled-score-refresh"
BUSINESS_EXECUTION_MODE = "formal"
MATERIALIZATION_QUERY_BATCH_SIZE = 10_000
CENTER_ORDER_BATCH_SIZE = 64
ANCHOR_UNAVAILABLE_REASONS = (
    "follow_poi_missing",
    "follow_poi_unmapped",
    "follow_poi_store_missing",
    "anchor_coordinates_invalid",
    "anchor_province_missing",
    "anchor_city_missing",
    "anchor_city_code_missing",
)
SCORE_SPARSE_PROTOCOL = "t344-score-sparse-v1"
MAX_SCORE_SPARSE_STORES = 8192
MAX_SCORE_SPARSE_RULES = 64


@dataclass(frozen=True)
class StatusResolution:
    raw_status: str | None
    normalized_status: str
    status_source: str
    closed_at: datetime | None


@dataclass(frozen=True)
class AnchorSnapshot:
    poi_id: str | None
    store_id: str | None
    unavailable_reason: str | None
    province: str | None
    city: str | None
    city_code: str | None
    longitude: Decimal | None
    latitude: Decimal | None


@dataclass
class StoreMetrics:
    conversion_numerator: int = 0
    conversion_denominator: int = 0
    follow_24h_numerator: int = 0
    follow_24h_denominator: int = 0

    def add(self, *, converted: bool, followed_within_24h: bool, has_full_follow_up_opportunity: bool) -> None:
        self.conversion_denominator += 1
        if converted:
            self.conversion_numerator += 1
        if has_full_follow_up_opportunity:
            self.follow_24h_denominator += 1
        if has_full_follow_up_opportunity and followed_within_24h:
            self.follow_24h_numerator += 1


@dataclass(frozen=True)
class StoreScoreConfig:
    rule_version_id: str | None
    lookback_days: int
    min_samples: int
    conversion_weight: Decimal
    follow_weight: Decimal
    store_weight: Decimal


@dataclass(frozen=True)
class ScoreManifest:
    generation_id: str
    base_generation_id: str
    snapshot_date: date
    rule_version_ids: tuple[str, ...]
    snapshot_run_ids: tuple[str, ...]
    partition_keys: tuple[str, ...]
    manifest_count: int
    row_count: int
    manifest_checksum: str
    resumed: bool


def materialize_clue_master_leads(
    session: Session,
    *,
    now: datetime | None = None,
    raw_clues: list[RawDouyinClue] | None = None,
    raw_clue_row_keys: set[str] | None = None,
    raw_page_clues: list[RawDouyinClue] | None = None,
    clue_ids: set[str] | None = None,
    order_ids: set[str] | None = None,
    poi_ids: set[str] | None = None,
    source_identity_keys: set[str] | None = None,
    existing_clue_ids: set[str] | None = None,
    existing_order_ids: set[str] | None = None,
    existing_poi_ids: set[str] | None = None,
    existing_source_identity_keys: set[str] | None = None,
) -> dict[str, object]:
    """Build clue master state from the full ledger or one bounded impact page."""
    now = _aware(now or utcnow())
    if not _try_transaction_lock(session, lock_name=MASTER_MATERIALIZATION_LOCK):
        return {"master_leads": 0, "closed_leads": 0, "headquarters_pool": 0, "skipped": "locked"}
    if raw_clues is not None and raw_page_clues is not None:
        raise ValueError("provide either raw_clues or raw_page_clues, not both")
    incremental = raw_clues is not None or raw_page_clues is not None or any(
        selector is not None
        for selector in (
            raw_clue_row_keys,
            clue_ids,
            order_ids,
            poi_ids,
            source_identity_keys,
            existing_clue_ids,
            existing_order_ids,
            existing_poi_ids,
            existing_source_identity_keys,
        )
    )
    if incremental:
        session.info.pop("clue_identifier_conflicts", None)
        session.info.pop("clue_identifier_conflict_counts", None)
    explicit_page = raw_page_clues if raw_page_clues is not None else raw_clues
    if explicit_page is not None:
        selected_raw_clues = list(explicit_page)
    elif incremental:
        selected_raw_clues = _bounded_raw_clues(
            session,
            raw_clue_row_keys=raw_clue_row_keys or set(),
            clue_ids=clue_ids or set(),
            order_ids=order_ids or set(),
            poi_ids=poi_ids or set(),
        )
    else:
        selected_raw_clues = session.scalars(select(RawDouyinClue)).all()
    raw_clues = selected_raw_clues
    if not raw_clues:
        return {"master_leads": 0, "closed_leads": 0, "headquarters_pool": 0}

    selected_order_ids = {_clean(row.order_id) for row in raw_clues}
    selected_order_ids.discard(None)
    selected_order_ids.update(
        _clean(value)
        for value in (
            existing_order_ids
            if explicit_page is not None and existing_order_ids is not None
            else order_ids or set()
        )
    )
    selected_order_ids.discard(None)
    raw_orders = _raw_orders_by_id(session, selected_order_ids)
    coupon_statuses_by_order = _coupon_statuses_by_order(session, selected_order_ids)
    verified_at_by_order = _verified_at_by_order(session, selected_order_ids)
    if incremental:
        mappings_by_poi, stores_by_id = _bounded_location_context(
            session,
            raw_clues,
            poi_ids=poi_ids or set(),
        )
    else:
        stores_by_id = {row.store_id: row for row in session.scalars(select(DimStore)).all()}
        mappings_by_poi = {row.poi_id: row for row in session.scalars(select(DimStorePoiMapping)).all()}
    _enrich_store_locations_from_raw_evidence(
        session,
        raw_clues,
        mappings_by_poi,
        stores_by_id,
        now,
    )
    source_links_by_record_key = _source_record_links_by_key(
        session,
        {row.clue_row_key for row in raw_clues if row.clue_row_key},
    )
    existing_rows = (
        _bounded_existing_masters(
            session,
            raw_clues,
            order_ids=(
                existing_order_ids
                if existing_order_ids is not None
                else selected_order_ids
            ),
            clue_ids=(existing_clue_ids if existing_clue_ids is not None else clue_ids),
            poi_ids=(existing_poi_ids if existing_poi_ids is not None else poi_ids),
            source_identity_keys=(
                existing_source_identity_keys
                if existing_source_identity_keys is not None
                else source_identity_keys
            ),
            source_link_lead_keys={
                source_link.lead_key
                for source_link in source_links_by_record_key.values()
            },
        )
        if incremental
        else session.scalars(select(ClueMasterLead)).all()
    )
    existing_by_lead_key = {row.lead_key: row for row in existing_rows}
    existing_by_source_clue_row_key = {row.source_clue_row_key: row for row in existing_rows}
    existing_by_identity = {row.source_identity_key: row for row in existing_rows}
    existing_by_canonical_clue_id = {
        row.canonical_clue_id: row for row in existing_rows if row.canonical_clue_id
    }
    masters_by_order_id: dict[str, list[ClueMasterLead]] = defaultdict(list)
    for row in existing_rows:
        if order_id := _clean(row.order_id):
            masters_by_order_id[order_id].append(row)
    existing_by_order_id = {
        order_id: rows[0] for order_id, rows in masters_by_order_id.items() if len(rows) == 1
    }
    if incremental:
        candidate_identifiers = {
            (identifier_type, identifier_value)
            for raw_clue in raw_clues
            for identifier_type, identifier_value in (
                ("clue_id", _clean(raw_clue.clue_id)),
                ("source_identity_key", _source_identity_key(raw_clue)),
            )
            if identifier_value
        }
        identifier_history_rows = _bounded_identifier_history(
            session,
            raw_clue_row_keys={row.clue_row_key for row in raw_clues},
            candidate_identifiers=candidate_identifiers,
        )
    else:
        identifier_history_rows = session.scalars(
            select(ClueSourceIdentifierHistory)
        ).all()
    identifier_history_by_key = {
        (row.source_clue_row_key, row.identifier_type, row.identifier_value): row
        for row in identifier_history_rows
    }
    current_identifier_history_by_source_type: dict[
        tuple[str, str], list[ClueSourceIdentifierHistory]
    ] = defaultdict(list)
    existing_by_identifier: dict[tuple[str, str], ClueMasterLead] = {}
    source_history_lead_keys: dict[str, set[str]] = defaultdict(set)
    for row in identifier_history_rows:
        source_history_lead_keys[row.source_clue_row_key].add(row.lead_key)
        lead = existing_by_lead_key.get(row.lead_key)
        if lead is not None:
            existing_by_identifier.setdefault(
                (row.identifier_type, row.identifier_value),
                lead,
            )
        if row.is_current:
            current_identifier_history_by_source_type[
                (row.source_clue_row_key, row.identifier_type)
            ].append(row)
    for source_clue_row_key, lead_keys in source_history_lead_keys.items():
        if len(lead_keys) != 1:
            continue
        lead = existing_by_lead_key.get(next(iter(lead_keys)))
        if lead is not None:
            existing_by_source_clue_row_key.setdefault(source_clue_row_key, lead)
    for source_record_key, source_link in source_links_by_record_key.items():
        linked_lead = existing_by_lead_key.get(source_link.lead_key)
        if linked_lead is not None:
            existing_by_source_clue_row_key.setdefault(source_record_key, linked_lead)
    current_round_ids = {
        row.current_assignment_round_id for row in existing_rows if row.current_assignment_round_id
    }
    current_rounds_by_id: dict[str, ClueAssignmentRound] = {}
    for round_id_batch in _materialization_order_id_batches(current_round_ids):
        current_rounds_by_id.update(
            {
                row.assignment_round_id: row
                for row in session.scalars(
                    select(ClueAssignmentRound).where(
                        ClueAssignmentRound.assignment_round_id.in_(round_id_batch)
                    )
                ).all()
            }
        )
    center_orders_by_id: dict[str, ClueCenterOrder] = {}
    for order_id_batch in _materialization_order_id_batches(selected_order_ids):
        center_orders_by_id.update(
            {
                row.order_id: row
                for row in session.scalars(
                    select(ClueCenterOrder).where(ClueCenterOrder.order_id.in_(order_id_batch))
                ).all()
            }
        )
    projected_round_ids = {
        row.current_assignment_round_id
        for row in center_orders_by_id.values()
        if row.current_assignment_round_id and row.current_assignment_round_id not in current_rounds_by_id
    }
    for round_id_batch in _materialization_order_id_batches(projected_round_ids):
        current_rounds_by_id.update(
            {
                row.assignment_round_id: row
                for row in session.scalars(
                    select(ClueAssignmentRound).where(
                        ClueAssignmentRound.assignment_round_id.in_(round_id_batch)
                    )
                ).all()
            }
        )
    affected_lead_keys = set(existing_by_lead_key)
    if incremental:
        active_headquarters_entries_by_lead = {
            row.lead_key: row
            for row in _bounded_active_headquarters_entries(session, affected_lead_keys)
        }
        anchor_issue_ids = set(_bounded_anchor_issue_ids(session, affected_lead_keys))
        status_event_ids = None
    else:
        active_headquarters_entries_by_lead = {
            row.lead_key: row
            for row in session.scalars(
                select(ClueHeadquartersPoolEntry).where(ClueHeadquartersPoolEntry.status == "active")
            ).all()
        }
        anchor_issue_ids = set(
            session.scalars(
                select(DataQualityIssue.issue_id).where(DataQualityIssue.issue_id.like("clue-anchor:%"))
            ).all()
        )
        status_event_ids = set(session.scalars(select(ClueOrderStatusEvent.event_id)).all())

    materialized_lead_keys: set[str] = set()
    closed_lead_keys: set[str] = set()
    headquarters_pool_keys: set[str] = set()
    for raw_clue in raw_clues:
        source_identity_key = _source_identity_key(raw_clue)
        canonical_clue_id = _clean(raw_clue.clue_id)
        order_id = _clean(raw_clue.order_id)
        source_match = existing_by_source_clue_row_key.get(raw_clue.clue_row_key)
        order_match = existing_by_order_id.get(order_id) if order_id else None
        identity_match = existing_by_identity.get(source_identity_key)
        canonical_match = (
            existing_by_canonical_clue_id.get(canonical_clue_id) if canonical_clue_id else None
        )
        history_match = (
            existing_by_identifier.get(("clue_id", canonical_clue_id))
            if canonical_clue_id
            else None
        )
        lead_key_match = existing_by_lead_key.get(_lead_key(source_identity_key))

        conflict_reason = _master_match_conflict_reason(
            source_match=source_match,
            order_match=order_match,
            identity_match=identity_match,
            order_id=order_id,
        )
        if conflict_reason:
            _record_identity_mapping_conflict(
                session,
                raw_clue=raw_clue,
                source_match=source_match,
                order_match=order_match,
                identity_match=identity_match,
                reason=conflict_reason,
                now=now,
            )
            conflict_lead = source_match or order_match or identity_match
            if conflict_lead is not None:
                _upsert_source_record_link(
                    session,
                    raw_clue=raw_clue,
                    lead=conflict_lead,
                    order_id=_clean(conflict_lead.order_id),
                    link_status=3,
                    link_method=2,
                    conflict_reason=conflict_reason,
                    observed_at=_observed_at(raw_clue, now),
                    now=now,
                    links_by_record_key=source_links_by_record_key,
                )
            continue

        collision_candidates = _raw_identifier_collision_candidates(
            raw_clue,
            source_identity_key=source_identity_key,
            canonical_clue_id=canonical_clue_id,
        )
        collision_counts = session.info.get("clue_identifier_conflict_counts", {})
        collisions = [
            (identifier_type, identifier_value, int(collision_counts[key]))
            for identifier_type, identifier_value in collision_candidates
            for key in ((identifier_type, identifier_value),)
            if key in collision_counts and int(collision_counts[key]) > 1
        ]
        stronger_match = any(
            match is not None
            for match in (source_match, order_match, identity_match, lead_key_match)
        )
        if incremental and collisions and not stronger_match:
            _record_identifier_collision_issue(
                session,
                raw_clue=raw_clue,
                collisions=collisions,
                now=now,
            )
            continue

        existing = source_match or order_match
        if existing is None:
            existing = next(
                (
                    candidate
                    for candidate in (
                        identity_match,
                        canonical_match,
                        history_match,
                        lead_key_match,
                    )
                    if candidate is not None and _master_order_is_compatible(candidate, order_id)
                ),
                None,
            )
        observed_at = _observed_at(raw_clue, now)
        if (
            incremental
            and existing is not None
            and existing.last_seen_at is not None
            and not _observation_is_newer(
                observed_at,
                _clean(raw_clue.observation_key),
                _aware(existing.last_seen_at),
                _clean(existing.last_observation_key),
            )
        ):
            _record_stale_source_identifiers(
                session,
                existing=existing,
                raw_clue=raw_clue,
                observed_at=observed_at,
                now=now,
                history_by_key=identifier_history_by_key,
                current_by_source_type=current_identifier_history_by_source_type,
            )
            continue
        is_new_master = existing is None
        resolution = _resolve_status(
            raw_clue,
            raw_orders.get(order_id or ""),
            verified_at_by_order.get(order_id or ""),
            now,
            coupon_statuses_by_order.get(order_id or "", ()),
        )
        anchor = _resolve_anchor(raw_clue, mappings_by_poi, stores_by_id)
        is_isolated_source = order_id is None
        lifecycle_status = (
            "isolated" if is_isolated_source else _lifecycle_status(resolution.normalized_status)
        )

        status_observed_at = _status_observed_at(
            raw_clue,
            raw_orders.get(order_id or ""),
            now,
        )
        if existing is None:
            lead_key = _lead_key(source_identity_key)
            existing = ClueMasterLead(
                lead_key=lead_key,
                source_clue_row_key=raw_clue.clue_row_key,
                source_identity_key=source_identity_key,
                master_kind=2 if is_isolated_source else 1,
                is_complete_pool=not is_isolated_source,
                state_version=1,
                created_at=now,
            )
            session.add(existing)
            existing_by_lead_key[lead_key] = existing
            status_changed = True
        else:
            status_changed = False

        state_before = None if is_new_master else _master_state_signature(existing)
        accepts_status_evidence = _accepts_status_evidence(
            existing,
            incoming_status=resolution.normalized_status,
            observed_at=status_observed_at,
        )
        if accepts_status_evidence:
            status_changed = (
                existing.raw_order_status != resolution.raw_status
                or existing.normalized_order_status != resolution.normalized_status
                or existing.status_source != resolution.status_source
            )
        else:
            lifecycle_status = existing.lifecycle_status

        current_self_owned_round = _active_self_owned_current_round(
            session,
            existing,
            current_rounds_by_id=current_rounds_by_id,
        )
        active_headquarters_entry = existing.lead_key in active_headquarters_entries_by_lead
        if is_isolated_source:
            pool_location = None
            allocation_state = "isolated"
        elif lifecycle_status in {"closed_verified", "closed_refunded", "closed_order"}:
            pool_location = "closed"
            allocation_state = "closed"
        elif lifecycle_status == "status_review":
            pool_location = "status_review"
            allocation_state = "status_review"
        elif current_self_owned_round is not None:
            pool_location = "store_follow_up_pool"
            allocation_state = "assigned"
        elif active_headquarters_entry or existing.pool_location == "headquarters_pool":
            # Re-entry to a store pool must be an explicit future operation.
            pool_location = "headquarters_pool"
            allocation_state = "headquarters"
        elif anchor.unavailable_reason:
            pool_location = "headquarters_pool"
            allocation_state = "headquarters"
        else:
            # M2 creates the first self-owned store assignment. This is not a business pool yet.
            pool_location = None
            allocation_state = "pending_allocation"

        existing.source_identity_key = source_identity_key
        existing.canonical_clue_id = canonical_clue_id or existing.canonical_clue_id
        existing.order_id = order_id
        existing.master_kind = 2 if is_isolated_source else 1
        existing.is_complete_pool = not is_isolated_source
        if accepts_status_evidence:
            existing.raw_order_status = resolution.raw_status
            existing.normalized_order_status = resolution.normalized_status
            existing.status_source = resolution.status_source
            existing.order_status_observed_at = status_observed_at
            existing.lifecycle_status = lifecycle_status
        existing.pool_location = pool_location
        existing.allocation_state = allocation_state
        existing.ended_without_assignment = (
            not is_isolated_source
            and lifecycle_status in {"closed_verified", "closed_refunded", "closed_order"}
            and existing.current_assignment_round_id is None
            and existing.allocation_cycle_id is None
        )
        if accepts_status_evidence:
            existing.closed_at = (
                resolution.closed_at
                if lifecycle_status in {"closed_verified", "closed_refunded", "closed_order"}
                else None
            )
            existing.closed_reason = (
                _closed_reason(resolution.normalized_status)
                if not is_isolated_source
                else None
            )
        existing.first_seen_at = existing.first_seen_at or _first_seen_at(raw_clue, now)
        existing.last_seen_at = max(
            filter(None, (_aware(existing.last_seen_at), observed_at)),
        )
        existing.last_observation_key = _clean(raw_clue.observation_key)
        existing.anchor_poi_id = anchor.poi_id
        existing.anchor_store_id = anchor.store_id
        existing.anchor_source = "douyin_follow_poi" if anchor.poi_id else None
        existing.anchor_unavailable_reason = anchor.unavailable_reason
        existing.anchor_province = anchor.province
        existing.anchor_city = anchor.city
        existing.anchor_city_code = anchor.city_code
        existing.anchor_longitude = anchor.longitude
        existing.anchor_latitude = anchor.latitude
        if state_before is not None and state_before != _master_state_signature(existing):
            existing.state_version = max(existing.state_version or 1, 1) + 1
        existing.updated_at = now

        materialized_lead_keys.add(existing.lead_key)
        existing_by_source_clue_row_key[raw_clue.clue_row_key] = existing
        if order_id:
            existing_by_order_id.setdefault(order_id, existing)
        existing_by_identity[source_identity_key] = existing
        if canonical_clue_id:
            existing_by_canonical_clue_id[canonical_clue_id] = existing
            existing_by_identifier[("clue_id", canonical_clue_id)] = existing
        existing_by_identifier[("source_identity_key", source_identity_key)] = existing
        payload_hash = _source_payload_hash(raw_clue.raw_payload)
        _upsert_source_record_link(
            session,
            raw_clue=raw_clue,
            lead=existing,
            order_id=order_id,
            link_status=2 if is_isolated_source else 1,
            link_method=2 if is_isolated_source else 1,
            conflict_reason="missing_order_id" if is_isolated_source else None,
            observed_at=observed_at,
            now=now,
            links_by_record_key=source_links_by_record_key,
        )
        _set_current_source_identifier(
            session,
            lead_key=existing.lead_key,
            source_clue_row_key=raw_clue.clue_row_key,
            identifier_type="clue_id",
            identifier_value=canonical_clue_id,
            source_payload_hash=payload_hash,
            observed_at=observed_at,
            now=now,
            history_by_key=identifier_history_by_key,
            current_by_source_type=current_identifier_history_by_source_type,
        )
        _set_current_source_identifier(
            session,
            lead_key=existing.lead_key,
            source_clue_row_key=raw_clue.clue_row_key,
            identifier_type="source_identity_key",
            identifier_value=source_identity_key,
            source_payload_hash=payload_hash,
            observed_at=observed_at,
            now=now,
            history_by_key=identifier_history_by_key,
            current_by_source_type=current_identifier_history_by_source_type,
        )
        if not is_isolated_source and lifecycle_status in {
            "closed_verified",
            "closed_refunded",
            "closed_order",
        }:
            closed_lead_keys.add(existing.lead_key)
        elif pool_location == "headquarters_pool":
            headquarters_pool_keys.add(existing.lead_key)

        if status_changed and not is_isolated_source:
            _record_status_event(
                session,
                lead_key=existing.lead_key,
                order_id=existing.order_id,
                resolution=resolution,
                observed_at=status_observed_at,
                created_at=now,
                known_event_ids=status_event_ids,
            )
        if anchor.unavailable_reason and not is_isolated_source:
            _record_anchor_quality_issue(session, existing.lead_key, anchor, now, anchor_issue_ids)
    session.flush()
    for lead_key in materialized_lead_keys:
        lead = existing_by_lead_key[lead_key]
        if lead.lifecycle_status == "active" and lead.pool_location == "headquarters_pool":
            ensure_active_headquarters_pool_entry(
                session,
                lead=lead,
                reason=lead.anchor_unavailable_reason or "headquarters_pool_retained",
                entered_at=now,
                _active_entries_by_lead=active_headquarters_entries_by_lead,
                _flush=False,
            )
        elif lead.lifecycle_status != "active" and lead.order_id:
            close_current_headquarters_pool_entry(
                session,
                lead.lead_key,
                closed_at=lead.closed_at or now,
                close_reason=lead.closed_reason or "order_closed",
                _active_entries_by_lead=active_headquarters_entries_by_lead,
            )
            _close_current_assignment(
                session,
                lead.order_id,
                lead.lifecycle_status,
                lead.closed_at or now,
                current_assignment_round_id=lead.current_assignment_round_id,
                center_orders_by_id=center_orders_by_id,
                current_rounds_by_id=current_rounds_by_id,
            )

    session.flush()
    return {
        "master_leads": len(materialized_lead_keys),
        "closed_leads": len(closed_lead_keys),
        "headquarters_pool": len(headquarters_pool_keys),
    }


def refresh_unknown_clue_master_statuses(
    session: Session,
    *,
    now: datetime | None = None,
    batch_size: int = 500,
    dry_run: bool = False,
) -> dict[str, int | bool | str]:
    """Refresh unresolved master statuses without loading the full clue ledger.

    This path is intentionally limited to existing ``unknown`` rows. It uses a
    session advisory lock and commits each keyset page independently so status
    repair cannot recreate the full materialization memory peak.
    """
    if batch_size <= 0:
        raise ValueError("batch_size must be positive")
    now = _aware(now or utcnow())
    lock_key = _advisory_lock_key(MASTER_STATUS_REPAIR_LOCK)
    if not _try_session_advisory_lock(session, lock_key):
        return {"scanned": 0, "updated": 0, "resolved": 0, "status_review": 0, "batches": 0, "dry_run": dry_run, "skipped": "locked"}

    stats: dict[str, int | bool | str] = {
        "scanned": 0,
        "updated": 0,
        "resolved": 0,
        "status_review": 0,
        "batches": 0,
        "dry_run": dry_run,
    }
    last_lead_key = ""
    try:
        while True:
            leads = session.scalars(
                select(ClueMasterLead)
                .where(ClueMasterLead.normalized_order_status == "unknown")
                .where(ClueMasterLead.lead_key > last_lead_key)
                .order_by(ClueMasterLead.lead_key)
                .limit(batch_size)
            ).all()
            if not leads:
                break

            last_lead_key = leads[-1].lead_key
            stats["batches"] = int(stats["batches"]) + 1
            stats["scanned"] = int(stats["scanned"]) + len(leads)
            source_keys = {lead.source_clue_row_key for lead in leads}
            order_ids = {_clean(lead.order_id) for lead in leads}
            order_ids.discard(None)
            raw_clues = {
                row.clue_row_key: row
                for row in session.scalars(
                    select(RawDouyinClue).where(RawDouyinClue.clue_row_key.in_(source_keys))
                ).all()
            }
            raw_orders = _raw_orders_by_id(session, order_ids)
            coupon_statuses_by_order = _coupon_statuses_by_order(session, order_ids)
            verified_at_by_order = _verified_at_by_order(session, order_ids)
            event_ids = set(
                session.scalars(
                    select(ClueOrderStatusEvent.event_id).where(
                        ClueOrderStatusEvent.lead_key.in_({lead.lead_key for lead in leads})
                    )
                ).all()
            )

            for lead in leads:
                raw_clue = raw_clues.get(lead.source_clue_row_key)
                if raw_clue is None:
                    stats["status_review"] = int(stats["status_review"]) + 1
                    continue
                order_id = _clean(lead.order_id)
                resolution = _resolve_status(
                    raw_clue,
                    raw_orders.get(order_id or ""),
                    verified_at_by_order.get(order_id or ""),
                    now,
                    coupon_statuses_by_order.get(order_id or "", ()),
                )
                lifecycle_status = _lifecycle_status(resolution.normalized_status)
                previous_state = (
                    lead.raw_order_status,
                    lead.normalized_order_status,
                    lead.status_source,
                    lead.lifecycle_status,
                    lead.pool_location,
                    lead.allocation_state,
                )
                if lifecycle_status in {"closed_verified", "closed_refunded", "closed_order"}:
                    pool_location = "closed"
                    allocation_state = "closed"
                elif lifecycle_status == "status_review":
                    pool_location = "status_review"
                    allocation_state = "status_review"
                elif lead.pool_location == "headquarters_pool":
                    pool_location = "headquarters_pool"
                    allocation_state = "headquarters"
                else:
                    current_round = _active_self_owned_current_round(session, lead)
                    pool_location = "store_follow_up_pool" if current_round else None
                    allocation_state = "assigned" if current_round else "pending_allocation"

                next_state = (
                    resolution.raw_status,
                    resolution.normalized_status,
                    resolution.status_source,
                    lifecycle_status,
                    pool_location,
                    allocation_state,
                )
                if next_state != previous_state:
                    stats["updated"] = int(stats["updated"]) + 1
                    if resolution.normalized_status == "unknown":
                        stats["status_review"] = int(stats["status_review"]) + 1
                    else:
                        stats["resolved"] = int(stats["resolved"]) + 1
                if dry_run or next_state == previous_state:
                    continue

                lead.raw_order_status = resolution.raw_status
                lead.normalized_order_status = resolution.normalized_status
                lead.status_source = resolution.status_source
                lead.lifecycle_status = lifecycle_status
                lead.pool_location = pool_location
                lead.allocation_state = allocation_state
                lead.ended_without_assignment = (
                    lifecycle_status in {"closed_verified", "closed_refunded", "closed_order"}
                    and lead.current_assignment_round_id is None
                    and lead.allocation_cycle_id is None
                )
                lead.closed_at = resolution.closed_at if lifecycle_status != "active" else None
                lead.closed_reason = _closed_reason(resolution.normalized_status)
                lead.updated_at = now
                if lifecycle_status != "active":
                    close_at = resolution.closed_at or now
                    close_current_headquarters_pool_entry(
                        session,
                        lead.lead_key,
                        closed_at=close_at,
                        close_reason=lead.closed_reason or "order_status_unknown",
                    )
                    _close_current_assignment(
                        session,
                        order_id or "",
                        lifecycle_status,
                        close_at,
                        current_assignment_round_id=lead.current_assignment_round_id,
                    )
                if resolution.normalized_status != "unknown" or previous_state[3] != lifecycle_status:
                    _record_status_event(
                        session,
                        lead_key=lead.lead_key,
                        order_id=order_id,
                        resolution=resolution,
                        observed_at=_observed_at(raw_clue, now),
                        created_at=now,
                        known_event_ids=event_ids,
                    )

            if not dry_run:
                session.flush()
                session.commit()
                session.expunge_all()
    finally:
        _release_session_advisory_lock(session, lock_key)
    return stats


def synchronize_non_active_clue_states(
    session: Session,
    *,
    now: datetime | None = None,
    batch_size: int = 500,
    dry_run: bool = False,
) -> dict[str, int | bool | str]:
    """Close stale center snapshots and rounds for terminal master leads.

    This is deliberately separate from full materialization. It scans the
    non-active master ledger in keyset pages, closes only already-existing
    rounds, and never creates a new assignment.
    """
    if batch_size <= 0:
        raise ValueError("batch_size must be positive")
    now = _aware(now or utcnow())
    lock_key = _advisory_lock_key(MASTER_TERMINAL_SYNC_LOCK)
    if not _try_session_advisory_lock(session, lock_key):
        return {
            "scanned": 0,
            "orders": 0,
            "rounds_closed": 0,
            "centers_closed": 0,
            "headquarters_closed": 0,
            "batches": 0,
            "dry_run": dry_run,
            "skipped": "locked",
        }

    stats: dict[str, int | bool | str] = {
        "scanned": 0,
        "orders": 0,
        "rounds_closed": 0,
        "centers_closed": 0,
        "headquarters_closed": 0,
        "batches": 0,
        "dry_run": dry_run,
    }
    last_lead_key = ""
    active_round_statuses = ("active_unfollowed", "active_followed")
    try:
        while True:
            leads = session.scalars(
                select(ClueMasterLead)
                .where(ClueMasterLead.lifecycle_status != "active")
                .where(ClueMasterLead.lead_key > last_lead_key)
                .order_by(ClueMasterLead.lead_key)
                .limit(batch_size)
            ).all()
            if not leads:
                break

            last_lead_key = leads[-1].lead_key
            stats["batches"] = int(stats["batches"]) + 1
            stats["scanned"] = int(stats["scanned"]) + len(leads)
            lead_keys = {lead.lead_key for lead in leads}
            leads_by_order: dict[str, list[ClueMasterLead]] = defaultdict(list)
            for lead in leads:
                if order_id := _clean(lead.order_id):
                    leads_by_order[order_id].append(lead)

            order_ids = set(leads_by_order)
            centers = {
                row.order_id: row
                for row in session.scalars(
                    select(ClueCenterOrder).where(ClueCenterOrder.order_id.in_(order_ids))
                ).all()
            }
            rounds_by_order: dict[str, list[ClueAssignmentRound]] = defaultdict(list)
            if order_ids:
                for round_row in session.scalars(
                    select(ClueAssignmentRound)
                    .where(ClueAssignmentRound.order_id.in_(order_ids))
                    .where(ClueAssignmentRound.execution_mode == BUSINESS_EXECUTION_MODE)
                    .where(ClueAssignmentRound.round_status.in_(active_round_statuses))
                ).all():
                    rounds_by_order[round_row.order_id].append(round_row)
            active_hq_by_lead = {
                row.lead_key: row
                for row in session.scalars(
                    select(ClueHeadquartersPoolEntry)
                    .where(ClueHeadquartersPoolEntry.lead_key.in_(lead_keys))
                    .where(ClueHeadquartersPoolEntry.status == "active")
                ).all()
            }

            for lead in leads:
                if lead.lead_key not in active_hq_by_lead:
                    continue
                stats["headquarters_closed"] = int(stats["headquarters_closed"]) + 1
                if not dry_run:
                    close_current_headquarters_pool_entry(
                        session,
                        lead.lead_key,
                        closed_at=_aware(lead.closed_at) if lead.closed_at else now,
                        close_reason=lead.closed_reason or "order_status_terminal",
                        _active_entries_by_lead=active_hq_by_lead,
                    )

            for order_id, order_leads in leads_by_order.items():
                lifecycle_status = _terminal_lifecycle_for_order(order_leads)
                closed_at = _terminal_closed_at(order_leads, now)
                order_rounds = rounds_by_order.get(order_id, [])
                stats["orders"] = int(stats["orders"]) + 1
                stats["rounds_closed"] = int(stats["rounds_closed"]) + len(order_rounds)
                center = centers.get(order_id)
                center_needs_update = center is not None and (
                    center.lead_status
                    != _terminal_center_lead_status(lifecycle_status)
                    or center.current_round_status
                    != _terminal_round_status(lifecycle_status)
                )
                if center_needs_update:
                    stats["centers_closed"] = int(stats["centers_closed"]) + 1

                if dry_run:
                    continue
                for round_row in order_rounds:
                    _close_current_assignment(
                        session,
                        order_id,
                        lifecycle_status,
                        closed_at,
                        current_assignment_round_id=round_row.assignment_round_id,
                    )
                if center_needs_update:
                    _close_current_assignment(
                        session,
                        order_id,
                        lifecycle_status,
                        closed_at,
                    )

            if not dry_run:
                session.flush()
                session.commit()
            session.expunge_all()
    finally:
        _release_session_advisory_lock(session, lock_key)
    return stats


def run_incremental_clue_materialization(
    session_factory: object,
    *,
    scope: str = "clue_materialization",
    batch_size: int = 100,
    raw_batch_size: int = 1000,
    lease_token: str | None = None,
    lease_seconds: int = 300,
    max_batches: int | None = None,
    now: datetime | None = None,
    phone_plain_resolver: Any | None = None,
    page_fence: Any | None = None,
) -> dict[str, object]:
    """Consume frozen JobImpact work with durable raw-page checkpoints.

    ``session_factory`` must create a new SQLAlchemy Session on every call.  The
    claim itself is short-lived, then each raw fanout page is processed in its
    own transaction and Session.  ``ClueMaterializationWorkItem.raw_cursor`` is
    committed together with the business projection, so a process crash or an
    expired short lease resumes at the next page.  The completion CAS remains
    fenced by the attempt token and never silently completes lock contention.
    """

    if not callable(session_factory):
        raise TypeError("session_factory must be callable")
    token = str(lease_token or "").strip()
    if not token:
        raise ValueError("lease_token is required and must be non-empty")
    safe_batch_size = max(1, int(batch_size))
    safe_raw_batch_size = max(1, int(raw_batch_size))
    safe_max_batches = None if max_batches is None else max(0, int(max_batches))
    fixed_now = _aware(now or utcnow())
    summary: dict[str, object] = {
        "scope": scope,
        "work_items": 0,
        "batches": 0,
        "master_leads": 0,
        "closed_leads": 0,
        "headquarters_pool": 0,
        "center_orders": 0,
        "raw_rows": 0,
        "frozen_upper_bound_id": 0,
    }
    checkpoint_started = False
    cycle_id: str | None = None
    phase = "raw"

    while safe_max_batches is None or int(summary["batches"]) < safe_max_batches:
        # Claim in a short transaction.  Business work never shares this
        # Session, which bounds the identity map before the first raw page.
        claim_session = session_factory()
        try:
            claim_session.begin()
            if not checkpoint_started:
                checkpoint = begin_clue_materialization_cycle(claim_session, scope=scope)
                checkpoint_started = True
            else:
                checkpoint = claim_session.get(JobImpactWatermark, scope)
                if checkpoint is None:
                    raise RuntimeError("clue materialization checkpoint disappeared")
            cycle_id = str(checkpoint.cycle_id)
            summary["frozen_upper_bound_id"] = int(checkpoint.frozen_upper_bound_id)
            batch = claim_clue_materialization_batch(
                claim_session,
                scope=scope,
                limit=safe_batch_size,
                lease_token=token,
                lease_seconds=lease_seconds,
                phase=phase,
            )
            work_item_ids = [int(item.work_item_id) for item in batch]
            if not work_item_ids:
                if phase == "raw":
                    unfinished_raw_count = claim_session.scalar(
                        select(func.count(ClueMaterializationWorkItem.work_item_id)).where(
                            ClueMaterializationWorkItem.scope == scope,
                            ClueMaterializationWorkItem.impact_id
                            <= checkpoint.frozen_upper_bound_id,
                            ClueMaterializationWorkItem.state != "completed",
                            ClueMaterializationWorkItem.raw_page_complete.is_(False),
                        )
                    )
                    if int(unfinished_raw_count or 0) > 0:
                        raise RuntimeError(
                            "clue materialization has unfinished raw work within frozen bound"
                        )
                    # Raw/master fanout is complete for every impact.  Only
                    # now can center projection consume the cycle-wide target
                    # set, which prevents a late impact from being missed.
                    phase = "center"
                    claim_session.commit()
                    claim_session.close()
                    continue
                unfinished_count = claim_session.scalar(
                    select(func.count(ClueMaterializationWorkItem.work_item_id)).where(
                        ClueMaterializationWorkItem.scope == scope,
                        ClueMaterializationWorkItem.impact_id
                        <= checkpoint.frozen_upper_bound_id,
                        ClueMaterializationWorkItem.state != "completed",
                    )
                )
                if int(unfinished_count or 0) > 0:
                    raise RuntimeError(
                        "clue materialization has unfinished work within frozen bound"
                    )
            claim_session.commit()
        except BaseException:
            try:
                claim_session.rollback()
            finally:
                claim_session.close()
            raise
        else:
            claim_session.close()

        if not work_item_ids:
            break

        if phase == "raw":
            summary["batches"] = int(summary["batches"]) + 1
        for index, work_item_id in enumerate(work_item_ids):
            try:
                item_result = _process_incremental_clue_work_item(
                    session_factory,
                    work_item_id=work_item_id,
                    scope=scope,
                    raw_batch_size=safe_raw_batch_size,
                    lease_token=token,
                    lease_seconds=lease_seconds,
                    now=fixed_now,
                    phone_plain_resolver=phone_plain_resolver,
                    page_fence=page_fence,
                    cycle_id=str(cycle_id),
                    center_enabled=phase == "center",
                )
            except Exception:
                # Ordinary application failures are recoverable immediately.
                # Release the current and remaining claims from this batch so
                # the retry backoff does not wait for the materialization lease
                # (normally 300s).  BaseException deliberately bypasses this
                # path: a hard crash relies on finished-at/lease expiry
                # reclamation because no cleanup code is guaranteed to run.
                for remaining_id in work_item_ids[index:]:
                    _release_incremental_clue_work_item(
                        session_factory,
                        work_item_id=remaining_id,
                        lease_token=token,
                    )
                raise
            if phase == "center":
                summary["work_items"] = int(summary["work_items"]) + 1
            summary["raw_rows"] = int(summary["raw_rows"]) + int(item_result.get("raw_rows", 0) or 0)
            for key in ("master_leads", "closed_leads", "headquarters_pool", "center_orders"):
                summary[key] = int(summary[key]) + int(item_result.get(key, 0) or 0)
    return summary


def _process_incremental_clue_work_item(
    session_factory: object,
    *,
    work_item_id: int,
    scope: str,
    raw_batch_size: int,
    lease_token: str,
    lease_seconds: int,
    now: datetime,
    phone_plain_resolver: Any | None,
    page_fence: Any | None,
    cycle_id: str,
    center_enabled: bool,
) -> dict[str, int]:
    """Process one leased impact until its durable raw cursor reaches EOF."""

    totals = {
        "master_leads": 0,
        "closed_leads": 0,
        "headquarters_pool": 0,
        "center_orders": 0,
        "raw_rows": 0,
    }
    while True:
        session = session_factory()
        try:
            session.begin()
            item = session.get(ClueMaterializationWorkItem, int(work_item_id))
            if item is None or item.scope != scope:
                raise RuntimeError("clue materialization work item disappeared")
            if item.state != "processing" or item.lease_owner != lease_token:
                raise RuntimeError("clue materialization work item lease lost")
            impact = session.scalar(
                select(JobImpact)
                .where(JobImpact.id == item.impact_id)
            )
            if impact is None:
                raise RuntimeError("clue materialization work item impact disappeared")
            closure = _merge_incremental_closure([impact])
            if item.raw_page_complete:
                if not center_enabled:
                    paused = retry_clue_materialization_batch(
                        session,
                        [item.work_item_id],
                        lease_token=lease_token,
                    )
                    if paused != 1:
                        raise RuntimeError("clue materialization raw-phase lease lost")
                    session.commit()
                    return totals
                center_order_ids = _bounded_center_order_ids(
                    session,
                    raw_clue_row_keys=closure["raw_clue_row_keys"],
                    clue_ids=closure["clue_ids"],
                    order_ids=closure["order_ids"],
                    poi_ids=closure["poi_ids"],
                    after_order_id=item.center_cursor,
                    limit=CENTER_ORDER_BATCH_SIZE,
                    target_scope=scope,
                    target_cycle_id=cycle_id,
                )
                if center_order_ids:
                    center_result = refresh_clue_center_projection(
                        session,
                        now=now,
                        phone_plain_resolver=(
                            phone_plain_resolver
                            if callable(phone_plain_resolver)
                            else None
                        ),
                        order_ids=set(center_order_ids),
                    )
                    _record_cycle_targets(
                        session,
                        scope=scope,
                        cycle_id=cycle_id,
                        target_type="center",
                        target_keys=center_order_ids,
                    )
                    totals["center_orders"] += int(
                        center_result.get("eligible_orders", 0) or 0
                    )
                    item.center_cursor = center_order_ids[-1]
                    _assert_incremental_page_fence(page_fence, session)
                    renewed = renew_clue_materialization_batch(
                        session,
                        [item.work_item_id],
                        lease_token=lease_token,
                        lease_seconds=lease_seconds,
                    )
                    if renewed != 1:
                        raise RuntimeError("clue materialization center lease expired")
                    session.commit()
                    continue

                _assert_incremental_page_fence(page_fence, session)
                completed = complete_clue_materialization_batch(
                    session, [item.work_item_id], lease_token=lease_token
                )
                if completed != 1:
                    raise RuntimeError("clue materialization completion lost its lease")
                session.commit()
                return totals

            raw_page = _bounded_raw_clues(
                session,
                raw_clue_row_keys=closure["raw_clue_row_keys"],
                clue_ids=closure["clue_ids"],
                order_ids=closure["order_ids"],
                poi_ids=closure["poi_ids"],
                limit=raw_batch_size,
                after_row_key=item.raw_cursor,
                target_scope=scope,
                target_cycle_id=cycle_id,
            )
            if not raw_page:
                item.raw_page_complete = True
                _assert_incremental_page_fence(page_fence, session)
                renewed = renew_clue_materialization_batch(
                    session,
                    [item.work_item_id],
                    lease_token=lease_token,
                    lease_seconds=lease_seconds,
                )
                if renewed != 1:
                    raise RuntimeError("clue materialization raw completion lease expired")
                session.commit()
                continue

            page_clue_ids = {
                value for row in raw_page if (value := _clean(row.clue_id)) is not None
            }
            page_order_ids = {
                value for row in raw_page if (value := _clean(row.order_id)) is not None
            }
            page_poi_ids = {
                value
                for row in raw_page
                for candidate in (row.follow_poi_id, row.intention_poi_id)
                if (value := _clean(candidate)) is not None
            }
            # Only old/new values from this impact are context selectors.  The
            # full closure may contain thousands of sibling rows (for example,
            # an order or POI fan-out); passing it to the master lookup would
            # defeat raw-page isolation.
            context_clue_ids = set(page_clue_ids)
            context_order_ids = set(page_order_ids)
            context_poi_ids = set(page_poi_ids)
            for observed_values in (
                impact.old_values_json or {},
                impact.new_values_json or {},
            ):
                for field_name, target in (
                    ("clue_id", context_clue_ids),
                    ("order_id", context_order_ids),
                    ("follow_poi_id", context_poi_ids),
                    ("intention_poi_id", context_poi_ids),
                ):
                    value = _clean(observed_values.get(field_name))
                    if value is not None:
                        target.add(value)
            page_result = materialize_clue_master_leads(
                session,
                now=now,
                raw_clue_row_keys={row.clue_row_key for row in raw_page},
                raw_page_clues=raw_page,
                clue_ids=closure["clue_ids"],
                order_ids=closure["order_ids"],
                poi_ids=closure["poi_ids"],
                source_identity_keys=closure["source_identity_keys"],
                existing_clue_ids=context_clue_ids,
                existing_order_ids=context_order_ids,
                existing_poi_ids=context_poi_ids,
                existing_source_identity_keys=closure[
                    "source_identity_keys"
                ],
            )
            if page_result.get("skipped") == "locked":
                raise RuntimeError("clue master materialization lock unavailable")
            _record_cycle_targets(
                session,
                scope=scope,
                cycle_id=cycle_id,
                target_type="raw",
                target_keys=[row.clue_row_key for row in raw_page],
            )
            for key in ("master_leads", "closed_leads", "headquarters_pool"):
                totals[key] += int(page_result.get(key, 0) or 0)
            item.raw_cursor = raw_page[-1].clue_row_key
            item.raw_page_complete = len(raw_page) < raw_batch_size
            _assert_incremental_page_fence(page_fence, session)
            renewed = renew_clue_materialization_batch(
                session,
                [item.work_item_id],
                lease_token=lease_token,
                lease_seconds=lease_seconds,
            )
            if renewed != 1:
                raise RuntimeError("clue materialization page lease expired")
            totals["raw_rows"] += len(raw_page)
            session.commit()
        except Exception:
            session.rollback()
            _release_incremental_clue_work_item(
                session_factory,
                work_item_id=work_item_id,
                lease_token=lease_token,
            )
            raise
        except BaseException:
            # A hard process interruption (SIGTERM/KeyboardInterrupt/OOM kill)
            # cannot run the release path.  Leave the durable cursor and lease
            # for expiry so a later attempt can reclaim it safely.
            session.rollback()
            raise
        finally:
            session.close()


def _release_incremental_clue_work_item(
    session_factory: object,
    *,
    work_item_id: int,
    lease_token: str,
) -> None:
    """Return an application-failed item to pending without masking the error."""

    cleanup_session = session_factory()
    try:
        cleanup_session.begin()
        retry_clue_materialization_batch(
            cleanup_session,
            [int(work_item_id)],
            lease_token=lease_token,
        )
        cleanup_session.commit()
    except BaseException:
        try:
            cleanup_session.rollback()
        finally:
            cleanup_session.close()
    else:
        cleanup_session.close()


def _assert_incremental_page_fence(
    page_fence: Any | None,
    session: Session,
) -> None:
    if page_fence is not None:
        if not callable(page_fence) or not bool(page_fence(session)):
            raise RuntimeError("daily execution lease is no longer valid")


def _merge_incremental_closure(impacts: list[JobImpact]) -> dict[str, set[str]]:
    closure: dict[str, set[str]] = {
        "raw_clue_row_keys": set(),
        "clue_ids": set(),
        "order_ids": set(),
        "poi_ids": set(),
        "source_identity_keys": set(),
    }
    for impact in impacts:
        entity_type = str(impact.entity_type or "")
        if entity_type == "clue" and impact.entity_key:
            closure["raw_clue_row_keys"].add(str(impact.entity_key))
            closure["clue_ids"].add(str(impact.entity_key))
        payload = impact.affected_closure_json or {}
        for value in payload.get("clue_ids", []) or []:
            if value not in (None, ""):
                value = str(value)
                closure["clue_ids"].add(value)
                # A clue closure contains both row keys and canonical IDs; the
                # raw-row selector is harmlessly deduplicated if the value is
                # not a row key.
                closure["raw_clue_row_keys"].add(value)
        for field_name in ("order_ids", "poi_ids", "source_identity_keys"):
            for value in payload.get(field_name, []) or []:
                if value not in (None, ""):
                    closure[field_name].add(str(value))
        if entity_type == "order" and impact.entity_key:
            closure["order_ids"].add(str(impact.entity_key))
        elif entity_type == "store_poi_mapping" and impact.entity_key:
            closure["poi_ids"].add(str(impact.entity_key))
    return closure



def import_store_locations(
    session: Session,
    workbook_path: Path,
    *,
    enable_participation: bool = False,
    now: datetime | None = None,
) -> dict[str, int]:
    """Load the business location workbook through POI mappings, never by assuming POI equals store id."""
    now = _aware(now or utcnow())
    workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        raise ValueError("store location workbook has no header row")
    columns = {_clean_header(value): index for index, value in enumerate(header) if _clean_header(value)}
    required = {"门店ID", "经度", "纬度", "门店所在城市"}
    missing = sorted(required.difference(columns))
    if missing:
        raise ValueError(f"store location workbook missing required columns: {', '.join(missing)}")

    mappings_by_poi = {row.poi_id: row for row in session.scalars(select(DimStorePoiMapping)).all()}
    rows_seen = 0
    updated = 0
    unmapped = 0
    invalid = 0
    imported_stores_by_poi: dict[str, DimStore] = {}
    for row in rows:
        if not row:
            continue
        poi_id = _text_cell(_cell(row, columns, "门店ID"))
        if not poi_id:
            continue
        rows_seen += 1
        mapping = mappings_by_poi.get(poi_id)
        if mapping is None:
            unmapped += 1
            _record_store_location_issue(session, poi_id, "store_location_unmapped_poi", now)
            continue
        store = session.get(DimStore, mapping.store_id)
        if store is None:
            unmapped += 1
            _record_store_location_issue(session, poi_id, "store_location_missing_store", now)
            continue

        longitude = _decimal(_cell(row, columns, "经度"))
        latitude = _decimal(_cell(row, columns, "纬度"))
        province = _text_cell(_cell(row, columns, "门店所在省份")) or _text_cell(_cell(row, columns, "省份"))
        city = _text_cell(_cell(row, columns, "门店所在城市"))
        city_code = normalize_city_code(city)
        status_note = _text_cell(_cell(row, columns, "状态备注"))
        has_coordinates_and_city = _valid_coordinates(latitude, longitude) and bool(city_code)
        if not has_coordinates_and_city:
            invalid += 1
            _record_store_location_issue(session, poi_id, "store_location_invalid_coordinates_or_city", now)

        store.standard_province = province or _clean(store.standard_province)
        store.standard_city = city or _clean(store.standard_city)
        store.city_code = city_code
        store.longitude = longitude if _valid_coordinates(latitude, longitude) else None
        store.latitude = latitude if _valid_coordinates(latitude, longitude) else None
        store.location_source = workbook_path.name
        store.location_status_note = status_note
        store.location_updated_at = now
        imported_stores_by_poi[poi_id] = store
        updated += 1

    stores_by_id = {store.store_id: store for store in imported_stores_by_poi.values()}
    _enrich_store_locations_from_raw_evidence(
        session,
        session.scalars(select(RawDouyinClue)).all(),
        mappings_by_poi,
        stores_by_id,
        now,
    )
    for poi_id, store in imported_stores_by_poi.items():
        store.location_status = _store_location_status(store)
        if store.location_status == "valid":
            store.is_douyin_clue_applicable = True
            store.participates_in_clue_allocation = bool(enable_participation or store.participates_in_clue_allocation)
        else:
            store.is_douyin_clue_applicable = False
            store.participates_in_clue_allocation = False
            if store.location_status == "partial":
                _record_store_location_issue(session, poi_id, "store_location_missing_province", now)

    session.flush()
    return {"rows": rows_seen, "updated": updated, "unmapped": unmapped, "invalid": invalid}


def _enrich_store_locations_from_raw_evidence(
    session: Session,
    raw_clues: list[RawDouyinClue],
    mappings_by_poi: dict[str, DimStorePoiMapping],
    stores_by_id: dict[str, DimStore],
    now: datetime,
) -> None:
    """Fill missing store geography from unambiguous raw clue evidence."""
    evidence_by_poi: dict[str, set[tuple[str, str]]] = defaultdict(set)
    provinces_by_city_code: dict[str, set[str]] = defaultdict(set)
    for raw_clue in raw_clues:
        poi_id = _clean(raw_clue.follow_poi_id)
        province = _clean(raw_clue.auto_province_name)
        city = _clean(raw_clue.auto_city_name)
        if poi_id and province and city:
            evidence_by_poi[poi_id].add((province, city))
        city_code = normalize_city_code(city)
        if province and city_code:
            provinces_by_city_code[city_code].add(province)

    for poi_id, candidates in evidence_by_poi.items():
        mapping = mappings_by_poi.get(poi_id)
        store = stores_by_id.get(mapping.store_id) if mapping else None
        if store is None:
            continue
        if len(candidates) != 1:
            _record_store_location_issue(session, poi_id, "store_location_conflicting_raw_evidence", now)
            continue
        province, city = next(iter(candidates))
        city_code = normalize_city_code(city)
        existing_city_code = normalize_city_code(store.standard_city) or _clean(store.city_code)
        if existing_city_code and city_code and existing_city_code != city_code:
            _record_store_location_issue(session, poi_id, "store_location_raw_city_mismatch", now)
            continue
        changed = False
        if not _clean(store.standard_city):
            store.standard_city = city
            changed = True
        if not _clean(store.city_code) and city_code:
            store.city_code = city_code
            changed = True
        if not _clean(store.standard_province):
            store.standard_province = province
            changed = True
        location_status = _store_location_status(store)
        if store.location_status != location_status:
            store.location_status = location_status
            changed = True
        if location_status == "valid" and not store.is_douyin_clue_applicable:
            store.is_douyin_clue_applicable = True
            changed = True
        if location_status != "valid" and store.is_douyin_clue_applicable:
            store.is_douyin_clue_applicable = False
            store.participates_in_clue_allocation = False
            changed = True
        if changed:
            store.location_updated_at = now

    province_by_city_code = {
        city_code: next(iter(provinces))
        for city_code, provinces in provinces_by_city_code.items()
        if len(provinces) == 1
    }
    for store in stores_by_id.values():
        if _clean(store.standard_province):
            continue
        city_code = normalize_city_code(store.standard_city) or normalize_city_code(store.city_code)
        province = province_by_city_code.get(city_code or "")
        if not province:
            continue
        store.standard_province = province
        location_status = _store_location_status(store)
        store.location_status = location_status
        store.is_douyin_clue_applicable = location_status == "valid"
        if location_status != "valid":
            store.participates_in_clue_allocation = False
        store.location_updated_at = now


def normalize_city_code(value: str | None) -> str | None:
    """Return the M1 canonical city key; it is not a government administrative code."""
    city = _clean(value)
    if not city:
        return None
    city = "".join(city.split())
    return city[:-1] if city.endswith("市") else city


def eligible_candidate_stores(session: Session, *, city_code: str | None = None) -> list[DimStore]:
    rows = session.scalars(select(DimStore).order_by(DimStore.store_id)).all()
    normalized_city = normalize_city_code(city_code)
    return [
        row
        for row in rows
        if _is_candidate_eligible(row)
        and (normalized_city is None or row.city_code == normalized_city)
    ]


def haversine_km(latitude_a: float, longitude_a: float, latitude_b: float, longitude_b: float) -> float:
    """Return the great-circle distance between two coordinates in kilometres."""
    earth_radius_km = 6371.0088
    latitude_delta = radians(float(latitude_b) - float(latitude_a))
    longitude_delta = radians(float(longitude_b) - float(longitude_a))
    latitude_a_radians = radians(float(latitude_a))
    latitude_b_radians = radians(float(latitude_b))
    haversine = sin(latitude_delta / 2) ** 2 + cos(latitude_a_radians) * cos(latitude_b_radians) * sin(
        longitude_delta / 2
    ) ** 2
    return earth_radius_km * 2 * asin(sqrt(haversine))


def _score_sparse_json(value: object) -> bytes:
    def normalize(item: object) -> object:
        if isinstance(item, datetime):
            aware = _aware(item)
            assert aware is not None
            return aware.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
        if isinstance(item, date):
            return item.isoformat()
        if isinstance(item, Decimal):
            rendered = format(item, "f")
            return "0" if not rendered or item == 0 else rendered
        if isinstance(item, Mapping):
            return {str(key): normalize(child) for key, child in item.items()}
        if isinstance(item, (list, tuple)):
            return [normalize(child) for child in item]
        return item

    try:
        return json.dumps(
            normalize(value),
            ensure_ascii=False,
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        ).encode("utf-8")
    except (TypeError, ValueError, OverflowError) as exc:
        raise ValueError("score sparse metadata is not canonical JSON") from exc


def _score_sparse_identity(value: object, *, label: str) -> str:
    if not isinstance(value, str) or not value or value.strip() != value:
        raise ValueError(f"{label} must be a non-empty canonical string")
    return value


def _score_sparse_date(value: date | str) -> date:
    if isinstance(value, datetime):
        raise ValueError("snapshot_date must be a date")
    if isinstance(value, date):
        return value
    if not isinstance(value, str) or value.strip() != value:
        raise ValueError("snapshot_date must use canonical ISO format")
    try:
        parsed = date.fromisoformat(value)
    except ValueError as exc:
        raise ValueError("snapshot_date must use canonical ISO format") from exc
    if parsed.isoformat() != value:
        raise ValueError("snapshot_date must use canonical ISO format")
    return parsed


def _score_sparse_rule_versions(
    session: Session, published_rule_ids: Iterable[str]
) -> tuple[ClueAllocationRuleVersion, ...]:
    requested = {
        _score_sparse_identity(value, label="published rule id")
        for value in published_rule_ids
    }
    if not requested:
        raise ValueError("published_rule_ids must not be empty")
    if len(requested) > MAX_SCORE_SPARSE_RULES:
        raise ValueError("published rule closure exceeds sparse limit")
    if "legacy-unversioned" in requested:
        raise ValueError("legacy-unversioned is reserved for legacy score fallback")
    rows = list(
        session.scalars(
            select(ClueAllocationRuleVersion)
            .where(
                ClueAllocationRuleVersion.status == "published",
                or_(
                    ClueAllocationRuleVersion.rule_version_id.in_(requested),
                    ClueAllocationRuleVersion.rule_id.in_(requested),
                ),
            )
            .order_by(ClueAllocationRuleVersion.rule_version_id)
            .limit(MAX_SCORE_SPARSE_RULES + 1)
        )
    )
    by_version = {row.rule_version_id: row for row in rows}
    by_rule: dict[str, list[ClueAllocationRuleVersion]] = defaultdict(list)
    for row in rows:
        by_rule[row.rule_id].append(row)
    resolved: dict[str, ClueAllocationRuleVersion] = {}
    for identity in sorted(requested):
        if identity in by_version:
            row = by_version[identity]
        else:
            candidates = by_rule.get(identity, [])
            if len(candidates) != 1:
                raise ValueError(f"published rule closure is incomplete: {identity}")
            row = candidates[0]
        resolved[row.rule_version_id] = row
    if len(resolved) > MAX_SCORE_SPARSE_RULES:
        raise ValueError("published rule closure exceeds sparse limit")
    return tuple(resolved[key] for key in sorted(resolved))


def _score_sparse_stores(
    session: Session, affected_store_ids: Iterable[str]
) -> tuple[tuple[str, ...], list[DimStore]]:
    requested = tuple(
        sorted(
            {
                _score_sparse_identity(value, label="affected store id")
                for value in affected_store_ids
            }
        )
    )
    if not requested:
        raise ValueError("affected_store_ids must not be empty")
    if len(requested) > MAX_SCORE_SPARSE_STORES:
        raise ValueError("affected store closure exceeds sparse limit")
    stores: list[DimStore] = []
    for index in range(0, len(requested), 400):
        stores.extend(
            session.scalars(
                select(DimStore)
                .where(DimStore.store_id.in_(requested[index : index + 400]))
                .order_by(DimStore.store_id)
            )
        )
    eligible = sorted(
        (store for store in stores if _is_candidate_eligible(store)),
        key=lambda store: store.store_id,
    )
    return requested, eligible


def _score_sparse_contract(
    *,
    generation_id: str,
    base_generation_id: str,
    affected_store_ids: tuple[str, ...],
    rule_version_ids: tuple[str, ...],
    snapshot_date: date,
    batch_size: int,
    closure_policy_hash: str,
) -> dict[str, object]:
    return {
        "protocol": SCORE_SPARSE_PROTOCOL,
        "projection": "settlement",
        "operation": "build_score_sparse_overlay",
        "generation_id": generation_id,
        "base_generation_id": base_generation_id,
        "affected_store_ids": list(affected_store_ids),
        "rule_version_ids": list(rule_version_ids),
        "snapshot_date": snapshot_date.isoformat(),
        "batch_size": batch_size,
        "closure_policy_hash": closure_policy_hash,
    }


def _score_sparse_preflight(
    session: Session,
    *,
    generation_id: str,
    base_generation_id: str,
    affected_store_ids: Iterable[str],
    published_rule_ids: Iterable[str],
    snapshot_date: date | str,
    batch_size: int,
    closure_policy_hash: str,
) -> tuple[
    SettlementProjectionGeneration,
    tuple[str, ...],
    list[DimStore],
    tuple[ClueAllocationRuleVersion, ...],
    date,
    dict[str, object],
    bool,
]:
    generation_id = _score_sparse_identity(generation_id, label="generation_id")
    base_generation_id = _score_sparse_identity(
        base_generation_id, label="base_generation_id"
    )
    if generation_id == base_generation_id:
        raise ValueError("score sparse generation cannot reference itself")
    if isinstance(batch_size, bool) or not isinstance(batch_size, int) or not 1 <= batch_size <= 400:
        raise ValueError("batch_size must be an integer between 1 and 400")
    if (
        not isinstance(closure_policy_hash, str)
        or re.fullmatch(r"[0-9a-f]{64}", closure_policy_hash) is None
    ):
        raise ValueError("closure_policy_hash must be 64 lowercase hexadecimal characters")
    target_date = _score_sparse_date(snapshot_date)
    active = session.get(SettlementProjectionActive, "settlement")
    if active is None or active.generation_id != base_generation_id:
        raise ValueError("active settlement pointer does not match score sparse base")
    base = session.get(SettlementProjectionGeneration, base_generation_id)
    if base is None or base.state != "published" or base.projection_name != "settlement":
        raise ValueError("score sparse base generation is not published")
    generation = session.get(SettlementProjectionGeneration, generation_id)
    if (
        generation is None
        or generation.state != "staging"
        or generation.projection_name != "settlement"
        or generation.generation_kind != "lineage"
        or generation.base_generation_id != base_generation_id
    ):
        raise ValueError("score sparse generation is not writable")
    if int(generation.lineage_depth) != int(base.lineage_depth) + 1:
        raise ValueError("score sparse generation lineage is inconsistent")
    requested_stores, eligible_stores = _score_sparse_stores(
        session, affected_store_ids
    )
    versions = _score_sparse_rule_versions(session, published_rule_ids)
    rule_version_ids = tuple(row.rule_version_id for row in versions)
    contract = _score_sparse_contract(
        generation_id=generation_id,
        base_generation_id=base_generation_id,
        affected_store_ids=requested_stores,
        rule_version_ids=rule_version_ids,
        snapshot_date=target_date,
        batch_size=batch_size,
        closure_policy_hash=closure_policy_hash,
    )
    source_input = generation.source_input_json
    if source_input is not None and not isinstance(source_input, Mapping):
        raise ValueError("score sparse generation source input is malformed")
    existing_contract = (source_input or {}).get("score")
    if existing_contract is not None and existing_contract != contract:
        raise ValueError("score sparse generation input conflicts with prior attempt")
    resumed = bool(
        session.scalar(
            select(func.count())
            .select_from(SettlementProjectionPartitionManifest)
            .where(
                SettlementProjectionPartitionManifest.generation_id == generation_id,
                SettlementProjectionPartitionManifest.artifact == "score",
            )
        )
    )
    return (
        generation,
        requested_stores,
        eligible_stores,
        versions,
        target_date,
        contract,
        resumed,
    )


def _score_sparse_run_payload(
    *,
    generation: SettlementProjectionGeneration,
    base_generation_id: str,
    rule_version_id: str,
    snapshot_date: date,
    window_start: datetime,
    window_end: datetime,
    affected_store_ids: tuple[str, ...],
    closure_policy_hash: str,
) -> dict[str, object]:
    return {
        "protocol": SCORE_SPARSE_PROTOCOL,
        "mode": "projection_sparse",
        "generation_id": generation.generation_id,
        "parent_generation_id": base_generation_id,
        "input_fingerprint": generation.input_fingerprint,
        "snapshot_date": snapshot_date.isoformat(),
        "rule_version_id": rule_version_id,
        "window_start": window_start,
        "window_end": window_end,
        "affected_store_ids": list(affected_store_ids),
        "closure_policy_hash": closure_policy_hash,
    }


def _score_sparse_materialize_run(
    session: Session,
    *,
    generation: SettlementProjectionGeneration,
    base_generation_id: str,
    version: ClueAllocationRuleVersion,
    snapshot_date: date,
    affected_store_ids: tuple[str, ...],
    eligible_stores: list[DimStore],
    closure_policy_hash: str,
    batch_size: int,
) -> tuple[StoreScoreSnapshotRun, dict[str, StoreScoreSnapshot]]:
    score_config = _resolve_store_score_config(
        session, rule_version_id=version.rule_version_id
    )
    local_end = datetime.combine(snapshot_date, SCHEDULED_SCORE_REFRESH_TIME, SHANGHAI)
    window_end = local_end.astimezone(timezone.utc)
    window_start = window_end - timedelta(days=score_config.lookback_days)
    payload = _score_sparse_run_payload(
        generation=generation,
        base_generation_id=base_generation_id,
        rule_version_id=version.rule_version_id,
        snapshot_date=snapshot_date,
        window_start=window_start,
        window_end=window_end,
        affected_store_ids=affected_store_ids,
        closure_policy_hash=closure_policy_hash,
    )
    digest = sha256(_score_sparse_json(payload)).hexdigest()
    scheduled_key = f"projection-score:{digest}"
    snapshot_run_id = f"score-projection-{digest}"
    config_json = {
        "projection_generation_id": generation.generation_id,
        "projection_base_generation_id": base_generation_id,
        "closure_policy_hash": closure_policy_hash,
        "rule_version_id": version.rule_version_id,
        "lookback_days": score_config.lookback_days,
        "min_samples": score_config.min_samples,
        "execution_mode": "formal",
        "conversion_weight": str(score_config.conversion_weight),
        "follow_24h_weight": str(score_config.follow_weight),
        "store_weight": str(score_config.store_weight),
        "scheduled_payload_sha256": digest,
    }
    values = {
        "snapshot_run_id": snapshot_run_id,
        "snapshot_date": snapshot_date,
        "run_mode": "projection_sparse",
        "scheduled_key": scheduled_key,
        "window_start": window_start,
        "window_end": window_end,
        "candidate_store_count": len(eligible_stores),
        "snapshot_count": len(eligible_stores),
        "triggered_by": "settlement-finalize",
        "config_json": config_json,
        "computed_at": window_end,
    }
    dialect_name = getattr(getattr(session.bind, "dialect", None), "name", "")
    if dialect_name == "sqlite":
        from sqlalchemy.dialects.sqlite import insert as dialect_insert

        statement = dialect_insert(StoreScoreSnapshotRun).values(**values)
        statement = statement.on_conflict_do_nothing(index_elements=["scheduled_key"])
        session.execute(statement)
    elif dialect_name == "postgresql":
        from sqlalchemy.dialects.postgresql import insert as dialect_insert

        statement = dialect_insert(StoreScoreSnapshotRun).values(**values)
        statement = statement.on_conflict_do_nothing(index_elements=["scheduled_key"])
        session.execute(statement)
    else:
        existing = session.scalar(
            select(StoreScoreSnapshotRun).where(
                StoreScoreSnapshotRun.scheduled_key == scheduled_key
            )
        )
        if existing is None:
            session.add(StoreScoreSnapshotRun(**values))
    session.flush()
    run = session.scalar(
        select(StoreScoreSnapshotRun)
        .where(StoreScoreSnapshotRun.scheduled_key == scheduled_key)
        .with_for_update()
    )
    if run is None:
        raise ValueError("score sparse run could not be materialized")
    if (
        run.snapshot_run_id != snapshot_run_id
        or run.snapshot_date != snapshot_date
        or run.run_mode != "projection_sparse"
        or run.candidate_store_count != len(eligible_stores)
        or run.snapshot_count != len(eligible_stores)
        or _aware(run.window_start) != window_start
        or _aware(run.window_end) != window_end
        or run.config_json != config_json
    ):
        raise ValueError("existing score sparse run conflicts with canonical input")

    expected_store_ids = {store.store_id for store in eligible_stores}
    existing_snapshots = {
        row.store_id: row
        for row in session.scalars(
            select(StoreScoreSnapshot)
            .where(StoreScoreSnapshot.snapshot_run_id == snapshot_run_id)
            .order_by(StoreScoreSnapshot.store_id)
        )
    }
    if set(existing_snapshots) == expected_store_ids:
        for row in existing_snapshots.values():
            if (
                row.snapshot_date != snapshot_date
                or row.run_mode != "projection_sparse"
                or _aware(row.window_start) != window_start
                or _aware(row.window_end) != window_end
                or row.config_json != config_json
            ):
                raise ValueError("existing score snapshot conflicts with canonical run")
        return run, existing_snapshots
    foreign_sidecars = session.scalar(
        select(func.count())
        .select_from(StoreScoreSnapshotGeneration)
        .where(
            StoreScoreSnapshotGeneration.snapshot_run_id == snapshot_run_id,
            StoreScoreSnapshotGeneration.generation_id != generation.generation_id,
        )
    )
    if foreign_sidecars:
        raise ValueError("partial score run is referenced by another generation")
    session.execute(
        delete(StoreScoreSnapshotGeneration).where(
            StoreScoreSnapshotGeneration.snapshot_run_id == snapshot_run_id,
            StoreScoreSnapshotGeneration.generation_id == generation.generation_id,
        )
    )
    session.execute(
        delete(StoreScoreSnapshot).where(
            StoreScoreSnapshot.snapshot_run_id == snapshot_run_id
        )
    )

    metrics_by_store = _formal_store_metrics(
        session, eligible_stores, window_start, window_end
    )
    city_metrics = _aggregate_city_metrics(eligible_stores, metrics_by_store)
    global_metrics = _sum_metrics(metrics_by_store.values())
    pending: list[dict[str, object]] = []
    for store in eligible_stores:
        own_metrics = metrics_by_store.get(store.store_id, StoreMetrics())
        city_metric = city_metrics.get(store.city_code or "", StoreMetrics())
        conversion_rate, conversion_source = _resolved_rate(
            own_metrics.conversion_numerator,
            own_metrics.conversion_denominator,
            city_metric.conversion_numerator,
            city_metric.conversion_denominator,
            global_metrics.conversion_numerator,
            global_metrics.conversion_denominator,
            score_config.min_samples,
        )
        follow_rate, follow_source = _resolved_rate(
            own_metrics.follow_24h_numerator,
            own_metrics.follow_24h_denominator,
            city_metric.follow_24h_numerator,
            city_metric.follow_24h_denominator,
            global_metrics.follow_24h_numerator,
            global_metrics.follow_24h_denominator,
            score_config.min_samples,
        )
        score = (
            conversion_rate * score_config.conversion_weight
            + follow_rate * score_config.follow_weight
        ) * score_config.store_weight
        pending.append(
            {
                "snapshot_id": f"{snapshot_run_id}-{store.store_id}",
                "snapshot_run_id": snapshot_run_id,
                "snapshot_date": snapshot_date,
                "run_mode": "projection_sparse",
                "store_id": store.store_id,
                "city_code": store.city_code,
                "window_start": window_start,
                "window_end": window_end,
                "conversion_numerator": own_metrics.conversion_numerator,
                "conversion_denominator": own_metrics.conversion_denominator,
                "conversion_rate": conversion_rate,
                "conversion_value_source": conversion_source,
                "follow_24h_numerator": own_metrics.follow_24h_numerator,
                "follow_24h_denominator": own_metrics.follow_24h_denominator,
                "follow_24h_rate": follow_rate,
                "follow_24h_value_source": follow_source,
                "conversion_weight": score_config.conversion_weight,
                "follow_24h_weight": score_config.follow_weight,
                "store_weight": score_config.store_weight,
                "composite_score": score,
                "config_json": config_json,
                "computed_at": window_end,
            }
        )
        if len(pending) == batch_size:
            session.execute(insert(StoreScoreSnapshot), pending)
            pending.clear()
    if pending:
        session.execute(insert(StoreScoreSnapshot), pending)
    session.flush()
    snapshots = {
        row.store_id: row
        for row in session.scalars(
            select(StoreScoreSnapshot)
            .where(StoreScoreSnapshot.snapshot_run_id == snapshot_run_id)
            .order_by(StoreScoreSnapshot.store_id)
        )
    }
    if set(snapshots) != expected_store_ids:
        raise ValueError("score sparse run is incomplete")
    return run, snapshots


def _score_sparse_snapshot_envelope(row: StoreScoreSnapshot) -> dict[str, object]:
    return {
        "snapshot_id": row.snapshot_id,
        "snapshot_run_id": row.snapshot_run_id,
        "snapshot_date": row.snapshot_date,
        "run_mode": row.run_mode,
        "store_id": row.store_id,
        "city_code": row.city_code,
        "window_start": row.window_start,
        "window_end": row.window_end,
        "conversion_numerator": row.conversion_numerator,
        "conversion_denominator": row.conversion_denominator,
        "conversion_rate": row.conversion_rate,
        "conversion_value_source": row.conversion_value_source,
        "follow_24h_numerator": row.follow_24h_numerator,
        "follow_24h_denominator": row.follow_24h_denominator,
        "follow_24h_rate": row.follow_24h_rate,
        "follow_24h_value_source": row.follow_24h_value_source,
        "conversion_weight": row.conversion_weight,
        "follow_24h_weight": row.follow_24h_weight,
        "store_weight": row.store_weight,
        "composite_score": row.composite_score,
        "config_json": row.config_json,
        "computed_at": row.computed_at,
    }


def _score_sparse_empty_digest() -> str:
    return sha256(_score_sparse_json({"rows": []})).hexdigest()


def _score_sparse_row_digest(row: StoreScoreSnapshot) -> str:
    return sha256(
        bytes.fromhex(_score_sparse_empty_digest())
        + _score_sparse_json(_score_sparse_snapshot_envelope(row))
    ).hexdigest()


def _score_sparse_last_key(
    *,
    snapshot_date: date,
    rule_version_id: str,
    store_id: str,
    snapshot_run_id: str,
    snapshot_id: str,
) -> str:
    return _score_sparse_json(
        {
            "artifact": "score",
            "cursor": {
                "snapshot_date": snapshot_date.isoformat(),
                "rule_version_id": rule_version_id,
                "store_id": store_id,
                "snapshot_run_id": snapshot_run_id,
                "snapshot_id": snapshot_id,
            },
        }
    ).decode("utf-8")


def _score_sparse_write_rule(
    session_factory: Callable[[], Session],
    *,
    generation_id: str,
    base_generation_id: str,
    version_id: str,
    snapshot_date: date,
    affected_store_ids: tuple[str, ...],
    closure_policy_hash: str,
    batch_size: int,
    contract: Mapping[str, object],
) -> tuple[str, tuple[str, ...]]:
    with session_factory() as session:
        generation = session.scalar(
            select(SettlementProjectionGeneration)
            .where(SettlementProjectionGeneration.generation_id == generation_id)
            .with_for_update()
        )
        if (
            generation is None
            or generation.state != "staging"
            or generation.base_generation_id != base_generation_id
        ):
            raise ValueError("score sparse generation is not writable")
        active = session.get(SettlementProjectionActive, "settlement")
        if active is None or active.generation_id != base_generation_id:
            raise ValueError("active settlement pointer changed during score build")
        source_input = dict(generation.source_input_json or {})
        existing_contract = source_input.get("score")
        if existing_contract is not None and existing_contract != dict(contract):
            raise ValueError("score sparse generation input conflicts with prior attempt")
        source_input["score"] = dict(contract)
        generation.source_input_json = source_input

        version = session.get(ClueAllocationRuleVersion, version_id)
        if version is None or version.status != "published":
            raise ValueError("published score rule version disappeared")
        stores: list[DimStore] = []
        for index in range(0, len(affected_store_ids), 400):
            stores.extend(
                session.scalars(
                    select(DimStore)
                    .where(
                        DimStore.store_id.in_(
                            affected_store_ids[index : index + 400]
                        )
                    )
                    .order_by(DimStore.store_id)
                )
            )
        eligible_stores = sorted(
            (store for store in stores if _is_candidate_eligible(store)),
            key=lambda store: store.store_id,
        )
        run, snapshots = _score_sparse_materialize_run(
            session,
            generation=generation,
            base_generation_id=base_generation_id,
            version=version,
            snapshot_date=snapshot_date,
            affected_store_ids=affected_store_ids,
            eligible_stores=eligible_stores,
            closure_policy_hash=closure_policy_hash,
            batch_size=batch_size,
        )
        partition_keys = tuple(
            canonical_score_partition_key(snapshot_date, version_id, store_id)
            for store_id in affected_store_ids
        )
        session.execute(
            delete(StoreScoreSnapshotGeneration).where(
                StoreScoreSnapshotGeneration.generation_id == generation_id,
                StoreScoreSnapshotGeneration.rule_version_id == version_id,
                StoreScoreSnapshotGeneration.snapshot_date == snapshot_date,
                StoreScoreSnapshotGeneration.store_id.in_(affected_store_ids),
            )
        )
        session.execute(
            delete(SettlementProjectionPartitionManifest).where(
                SettlementProjectionPartitionManifest.generation_id == generation_id,
                SettlementProjectionPartitionManifest.artifact == "score",
                SettlementProjectionPartitionManifest.partition_key.in_(partition_keys),
            )
        )
        sidecars: list[dict[str, object]] = []
        manifests: list[dict[str, object]] = []
        for store_id, partition_key in zip(affected_store_ids, partition_keys):
            snapshot = snapshots.get(store_id)
            if snapshot is None:
                manifests.append(
                    {
                        "generation_id": generation_id,
                        "artifact": "score",
                        "partition_key": partition_key,
                        "owner_state": "tombstone",
                        "source_kind": "tombstone",
                        "data_generation_id": None,
                        "reference_head_generation_id": None,
                        "base_generation_id": base_generation_id,
                        "row_count": 0,
                        "amount_total_cent": 0,
                        "status_counts_json": {},
                        "checksum": _score_sparse_empty_digest(),
                        "last_key": None,
                    }
                )
                continue
            partition_checksum = _score_sparse_row_digest(snapshot)
            sidecars.append(
                {
                    "generation_id": generation_id,
                    "snapshot_run_id": run.snapshot_run_id,
                    "store_id": store_id,
                    "rule_version_id": version_id,
                    "snapshot_date": snapshot_date,
                    "partition_key": partition_key,
                    "owner_state": "owned",
                    "checksum": partition_checksum,
                }
            )
            manifests.append(
                {
                    "generation_id": generation_id,
                    "artifact": "score",
                    "partition_key": partition_key,
                    "owner_state": "owned",
                    "source_kind": "overlay",
                    "data_generation_id": generation_id,
                    "reference_head_generation_id": None,
                    "base_generation_id": base_generation_id,
                    "row_count": 1,
                    "amount_total_cent": 0,
                    "status_counts_json": {},
                    "checksum": partition_checksum,
                    "last_key": _score_sparse_last_key(
                        snapshot_date=snapshot_date,
                        rule_version_id=version_id,
                        store_id=store_id,
                        snapshot_run_id=run.snapshot_run_id,
                        snapshot_id=snapshot.snapshot_id,
                    ),
                }
            )
        for index in range(0, len(sidecars), batch_size):
            session.execute(
                insert(StoreScoreSnapshotGeneration),
                sidecars[index : index + batch_size],
            )
        for index in range(0, len(manifests), batch_size):
            session.execute(
                insert(SettlementProjectionPartitionManifest),
                manifests[index : index + batch_size],
            )
        session.commit()
        return run.snapshot_run_id, partition_keys


def _score_sparse_finalize(
    session_factory: Callable[[], Session],
    *,
    generation_id: str,
    base_generation_id: str,
    target_date: date,
    rule_version_ids: tuple[str, ...],
    snapshot_run_ids: tuple[str, ...],
    contract: Mapping[str, object],
    resumed: bool,
) -> ScoreManifest:
    from apps.worker.legacy_projection_bootstrap import _manifest_checksum

    with session_factory() as session:
        generation = session.scalar(
            select(SettlementProjectionGeneration)
            .where(SettlementProjectionGeneration.generation_id == generation_id)
            .with_for_update()
        )
        if (
            generation is None
            or generation.state != "staging"
            or generation.base_generation_id != base_generation_id
        ):
            raise ValueError("score sparse generation is not writable")
        active = session.get(SettlementProjectionActive, "settlement")
        if active is None or active.generation_id != base_generation_id:
            raise ValueError("active settlement pointer changed before score finalize")
        manifest_rows = [
            dict(row)
            for row in session.execute(
                select(
                    SettlementProjectionPartitionManifest.artifact,
                    SettlementProjectionPartitionManifest.partition_key,
                    SettlementProjectionPartitionManifest.owner_state,
                    SettlementProjectionPartitionManifest.source_kind,
                    SettlementProjectionPartitionManifest.data_generation_id,
                    SettlementProjectionPartitionManifest.base_generation_id,
                    SettlementProjectionPartitionManifest.row_count,
                    SettlementProjectionPartitionManifest.amount_total_cent,
                    SettlementProjectionPartitionManifest.status_counts_json,
                    SettlementProjectionPartitionManifest.checksum,
                )
                .where(
                    SettlementProjectionPartitionManifest.generation_id == generation_id
                )
                .order_by(
                    SettlementProjectionPartitionManifest.artifact,
                    SettlementProjectionPartitionManifest.partition_key,
                )
            ).mappings()
        ]
        manifest_checksum = _manifest_checksum(manifest_rows)
        manifest_count = len(manifest_rows)
        row_count = sum(int(row["row_count"]) for row in manifest_rows)
        write_rows = 1 + manifest_count + row_count
        write_bytes = 16_384 + 4_096 * (manifest_count + row_count)
        generation.estimated_write_rows = write_rows
        generation.estimated_write_bytes = write_bytes
        generation.estimated_wal_bytes = 2 * write_bytes
        generation.estimated_disk_headroom_bytes = 0
        checkpoint = dict(generation.checkpoint_json or {})
        checkpoint.update(
            {
                "phase": "score_ready",
                "score": dict(contract),
                "score_manifest_count": sum(
                    1 for row in manifest_rows if row["artifact"] == "score"
                ),
                "score_row_count": sum(
                    int(row["row_count"])
                    for row in manifest_rows
                    if row["artifact"] == "score"
                ),
                "manifest_count": manifest_count,
                "row_count": row_count,
                "expected_active_pointer": base_generation_id,
            }
        )
        generation.checkpoint_json = checkpoint
        generation.manifest_checksum = manifest_checksum
        source_input = dict(generation.source_input_json or {})
        source_input["score"] = dict(contract)
        generation.source_input_json = source_input
        session.commit()

        score_manifests = list(
            session.scalars(
                select(SettlementProjectionPartitionManifest)
                .where(
                    SettlementProjectionPartitionManifest.generation_id == generation_id,
                    SettlementProjectionPartitionManifest.artifact == "score",
                )
                .order_by(SettlementProjectionPartitionManifest.partition_key)
            )
        )
        return ScoreManifest(
            generation_id=generation_id,
            base_generation_id=base_generation_id,
            snapshot_date=target_date,
            rule_version_ids=rule_version_ids,
            snapshot_run_ids=snapshot_run_ids,
            partition_keys=tuple(row.partition_key for row in score_manifests),
            manifest_count=len(score_manifests),
            row_count=sum(int(row.row_count) for row in score_manifests),
            manifest_checksum=manifest_checksum,
            resumed=resumed,
        )


def build_score_sparse_overlay(
    session_factory: Callable[[], Session],
    *,
    generation_id: str,
    base_generation_id: str,
    affected_store_ids: Iterable[str],
    published_rule_ids: Iterable[str],
    snapshot_date: date | str,
    batch_size: int,
    closure_policy_hash: str,
) -> ScoreManifest:
    """Build generation-owned score partitions without mutating legacy rows."""

    if not callable(session_factory):
        raise TypeError("session_factory must be callable")
    with session_factory() as session:
        (
            _generation,
            stores,
            _eligible,
            versions,
            target_date,
            contract,
            resumed,
        ) = _score_sparse_preflight(
            session,
            generation_id=generation_id,
            base_generation_id=base_generation_id,
            affected_store_ids=affected_store_ids,
            published_rule_ids=published_rule_ids,
            snapshot_date=snapshot_date,
            batch_size=batch_size,
            closure_policy_hash=closure_policy_hash,
        )
    run_ids: list[str] = []
    for version in versions:
        run_id, _partition_keys = _score_sparse_write_rule(
            session_factory,
            generation_id=generation_id,
            base_generation_id=base_generation_id,
            version_id=version.rule_version_id,
            snapshot_date=target_date,
            affected_store_ids=stores,
            closure_policy_hash=closure_policy_hash,
            batch_size=batch_size,
            contract=contract,
        )
        run_ids.append(run_id)
    return _score_sparse_finalize(
        session_factory,
        generation_id=generation_id,
        base_generation_id=base_generation_id,
        target_date=target_date,
        rule_version_ids=tuple(row.rule_version_id for row in versions),
        snapshot_run_ids=tuple(run_ids),
        contract=contract,
        resumed=resumed,
    )


def refresh_store_score_snapshots(
    session: Session,
    *,
    rule_version_id: str,
    now: datetime | None = None,
    run_mode: str = "scheduled",
    triggered_by: str | None = None,
    _scheduled_lock_acquired: bool = False,
) -> dict[str, object]:
    """Create an immutable score run for eligible stores using only formal, mature rounds.

    A version-bound run takes every scoring setting from that immutable rule
    version. The unbound configuration remains only for legacy/manual callers;
    allocation never selects those snapshots for a bound lead.
    """
    now = _aware(now or utcnow())
    score_config = _resolve_store_score_config(
        session,
        rule_version_id=rule_version_id,
    )

    snapshot_date = now.astimezone(SHANGHAI).date()
    scheduled_key: str | None = None
    if run_mode == "scheduled":
        if not _scheduled_lock_acquired and not _try_transaction_lock(session, lock_name=SCHEDULED_SCORE_REFRESH_LOCK):
            return {"snapshot_run_id": None, "snapshots": 0, "skipped": "locked"}
        scheduled_key = _scheduled_score_refresh_key(snapshot_date, score_config.rule_version_id)
        existing = session.scalar(
            select(StoreScoreSnapshotRun.snapshot_run_id)
            .where(StoreScoreSnapshotRun.scheduled_key == scheduled_key)
            .limit(1)
        )
        if existing:
            return {"snapshot_run_id": None, "snapshots": 0, "skipped": "already_refreshed"}

    snapshot_run_id = f"score-{now.strftime('%Y%m%dT%H%M%S%f')}-{uuid4().hex[:8]}"
    window_end = now
    window_start = now - timedelta(days=score_config.lookback_days)
    stores = eligible_candidate_stores(session)
    score_config_json = {
        "rule_version_id": score_config.rule_version_id,
        "lookback_days": score_config.lookback_days,
        "min_samples": score_config.min_samples,
        "execution_mode": "formal",
        "conversion_weight": str(score_config.conversion_weight),
        "follow_24h_weight": str(score_config.follow_weight),
        "store_weight": str(score_config.store_weight),
    }
    session.add(
        StoreScoreSnapshotRun(
            snapshot_run_id=snapshot_run_id,
            snapshot_date=snapshot_date,
            run_mode=run_mode,
            scheduled_key=scheduled_key,
            window_start=window_start,
            window_end=window_end,
            candidate_store_count=len(stores),
            snapshot_count=len(stores),
            triggered_by=_clean(triggered_by),
            config_json=score_config_json,
            computed_at=now,
        )
    )
    session.flush()
    if not stores:
        return {"snapshot_run_id": snapshot_run_id, "snapshots": 0}

    metrics_by_store = _formal_store_metrics(session, stores, window_start, window_end)
    city_metrics = _aggregate_city_metrics(stores, metrics_by_store)
    global_metrics = _sum_metrics(metrics_by_store.values())
    for store in stores:
        own_metrics = metrics_by_store.get(store.store_id, StoreMetrics())
        city_metric = city_metrics.get(store.city_code or "", StoreMetrics())
        conversion_rate, conversion_source = _resolved_rate(
            own_metrics.conversion_numerator,
            own_metrics.conversion_denominator,
            city_metric.conversion_numerator,
            city_metric.conversion_denominator,
            global_metrics.conversion_numerator,
            global_metrics.conversion_denominator,
            score_config.min_samples,
        )
        follow_rate, follow_source = _resolved_rate(
            own_metrics.follow_24h_numerator,
            own_metrics.follow_24h_denominator,
            city_metric.follow_24h_numerator,
            city_metric.follow_24h_denominator,
            global_metrics.follow_24h_numerator,
            global_metrics.follow_24h_denominator,
            score_config.min_samples,
        )
        score = (conversion_rate * score_config.conversion_weight + follow_rate * score_config.follow_weight) * score_config.store_weight
        session.add(
            StoreScoreSnapshot(
                snapshot_id=f"{snapshot_run_id}-{store.store_id}",
                snapshot_run_id=snapshot_run_id,
                snapshot_date=snapshot_date,
                run_mode=run_mode,
                store_id=store.store_id,
                city_code=store.city_code,
                window_start=window_start,
                window_end=window_end,
                conversion_numerator=own_metrics.conversion_numerator,
                conversion_denominator=own_metrics.conversion_denominator,
                conversion_rate=conversion_rate,
                conversion_value_source=conversion_source,
                follow_24h_numerator=own_metrics.follow_24h_numerator,
                follow_24h_denominator=own_metrics.follow_24h_denominator,
                follow_24h_rate=follow_rate,
                follow_24h_value_source=follow_source,
                conversion_weight=score_config.conversion_weight,
                follow_24h_weight=score_config.follow_weight,
                store_weight=score_config.store_weight,
                composite_score=score,
                config_json=score_config_json,
                computed_at=now,
            )
        )
    session.flush()
    return {"snapshot_run_id": snapshot_run_id, "snapshots": len(stores)}


def refresh_due_store_score_snapshots(
    session: Session,
    *,
    now: datetime | None = None,
) -> dict[str, object]:
    """Run at most one scheduled score refresh per Shanghai calendar day after 03:00."""
    now = _aware(now or utcnow())
    local_now = now.astimezone(SHANGHAI)
    if local_now.time() < SCHEDULED_SCORE_REFRESH_TIME:
        return {"snapshot_run_id": None, "snapshots": 0, "skipped": "before_schedule"}
    if not _try_transaction_lock(session, lock_name=SCHEDULED_SCORE_REFRESH_LOCK):
        return {"snapshot_run_id": None, "snapshots": 0, "skipped": "locked"}

    rule_version_ids = _score_rule_version_ids_for_refresh(session)
    if not rule_version_ids:
        return {"snapshot_run_id": None, "snapshots": 0, "skipped": "no_rule_versions"}

    results = [
        refresh_store_score_snapshots(
            session,
            now=now,
            run_mode="scheduled",
            rule_version_id=rule_version_id,
            _scheduled_lock_acquired=True,
        )
        for rule_version_id in rule_version_ids
    ]
    snapshot_run_ids = [str(result["snapshot_run_id"]) for result in results if result.get("snapshot_run_id")]
    if not snapshot_run_ids:
        return {"snapshot_run_id": None, "snapshots": 0, "skipped": "already_refreshed"}
    return {
        "snapshot_run_id": snapshot_run_ids[0],
        "snapshot_run_ids": snapshot_run_ids,
        "snapshots": sum(int(result.get("snapshots", 0) or 0) for result in results),
    }


def _resolve_store_score_config(
    session: Session,
    *,
    rule_version_id: str,
) -> StoreScoreConfig:
    normalized_rule_version_id = rule_version_id.strip() if isinstance(rule_version_id, str) else ""
    if not normalized_rule_version_id:
        raise ValueError("rule_version_id is required")
    version = session.get(ClueAllocationRuleVersion, normalized_rule_version_id)
    if version is None:
        raise ValueError("clue allocation rule version was not found")
    if (
        version.lookback_days is None
        or version.min_samples is None
        or version.conversion_weight is None
        or version.follow_24h_weight is None
    ):
        raise ValueError("clue allocation rule version is missing score configuration")
    return _validated_store_score_config(
        rule_version_id=version.rule_version_id,
        min_samples=version.min_samples,
        lookback_days=version.lookback_days,
        conversion_weight=version.conversion_weight,
        follow_weight=version.follow_24h_weight,
        store_weight=DEFAULT_STORE_WEIGHT,
    )


def _validated_store_score_config(
    *,
    rule_version_id: str | None,
    min_samples: int,
    lookback_days: int,
    conversion_weight: Decimal,
    follow_weight: Decimal,
    store_weight: Decimal,
) -> StoreScoreConfig:
    if min_samples <= 0:
        raise ValueError("min_samples must be positive")
    if lookback_days <= 0:
        raise ValueError("lookback_days must be positive")
    conversion = Decimal(str(conversion_weight))
    follow = Decimal(str(follow_weight))
    if conversion + follow != Decimal("1"):
        raise ValueError("conversion_weight and follow_weight must add up to 1")
    return StoreScoreConfig(
        rule_version_id=rule_version_id,
        min_samples=min_samples,
        lookback_days=lookback_days,
        conversion_weight=conversion,
        follow_weight=follow,
        store_weight=Decimal(str(store_weight)),
    )


def _score_rule_version_ids_for_refresh(session: Session) -> list[str]:
    published_version_ids = session.scalars(
        select(ClueAllocationRuleVersion.rule_version_id)
        .where(ClueAllocationRuleVersion.status == "published")
        .order_by(ClueAllocationRuleVersion.rule_version_id)
    ).all()
    bound_version_ids = session.scalars(
        select(ClueLeadRuleVersionBinding.rule_version_id)
        .join(ClueMasterLead, ClueMasterLead.lead_key == ClueLeadRuleVersionBinding.lead_key)
        .where(ClueMasterLead.lifecycle_status == "active")
        .where(ClueMasterLead.normalized_order_status == "active")
        .order_by(ClueLeadRuleVersionBinding.rule_version_id)
    ).all()
    return sorted(set(published_version_ids).union(bound_version_ids))


def _scheduled_score_refresh_key(snapshot_date: date, rule_version_id: str | None) -> str:
    target = rule_version_id or "legacy"
    return f"scheduled-{snapshot_date.isoformat()}-{target}"


def _master_order_is_compatible(master: ClueMasterLead, order_id: str | None) -> bool:
    master_order_id = _clean(master.order_id)
    if order_id is None:
        return master_order_id is None
    return master_order_id is None or master_order_id == order_id


def _master_match_conflict_reason(
    *,
    source_match: ClueMasterLead | None,
    order_match: ClueMasterLead | None,
    identity_match: ClueMasterLead | None,
    order_id: str | None,
) -> str | None:
    if source_match is not None and not _master_order_is_compatible(source_match, order_id):
        return "source_record_order_changed"
    if source_match is not None and order_match is not None and source_match is not order_match:
        return "source_record_and_order_resolve_to_different_leads"
    selected = source_match or order_match
    if selected is not None and identity_match is not None and selected is not identity_match:
        return "source_identity_already_bound_to_another_lead"
    if identity_match is not None and not _master_order_is_compatible(identity_match, order_id):
        return "source_identity_order_conflict"
    return None


def _record_identity_mapping_conflict(
    session: Session,
    *,
    raw_clue: RawDouyinClue,
    source_match: ClueMasterLead | None,
    order_match: ClueMasterLead | None,
    identity_match: ClueMasterLead | None,
    reason: str,
    now: datetime,
) -> None:
    source = "|".join(
        [
            raw_clue.clue_row_key,
            _clean(raw_clue.order_id) or "",
            source_match.lead_key if source_match else "",
            order_match.lead_key if order_match else "",
            identity_match.lead_key if identity_match else "",
            reason,
        ]
    )
    issue_id = f"clue-identity-conflict:{sha256(source.encode('utf-8')).hexdigest()[:32]}"
    upsert_data_quality_issue(
        session,
        issue_id,
        issue_type="clue_identity_conflict",
        message="clue source identifiers resolve to conflicting master leads",
        order_id=_clean(raw_clue.order_id),
        severity="error",
        raw_context_json={
            "source_clue_row_key": raw_clue.clue_row_key,
            "canonical_clue_id": _clean(raw_clue.clue_id),
            "source_lead_key": source_match.lead_key if source_match else None,
            "source_order_id": _clean(source_match.order_id) if source_match else None,
            "order_lead_key": order_match.lead_key if order_match else None,
            "identity_lead_key": identity_match.lead_key if identity_match else None,
            "reason": reason,
            "observed_at": now.isoformat(),
        },
        source_run_id=None,
        flush=False,
    )


def _mark_identifier_conflicts(
    session: Session,
    rows: list[tuple[str, str, int]],
) -> None:
    if not rows:
        return
    conflicts: set[tuple[str, str]] = session.info.setdefault(
        "clue_identifier_conflicts",
        set(),
    )
    counts: dict[tuple[str, str], int] = session.info.setdefault(
        "clue_identifier_conflict_counts",
        {},
    )
    for identifier_type, identifier_value, lead_count in rows:
        if int(lead_count) <= 1:
            continue
        key = (str(identifier_type), str(identifier_value))
        conflicts.add(key)
        counts[key] = max(int(counts.get(key, 0)), int(lead_count))


def _raw_identifier_collision_candidates(
    raw_clue: RawDouyinClue,
    *,
    source_identity_key: str,
    canonical_clue_id: str | None,
) -> tuple[tuple[str, str], ...]:
    candidates: list[tuple[str, str]] = []
    if canonical_clue_id:
        candidates.append(("clue_id", canonical_clue_id))
    if source_identity_key:
        candidates.append(("source_identity_key", source_identity_key))
    if raw_clue.clue_row_key:
        candidates.append(("source_clue_row_key", raw_clue.clue_row_key))
    return tuple(candidates)


def _record_identifier_collision_issue(
    session: Session,
    *,
    raw_clue: RawDouyinClue,
    collisions: list[tuple[str, str, int]],
    now: datetime,
) -> None:
    descriptors = sorted(
        (
            identifier_type,
            sha256(identifier_value.encode("utf-8")).hexdigest()[:16],
            int(lead_count),
        )
        for identifier_type, identifier_value, lead_count in collisions
    )
    if not descriptors:
        return
    digest_source = "|".join(
        [
            raw_clue.clue_row_key,
            *(
                f"{identifier_type}:{value_hash}"
                for identifier_type, value_hash, _ in descriptors
            ),
        ]
    )
    issue_id = (
        "clue-identifier-collision:"
        f"{sha256(digest_source.encode('utf-8')).hexdigest()[:32]}"
    )
    upsert_data_quality_issue(
        session,
        issue_id,
        issue_type="clue_identifier_collision",
        message="clue identifier collision requires manual resolution",
        order_id=None,
        severity="error",
        raw_context_json={
            "source_row_hash": sha256(
                raw_clue.clue_row_key.encode("utf-8")
            ).hexdigest()[:16],
            "identifier_types": sorted(
                {identifier_type for identifier_type, _, _ in descriptors}
            ),
            "candidate_hashes": [
                f"{identifier_type}:{value_hash}"
                for identifier_type, value_hash, _ in descriptors
            ],
            "distinct_lead_count": max(
                lead_count for _, _, lead_count in descriptors
            ),
            "observed_at": now.isoformat(),
        },
        source_run_id=None,
        flush=False,
    )


def _source_record_links_by_key(
    session: Session,
    source_record_keys: set[str],
) -> dict[str, ClueSourceRecordLink]:
    links: dict[str, ClueSourceRecordLink] = {}
    for key_batch in _materialization_order_id_batches(source_record_keys):
        rows = session.scalars(
            select(ClueSourceRecordLink)
            .where(ClueSourceRecordLink.source_table == "raw_douyin_clues")
            .where(ClueSourceRecordLink.source_record_key.in_(key_batch))
        ).all()
        links.update({row.source_record_key: row for row in rows})
    return links


def _upsert_source_record_link(
    session: Session,
    *,
    raw_clue: RawDouyinClue,
    lead: ClueMasterLead,
    order_id: str | None,
    link_status: int,
    link_method: int,
    conflict_reason: str | None,
    observed_at: datetime,
    now: datetime,
    links_by_record_key: dict[str, ClueSourceRecordLink],
) -> ClueSourceRecordLink:
    source_record_key = raw_clue.clue_row_key
    link = links_by_record_key.get(source_record_key)
    if link is None:
        link = ClueSourceRecordLink(
            source_system=1,
            source_table="raw_douyin_clues",
            source_record_key=source_record_key,
            lead_key=lead.lead_key,
            link_status=link_status,
            link_method=link_method,
            link_version=1,
            linked_at=now,
            created_at=now,
            updated_at=now,
        )
        session.add(link)
        links_by_record_key[source_record_key] = link
    else:
        target_lead_key = lead.lead_key
        if link.lead_key != target_lead_key:
            link_status = 3
            conflict_reason = conflict_reason or "source_record_already_linked_to_another_lead"
            target_lead_key = link.lead_key
        mapping_before = (
            link.lead_key,
            link.order_id,
            link.link_status,
            link.link_method,
        )
        mapping_after = (target_lead_key, order_id, link_status, link_method)
        if mapping_before != mapping_after:
            link.link_version = max(link.link_version or 1, 1) + 1
        link.lead_key = target_lead_key

    link.source_clue_id = _clean(raw_clue.clue_id)
    link.source_order_id = _clean(raw_clue.order_id)
    link.order_id = order_id if link_status == 1 else _clean(lead.order_id)
    link.link_status = link_status
    link.link_method = link_method
    link.source_observed_at = max(
        filter(None, (_aware(link.source_observed_at), observed_at)),
    )
    link.source_payload_hash = _source_payload_hash(raw_clue.raw_payload)
    link.conflict_reason = conflict_reason
    link.updated_at = now
    return link


def _accepts_status_evidence(
    lead: ClueMasterLead,
    *,
    incoming_status: str,
    observed_at: datetime,
) -> bool:
    previous_observed_at = _aware(lead.order_status_observed_at)
    if previous_observed_at is not None and observed_at < previous_observed_at:
        return False
    terminal_statuses = {"verified", "refunded", "closed"}
    if (
        lead.normalized_order_status in terminal_statuses
        and incoming_status not in terminal_statuses
    ):
        return False
    return True


def _master_state_signature(lead: ClueMasterLead) -> tuple[object, ...]:
    return (
        lead.master_kind,
        lead.normalized_order_status,
        lead.status_source,
        lead.lifecycle_status,
        lead.pool_location,
        lead.allocation_state,
        lead.current_assignment_round_id,
        lead.allocation_cycle_id,
        lead.ended_without_assignment,
        lead.closed_reason,
        lead.is_complete_pool,
        lead.anchor_poi_id,
        lead.anchor_store_id,
        lead.anchor_unavailable_reason,
    )


def _source_payload_hash(raw_payload: dict[str, object] | None) -> str:
    serialized = json.dumps(
        raw_payload or {},
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    )
    return sha256(serialized.encode("utf-8")).hexdigest()


def _source_identifier_history_id(
    source_clue_row_key: str,
    identifier_type: str,
    identifier_value: str,
) -> str:
    source = f"{source_clue_row_key}|{identifier_type}|{identifier_value}"
    return f"clue-identifier-{sha256(source.encode('utf-8')).hexdigest()[:32]}"


def _set_current_source_identifier(
    session: Session,
    *,
    lead_key: str,
    source_clue_row_key: str,
    identifier_type: str,
    identifier_value: str | None,
    source_payload_hash: str,
    observed_at: datetime,
    now: datetime,
    history_by_key: dict[tuple[str, str, str], ClueSourceIdentifierHistory],
    current_by_source_type: dict[tuple[str, str], list[ClueSourceIdentifierHistory]],
) -> None:
    source_type_key = (source_clue_row_key, identifier_type)
    for current in current_by_source_type.get(source_type_key, []):
        if current.identifier_value != identifier_value:
            current.is_current = False
            current.updated_at = now

    if identifier_value is None:
        current_by_source_type[source_type_key] = []
        return

    history_key = (source_clue_row_key, identifier_type, identifier_value)
    history = history_by_key.get(history_key)
    if history is None:
        history = ClueSourceIdentifierHistory(
            identifier_history_id=_source_identifier_history_id(
                source_clue_row_key,
                identifier_type,
                identifier_value,
            ),
            lead_key=lead_key,
            source_clue_row_key=source_clue_row_key,
            identifier_type=identifier_type,
            identifier_value=identifier_value,
            source_payload_hash=source_payload_hash,
            first_seen_at=observed_at,
            last_seen_at=observed_at,
            is_current=True,
            created_at=now,
            updated_at=now,
        )
        session.add(history)
        history_by_key[history_key] = history
    else:
        if history.lead_key != lead_key:
            raise ValueError("source identifier history is already linked to another lead")
        history.source_payload_hash = source_payload_hash
        history.last_seen_at = observed_at
        history.is_current = True
        history.updated_at = now
    current_by_source_type[source_type_key] = [history]


def _record_stale_source_identifiers(
    session: Session,
    *,
    existing: ClueMasterLead,
    raw_clue: RawDouyinClue,
    observed_at: datetime,
    now: datetime,
    history_by_key: dict[tuple[str, str, str], ClueSourceIdentifierHistory],
    current_by_source_type: dict[
        tuple[str, str],
        list[ClueSourceIdentifierHistory],
    ],
) -> None:
    payload_hash = _source_payload_hash(raw_clue.raw_payload)
    for identifier_type, identifier_value in (
        ("clue_id", _clean(raw_clue.clue_id)),
        ("source_identity_key", _source_identity_key(raw_clue)),
    ):
        if identifier_value is None:
            continue
        history_key = (raw_clue.clue_row_key, identifier_type, identifier_value)
        history = history_by_key.get(history_key)
        if history is None:
            history = ClueSourceIdentifierHistory(
                identifier_history_id=_source_identifier_history_id(
                    raw_clue.clue_row_key,
                    identifier_type,
                    identifier_value,
                ),
                lead_key=existing.lead_key,
                source_clue_row_key=raw_clue.clue_row_key,
                identifier_type=identifier_type,
                identifier_value=identifier_value,
                source_payload_hash=payload_hash,
                first_seen_at=observed_at,
                last_seen_at=observed_at,
                is_current=False,
                created_at=now,
                updated_at=now,
            )
            session.add(history)
            history_by_key[history_key] = history
            continue
        if history.lead_key != existing.lead_key:
            raise ValueError("source identifier history is already linked to another lead")
        history.first_seen_at = min(
            _aware(history.first_seen_at) or observed_at,
            observed_at,
        )
        history.last_seen_at = max(
            _aware(history.last_seen_at) or observed_at,
            observed_at,
        )
        if history.source_payload_hash is None:
            history.source_payload_hash = payload_hash
        history.updated_at = now
        if history.is_current:
            current_by_source_type.setdefault(
                (raw_clue.clue_row_key, identifier_type),
                [],
            )


def _source_identity_key(raw_clue: RawDouyinClue) -> str:
    order_id = _clean(raw_clue.order_id)
    contact_value = _clean(raw_clue.telephone) or _clean(raw_clue.enc_telephone)
    canonical_clue_id = _clean(raw_clue.clue_id)
    if order_id and contact_value:
        source = f"order-contact|{order_id}|{contact_value}"
    elif canonical_clue_id:
        source = f"clue|{canonical_clue_id}"
    else:
        source = f"raw|{raw_clue.clue_row_key}"
    return f"identity-{sha256(source.encode('utf-8')).hexdigest()[:32]}"


def _lead_key(source_identity_key: str) -> str:
    return f"lead-{source_identity_key.removeprefix('identity-')}"


def _try_transaction_lock(session: Session, *, lock_name: str) -> bool:
    """Prevent parallel PostgreSQL workers from materializing the same M1 state."""
    if session.get_bind().dialect.name != "postgresql":
        return True
    lock_key = _advisory_lock_key(lock_name)
    return bool(session.scalar(select(func.pg_try_advisory_xact_lock(lock_key))))


def _advisory_lock_key(lock_name: str) -> int:
    return int.from_bytes(
        sha256(lock_name.encode("utf-8")).digest()[:8],
        byteorder="big",
        signed=True,
    )


def _try_session_advisory_lock(session: Session, lock_key: int) -> bool:
    if session.get_bind().dialect.name != "postgresql":
        return True
    return bool(session.scalar(select(func.pg_try_advisory_lock(lock_key))))


def _release_session_advisory_lock(session: Session, lock_key: int) -> None:
    if session.get_bind().dialect.name != "postgresql":
        return
    try:
        session.scalar(select(func.pg_advisory_unlock(lock_key)))
    except Exception:
        # The surrounding session scope will roll back a failed transaction.
        pass


def _raw_orders_by_id(session: Session, order_ids: set[str]) -> dict[str, RawDouyinOrder]:
    if not order_ids:
        return {}
    values: dict[str, RawDouyinOrder] = {}
    for order_id_batch in _materialization_order_id_batches(order_ids):
        rows = session.scalars(
            select(RawDouyinOrder).where(RawDouyinOrder.order_id.in_(order_id_batch))
        ).all()
        values.update({row.order_id: row for row in rows})
    return values


def _coupon_statuses_by_order(
    session: Session,
    order_ids: set[str],
) -> dict[str, list[str]]:
    if not order_ids:
        return {}
    values: dict[str, list[str]] = defaultdict(list)
    for order_id_batch in _materialization_order_id_batches(order_ids):
        rows = session.execute(
            select(
                RawDouyinOrderCoupon.order_id,
                RawDouyinOrderCoupon.coupon_status,
                RawDouyinOrderCoupon.coupon_status_raw,
                RawDouyinOrderCoupon.coupon_status_normalized,
            ).where(RawDouyinOrderCoupon.order_id.in_(order_id_batch))
        ).all()
        for order_id, coupon_status, coupon_status_raw, normalized in rows:
            if not order_id:
                continue
            values[order_id].append(
                _clean(normalized)
                or normalize_coupon_status(coupon_status_raw or coupon_status)
            )
    return values


def _verified_at_by_order(session: Session, order_ids: set[str]) -> dict[str, datetime]:
    if not order_ids:
        return {}
    values: dict[str, datetime] = {}
    for order_id_batch in _materialization_order_id_batches(order_ids):
        rows = session.execute(
            select(
                SettlementOrderDetail.order_id,
                func.min(SettlementOrderDetail.verify_time).label("verify_time"),
            )
            .where(SettlementOrderDetail.order_id.in_(order_id_batch))
            .where(SettlementOrderDetail.is_verified.is_(True))
            .where(SettlementOrderDetail.verify_time.is_not(None))
            .group_by(SettlementOrderDetail.order_id)
        ).all()
        for order_id, verify_time in rows:
            if not order_id:
                continue
            candidate = _aware(verify_time)
            if candidate is None:
                continue
            values[order_id] = candidate
    return values


def _materialization_order_id_batches(order_ids: set[str]) -> list[list[str]]:
    ordered_ids = sorted(order_id for order_id in order_ids if order_id)
    return [
        ordered_ids[offset : offset + MATERIALIZATION_QUERY_BATCH_SIZE]
        for offset in range(0, len(ordered_ids), MATERIALIZATION_QUERY_BATCH_SIZE)
    ]


def _bounded_raw_clues(
    session: Session,
    *,
    raw_clue_row_keys: set[str],
    clue_ids: set[str],
    order_ids: set[str],
    poi_ids: set[str],
    limit: int | None = None,
    after_row_key: str | None = None,
    target_scope: str | None = None,
    target_cycle_id: str | None = None,
) -> list[RawDouyinClue]:
    """Read only the raw rows in one impact closure.

    This helper intentionally returns an empty list when all selectors are
    empty.  The incremental caller can therefore never accidentally turn an
    empty impact into a legacy full-table scan.
    """

    row_keys = {str(value) for value in raw_clue_row_keys if _clean(value)}
    canonical_ids = {str(value) for value in clue_ids if _clean(value)}
    selected_orders = {str(value) for value in order_ids if _clean(value)}
    selected_pois = {str(value) for value in poi_ids if _clean(value)}
    selectors = []
    if row_keys:
        selectors.append(RawDouyinClue.clue_row_key.in_(row_keys))
    if canonical_ids:
        selectors.append(RawDouyinClue.clue_id.in_(canonical_ids))
    if selected_orders:
        selectors.append(RawDouyinClue.order_id.in_(selected_orders))
    if selected_pois:
        selectors.extend(
            (
                RawDouyinClue.follow_poi_id.in_(selected_pois),
                RawDouyinClue.intention_poi_id.in_(selected_pois),
            )
        )
    if not selectors:
        return []
    stmt = (
        select(RawDouyinClue)
        .where(or_(*selectors))
        .order_by(RawDouyinClue.clue_row_key)
    )
    if target_scope and target_cycle_id:
        stmt = stmt.where(
            ~exists(
                select(1).where(
                    ClueMaterializationTarget.scope == target_scope,
                    ClueMaterializationTarget.cycle_id == target_cycle_id,
                    ClueMaterializationTarget.target_type == "raw",
                    ClueMaterializationTarget.target_key
                    == RawDouyinClue.clue_row_key,
                )
            )
        )
    if after_row_key:
        stmt = stmt.where(RawDouyinClue.clue_row_key > str(after_row_key))
    if limit is not None:
        stmt = stmt.limit(max(1, int(limit)))
    return list(
        session.scalars(stmt).yield_per(
            max(1, min(int(limit or MATERIALIZATION_QUERY_BATCH_SIZE), MATERIALIZATION_QUERY_BATCH_SIZE))
        )
    )


def _bounded_center_order_ids(
    session: Session,
    *,
    raw_clue_row_keys: set[str],
    clue_ids: set[str],
    order_ids: set[str],
    poi_ids: set[str],
    after_order_id: str | None = None,
    limit: int = CENTER_ORDER_BATCH_SIZE,
    target_scope: str | None = None,
    target_cycle_id: str | None = None,
) -> list[str]:
    """Return one deterministic order-id page for the center phase.

    The query returns only distinct order keys.  Center projection then owns
    the bounded order page and may read the source clues necessary to rebuild
    that order; it never re-reads the same order once per raw clue page.
    """

    row_keys = {str(value) for value in raw_clue_row_keys if _clean(value)}
    canonical_ids = {str(value) for value in clue_ids if _clean(value)}
    selected_orders = {str(value) for value in order_ids if _clean(value)}
    selected_pois = {str(value) for value in poi_ids if _clean(value)}
    selectors = []
    if row_keys:
        selectors.append(RawDouyinClue.clue_row_key.in_(row_keys))
    if canonical_ids:
        selectors.append(RawDouyinClue.clue_id.in_(canonical_ids))
    if selected_orders:
        selectors.append(RawDouyinClue.order_id.in_(selected_orders))
    if selected_pois:
        selectors.extend(
            (
                RawDouyinClue.follow_poi_id.in_(selected_pois),
                RawDouyinClue.intention_poi_id.in_(selected_pois),
            )
        )
    if not selectors:
        return []
    stmt = (
        select(RawDouyinClue.order_id)
        .where(or_(*selectors))
        .where(RawDouyinClue.order_status == "\u5c65\u7ea6\u4e2d")
        .where(RawDouyinClue.order_id.is_not(None))
        .where(RawDouyinClue.order_id != "")
        .where(RawDouyinClue.order_id != "0")
    )
    if target_scope and target_cycle_id:
        stmt = stmt.where(
            ~exists(
                select(1).where(
                    ClueMaterializationTarget.scope == target_scope,
                    ClueMaterializationTarget.cycle_id == target_cycle_id,
                    ClueMaterializationTarget.target_type == "center",
                    ClueMaterializationTarget.target_key == RawDouyinClue.order_id,
                )
            )
        )
    if after_order_id:
        stmt = stmt.where(RawDouyinClue.order_id > str(after_order_id))
    stmt = stmt.distinct().order_by(RawDouyinClue.order_id).limit(max(1, int(limit)))
    return [
        str(value)
        for value in session.scalars(stmt).yield_per(max(1, min(int(limit), CENTER_ORDER_BATCH_SIZE)))
        if _clean(value)
    ]


def _record_cycle_targets(
    session: Session,
    *,
    scope: str,
    cycle_id: str,
    target_type: str,
    target_keys: list[str] | set[str] | tuple[str, ...],
) -> None:
    """Persist bounded completion markers in the same projection transaction."""

    keys = {str(value) for value in target_keys if _clean(value)}
    if not keys:
        return
    existing = set(
        session.scalars(
            select(ClueMaterializationTarget.target_key).where(
                ClueMaterializationTarget.scope == scope,
                ClueMaterializationTarget.cycle_id == cycle_id,
                ClueMaterializationTarget.target_type == target_type,
                ClueMaterializationTarget.target_key.in_(keys),
            )
        ).all()
    )
    for target_key in sorted(keys.difference(existing)):
        session.add(
            ClueMaterializationTarget(
                scope=scope,
                cycle_id=cycle_id,
                target_type=target_type,
                target_key=target_key,
            )
        )
    session.flush()



def _bounded_location_context(
    session: Session,
    raw_clues: list[RawDouyinClue],
    *,
    poi_ids: set[str],
) -> tuple[dict[str, DimStorePoiMapping], dict[str, DimStore]]:
    selected_pois = {
        str(value)
        for value in poi_ids
        if _clean(value)
    }
    selected_pois.update(
        value
        for raw_clue in raw_clues
        for value in (_clean(raw_clue.follow_poi_id), _clean(raw_clue.intention_poi_id))
        if value
    )
    if not selected_pois:
        return {}, {}
    mappings = list(
        session.scalars(
            select(DimStorePoiMapping)
            .where(DimStorePoiMapping.poi_id.in_(selected_pois))
            .order_by(DimStorePoiMapping.poi_id)
        ).yield_per(MATERIALIZATION_QUERY_BATCH_SIZE)
    )
    mappings_by_poi = {row.poi_id: row for row in mappings}
    store_ids = {row.store_id for row in mappings if row.store_id}
    if not store_ids:
        return mappings_by_poi, {}
    stores = list(
        session.scalars(
            select(DimStore)
            .where(DimStore.store_id.in_(store_ids))
            .order_by(DimStore.store_id)
        ).yield_per(MATERIALIZATION_QUERY_BATCH_SIZE)
    )
    return mappings_by_poi, {row.store_id: row for row in stores}


def _bounded_existing_masters(
    session: Session,
    raw_clues: list[RawDouyinClue],
    *,
    order_ids: set[str],
    clue_ids: set[str] | None = None,
    poi_ids: set[str] | None = None,
    source_identity_keys: set[str] | None = None,
    source_link_lead_keys: set[str] | None = None,
) -> list[ClueMasterLead]:
    row_keys = {row.clue_row_key for row in raw_clues if row.clue_row_key}
    canonical_ids = {row.clue_id for row in raw_clues if row.clue_id}
    canonical_ids.update(value for value in (clue_ids or set()) if value)
    selected_orders = {value for value in order_ids if value}
    selected_pois = {value for value in (poi_ids or set()) if value}
    selected_identities = {
        value for value in (source_identity_keys or set()) if value
    }
    selected_identities.update(
        _source_identity_key(raw_clue)
        for raw_clue in raw_clues
        if _source_identity_key(raw_clue)
    )

    canonical_lead_keys: set[str] = set()
    if canonical_ids:
        canonical_rows = session.execute(
            select(
                ClueMasterLead.canonical_clue_id,
                func.count(func.distinct(ClueMasterLead.lead_key)).label(
                    "lead_count"
                ),
                func.min(ClueMasterLead.lead_key).label(
                    "representative_lead_key"
                ),
            )
            .where(ClueMasterLead.canonical_clue_id.in_(canonical_ids))
            .group_by(ClueMasterLead.canonical_clue_id)
            .order_by(ClueMasterLead.canonical_clue_id)
        ).all()
        _mark_identifier_conflicts(
            session,
            [
                ("clue_id", str(canonical_id), int(lead_count))
                for canonical_id, lead_count, _ in canonical_rows
                if canonical_id and int(lead_count) > 1
            ],
        )
        canonical_lead_keys.update(
            str(representative_lead_key)
            for canonical_id, lead_count, representative_lead_key in canonical_rows
            if canonical_id and int(lead_count) == 1 and representative_lead_key
        )

    selectors = []
    if row_keys:
        selectors.append(ClueMasterLead.source_clue_row_key.in_(row_keys))
    if canonical_lead_keys:
        selectors.append(ClueMasterLead.lead_key.in_(canonical_lead_keys))
    has_strong_page_selector = bool(
        row_keys or canonical_ids or selected_identities
    )
    if selected_orders and not has_strong_page_selector:
        selectors.append(ClueMasterLead.order_id.in_(selected_orders))
    if selected_pois and not has_strong_page_selector:
        selectors.append(ClueMasterLead.anchor_poi_id.in_(selected_pois))
    if selected_identities:
        selectors.append(
            ClueMasterLead.source_identity_key.in_(selected_identities)
        )

    history_lead_keys: set[str] = set()
    if row_keys:
        source_history_rows = session.execute(
            select(
                ClueSourceIdentifierHistory.source_clue_row_key,
                func.count(
                    func.distinct(ClueSourceIdentifierHistory.lead_key)
                ).label("lead_count"),
                func.min(ClueSourceIdentifierHistory.lead_key).label(
                    "representative_lead_key"
                ),
            )
            .where(
                ClueSourceIdentifierHistory.source_clue_row_key.in_(row_keys)
            )
            .group_by(ClueSourceIdentifierHistory.source_clue_row_key)
            .order_by(ClueSourceIdentifierHistory.source_clue_row_key)
        ).all()
        _mark_identifier_conflicts(
            session,
            [
                (
                    "source_clue_row_key",
                    str(source_row_key),
                    int(lead_count),
                )
                for source_row_key, lead_count, _ in source_history_rows
                if source_row_key and int(lead_count) > 1
            ],
        )
        history_lead_keys.update(
            str(representative_lead_key)
            for source_row_key, lead_count, representative_lead_key in source_history_rows
            if source_row_key and int(lead_count) == 1 and representative_lead_key
        )

    candidate_identifier_groups = {
        "clue_id": canonical_ids,
        "source_identity_key": selected_identities,
    }
    for identifier_type, identifier_values in sorted(
        candidate_identifier_groups.items()
    ):
        values = {
            str(value) for value in identifier_values if _clean(value)
        }
        if not values:
            continue
        history_rows = session.execute(
            select(
                ClueSourceIdentifierHistory.identifier_value,
                func.count(
                    func.distinct(ClueSourceIdentifierHistory.lead_key)
                ).label("lead_count"),
                func.min(ClueSourceIdentifierHistory.lead_key).label(
                    "representative_lead_key"
                ),
            )
            .where(
                ClueSourceIdentifierHistory.identifier_type == identifier_type
            )
            .where(
                ClueSourceIdentifierHistory.identifier_value.in_(values)
            )
            .group_by(ClueSourceIdentifierHistory.identifier_value)
            .order_by(ClueSourceIdentifierHistory.identifier_value)
        ).all()
        _mark_identifier_conflicts(
            session,
            [
                (identifier_type, str(identifier_value), int(lead_count))
                for identifier_value, lead_count, _ in history_rows
                if identifier_value and int(lead_count) > 1
            ],
        )
        history_lead_keys.update(
            str(representative_lead_key)
            for identifier_value, lead_count, representative_lead_key in history_rows
            if identifier_value and int(lead_count) == 1 and representative_lead_key
        )
    if history_lead_keys:
        selectors.append(ClueMasterLead.lead_key.in_(history_lead_keys))
    if source_link_lead_keys:
        selectors.append(
            ClueMasterLead.lead_key.in_(
                {value for value in source_link_lead_keys if value}
            )
        )

    rows: list[ClueMasterLead] = []
    if selectors:
        stmt = (
            select(ClueMasterLead)
            .where(or_(*selectors))
            .order_by(ClueMasterLead.lead_key)
        )
        rows = list(
            session.scalars(stmt).yield_per(MATERIALIZATION_QUERY_BATCH_SIZE)
        )
    if selected_orders and has_strong_page_selector:
        ranked_order_leads = (
            select(
                ClueMasterLead.lead_key.label("lead_key"),
                func.row_number()
                .over(
                    partition_by=ClueMasterLead.order_id,
                    order_by=ClueMasterLead.lead_key,
                )
                .label("order_rank"),
            )
            .where(ClueMasterLead.order_id.in_(selected_orders))
            .subquery()
        )
        order_candidates = list(
            session.scalars(
                select(ClueMasterLead)
                .join(
                    ranked_order_leads,
                    ClueMasterLead.lead_key == ranked_order_leads.c.lead_key,
                )
                .where(ranked_order_leads.c.order_rank <= 2)
                .order_by(
                    ClueMasterLead.order_id,
                    ClueMasterLead.lead_key,
                )
            )
        )
        candidates_by_order: dict[str, list[ClueMasterLead]] = defaultdict(list)
        for row in order_candidates:
            if row.order_id:
                candidates_by_order[row.order_id].append(row)
        existing_lead_keys = {row.lead_key for row in rows}
        for candidates in candidates_by_order.values():
            for candidate in candidates:
                if candidate.lead_key in existing_lead_keys:
                    continue
                rows.append(candidate)
                existing_lead_keys.add(candidate.lead_key)
    return rows


def _bounded_identifier_history(
    session: Session,
    *,
    raw_clue_row_keys: set[str],
    candidate_identifiers: set[tuple[str, str]],
) -> list[ClueSourceIdentifierHistory]:
    row_keys = {
        str(value) for value in raw_clue_row_keys if _clean(value)
    }
    identifiers = {
        (str(identifier_type), str(identifier_value))
        for identifier_type, identifier_value in candidate_identifiers
        if _clean(identifier_type) and _clean(identifier_value)
    }
    selectors = []
    if row_keys:
        selectors.append(
            and_(
                ClueSourceIdentifierHistory.source_clue_row_key.in_(row_keys),
                ClueSourceIdentifierHistory.is_current.is_(True),
            )
        )
    if row_keys and identifiers:
        identifiers_by_type: dict[str, set[str]] = defaultdict(set)
        for identifier_type, identifier_value in identifiers:
            identifiers_by_type[identifier_type].add(identifier_value)
        selectors.extend(
            and_(
                ClueSourceIdentifierHistory.source_clue_row_key.in_(row_keys),
                ClueSourceIdentifierHistory.identifier_type == identifier_type,
                ClueSourceIdentifierHistory.identifier_value.in_(values),
            )
            for identifier_type, values in sorted(
                identifiers_by_type.items()
            )
            if values
        )

    representative_history_ids: set[str] = set()
    identifier_values_by_type: dict[str, set[str]] = defaultdict(set)
    for identifier_type, identifier_value in identifiers:
        identifier_values_by_type[identifier_type].add(identifier_value)
    for identifier_type, values in sorted(
        identifier_values_by_type.items()
    ):
        if not values:
            continue
        aggregate_rows = session.execute(
            select(
                ClueSourceIdentifierHistory.identifier_value,
                func.count(
                    func.distinct(ClueSourceIdentifierHistory.lead_key)
                ).label("lead_count"),
                func.min(
                    ClueSourceIdentifierHistory.identifier_history_id
                ).label("representative_history_id"),
            )
            .where(
                ClueSourceIdentifierHistory.identifier_type == identifier_type
            )
            .where(
                ClueSourceIdentifierHistory.identifier_value.in_(values)
            )
            .group_by(ClueSourceIdentifierHistory.identifier_value)
            .order_by(ClueSourceIdentifierHistory.identifier_value)
        ).all()
        _mark_identifier_conflicts(
            session,
            [
                (identifier_type, str(identifier_value), int(lead_count))
                for identifier_value, lead_count, _ in aggregate_rows
                if identifier_value and int(lead_count) > 1
            ],
        )
        representative_history_ids.update(
            str(representative_history_id)
            for identifier_value, lead_count, representative_history_id in aggregate_rows
            if identifier_value and int(lead_count) == 1 and representative_history_id
        )
    if representative_history_ids:
        selectors.append(
            ClueSourceIdentifierHistory.identifier_history_id.in_(
                representative_history_ids
            )
        )
    if not selectors:
        return []
    stmt = (
        select(ClueSourceIdentifierHistory)
        .where(or_(*selectors))
        .order_by(
            ClueSourceIdentifierHistory.source_clue_row_key,
            ClueSourceIdentifierHistory.identifier_type,
            ClueSourceIdentifierHistory.identifier_value,
        )
    )
    return list(
        session.scalars(stmt).yield_per(MATERIALIZATION_QUERY_BATCH_SIZE)
    )


def _bounded_active_headquarters_entries(
    session: Session,
    lead_keys: set[str],
) -> list[ClueHeadquartersPoolEntry]:
    if not lead_keys:
        return []
    return list(
        session.scalars(
            select(ClueHeadquartersPoolEntry).where(
                ClueHeadquartersPoolEntry.status == "active",
                ClueHeadquartersPoolEntry.lead_key.in_(lead_keys),
            )
        ).all()
    )


def _bounded_anchor_issue_ids(session: Session, lead_keys: set[str]) -> list[str]:
    if not lead_keys:
        return []
    candidate_ids = [
        f"clue-anchor:{lead_key}:{reason}"
        for lead_key in lead_keys
        for reason in ANCHOR_UNAVAILABLE_REASONS
    ]
    issue_ids: list[str] = []
    for offset in range(0, len(candidate_ids), MATERIALIZATION_QUERY_BATCH_SIZE):
        issue_ids.extend(
            session.scalars(
                select(DataQualityIssue.issue_id).where(
                    DataQualityIssue.issue_id.in_(
                        candidate_ids[offset : offset + MATERIALIZATION_QUERY_BATCH_SIZE]
                    )
                )
            ).all()
        )
    return issue_ids


def _resolve_status(
    raw_clue: RawDouyinClue,
    raw_order: RawDouyinOrder | None,
    verified_at: datetime | None,
    now: datetime,
    coupon_statuses: list[str] | tuple[str, ...] = (),
) -> StatusResolution:
    if verified_at is not None:
        return StatusResolution(
            raw_status=_clean(raw_order.order_status) if raw_order else _clean(raw_clue.order_status),
            normalized_status="verified",
            status_source="settlement_verification",
            closed_at=verified_at,
        )
    order_status = _clean(raw_order.order_status) if raw_order is not None else None
    source = "order" if order_status else "clue"
    raw_status = order_status or _clean(raw_clue.order_status)
    payload = raw_order.raw_payload if order_status and raw_order is not None else raw_clue.raw_payload
    normalized_status = resolve_clue_order_status(
        raw_status,
        payload,
        normalized_order_status=(raw_order.order_status_normalized if raw_order else None),
        coupon_statuses=coupon_statuses,
    )
    closed_at = (
        _status_observed_at(raw_clue, raw_order, now)
        if normalized_status in {"verified", "refunded", "closed"}
        else None
    )
    return StatusResolution(raw_status, normalized_status, source, closed_at)


def _resolve_anchor(
    raw_clue: RawDouyinClue,
    mappings_by_poi: dict[str, DimStorePoiMapping],
    stores_by_id: dict[str, DimStore],
) -> AnchorSnapshot:
    poi_id = _clean(raw_clue.follow_poi_id)
    if not poi_id:
        return AnchorSnapshot(None, None, "follow_poi_missing", None, None, None, None, None)
    mapping = mappings_by_poi.get(poi_id)
    if mapping is None:
        return AnchorSnapshot(poi_id, None, "follow_poi_unmapped", None, None, None, None, None)
    store = stores_by_id.get(mapping.store_id)
    if store is None:
        return AnchorSnapshot(poi_id, None, "follow_poi_store_missing", None, None, None, None, None)
    longitude = _decimal(store.longitude)
    latitude = _decimal(store.latitude)
    if not _valid_coordinates(latitude, longitude):
        return AnchorSnapshot(poi_id, store.store_id, "anchor_coordinates_invalid", None, None, None, None, None)
    province = _clean(store.standard_province)
    if not province:
        return AnchorSnapshot(poi_id, store.store_id, "anchor_province_missing", None, None, None, None, None)
    city = _clean(store.standard_city)
    if not city:
        return AnchorSnapshot(poi_id, store.store_id, "anchor_city_missing", None, None, None, None, None)
    if not _clean(store.city_code):
        return AnchorSnapshot(poi_id, store.store_id, "anchor_city_code_missing", None, None, None, None, None)
    return AnchorSnapshot(
        poi_id,
        store.store_id,
        None,
        province,
        city,
        _clean(store.city_code),
        longitude,
        latitude,
    )


def _lifecycle_status(normalized_status: str) -> str:
    return {
        "verified": "closed_verified",
        "refunded": "closed_refunded",
        "closed": "closed_order",
        "active": "active",
    }.get(normalized_status, "status_review")


def _closed_reason(normalized_status: str) -> str | None:
    return {
        "verified": "order_verified",
        "refunded": "order_refunded",
        "closed": "order_closed",
        "unknown": "order_status_unknown",
    }.get(normalized_status)


def _first_seen_at(raw_clue: RawDouyinClue, now: datetime) -> datetime:
    return _aware(raw_clue.create_time_detail) or _aware(raw_clue.fetched_at) or now


def _observed_at(raw_clue: RawDouyinClue, now: datetime) -> datetime:
    return (
        _aware(raw_clue.source_observed_at)
        or _aware(raw_clue.modify_time)
        or _aware(raw_clue.fetched_at)
        or _aware(raw_clue.updated_at)
        or now
    )


def _observation_is_newer(
    candidate_at: datetime,
    candidate_key: str | None,
    existing_at: datetime | None,
    existing_key: str | None,
) -> bool:
    if existing_at is None:
        return True
    candidate_at = _aware(candidate_at) or candidate_at
    existing_at = _aware(existing_at) or existing_at
    if candidate_at != existing_at:
        return candidate_at > existing_at
    if candidate_key is None or existing_key is None:
        return False
    return candidate_key > existing_key


def _status_observed_at(raw_clue: RawDouyinClue, raw_order: RawDouyinOrder | None, now: datetime) -> datetime:
    if raw_order is not None:
        return _aware(raw_order.updated_at) or _observed_at(raw_clue, now)
    return _observed_at(raw_clue, now)


def _record_status_event(
    session: Session,
    *,
    lead_key: str,
    order_id: str | None,
    resolution: StatusResolution,
    observed_at: datetime,
    created_at: datetime,
    known_event_ids: set[str] | None = None,
) -> None:
    event_key = "|".join(
        (
            lead_key,
            resolution.status_source,
            resolution.raw_status or "",
            resolution.normalized_status,
            observed_at.isoformat(),
        )
    )
    digest = sha256(event_key.encode("utf-8")).hexdigest()
    event_id = f"status-{digest[:24]}"
    if known_event_ids is not None:
        if event_id in known_event_ids:
            return
        known_event_ids.add(event_id)
    elif session.get(ClueOrderStatusEvent, event_id) is not None:
        return
    session.add(
        ClueOrderStatusEvent(
            event_id=event_id,
            event_key=digest,
            lead_key=lead_key,
            order_id=order_id,
            raw_status=resolution.raw_status,
            normalized_status=resolution.normalized_status,
            status_source=resolution.status_source,
            observed_at=observed_at,
            created_at=created_at,
        )
    )


def _record_anchor_quality_issue(
    session: Session,
    lead_key: str,
    anchor: AnchorSnapshot,
    now: datetime,
    known_issue_ids: set[str],
) -> None:
    if not anchor.unavailable_reason:
        return
    issue_id = f"clue-anchor:{lead_key}:{anchor.unavailable_reason}"
    if issue_id in known_issue_ids:
        return
    known_issue_ids.add(issue_id)
    session.add(
        DataQualityIssue(
            issue_id=issue_id,
            issue_type="clue_anchor_unavailable",
            message="clue anchor is unavailable for allocation",
            severity="warning",
            raw_context_json={"anchor_poi_id": anchor.poi_id, "reason": anchor.unavailable_reason},
            source_run_id=None,
            created_at=now,
        )
    )


def _record_store_location_issue(session: Session, poi_id: str, reason: str, now: datetime) -> None:
    _ = now
    upsert_data_quality_issue(
        session,
        f"store-location:{poi_id}:{reason}",
        issue_type=reason,
        message="store location import requires attention",
        severity="warning",
        raw_context_json={"poi_id": poi_id, "reason": reason},
        source_run_id=None,
        flush=False,
    )


def _active_self_owned_current_round(
    session: Session,
    lead: ClueMasterLead,
    *,
    current_rounds_by_id: dict[str, ClueAssignmentRound] | None = None,
) -> ClueAssignmentRound | None:
    if not lead.current_assignment_round_id:
        return None
    round_row = (
        current_rounds_by_id.get(lead.current_assignment_round_id)
        if current_rounds_by_id is not None
        else session.get(ClueAssignmentRound, lead.current_assignment_round_id)
    )
    if round_row is None or round_row.execution_mode != BUSINESS_EXECUTION_MODE:
        return None
    if round_row.round_status not in {"active_unfollowed", "active_followed"}:
        return None
    return round_row


def _close_current_assignment(
    session: Session,
    order_id: str,
    lifecycle_status: str,
    closed_at: datetime,
    *,
    current_assignment_round_id: str | None = None,
    center_orders_by_id: dict[str, ClueCenterOrder] | None = None,
    current_rounds_by_id: dict[str, ClueAssignmentRound] | None = None,
) -> None:
    center_order = (
        center_orders_by_id.get(order_id)
        if center_orders_by_id is not None
        else session.get(ClueCenterOrder, order_id)
    )
    round_status = _terminal_round_status(lifecycle_status)
    terminal_reason = _terminal_reason(lifecycle_status)
    lead_status = _terminal_center_lead_status(lifecycle_status)
    round_id = current_assignment_round_id or (
        center_order.current_assignment_round_id if center_order is not None else None
    )
    round_row = (
        current_rounds_by_id.get(round_id)
        if round_id and current_rounds_by_id is not None
        else (session.get(ClueAssignmentRound, round_id) if round_id else None)
    )
    if round_row is not None and round_row.execution_mode == BUSINESS_EXECUTION_MODE:
        round_row.round_status = round_status
        round_row.terminal_reason = terminal_reason
        round_row.matured_at = closed_at
        round_row.updated_at = closed_at
    if center_order is not None and (
        current_assignment_round_id is None
        or center_order.current_assignment_round_id == current_assignment_round_id
    ):
        center_order.lead_status = lead_status
        center_order.current_round_status = round_status
        center_order.reassign_reason = terminal_reason
        center_order.updated_at = closed_at


def _terminal_lifecycle_for_order(leads: list[ClueMasterLead]) -> str:
    """Choose one deterministic terminal state for an order-level snapshot."""
    lifecycle_statuses = {lead.lifecycle_status for lead in leads}
    for lifecycle_status in (
        "closed_verified",
        "closed_refunded",
        "closed_order",
        "status_review",
    ):
        if lifecycle_status in lifecycle_statuses:
            return lifecycle_status
    return "status_review"


def _terminal_round_status(lifecycle_status: str) -> str:
    return {
        "closed_verified": "closed_order_verified",
        "closed_refunded": "closed_order_refunded",
        "closed_order": "closed_order_closed",
    }.get(lifecycle_status, "closed_order_status_unknown")


def _terminal_reason(lifecycle_status: str) -> str:
    return {
        "closed_verified": "order_verified",
        "closed_refunded": "order_refunded",
        "closed_order": "order_closed",
    }.get(lifecycle_status, "order_status_unknown")


def _terminal_center_lead_status(lifecycle_status: str) -> str:
    return {
        "closed_verified": "converted",
        "closed_refunded": "refunded",
        "closed_order": "closed",
    }.get(lifecycle_status, "status_review")


def _terminal_closed_at(leads: list[ClueMasterLead], fallback: datetime) -> datetime:
    timestamps = [_aware(lead.closed_at) for lead in leads if lead.closed_at is not None]
    return min(timestamps) if timestamps else fallback


def _is_candidate_eligible(store: DimStore) -> bool:
    return bool(
        store.is_active
        and store.is_douyin_clue_applicable
        and store.participates_in_clue_allocation
        and store.location_status == "valid"
        and _clean(store.standard_province)
        and _clean(store.standard_city)
        and _clean(store.city_code)
        and _valid_coordinates(_decimal(store.latitude), _decimal(store.longitude))
    )


def _formal_store_metrics(
    session: Session,
    stores: list[DimStore],
    window_start: datetime,
    window_end: datetime,
) -> dict[str, StoreMetrics]:
    store_ids = {store.store_id for store in stores}
    rows = session.scalars(
        select(ClueAssignmentRound)
        .where(ClueAssignmentRound.execution_mode == "formal")
        .where(ClueAssignmentRound.matured_at.is_not(None))
        .where(ClueAssignmentRound.matured_at >= window_start)
        .where(ClueAssignmentRound.matured_at <= window_end)
        .where(ClueAssignmentRound.assigned_store_id.in_(store_ids))
    ).all()
    metrics_by_store: dict[str, StoreMetrics] = defaultdict(StoreMetrics)
    if not rows:
        return metrics_by_store
    round_ids = {row.assignment_round_id for row in rows}
    order_ids = {row.order_id for row in rows}
    follow_rows = session.scalars(
        select(ClueFollowUpRecord).where(ClueFollowUpRecord.assignment_round_id.in_(round_ids))
    ).all()
    follows_by_round: dict[str, list[ClueFollowUpRecord]] = defaultdict(list)
    for row in follow_rows:
        follows_by_round[row.assignment_round_id].append(row)
    verify_rows = session.execute(
        select(SettlementOrderDetail.order_id, SettlementOrderDetail.verify_time)
        .where(SettlementOrderDetail.order_id.in_(order_ids))
        .where(SettlementOrderDetail.is_verified.is_(True))
    ).all()
    verifies_by_order: dict[str, list[datetime | None]] = defaultdict(list)
    for order_id, verify_time in verify_rows:
        verifies_by_order[order_id].append(_aware(verify_time))
    all_formal_rounds_by_order: dict[str, list[ClueAssignmentRound]] = defaultdict(list)
    for formal_round in session.scalars(
        select(ClueAssignmentRound)
        .where(ClueAssignmentRound.execution_mode == "formal")
        .where(ClueAssignmentRound.order_id.in_(order_ids))
    ).all():
        all_formal_rounds_by_order[formal_round.order_id].append(formal_round)

    for round_row in rows:
        assigned_at = _aware(round_row.assigned_at)
        if not round_row.assigned_store_id or assigned_at is None:
            continue
        followed = _has_follow_within_24_hours(follows_by_round.get(round_row.assignment_round_id, []), assigned_at)
        converted = _has_verification_attributed_to_round(
            round_row,
            verifies_by_order.get(round_row.order_id, []),
            all_formal_rounds_by_order.get(round_row.order_id, []),
        )
        has_full_follow_up_opportunity = not _completed_within_24_hours(
            round_row,
            verifies_by_order.get(round_row.order_id, []),
            assigned_at,
        )
        metrics_by_store[round_row.assigned_store_id].add(
            converted=converted,
            followed_within_24h=followed,
            has_full_follow_up_opportunity=has_full_follow_up_opportunity,
        )
    return metrics_by_store


def _aggregate_city_metrics(
    stores: list[DimStore], metrics_by_store: dict[str, StoreMetrics]
) -> dict[str, StoreMetrics]:
    result: dict[str, StoreMetrics] = defaultdict(StoreMetrics)
    for store in stores:
        city_code = store.city_code or ""
        metric = metrics_by_store.get(store.store_id)
        if metric is None:
            continue
        target = result[city_code]
        target.conversion_numerator += metric.conversion_numerator
        target.conversion_denominator += metric.conversion_denominator
        target.follow_24h_numerator += metric.follow_24h_numerator
        target.follow_24h_denominator += metric.follow_24h_denominator
    return result


def _completed_within_24_hours(
    round_row: ClueAssignmentRound,
    verification_times: list[datetime | None],
    assigned_at: datetime,
) -> bool:
    cutoff = assigned_at + timedelta(hours=24)
    for candidate in [round_row.verified_at, *verification_times]:
        completed_at = _aware(candidate)
        if completed_at is not None and assigned_at <= completed_at <= cutoff:
            return True
    if round_row.terminal_reason in {"order_verified", "order_refunded"}:
        completed_at = _aware(round_row.matured_at)
        return completed_at is not None and assigned_at <= completed_at <= cutoff
    return False


def _sum_metrics(metrics: object) -> StoreMetrics:
    result = StoreMetrics()
    for metric in metrics:
        result.conversion_numerator += metric.conversion_numerator
        result.conversion_denominator += metric.conversion_denominator
        result.follow_24h_numerator += metric.follow_24h_numerator
        result.follow_24h_denominator += metric.follow_24h_denominator
    return result


def _resolved_rate(
    own_numerator: int,
    own_denominator: int,
    city_numerator: int,
    city_denominator: int,
    global_numerator: int,
    global_denominator: int,
    min_samples: int,
) -> tuple[Decimal, str]:
    for numerator, denominator, source in (
        (own_numerator, own_denominator, "store"),
        (city_numerator, city_denominator, "city"),
        (global_numerator, global_denominator, "global"),
    ):
        if denominator >= min_samples:
            return (Decimal(numerator) / Decimal(denominator)).quantize(Decimal("0.000001")), source
    return Decimal("0"), "cold_start_empty"


def _has_follow_within_24_hours(records: list[ClueFollowUpRecord], assigned_at: datetime | None) -> bool:
    assigned_at = _aware(assigned_at)
    if assigned_at is None:
        return False
    deadline = assigned_at + timedelta(hours=24)
    return any(
        (created_at := _aware(record.created_at)) is not None and assigned_at <= created_at <= deadline
        for record in records
    )


def _has_verification_attributed_to_round(
    round_row: ClueAssignmentRound,
    verify_times: list[datetime | None],
    formal_rounds: list[ClueAssignmentRound],
) -> bool:
    """Attribute each verification to the latest formal assignment effective at that time."""
    for verify_time in verify_times:
        verified_at = _aware(verify_time)
        if verified_at is None:
            continue
        effective_rounds = [
            candidate
            for candidate in formal_rounds
            if (candidate_assigned_at := _aware(candidate.assigned_at)) is not None
            and candidate_assigned_at <= verified_at
        ]
        if not effective_rounds:
            continue
        current_round = max(
            effective_rounds,
            key=lambda candidate: (
                _aware(candidate.assigned_at),
                candidate.round_no,
                candidate.assignment_round_id,
            ),
        )
        if current_round.assignment_round_id == round_row.assignment_round_id:
            return True
    return False


def _clean(value: object | None) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _aware(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    return value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)


def _decimal(value: object | None) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _valid_coordinates(latitude: Decimal | None, longitude: Decimal | None) -> bool:
    return bool(
        latitude is not None
        and longitude is not None
        and Decimal("-90") <= latitude <= Decimal("90")
        and Decimal("-180") <= longitude <= Decimal("180")
    )


def _store_location_status(store: DimStore) -> str:
    if _is_closed_store_note(store.location_status_note):
        return "closed"
    if not _valid_coordinates(_decimal(store.latitude), _decimal(store.longitude)) or not _clean(store.city_code):
        return "invalid"
    if not _clean(store.standard_city) or not _clean(store.standard_province):
        return "partial"
    return "valid"


def _clean_header(value: object | None) -> str:
    return "".join(str(value or "").strip().split())


def _cell(row: tuple[object, ...], columns: dict[str, int], name: str) -> object | None:
    index = columns.get(name)
    return row[index] if index is not None and index < len(row) else None


def _text_cell(value: object | None) -> str | None:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean(value)


def _is_closed_store_note(value: str | None) -> bool:
    note = _clean(value) or ""
    return "关闭" in note or "撤店" in note
