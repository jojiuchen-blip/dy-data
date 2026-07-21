from __future__ import annotations

from collections import Counter
import csv
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from email import policy
from email.parser import BytesParser
from hashlib import sha256
from io import BytesIO, StringIO
import json
from pathlib import Path
import re
from typing import Any
from uuid import uuid4
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, Header, HTTPException, Query, Request, Response, status
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, ValidationError
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError

from apps.api.dy_api.models import (
    DimSkuProductRule,
    SettlementScopeRule,
    SkuFeeRule,
    SkuFeeRuleImportBatch,
    SkuFeeRuleImportRow,
    utcnow,
)
from dy_api.auth import get_current_admin, get_current_super_admin
from dy_api.routes._data import generated_at, get_data_store
from dy_api.schemas import (
    SettlementScopeRuleCreateRequest,
    SkuFeeRuleCreateRequest,
    SkuFeeRuleImportCommitRequest,
    SkuProductManualUpdateRequest,
)


router = APIRouter()
SHANGHAI_TZ = ZoneInfo("Asia/Shanghai")
IMPORT_HEADERS = (
    "skuName",
    "skuId",
    "promotionServiceFeeRate",
    "managementServiceFeeRate",
)
MAX_IMPORT_BYTES = 10 * 1024 * 1024
MAX_IMPORT_ROWS = 5000
FORMAL_EFFECTIVE_START_DATE = date(2026, 8, 1)
BATCH_STATUS_NAMES = {
    1: "UPLOADED",
    2: "VALIDATION_FAILED",
    3: "PENDING_COMMIT",
    4: "COMMITTING",
    5: "COMPLETED",
    6: "FAILED",
}
ROW_STATUS_NAMES = {
    1: "PENDING",
    2: "VALID",
    3: "INVALID",
    4: "COMMITTED",
    5: "COMMIT_FAILED",
}


@router.get("/sku-products")
def list_sku_products(
    request: Request,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200, alias="pageSize"),
    q: str | None = None,
    product_scope: str | None = Query(default=None, alias="productScope"),
    product_type: str | None = Query(default=None, alias="productType"),
    product_status: str | None = Query(default=None, alias="productStatus"),
    is_active_product: bool | None = Query(default=None, alias="isActiveProduct"),
    _username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    conditions = []
    if q and q.strip():
        pattern = f"%{q.strip()}%"
        conditions.append(
            or_(
                DimSkuProductRule.sku_id.ilike(pattern),
                DimSkuProductRule.sku_name.ilike(pattern),
                DimSkuProductRule.product_name.ilike(pattern),
            )
        )
    if product_scope is not None:
        conditions.append(DimSkuProductRule.product_scope == product_scope.strip())
    if product_type is not None:
        conditions.append(DimSkuProductRule.product_type == product_type.strip())
    if product_status:
        normalized_status = product_status.strip().upper()
        if normalized_status not in {"ACTIVE", "INACTIVE", "DELETED", "UNKNOWN"}:
            raise _error(request, 422, "VALIDATION_FAILED", "商品状态不合法")
        conditions.append(DimSkuProductRule.product_status_normalized == normalized_status)
    if is_active_product is not None:
        conditions.append(DimSkuProductRule.is_active_product == is_active_product)

    total = store.session.scalar(
        select(func.count()).select_from(DimSkuProductRule).where(*conditions)
    ) or 0
    rows = list(
        store.session.scalars(
            select(DimSkuProductRule)
            .where(*conditions)
            .order_by(DimSkuProductRule.sku_id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    return _success(
        request,
        {
            "list": [_sku_product_item(row) for row in rows],
            "total": total,
            "page": page,
            "pageSize": page_size,
        },
    )


@router.put("/sku-products/{sku_id}")
def update_sku_product(
    sku_id: str,
    payload: dict[str, Any],
    request: Request,
    username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    parsed = _validate_model(SkuProductManualUpdateRequest, payload, request)
    row = store.session.scalar(
        select(DimSkuProductRule).where(DimSkuProductRule.sku_id == sku_id)
    )
    if row is None:
        raise _error(request, 404, "RESOURCE_NOT_FOUND", "SKU 不存在")
    row.product_scope = parsed.product_scope
    row.product_type = parsed.product_type
    row.is_service_product = parsed.is_service_product
    row.manual_modified_by = username
    row.manual_modified_at = utcnow()
    store.session.commit()
    store.session.refresh(row)
    return _success(request, _sku_product_item(row))


@router.get("/sku-fee-rules")
def list_sku_fee_rules(
    request: Request,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200, alias="pageSize"),
    q: str | None = None,
    sku_id: str | None = Query(default=None, alias="skuId"),
    product_scope: str | None = Query(default=None, alias="productScope"),
    product_type: str | None = Query(default=None, alias="productType"),
    rule_status: str | None = Query(default=None, alias="ruleStatus"),
    effective_date_from: date | None = Query(default=None, alias="effectiveDateFrom"),
    effective_date_to: date | None = Query(default=None, alias="effectiveDateTo"),
    as_of_date: date | None = Query(default=None, alias="asOfDate"),
    _username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    conditions = []
    if q and q.strip():
        pattern = f"%{q.strip()}%"
        conditions.append(
            or_(
                SkuFeeRule.rule_version.ilike(pattern),
                SkuFeeRule.sku_id.ilike(pattern),
                SkuFeeRule.sku_name_snapshot.ilike(pattern),
            )
        )
    if sku_id:
        conditions.append(SkuFeeRule.sku_id == sku_id.strip())
    if product_scope is not None:
        conditions.append(SkuFeeRule.product_scope_snapshot == product_scope.strip())
    if product_type is not None:
        conditions.append(SkuFeeRule.product_type_snapshot == product_type.strip())
    if rule_status:
        status_value = rule_status.strip().upper()
        if status_value not in {"ACTIVE", "INACTIVE"}:
            raise _error(request, 422, "VALIDATION_FAILED", "费率规则状态不合法")
        conditions.append(SkuFeeRule.rule_status == (1 if status_value == "ACTIVE" else 2))
    if effective_date_from:
        conditions.append(SkuFeeRule.effective_date >= effective_date_from)
    if effective_date_to:
        conditions.append(SkuFeeRule.effective_date <= effective_date_to)

    total = store.session.scalar(
        select(func.count()).select_from(SkuFeeRule).where(*conditions)
    ) or 0
    rows = list(
        store.session.scalars(
            select(SkuFeeRule)
            .where(*conditions)
            .order_by(SkuFeeRule.effective_date.desc(), SkuFeeRule.published_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    matched_versions = _matched_rule_versions(store.session, rows, as_of_date)
    return _success(
        request,
        {
            "list": [
                _sku_fee_rule_item(
                    row,
                    is_matched=(row.rule_version in matched_versions)
                    if as_of_date is not None
                    else None,
                )
                for row in rows
            ],
            "total": total,
            "page": page,
            "pageSize": page_size,
        },
    )


@router.get("/sku-fee-rules/{rule_version}")
def get_sku_fee_rule(
    rule_version: str,
    request: Request,
    _username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    row = store.session.scalar(
        select(SkuFeeRule).where(SkuFeeRule.rule_version == rule_version)
    )
    if row is None:
        raise _error(request, 404, "RESOURCE_NOT_FOUND", "费率规则版本不存在")
    return _success(request, _sku_fee_rule_item(row))


@router.post("/sku-fee-rules")
def create_sku_fee_rule(
    payload: dict[str, Any],
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    username: str = Depends(get_current_super_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    parsed = _validate_model(SkuFeeRuleCreateRequest, payload, request)
    if parsed.effective_date < FORMAL_EFFECTIVE_START_DATE:
        raise _error(
            request,
            422,
            "VALIDATION_FAILED",
            "费率规则生效日不得早于 2026-08-01",
        )
    key_hash = _idempotency_key_hash(idempotency_key, request)
    request_hash = _canonical_sha256(
        {
            "skuId": parsed.sku_id,
            "promotionServiceFeeRate": _decimal_string(parsed.promotion_service_fee_rate),
            "managementServiceFeeRate": _decimal_string(parsed.management_service_fee_rate),
            "effectiveDate": parsed.effective_date.isoformat(),
            "ruleStatus": parsed.rule_status,
            "changeReason": parsed.change_reason,
        }
    )
    idempotent = store.session.scalar(
        select(SkuFeeRule).where(
            SkuFeeRule.idempotency_key_hash == key_hash,
        )
    )
    if idempotent is not None:
        if idempotent.request_payload_sha256 != request_hash:
            raise _error(
                request,
                409,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        return _success(request, _sku_fee_rule_item(idempotent))

    sku = store.session.scalar(
        select(DimSkuProductRule).where(DimSkuProductRule.sku_id == parsed.sku_id)
    )
    if sku is None:
        raise _error(request, 404, "RESOURCE_NOT_FOUND", "SKU 不存在")
    date_conflict = store.session.scalar(
        select(SkuFeeRule.id).where(
            SkuFeeRule.sku_id == parsed.sku_id,
            SkuFeeRule.effective_date == parsed.effective_date,
        )
    )
    if date_conflict is not None:
        raise _error(
            request,
            409,
            "SKU_FEE_RULE_DATE_CONFLICT",
            "该 SKU 在所选生效日已存在规则",
        )
    previous = store.session.scalar(
        select(SkuFeeRule)
        .where(
            SkuFeeRule.sku_id == parsed.sku_id,
            SkuFeeRule.effective_date < parsed.effective_date,
        )
        .order_by(SkuFeeRule.effective_date.desc(), SkuFeeRule.published_at.desc())
        .limit(1)
    )
    now = utcnow()
    row = SkuFeeRule(
        rule_version=_new_rule_version("sfr", parsed.effective_date),
        idempotency_key_hash=key_hash,
        request_payload_sha256=request_hash,
        sku_id=parsed.sku_id,
        sku_name_snapshot=sku.sku_name,
        product_scope_snapshot=sku.product_scope,
        product_type_snapshot=sku.product_type,
        promotion_service_fee_rate=parsed.promotion_service_fee_rate,
        management_service_fee_rate=parsed.management_service_fee_rate,
        effective_date=parsed.effective_date,
        effective_at=datetime.combine(parsed.effective_date, time.min, SHANGHAI_TZ),
        rule_status=1 if parsed.rule_status == "ACTIVE" else 2,
        previous_rule_version=previous.rule_version if previous else None,
        created_by=username,
        change_reason=parsed.change_reason,
        published_at=now,
    )
    store.session.add(row)
    try:
        store.session.commit()
    except IntegrityError as exc:
        store.session.rollback()
        raise _error(
            request,
            409,
            "SKU_FEE_RULE_DATE_CONFLICT",
            "该 SKU 在所选生效日已存在规则",
        ) from exc
    store.session.refresh(row)
    return _success(request, _sku_fee_rule_item(row))


@router.get("/settlement-scope-rules")
def list_settlement_scope_rules(
    request: Request,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200, alias="pageSize"),
    effective_month: str | None = Query(default=None, alias="effectiveMonth"),
    owner_account_id: str | None = Query(default=None, alias="ownerAccountId"),
    sale_channel: str | None = Query(default=None, alias="saleChannel"),
    is_active: bool | None = Query(default=None, alias="isActive"),
    _username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    conditions = []
    if effective_month:
        conditions.append(SettlementScopeRule.effective_month == effective_month)
    if owner_account_id:
        conditions.append(SettlementScopeRule.owner_account_id == owner_account_id)
    if sale_channel:
        normalized_channel = sale_channel.strip().upper()
        if normalized_channel not in {"LIVE", "SHORT_VIDEO"}:
            raise _error(request, 422, "VALIDATION_FAILED", "销售渠道不合法")
        conditions.append(
            SettlementScopeRule.sale_channel_normalized == normalized_channel.lower()
        )
    if is_active is not None:
        conditions.append(SettlementScopeRule.is_active == is_active)
    total = store.session.scalar(
        select(func.count()).select_from(SettlementScopeRule).where(*conditions)
    ) or 0
    rows = list(
        store.session.scalars(
            select(SettlementScopeRule)
            .where(*conditions)
            .order_by(
                SettlementScopeRule.effective_month.desc(),
                SettlementScopeRule.owner_account_id,
                SettlementScopeRule.sale_channel_normalized,
            )
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    return _success(
        request,
        {
            "list": [_settlement_scope_item(row) for row in rows],
            "total": total,
            "page": page,
            "pageSize": page_size,
        },
    )


@router.post("/settlement-scope-rules")
def create_settlement_scope_rules(
    payload: dict[str, Any],
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    username: str = Depends(get_current_super_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    parsed = _validate_model(SettlementScopeRuleCreateRequest, payload, request)
    if parsed.effective_month < "2026-08":
        raise _error(
            request,
            422,
            "VALIDATION_FAILED",
            "结算范围生效月份不得早于 2026-08",
        )
    key_hash = _idempotency_key_hash(idempotency_key, request)
    channels = list(parsed.allowed_sale_channels)
    request_hash = _canonical_sha256(
        {
            "effectiveMonth": parsed.effective_month,
            "ownerAccountId": parsed.owner_account_id,
            "allowedSaleChannels": channels,
            "changeReason": parsed.change_reason,
        }
    )
    existing_key_rows = list(
        store.session.scalars(
            select(SettlementScopeRule)
            .where(SettlementScopeRule.idempotency_key_hash == key_hash)
            .order_by(SettlementScopeRule.id)
        )
    )
    if existing_key_rows:
        if any(row.request_payload_sha256 != request_hash for row in existing_key_rows):
            raise _error(
                request,
                409,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        return _success(
            request,
            {
                "scopeRuleVersions": [row.scope_rule_version for row in existing_key_rows],
                "effectiveMonth": parsed.effective_month,
                "ownerAccountId": parsed.owner_account_id,
                "allowedSaleChannels": [
                    row.sale_channel_normalized.upper() for row in existing_key_rows
                ],
            },
        )
    owner_exists = store.session.scalar(
        select(func.count())
        .select_from(DimSkuProductRule)
        .where(DimSkuProductRule.owner_account_id == parsed.owner_account_id)
    )
    if not owner_exists:
        raise _error(
            request,
            422,
            "OWNER_ACCOUNT_NOT_FOUND",
            "归属账号稳定 ID 不存在于商品事实源",
        )
    normalized_channels = [channel.lower() for channel in channels]
    slot_conflict = store.session.scalar(
        select(SettlementScopeRule.id).where(
            SettlementScopeRule.effective_month == parsed.effective_month,
            SettlementScopeRule.owner_account_id == parsed.owner_account_id,
            SettlementScopeRule.sale_channel_normalized.in_(normalized_channels),
        )
    )
    if slot_conflict is not None:
        raise _error(
            request,
            409,
            "SETTLEMENT_SCOPE_SLOT_CONFLICT",
            "同月、同归属账号和渠道已存在范围规则",
        )
    rows = []
    for channel, normalized_channel in zip(channels, normalized_channels):
        row = SettlementScopeRule(
            scope_rule_version=_new_scope_version(parsed.effective_month, channel),
            idempotency_key_hash=key_hash,
            request_payload_sha256=request_hash,
            effective_month=parsed.effective_month,
            owner_account_id=parsed.owner_account_id,
            sale_channel_normalized=normalized_channel,
            is_active=True,
            created_by=username,
            change_reason=parsed.change_reason,
        )
        store.session.add(row)
        rows.append(row)
    try:
        store.session.commit()
    except IntegrityError as exc:
        store.session.rollback()
        raise _error(
            request,
            409,
            "SETTLEMENT_SCOPE_SLOT_CONFLICT",
            "同月、同归属账号和渠道已存在范围规则",
        ) from exc
    return _success(
        request,
        {
            "scopeRuleVersions": [row.scope_rule_version for row in rows],
            "effectiveMonth": parsed.effective_month,
            "ownerAccountId": parsed.owner_account_id,
            "allowedSaleChannels": channels,
        },
    )


@router.get("/sku-fee-rule-imports/template")
def download_sku_fee_rule_import_template(
    request: Request,
    _username: str = Depends(get_current_admin),
) -> Response:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(IMPORT_HEADERS)
    content = ("\ufeff" + output.getvalue()).encode("utf-8")
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="sku-fee-rule-template.csv"',
            "X-Request-ID": _request_id(request),
        },
    )


@router.get("/sku-fee-rule-imports")
def list_sku_fee_rule_imports(
    request: Request,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200, alias="pageSize"),
    batch_status: str | None = Query(default=None, alias="batchStatus"),
    effective_date_from: date | None = Query(default=None, alias="effectiveDateFrom"),
    effective_date_to: date | None = Query(default=None, alias="effectiveDateTo"),
    uploaded_by: str | None = Query(default=None, alias="uploadedBy"),
    _username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    conditions = []
    if batch_status:
        reverse_statuses = {value: key for key, value in BATCH_STATUS_NAMES.items()}
        status_value = reverse_statuses.get(batch_status.strip().upper())
        if status_value is None:
            raise _error(request, 422, "VALIDATION_FAILED", "导入批次状态不合法")
        conditions.append(SkuFeeRuleImportBatch.batch_status == status_value)
    if effective_date_from:
        conditions.append(SkuFeeRuleImportBatch.effective_date >= effective_date_from)
    if effective_date_to:
        conditions.append(SkuFeeRuleImportBatch.effective_date <= effective_date_to)
    if uploaded_by:
        conditions.append(SkuFeeRuleImportBatch.uploaded_by == uploaded_by.strip())
    total = store.session.scalar(
        select(func.count()).select_from(SkuFeeRuleImportBatch).where(*conditions)
    ) or 0
    rows = list(
        store.session.scalars(
            select(SkuFeeRuleImportBatch)
            .where(*conditions)
            .order_by(SkuFeeRuleImportBatch.created_at.desc(), SkuFeeRuleImportBatch.id.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    return _success(
        request,
        {
            "list": [_import_batch_item(row) for row in rows],
            "total": total,
            "page": page,
            "pageSize": page_size,
        },
    )


@router.post("/sku-fee-rule-imports")
async def upload_sku_fee_rule_import(
    request: Request,
    username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    filename, content, effective_date_text = await _multipart_import_payload(request)
    try:
        effective_date = date.fromisoformat(effective_date_text)
    except (TypeError, ValueError) as exc:
        raise _error(
            request, 422, "VALIDATION_FAILED", "effectiveDate 必须为 YYYY-MM-DD"
        ) from exc
    if effective_date < FORMAL_EFFECTIVE_START_DATE:
        raise _error(
            request,
            422,
            "VALIDATION_FAILED",
            "导入批次生效日不得早于 2026-08-01",
        )
    safe_filename = _safe_filename(filename)
    extension = Path(safe_filename).suffix.lower()
    if extension not in {".csv", ".xlsx"}:
        raise _error(request, 422, "VALIDATION_FAILED", "仅支持 .xlsx 或 UTF-8 .csv")
    if len(content) > MAX_IMPORT_BYTES:
        raise _error(request, 422, "IMPORT_FILE_TOO_LARGE", "导入文件不能超过 10 MiB")
    try:
        source_rows = _read_import_rows(content, extension)
    except ValueError as exc:
        batch_id = f"fee-import-{uuid4().hex}"
        validation_error = _row_error(
            "template", "IMPORT_TEMPLATE_INVALID", str(exc)
        )
        batch = SkuFeeRuleImportBatch(
            batch_id=batch_id,
            file_name=safe_filename,
            file_sha256=sha256(content).hexdigest(),
            batch_status=2,
            commit_mode=1,
            effective_date=effective_date,
            total_count=0,
            valid_count=0,
            success_count=0,
            failed_count=1,
            uploaded_by=username,
            validated_at=utcnow(),
            result_file_key=f"fee-import-results/{batch_id}{extension}",
        )
        row = SkuFeeRuleImportRow(
            batch_id=batch_id,
            row_number=1,
            validation_status=3,
            error_count=1,
            error_field="template",
            error_code="IMPORT_TEMPLATE_INVALID",
            error_message=str(exc),
            validation_errors_json=[validation_error],
            source_row_json={},
        )
        store.session.add_all([batch, row])
        store.session.commit()
        return _success(
            request,
            {
                "batch": _import_batch_item(batch),
                "errorPreview": [_import_row_item(row)],
                "hasMoreErrors": False,
            },
        )
    if len(source_rows) > MAX_IMPORT_ROWS:
        raise _error(request, 422, "IMPORT_ROW_LIMIT_EXCEEDED", "导入数据不能超过 5000 行")

    batch_id = f"fee-import-{uuid4().hex}"
    batch = SkuFeeRuleImportBatch(
        batch_id=batch_id,
        file_name=safe_filename,
        file_sha256=sha256(content).hexdigest(),
        batch_status=1,
        commit_mode=1,
        effective_date=effective_date,
        total_count=len(source_rows),
        valid_count=0,
        success_count=0,
        failed_count=0,
        uploaded_by=username,
        result_file_key=f"fee-import-results/{batch_id}{extension}",
    )
    store.session.add(batch)
    sku_ids = [_clean_text(row.get("skuId")) for row in source_rows]
    sku_id_counts = Counter(sku_ids)
    duplicate_skus = {
        sku_id for sku_id, count in sku_id_counts.items() if sku_id and count > 1
    }
    unique_sku_ids = sorted({sku_id for sku_id in sku_ids if sku_id})
    skus = {
        row.sku_id: row
        for row in store.session.scalars(
            select(DimSkuProductRule).where(DimSkuProductRule.sku_id.in_(unique_sku_ids))
        )
    } if unique_sku_ids else {}
    conflicts = set(
        store.session.scalars(
            select(SkuFeeRule.sku_id).where(
                SkuFeeRule.sku_id.in_(unique_sku_ids),
                SkuFeeRule.effective_date == effective_date,
            )
        )
    ) if unique_sku_ids else set()
    row_models: list[SkuFeeRuleImportRow] = []
    failed_count = 0
    for row_number, source in enumerate(source_rows, start=2):
        row_model = _validate_import_row(
            batch_id=batch_id,
            row_number=row_number,
            source=source,
            skus=skus,
            duplicate_skus=duplicate_skus,
            database_conflicts=conflicts,
        )
        if row_model.validation_status == 3:
            failed_count += 1
        row_models.append(row_model)
        store.session.add(row_model)
    batch.failed_count = failed_count
    batch.valid_count = len(row_models) - failed_count
    batch.batch_status = 2 if failed_count else 3
    batch.validated_at = utcnow()
    store.session.commit()
    errors = [row for row in row_models if row.validation_status == 3]
    return _success(
        request,
        {
            "batch": _import_batch_item(batch),
            "errorPreview": [_import_row_item(row) for row in errors[:100]],
            "hasMoreErrors": len(errors) > 100,
        },
    )


@router.get("/sku-fee-rule-imports/{batch_id}")
def get_sku_fee_rule_import(
    batch_id: str,
    request: Request,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=200, alias="pageSize"),
    validation_status: str | None = Query(default=None, alias="validationStatus"),
    _username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    batch = _get_import_batch(store.session, batch_id, request)
    conditions = [SkuFeeRuleImportRow.batch_id == batch_id]
    if validation_status:
        reverse_statuses = {value: key for key, value in ROW_STATUS_NAMES.items()}
        status_value = reverse_statuses.get(validation_status.strip().upper())
        if status_value is None:
            raise _error(request, 422, "VALIDATION_FAILED", "逐行校验状态不合法")
        conditions.append(SkuFeeRuleImportRow.validation_status == status_value)
    total = store.session.scalar(
        select(func.count()).select_from(SkuFeeRuleImportRow).where(*conditions)
    ) or 0
    rows = list(
        store.session.scalars(
            select(SkuFeeRuleImportRow)
            .where(*conditions)
            .order_by(SkuFeeRuleImportRow.row_number)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    return _success(
        request,
        {
            "batch": _import_batch_item(batch),
            "rows": {
                "list": [_import_row_item(row) for row in rows],
                "total": total,
                "page": page,
                "pageSize": page_size,
            },
        },
    )


@router.post("/sku-fee-rule-imports/{batch_id}/commit")
def commit_sku_fee_rule_import(
    batch_id: str,
    payload: dict[str, Any],
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    username: str = Depends(get_current_super_admin),
    store=Depends(get_data_store),
):
    store = _require_store(store, request)
    parsed = _validate_model(SkuFeeRuleImportCommitRequest, payload, request)
    key_hash = _idempotency_key_hash(idempotency_key, request)
    request_hash = _canonical_sha256({"changeReason": parsed.change_reason})
    batch = _get_import_batch(store.session, batch_id, request)
    if batch.commit_idempotency_key_hash:
        if batch.commit_idempotency_key_hash != key_hash:
            raise _error(
                request,
                409,
                "IMPORT_BATCH_STATE_CONFLICT",
                "该批次已使用其他幂等键提交",
            )
        if batch.commit_payload_sha256 != request_hash:
            raise _error(
                request,
                409,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        if batch.batch_status == 5:
            versions = list(
                store.session.scalars(
                    select(SkuFeeRuleImportRow.created_rule_version)
                    .where(SkuFeeRuleImportRow.batch_id == batch_id)
                    .order_by(SkuFeeRuleImportRow.row_number)
                )
            )
            return _success(
                request,
                {
                    "batch": _import_batch_item(batch),
                    "createdRuleVersions": [version for version in versions if version],
                },
            )
    if batch.batch_status != 3:
        raise _error(
            request,
            409,
            "IMPORT_BATCH_STATE_CONFLICT",
            "仅待确认批次可以提交",
        )
    reused_key = store.session.scalar(
        select(SkuFeeRuleImportBatch.batch_id).where(
            SkuFeeRuleImportBatch.commit_idempotency_key_hash == key_hash,
            SkuFeeRuleImportBatch.batch_id != batch_id,
        )
    )
    if reused_key is not None:
        raise _error(
            request,
            409,
            "IDEMPOTENCY_KEY_REUSED",
            "Idempotency-Key 已用于其他导入批次",
        )
    rows = list(
        store.session.scalars(
            select(SkuFeeRuleImportRow)
            .where(SkuFeeRuleImportRow.batch_id == batch_id)
            .order_by(SkuFeeRuleImportRow.row_number)
        )
    )
    if not rows or any(row.validation_status != 2 for row in rows):
        raise _error(
            request,
            409,
            "IMPORT_BATCH_STATE_CONFLICT",
            "批次存在未通过预校验的行",
        )
    sku_ids = [row.sku_id for row in rows if row.sku_id]
    database_conflicts = list(
        store.session.scalars(
            select(SkuFeeRule.sku_id).where(
                SkuFeeRule.sku_id.in_(sku_ids),
                SkuFeeRule.effective_date == batch.effective_date,
            )
        )
    )
    if database_conflicts:
        batch.batch_status = 6
        batch.success_count = 0
        batch.commit_idempotency_key_hash = key_hash
        batch.commit_payload_sha256 = request_hash
        store.session.commit()
        raise _error(
            request,
            409,
            "SKU_FEE_RULE_DATE_CONFLICT",
            "提交期间发现同 SKU 和生效日规则冲突，整批未写入",
            errors=[
                {
                    "rowNumber": row.row_number,
                    "field": "effectiveDate",
                    "reason": f"SKU {row.sku_id} 在 {batch.effective_date.isoformat()} 已存在规则",
                }
                for row in rows
                if row.sku_id in set(database_conflicts)
            ],
        )
    skus = {
        sku.sku_id: sku
        for sku in store.session.scalars(
            select(DimSkuProductRule).where(DimSkuProductRule.sku_id.in_(sku_ids))
        )
    }
    if len(skus) != len(set(sku_ids)):
        raise _error(request, 409, "IMPORT_DATA_CHANGED", "SKU 事实源在提交前发生变化")
    batch.batch_status = 4
    batch.commit_idempotency_key_hash = key_hash
    batch.commit_payload_sha256 = request_hash
    now = utcnow()
    created_versions: list[str] = []
    try:
        with store.session.begin_nested():
            for import_row in rows:
                assert import_row.sku_id is not None
                sku = skus[import_row.sku_id]
                previous = store.session.scalar(
                    select(SkuFeeRule)
                    .where(
                        SkuFeeRule.sku_id == import_row.sku_id,
                        SkuFeeRule.effective_date < batch.effective_date,
                    )
                    .order_by(
                        SkuFeeRule.effective_date.desc(),
                        SkuFeeRule.published_at.desc(),
                    )
                    .limit(1)
                )
                rule_version = _new_rule_version("sfr", batch.effective_date)
                store.session.add(
                    SkuFeeRule(
                        rule_version=rule_version,
                        idempotency_key_hash=key_hash,
                        request_payload_sha256=_canonical_sha256(
                            {
                                "batchId": batch_id,
                                "rowNumber": import_row.row_number,
                                "skuId": import_row.sku_id,
                                "changeReason": parsed.change_reason,
                            }
                        ),
                        sku_id=import_row.sku_id,
                        sku_name_snapshot=sku.sku_name,
                        product_scope_snapshot=sku.product_scope,
                        product_type_snapshot=sku.product_type,
                        promotion_service_fee_rate=import_row.promotion_service_fee_rate,
                        management_service_fee_rate=import_row.management_service_fee_rate,
                        effective_date=batch.effective_date,
                        effective_at=datetime.combine(
                            batch.effective_date, time.min, SHANGHAI_TZ
                        ),
                        rule_status=1,
                        previous_rule_version=previous.rule_version if previous else None,
                        created_by=username,
                        change_reason=parsed.change_reason,
                        published_at=now,
                    )
                )
                import_row.validation_status = 4
                import_row.created_rule_version = rule_version
                created_versions.append(rule_version)
            store.session.flush()
    except IntegrityError as exc:
        batch.batch_status = 6
        batch.success_count = 0
        for import_row in rows:
            import_row.validation_status = 5
            import_row.created_rule_version = None
        store.session.commit()
        raise _error(
            request,
            409,
            "SKU_FEE_RULE_DATE_CONFLICT",
            "提交期间发生唯一冲突，整批已回滚",
        ) from exc
    batch.batch_status = 5
    batch.success_count = len(rows)
    batch.failed_count = 0
    batch.committed_at = now
    store.session.commit()
    return _success(
        request,
        {
            "batch": _import_batch_item(batch),
            "createdRuleVersions": created_versions,
        },
    )


@router.get("/sku-fee-rule-imports/{batch_id}/result-file")
def download_sku_fee_rule_import_result(
    batch_id: str,
    request: Request,
    _username: str = Depends(get_current_admin),
    store=Depends(get_data_store),
) -> Response:
    store = _require_store(store, request)
    batch = _get_import_batch(store.session, batch_id, request)
    if not batch.result_file_key:
        raise _error(request, 404, "RESOURCE_NOT_FOUND", "该批次没有结果文件")
    rows = list(
        store.session.scalars(
            select(SkuFeeRuleImportRow)
            .where(SkuFeeRuleImportRow.batch_id == batch_id)
            .order_by(SkuFeeRuleImportRow.row_number)
        )
    )
    extension = Path(batch.file_name).suffix.lower()
    content, media_type = _render_result_file(rows, extension)
    filename = f"{Path(batch.file_name).stem}-result{extension}"
    return Response(
        content=content,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "X-Request-ID": _request_id(request),
        },
    )


def _require_store(store, request: Request):
    if not store.available:
        raise _error(request, 503, "DATABASE_UNAVAILABLE", "数据库暂不可用")
    return store


def _success(request: Request, data: Any) -> dict[str, Any]:
    return {
        "data": data,
        "meta": {
            "generatedAt": generated_at(),
            "source": "postgres",
            "requestId": _request_id(request),
        },
    }


def _error(
    request: Request,
    http_status: int,
    code: str,
    message: str,
    *,
    errors: list[dict[str, Any]] | None = None,
) -> HTTPException:
    return HTTPException(
        status_code=http_status,
        detail={
            "code": code,
            "message": message,
            "errors": errors or [],
            "requestId": _request_id(request),
        },
    )


def _request_id(request: Request) -> str:
    existing = getattr(request.state, "fee_admin_request_id", None)
    if existing:
        return existing
    provided = (request.headers.get("X-Request-ID") or "").strip()
    if provided and re.fullmatch(r"[A-Za-z0-9._:-]{1,128}", provided):
        request_id = provided
    else:
        request_id = f"req_{uuid4().hex}"
    request.state.fee_admin_request_id = request_id
    return request_id


def _validate_model(
    model: type[BaseModel], payload: dict[str, Any], request: Request
) -> BaseModel:
    try:
        return model.model_validate(payload)
    except ValidationError as exc:
        errors = []
        for item in exc.errors(include_url=False):
            location = item.get("loc") or ()
            field = str(location[-1]) if location else "request"
            errors.append(
                {
                    "field": field,
                    "reason": str(item.get("msg") or "字段不合法"),
                }
            )
        raise _error(
            request,
            422,
            "VALIDATION_FAILED",
            "请求字段校验失败",
            errors=errors,
        ) from exc


def _idempotency_key_hash(value: str | None, request: Request) -> str:
    normalized = (value or "").strip()
    if not 16 <= len(normalized) <= 128:
        raise _error(
            request,
            422,
            "VALIDATION_FAILED",
            "Idempotency-Key 长度必须为 16～128",
        )
    return sha256(normalized.encode("utf-8")).hexdigest()


def _canonical_sha256(value: Any) -> str:
    encoded = json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return sha256(encoded).hexdigest()


def _decimal_string(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return f"{value.quantize(Decimal('0.000001')):.6f}"


def _shanghai_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=SHANGHAI_TZ)
    return value.astimezone(SHANGHAI_TZ)


def _utc_to_shanghai(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(SHANGHAI_TZ)


def _sku_product_item(row: DimSkuProductRule) -> dict[str, Any]:
    return {
        "skuId": row.sku_id,
        "skuName": row.sku_name,
        "productId": row.product_id,
        "productName": row.product_name,
        "spuId": row.spu_id,
        "productScope": row.product_scope,
        "productType": row.product_type,
        "isServiceProduct": row.is_service_product,
        "creatorAccountId": row.creator_account_id,
        "creatorAccountName": row.creator_account_name,
        "ownerAccountId": row.owner_account_id,
        "ownerAccountName": row.owner_account_name,
        "productStatus": row.product_status_normalized,
        "isActiveProduct": row.is_active_product,
        "lastSyncedAt": _utc_to_shanghai(row.last_synced_at),
        "manualModifiedAt": _utc_to_shanghai(row.manual_modified_at),
    }


def _sku_fee_rule_item(
    row: SkuFeeRule,
    *,
    is_matched: bool | None = None,
) -> dict[str, Any]:
    result = {
        "ruleVersion": row.rule_version,
        "skuId": row.sku_id,
        "skuName": row.sku_name_snapshot,
        "productScope": row.product_scope_snapshot,
        "productType": row.product_type_snapshot,
        "promotionServiceFeeRate": _decimal_string(row.promotion_service_fee_rate),
        "managementServiceFeeRate": _decimal_string(row.management_service_fee_rate),
        "effectiveDate": row.effective_date,
        "effectiveAt": _shanghai_datetime(row.effective_at),
        "ruleStatus": "ACTIVE" if row.rule_status == 1 else "INACTIVE",
        "previousRuleVersion": row.previous_rule_version,
        "createdBy": row.created_by,
        "changeReason": row.change_reason,
        "publishedAt": _utc_to_shanghai(row.published_at),
    }
    if is_matched is not None:
        result["isMatchedVersion"] = is_matched
    return result


def _matched_rule_versions(
    session,
    visible_rows: list[SkuFeeRule],
    as_of_date: date | None,
) -> set[str]:
    if as_of_date is None:
        return set()
    matched: set[str] = set()
    for sku_id in {row.sku_id for row in visible_rows}:
        row = session.scalar(
            select(SkuFeeRule)
            .where(
                SkuFeeRule.sku_id == sku_id,
                SkuFeeRule.rule_status == 1,
                SkuFeeRule.effective_date <= as_of_date,
            )
            .order_by(SkuFeeRule.effective_date.desc(), SkuFeeRule.published_at.desc())
            .limit(1)
        )
        if row is not None:
            matched.add(row.rule_version)
    return matched


def _settlement_scope_item(row: SettlementScopeRule) -> dict[str, Any]:
    return {
        "scopeRuleVersion": row.scope_rule_version,
        "effectiveMonth": row.effective_month,
        "ownerAccountId": row.owner_account_id,
        "saleChannel": row.sale_channel_normalized.upper(),
        "isActive": row.is_active,
        "createdBy": row.created_by,
        "changeReason": row.change_reason,
        "createdAt": _utc_to_shanghai(row.created_at),
    }


def _new_rule_version(prefix: str, effective_date: date) -> str:
    return f"{prefix}_{effective_date:%Y%m%d}_{uuid4().hex[:20]}"


def _new_scope_version(effective_month: str, channel: str) -> str:
    return f"ssr_{effective_month.replace('-', '')}_{channel.lower()}_{uuid4().hex[:16]}"


async def _multipart_import_payload(request: Request) -> tuple[str, bytes, str]:
    content_type = request.headers.get("content-type") or ""
    if "multipart/form-data" not in content_type.lower():
        raise _error(
            request, 422, "VALIDATION_FAILED", "Content-Type 必须为 multipart/form-data"
        )
    body = await request.body()
    if len(body) > MAX_IMPORT_BYTES + (1024 * 1024):
        raise _error(request, 422, "IMPORT_FILE_TOO_LARGE", "导入文件不能超过 10 MiB")
    message = BytesParser(policy=policy.default).parsebytes(
        (
            f"Content-Type: {content_type}\r\n"
            "MIME-Version: 1.0\r\n\r\n"
        ).encode("utf-8")
        + body
    )
    filename: str | None = None
    file_content: bytes | None = None
    effective_date: str | None = None
    for part in message.iter_parts():
        field_name = part.get_param("name", header="content-disposition")
        if field_name == "file":
            filename = part.get_filename()
            file_content = part.get_payload(decode=True) or b""
        elif field_name == "effectiveDate":
            payload = part.get_payload(decode=True) or b""
            effective_date = payload.decode(part.get_content_charset() or "utf-8").strip()
    if not filename or file_content is None or not effective_date:
        raise _error(
            request,
            422,
            "VALIDATION_FAILED",
            "file 与 effectiveDate 均为必填字段",
        )
    return filename, file_content, effective_date


def _safe_filename(value: str) -> str:
    normalized = value.replace("\\", "/").split("/")[-1].strip()
    return normalized[:512] or "fee-rules.csv"


def _read_import_rows(content: bytes, extension: str) -> list[dict[str, Any]]:
    if extension == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV 必须使用 UTF-8 编码") from exc
        reader = csv.DictReader(StringIO(text))
        if tuple(reader.fieldnames or ()) != IMPORT_HEADERS:
            raise ValueError("导入模板必须且只能包含四个标准业务列")
        rows = [dict(row) for row in reader]
        if not rows:
            raise ValueError("导入文件至少包含一条数据行")
        return rows
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - malformed workbook boundary.
        raise ValueError("XLSX 文件无法解析") from exc
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if tuple(_clean_text(value) for value in (header_row or ())) != IMPORT_HEADERS:
        raise ValueError("导入模板必须且只能包含四个标准业务列")
    rows = [dict(zip(IMPORT_HEADERS, row)) for row in iterator]
    if not rows:
        raise ValueError("导入文件至少包含一条数据行")
    return rows


def _validate_import_row(
    *,
    batch_id: str,
    row_number: int,
    source: dict[str, Any],
    skus: dict[str, DimSkuProductRule],
    duplicate_skus: set[str],
    database_conflicts: set[str],
) -> SkuFeeRuleImportRow:
    sku_name = _clean_text(source.get("skuName"))
    sku_id = _clean_text(source.get("skuId"))
    errors: list[dict[str, str]] = []
    if not sku_name:
        errors.append(_row_error("skuName", "REQUIRED", "SKU 名称不能为空"))
    if not sku_id:
        errors.append(_row_error("skuId", "REQUIRED", "SKU ID 不能为空"))
    sku = skus.get(sku_id) if sku_id else None
    if sku_id and sku is None:
        errors.append(_row_error("skuId", "SKU_NOT_FOUND", "SKU ID 不存在于商品事实源"))
    if sku is not None and sku_name and (sku.sku_name or "").strip() != sku_name:
        errors.append(_row_error("skuName", "SKU_NAME_MISMATCH", "SKU 名称与 SKU ID 不匹配"))
    if sku_id and sku_id in duplicate_skus:
        errors.append(_row_error("skuId", "DUPLICATE_SKU", "同一批次内 SKU ID 重复"))
    if sku_id and sku_id in database_conflicts:
        errors.append(
            _row_error("skuId", "SKU_FEE_RULE_DATE_CONFLICT", "该 SKU 生效日已存在规则")
        )
    promotion_rate, promotion_errors = _parse_import_rate(
        source.get("promotionServiceFeeRate"), "promotionServiceFeeRate", "推广服务费率"
    )
    management_rate, management_errors = _parse_import_rate(
        source.get("managementServiceFeeRate"), "managementServiceFeeRate", "管理服务费率"
    )
    errors.extend(promotion_errors)
    errors.extend(management_errors)
    first_error = errors[0] if errors else None
    return SkuFeeRuleImportRow(
        batch_id=batch_id,
        row_number=row_number,
        sku_name=sku_name,
        sku_id=sku_id,
        promotion_service_fee_rate=promotion_rate,
        management_service_fee_rate=management_rate,
        validation_status=3 if errors else 2,
        error_count=len(errors),
        error_field=first_error["field"] if first_error else None,
        error_code=first_error["code"] if first_error else None,
        error_message=first_error["message"] if first_error else None,
        validation_errors_json=errors,
        source_row_json={key: _clean_text(source.get(key)) for key in IMPORT_HEADERS},
    )


def _parse_import_rate(
    value: Any,
    field: str,
    label: str,
) -> tuple[Decimal | None, list[dict[str, str]]]:
    text = _clean_text(value)
    if not text:
        return None, [_row_error(field, "REQUIRED", f"{label}不能为空")]
    try:
        rate = Decimal(text)
    except (InvalidOperation, ValueError):
        return None, [_row_error(field, "INVALID_DECIMAL", f"{label}必须为数字")]
    if not rate.is_finite():
        return None, [_row_error(field, "INVALID_DECIMAL", f"{label}必须为有限数字")]
    errors = []
    if rate < 0 or rate > 1:
        errors.append(_row_error(field, "RATE_OUT_OF_RANGE", f"{label}必须在 0～1 之间"))
    if max(0, -rate.as_tuple().exponent) > 6:
        errors.append(_row_error(field, "RATE_SCALE_EXCEEDED", f"{label}最多六位小数"))
    return (rate if not errors else None), errors


def _row_error(field: str, code: str, message: str) -> dict[str, str]:
    return {"field": field, "code": code, "message": message}


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _import_batch_item(row: SkuFeeRuleImportBatch) -> dict[str, Any]:
    return {
        "batchId": row.batch_id,
        "fileName": row.file_name,
        "batchStatus": BATCH_STATUS_NAMES[row.batch_status],
        "commitMode": "ATOMIC",
        "effectiveDate": row.effective_date,
        "totalCount": row.total_count,
        "validCount": row.valid_count,
        "successCount": row.success_count,
        "failedCount": row.failed_count,
        "uploadedBy": row.uploaded_by,
        "validatedAt": _utc_to_shanghai(row.validated_at),
        "committedAt": _utc_to_shanghai(row.committed_at),
        "hasResultFile": bool(row.result_file_key),
    }


def _import_row_item(row: SkuFeeRuleImportRow) -> dict[str, Any]:
    return {
        "rowNumber": row.row_number,
        "skuName": row.sku_name,
        "skuId": row.sku_id,
        "promotionServiceFeeRate": _decimal_string(row.promotion_service_fee_rate),
        "managementServiceFeeRate": _decimal_string(row.management_service_fee_rate),
        "validationStatus": ROW_STATUS_NAMES[row.validation_status],
        "errors": row.validation_errors_json or [],
        "createdRuleVersion": row.created_rule_version,
    }


def _get_import_batch(session, batch_id: str, request: Request) -> SkuFeeRuleImportBatch:
    batch = session.scalar(
        select(SkuFeeRuleImportBatch).where(SkuFeeRuleImportBatch.batch_id == batch_id)
    )
    if batch is None:
        raise _error(request, 404, "RESOURCE_NOT_FOUND", "导入批次不存在")
    return batch


def _render_result_file(
    rows: list[SkuFeeRuleImportRow], extension: str
) -> tuple[bytes, str]:
    headers = (
        *IMPORT_HEADERS,
        "validationStatus",
        "errorFields",
        "errorMessages",
        "createdRuleVersion",
    )
    rendered_rows = []
    for row in rows:
        errors = row.validation_errors_json or []
        rendered_rows.append(
            (
                row.sku_name or "",
                row.sku_id or "",
                _decimal_string(row.promotion_service_fee_rate) or "",
                _decimal_string(row.management_service_fee_rate) or "",
                ROW_STATUS_NAMES[row.validation_status],
                "|".join(str(error.get("field") or "") for error in errors),
                "|".join(str(error.get("message") or "") for error in errors),
                row.created_rule_version or "",
            )
        )
    if extension == ".xlsx":
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(list(headers))
        for rendered in rendered_rows:
            sheet.append(list(rendered))
        output = BytesIO()
        workbook.save(output)
        return (
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rendered_rows)
    return ("\ufeff" + output.getvalue()).encode("utf-8"), "text/csv; charset=utf-8"
