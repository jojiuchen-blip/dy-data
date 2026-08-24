from __future__ import annotations

import json
import os
from csv import DictReader, DictWriter
from base64 import urlsafe_b64encode
from datetime import date, datetime, timezone
from hashlib import sha256
from io import StringIO, TextIOWrapper
from urllib.parse import quote
from uuid import uuid4
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, Request, UploadFile, status
from fastapi.responses import Response, StreamingResponse
from cryptography.fernet import Fernet, InvalidToken
from openpyxl import load_workbook
from sqlalchemy import and_, func, select, text, update
from sqlalchemy.exc import IntegrityError

from dy_api.auth import AuthContext, get_current_user
from dy_api.routes._data import (
    ReportingPermissionError,
    ReportingValidationError,
    camelize,
    generated_at,
    get_data_store,
    request_id,
    with_utf8_bom,
)
from dy_api.schemas import (
    CommissionRulesSummaryData,
    OrderDetailsData,
    SalesDashboardData,
    dump_model,
)
from apps.api.dy_api.models import (
    DimSkuProductRule,
    DimStore,
    FinanceImportBatch,
    FinanceImportRow,
    FinanceOperationAudit,
    InvoiceRecord,
    InvoiceStatusEvent,
    ManagementCarryforwardApplication,
    PromotionInvoice,
    PromotionInvoiceAllocation,
    PromotionInvoiceLifecycleEvent,
    PromotionInvoiceNumberRegistry,
    PromotionInvoiceReplacementSource,
    RawDouyinOrder,
    RawDouyinOrderCoupon,
    RawDouyinVerifyRecord,
    SapSuggestion,
    SettlementCarryforwardApplication,
    SettlementDispute,
    SettlementDisputeOrder,
    SettlementFeeAdjustment,
    SettlementFeeResult,
    SettlementStatement,
    SettlementStatementConfirmation,
    SettlementStatementEntry,
    SettlementStatementLine,
    StoreFinanceProfile,
    utcnow,
)


router = APIRouter()

FORMAL_PERIOD_START_MONTH = "2026-08"
PERIOD_TYPES = {"MONTHLY", "CUMULATIVE"}
METRIC_SCOPES = {"MONTH", "CUMULATIVE"}
FEE_DIRECTIONS = {"PROMOTION", "MANAGEMENT"}
DATA_STATUSES = {"VALID", "ADJUSTED", "BLOCKED", "LOCKED"}
RANKING_SORT_FIELDS = {
    "SALES_AMOUNT",
    "VERIFIED_AMOUNT",
    "PROMOTION_FEE",
    "MANAGEMENT_FEE",
    "NET_SETTLEMENT_REFERENCE",
}
SORT_ORDERS = {"ASC", "DESC"}
CONFIRMATION_DIRECTION_TO_DB = {"PROMOTION": 1, "MANAGEMENT": 2}
CONFIRMATION_DIRECTION_FROM_DB = {value: key for key, value in CONFIRMATION_DIRECTION_TO_DB.items()}
STATEMENT_STATUS_NAMES = {
    1: "GENERATING",
    2: "PENDING_CONFIRMATION",
    3: "CONFIRMED",
    4: "LOCKED",
}
CONFIRMATION_STATUS_NAMES = {1: "CONFIRMED", 2: "REVOKED"}
INVOICE_STATUS_NAMES = {
    1: "PENDING_INVOICE",
    2: "SUBMITTED_PENDING_FACTORY_REVIEW",
    3: "APPROVED_SETTLED",
    4: "REJECTED_REUPLOAD",
}
SAP_SUGGESTION_STATUS_NAMES = {
    1: "PENDING",
    2: "CONFIRMED",
    3: "CORRECTED",
    4: "REJECTED",
}
PROMOTION_LIFECYCLE_EVENT_TO_DB = {"RED_FLUSHED": 1, "VOIDED": 2}
PROMOTION_LIFECYCLE_EVENT_FROM_DB = {
    value: key for key, value in PROMOTION_LIFECYCLE_EVENT_TO_DB.items()
}
DISPUTE_TYPE_TO_DB = {
    "RATE_ERROR": 1,
    "DATA_MISSING": 2,
    "AMOUNT_ERROR": 3,
    "OTHER": 4,
}
DISPUTE_TYPE_FROM_DB = {value: key for key, value in DISPUTE_TYPE_TO_DB.items()}
DISPUTE_STATUS_NAMES = {
    1: "PENDING",
    2: "IN_REVIEW",
    3: "PENDING_ADMIN_APPROVAL",
    4: "ACCEPTED_WITH_ADJUSTMENT",
    5: "REJECTED",
    6: "WITHDRAWN",
}
DISPUTE_STATUS_TO_DB = {value: key for key, value in DISPUTE_STATUS_NAMES.items()}
DISPUTE_TRANSITIONS = {
    1: {2, 4, 5},
    2: {3, 5},
    3: {4, 5},
}
FINANCE_IMPORT_TYPE_TO_DB = {
    "BASIC_INFO": 1,
    "PROMOTION_FACTORY_RESULT": 2,
    "MANAGEMENT_FACTORY_RESULT": 3,
    "SAP_CONFIRMATION": 4,
}
FINANCE_IMPORT_TYPE_FROM_DB = {
    value: key for key, value in FINANCE_IMPORT_TYPE_TO_DB.items()
}
FINANCE_IMPORT_SCENARIO_FROM_STATUS = {
    1: "VALIDATING",
    2: "NO_CHANGE",
    3: "DIFF_CONFIRMATION_REQUIRED",
    4: "FIRST_IMPORT_READY",
    5: "COMMITTED",
    6: "BATCH_VALIDATION_FAILED",
    7: "VERSION_CONFLICT",
    8: "CORRECTED",
    9: "REVERSED",
}
MAX_FINANCE_IMPORT_BYTES = 10 * 1024 * 1024
PROMOTION_INVOICE_BUYER_NAME = "比亚迪汽车销售有限公司"
PROMOTION_INVOICE_TAX_RATE_PERCENT = 6
BEIJING_TIME_ZONE = ZoneInfo("Asia/Shanghai")
MAX_FINANCE_IMPORT_ROWS = 5000
MAX_FINANCE_ORDER_EXPORT_ROWS = 100_000

FINANCE_ORDER_DETAIL_DEFINITIONS = {
    "storeName": {
        "source": "settlement_statement.store_name_snapshot",
        "description": "账单版本首次生成时冻结的服务店名称，不读取当前门店主数据。",
    },
    "sapCode": {
        "source": "settlement_statement.sap_code_snapshot",
        "description": "账单版本首次生成时冻结的有效 SAP；合法空值与无法回填由快照状态区分。",
    },
    "frozenFeeBaseCent": {
        "source": "settlement_statement_entry",
        "description": "锁账分录冻结的计费基数；退款或取消核销以独立负数调整行展示。",
    },
    "actualFeeRate": {
        "source": "settlement_statement_entry.fee_rate_snapshot",
        "description": "锁账时冻结的实际费率；旧分录仅回退到其原费用结果，不由前端重算。",
    },
    "frozenFeeAmountCent": {
        "source": "settlement_statement_entry",
        "description": "锁账分录冻结的费用金额；页面和导出均直接使用该值。",
    },
    "invoiceStatus": {
        "source": "promotion_invoice + invoice_status_event",
        "description": "仅推广服务费使用当前有效发票及其外部处理结果；管理服务费不套用该状态链。",
    },
    "settlementStatus": {
        "source": "invoice_record",
        "description": "仅管理服务费使用当前有效发票/厂家扣款与系统导入事实。",
    },
}

FINANCE_ORDER_DETAIL_DEFINITIONS.update(
    {
        "statementEntryId": {"source": "settlement_statement_entry.statement_entry_id", "description": "冻结分录唯一标识。"},
        "statementId": {"source": "settlement_statement.statement_id", "description": "当前有效账单版本标识。"},
        "storeId": {"source": "settlement_statement.store_id", "description": "账单所属门店稳定标识。"},
        "storeName": {"source": "settlement_statement.store_name_snapshot", "description": "账单创建时冻结的门店名称，不读取当前门店主数据。"},
        "sapCode": {"source": "settlement_statement.sap_code_snapshot", "description": "账单创建时冻结的 SAP；墓碑版本的空值不会复活旧 SAP。"},
        "statementMonth": {"source": "settlement_statement.statement_month", "description": "该行归属的账单账期。"},
        "feeDirection": {"source": "settlement_statement_entry.fee_direction", "description": "费用方向：推广服务费或管理服务费。"},
        "orderId": {"source": "settlement_statement_entry.order_id", "description": "冻结分录关联的订单标识。"},
        "couponId": {"source": "settlement_statement_entry.coupon_id", "description": "冻结分录关联的券标识。"},
        "orderStatus": {"source": "settlement_statement_entry.order_status_snapshot", "description": "锁账时冻结的订单状态。"},
        "couponStatus": {"source": "settlement_statement_entry.coupon_status_snapshot", "description": "锁账时冻结的券状态。"},
        "productName": {"source": "settlement_statement_entry.product_name_snapshot", "description": "锁账时冻结的商品名称。"},
        "skuId": {"source": "settlement_statement_entry.sku_id_snapshot", "description": "锁账时冻结的 SKU 标识。"},
        "skuName": {"source": "settlement_statement_entry.sku_name_snapshot", "description": "锁账时冻结的 SKU 名称。"},
        "saleChannel": {"source": "settlement_statement_entry.sale_channel_snapshot", "description": "锁账时冻结的销售渠道。"},
        "saleStoreId": {"source": "settlement_statement_entry.sale_store_id_snapshot", "description": "锁账时冻结的销售门店标识。"},
        "saleStoreName": {"source": "settlement_statement_entry.sale_store_snapshot", "description": "锁账时冻结的销售门店名称。"},
        "verifyStoreId": {"source": "settlement_statement_entry.verify_store_id_snapshot", "description": "锁账时冻结的核销门店标识。"},
        "verifyStoreName": {"source": "settlement_statement_entry.verify_store_snapshot", "description": "锁账时冻结的核销门店名称。"},
        "saleTime": {"source": "settlement_statement_entry.sale_time_snapshot", "description": "锁账时冻结的销售时间。"},
        "verifyTime": {"source": "settlement_statement_entry.verify_time_snapshot", "description": "锁账时冻结的核销时间；合法空值不回退当前核销记录。"},
        "receivedAmountCent": {"source": "settlement_statement_entry.received_amount_cent_snapshot", "description": "锁账时冻结的实收金额，单位为分。"},
        "frozenFeeBaseCent": {"source": "settlement_statement_entry.base_amount_cent", "description": "冻结计费基数；调整以独立正负行展示。"},
        "actualFeeRate": {"source": "settlement_statement_entry.fee_rate_snapshot", "description": "锁账时冻结的实际费率。"},
        "frozenFeeAmountCent": {"source": "settlement_statement_entry.fee_amount_cent", "description": "冻结费用金额，单位为分。"},
        "refundTime": {"source": "settlement_statement_entry.refund_at_snapshot", "description": "冻结的退款或异议调整发生时间。"},
        "adjustmentType": {"source": "settlement_statement_entry.adjustment_type_snapshot", "description": "冻结的调整类型；原费用行为空。"},
        "rowType": {"source": "settlement_statement_entry.source_type", "description": "原费用行或独立调整行。"},
        "invoiceNumber": {"source": "promotion_invoice / invoice_record", "description": "当前有效的发票登记号码。"},
        "submittedAt": {"source": "promotion_invoice.registered_at", "description": "推广服务费发票登记时间。"},
        "invoiceStatus": {"source": "promotion_invoice.invoice_status", "description": "推广服务费当前厂端处理状态。"},
        "settledAt": {"source": "invoice_status_event.business_date / invoice_record.factory_deduction_date", "description": "推广结算日期或管理服务费厂家扣款日期。"},
        "rejectionReason": {"source": "invoice_status_event.result_reason", "description": "推广服务费厂端不通过原因。"},
        "importedAt": {"source": "invoice_record.registered_at", "description": "管理服务费结果导入系统的时间。"},
        "settlementStatus": {"source": "invoice_record.invoice_status", "description": "管理服务费当前结算状态。"},
        "factoryDeductionDate": {"source": "invoice_record.factory_deduction_date", "description": "管理服务费厂家扣款日期。"},
        "factoryDeductionAmountCent": {"source": "invoice_record.factory_deduction_amount_cent", "description": "管理服务费厂家扣款金额，单位为分。"},
    }
)


STORE_RANKING_DEFINITIONS = [
    {
        "key": "salesOrderCount",
        "label": "销售订单数量",
        "description": "当前完整筛选集合中的销售订单数量，不受当前分页影响。",
    },
    {
        "key": "salesAmountCent",
        "label": "销售金额",
        "description": "当前完整筛选集合中的销售金额，单位分。",
    },
    {
        "key": "verifiedOrderCount",
        "label": "核销订单数量",
        "description": "当前完整筛选集合中的核销订单数量。",
    },
    {
        "key": "verifiedAmountCent",
        "label": "核销金额",
        "description": "当前完整筛选集合中的核销金额，单位分。",
    },
    {
        "key": "promotionNetFeeCent",
        "label": "推广服务费净额",
        "description": "推广服务费原始金额加调整金额后的净额，单位分。",
    },
    {
        "key": "managementNetFeeCent",
        "label": "管理服务费净额",
        "description": "管理服务费原始金额加调整金额后的净额，单位分。",
    },
    {
        "key": "netSettlementReferenceCent",
        "label": "结算参考净额",
        "description": "推广服务费净额减管理服务费净额，仅作为经营与结算核对依据。",
    },
]

MONTHLY_SETTLEMENT_DEFINITIONS = [
    {
        "key": "promotionNetFeeCent",
        "label": "应收推广服务费净额",
        "description": "推广服务费原始金额与调整金额合计后的调整后净额。",
    },
    {
        "key": "managementNetFeeCent",
        "label": "应扣管理服务费净额",
        "description": "管理服务费原始金额与调整金额合计后的调整后净额。",
    },
]

SALES_DASHBOARD_DEFINITIONS = [
    {
        "key": "total_sales_order_count",
        "label": "总销售订单量",
        "description": "销售归属门店在所选期间卖出的有效订单数，按订单编号去重，退款订单不计入。",
    },
    {
        "key": "self_verify_order_count",
        "label": "自店核销数",
        "description": "销售归属门店和实际核销门店都是当前门店的订单数，按订单编号去重，退款订单不计入。",
    },
    {
        "key": "self_verify_rate",
        "label": "自店核销率",
        "description": "自店核销数 / 总销售订单量；总销售订单量为 0 时显示 0。",
    },
    {
        "key": "total_verify_order_count",
        "label": "实际核销总数",
        "description": "不管销售归属门店，只要在当前门店于所选期间完成核销即计入，按订单编号去重；一单核销多券也只算一单。",
    },
    {
        "key": "actual_verify_amount_cent",
        "label": "实际核销金额",
        "description": "当前门店产生核销后的实收金额合计，退款订单不计入。",
    },
    {
        "key": "avg_verify_cycle_days",
        "label": "平均核销周期",
        "description": "当前门店已核销订单从销售时间到核销时间的平均天数，按订单编号去重。",
    },
    {
        "key": "cycle_distribution",
        "label": "核销周期分布",
        "description": "按商品类型展示当前门店核销订单从销售时间到核销时间的周期，箱线图展示四分位，散点展示订单样本。",
    },
]


def _filters_from_query(
    *,
    product_scope: str,
    product_type: str,
    sale_store_id: str | None,
    exclude_sale_store_id: str | None,
    sale_month: str | None,
    is_verified: bool | None,
    verify_store_id: str | None,
    exclude_verify_store_id: str | None,
    verify_month: str | None,
    relation_type: str | None,
    is_commissionable: bool | None,
    q: str | None,
    page: int,
    page_size: int,
) -> dict:
    return {
        "product_scope": product_scope,
        "product_type": product_type,
        "sale_store_id": sale_store_id,
        "exclude_sale_store_id": exclude_sale_store_id,
        "sale_month": sale_month,
        "is_verified": is_verified,
        "verify_store_id": verify_store_id,
        "exclude_verify_store_id": exclude_verify_store_id,
        "verify_month": verify_month,
        "relation_type": relation_type,
        "is_commissionable": is_commissionable,
        "q": q,
        "page": page,
        "page_size": page_size,
    }


@router.get("/dashboard/store-ranking")
def store_ranking(
    request: Request,
    period_type: str = Query(default="MONTHLY", alias="periodType"),
    period_key: str = Query(alias="periodKey"),
    product_scope: str = Query(default="all", alias="productScope"),
    product_type: str = Query(default="all", alias="productType"),
    q: str | None = None,
    sort_by: str = Query(default="NET_SETTLEMENT_REFERENCE", alias="sortBy"),
    sort_order: str = Query(default="DESC", alias="sortOrder"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=50, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    period_type = period_type.upper()
    sort_by = sort_by.upper()
    sort_order = sort_order.upper()
    _validate_month(period_key, "periodKey", request)
    _validate_enum(period_type, PERIOD_TYPES, "periodType", request)
    _validate_enum(sort_by, RANKING_SORT_FIELDS, "sortBy", request)
    _validate_enum(sort_order, SORT_ORDERS, "sortOrder", request)
    _validate_product_selection(store, product_scope, product_type, request)
    scope_mode = (
        "AUTHORIZED"
        if current_user.has_global_data_access
        else "GLOBAL_TOP_20_EXCEPTION"
    )
    if scope_mode == "GLOBAL_TOP_20_EXCEPTION":
        page = 1
        page_size = min(page_size, 20)
    filters = {
        "period_type": period_type,
        "period_key": period_key,
        "product_scope": product_scope,
        "product_type": product_type,
        "q": (q or "").strip() or None,
        "sort_by": sort_by,
        "sort_order": sort_order,
        "page": page,
        "page_size": page_size,
        "scope_mode": scope_mode,
        "scope_store_ids": (
            None if current_user.has_global_data_access else current_user.store_ids
        ),
    }
    data = _call_reporting_store(request, store.store_ranking_report, filters)
    return _reporting_success(request, data, definitions=STORE_RANKING_DEFINITIONS)


@router.get("/commission-rules/summary")
def commission_rules_summary(
    _current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    data = CommissionRulesSummaryData(**store.commission_rules_summary())
    return {
        "data": dump_model(data),
        "meta": {"generated_at": generated_at(), "source": "postgres"},
    }


@router.get("/store-settlements")
def list_store_settlements(
    request: Request,
    store_id: str = Query(alias="storeId"),
    month: str = Query(),
    metric_scope: str = Query(alias="metricScope"),
    fee_direction: str | None = Query(default=None, alias="feeDirection"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=50, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_billing_store_scope(current_user, store_id, request)
    _validate_month(month, "month", request)
    metric_scope = metric_scope.upper()
    _validate_enum(metric_scope, METRIC_SCOPES, "metricScope", request)
    direction = _normalize_billing_direction(fee_direction, request)
    _require_billing_store(session, store_id, request)

    conditions = [
        SettlementStatement.store_id == store_id,
        SettlementStatement.statement_month == month,
        SettlementStatement.is_current.is_(True),
    ]
    total = session.scalar(
        select(func.count()).select_from(SettlementStatement).where(*conditions)
    ) or 0
    statements = list(
        session.scalars(
            select(SettlementStatement)
            .where(*conditions)
            .order_by(SettlementStatement.statement_month.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    promotion_projection, _ = _promotion_invoice_carryforward_projection(
        session,
        store_id=store_id,
    )
    data = {
        "list": [
            _statement_list_item(
                session,
                statement,
                direction=direction,
                promotion_projection=promotion_projection,
            )
            for statement in statements
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
        "metric_scope": metric_scope,
        "metrics": _statement_metrics(session, store_id, month, metric_scope),
    }
    return _reporting_success(request, data)


@router.get("/store-settlements/{statement_id}")
def get_store_settlement_detail(
    statement_id: str,
    request: Request,
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    statement = _get_billing_statement(session, statement_id, request)
    _require_billing_store_scope(current_user, statement.store_id, request)
    versions = list(
        session.scalars(
            select(SettlementStatement)
            .where(
                SettlementStatement.store_id == statement.store_id,
                SettlementStatement.statement_month == statement.statement_month,
            )
            .order_by(SettlementStatement.version_no.desc())
        )
    )
    lines = list(
        session.scalars(
            select(SettlementStatementLine)
            .where(SettlementStatementLine.statement_id == statement.statement_id)
            .order_by(
                SettlementStatementLine.fee_direction,
                SettlementStatementLine.product_scope,
                SettlementStatementLine.product_type,
            )
        )
    )
    entry_count = session.scalar(
        select(func.count())
        .select_from(SettlementStatementEntry)
        .where(SettlementStatementEntry.statement_id == statement.statement_id)
    ) or 0
    dispute_count = session.scalar(
        select(func.count())
        .select_from(SettlementDispute)
        .where(SettlementDispute.statement_id == statement.statement_id)
    ) or 0
    invoices = list(
        session.scalars(
            select(InvoiceRecord).where(InvoiceRecord.statement_id == statement.statement_id)
        )
    )
    data = {
        **_statement_header_item(session, statement),
        "versions": [_statement_version_item(row) for row in versions],
        "lines": [_statement_line_item(line) for line in lines],
        "source_summary": {"entry_count": entry_count},
        "dispute_summary": {"count": dispute_count},
        "invoice_summary": [
            {
                "invoice_id": invoice.invoice_id,
                "fee_direction": CONFIRMATION_DIRECTION_FROM_DB[invoice.fee_direction],
                "status": INVOICE_STATUS_NAMES[invoice.invoice_status],
                "invoice_amount_cent": invoice.invoice_amount_cent,
            }
            for invoice in invoices
        ],
    }
    return _reporting_success(request, data)


@router.post("/store-settlements/{statement_id}/confirmations")
def confirm_store_settlement(
    statement_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    statement = _get_billing_statement(session, statement_id, request)
    _require_billing_store_scope(current_user, statement.store_id, request)
    parsed = _parse_confirmation_payload(payload, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(parsed)
    replay = session.scalar(
        select(SettlementStatementConfirmation).where(
            SettlementStatementConfirmation.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if replay.request_payload_sha256 != payload_hash:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        return _reporting_success(request, _confirmation_item(replay, statement))

    if not statement.is_current or statement.version_no != parsed["read_version"]:
        _raise_statement_version_conflict(request, statement)
    direction = parsed["fee_direction"]
    expected_amount = (
        statement.promotion_net_fee_cent
        if direction == "PROMOTION"
        else statement.management_net_fee_cent
    )
    frozen_amount = int(
        session.scalar(
            select(func.coalesce(func.sum(SettlementDispute.disputed_amount_cent), 0)).where(
                SettlementDispute.statement_id == statement.statement_id,
                SettlementDispute.fee_direction
                == CONFIRMATION_DIRECTION_TO_DB[direction],
                SettlementDispute.status.in_((1, 2, 3)),
            )
        )
        or 0
    )
    expected_amount -= frozen_amount
    if direction == "MANAGEMENT":
        expected_amount = max(expected_amount, 0)
    if parsed["confirmed_amount_cent"] != expected_amount:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "确认金额必须等于当前账单净额",
            field="confirmedAmountCent",
        )
    existing = session.scalar(
        select(SettlementStatementConfirmation).where(
            SettlementStatementConfirmation.statement_id == statement.statement_id,
            SettlementStatementConfirmation.fee_direction
            == CONFIRMATION_DIRECTION_TO_DB[direction],
        )
    )
    if existing is not None:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "CONFIRMATION_ALREADY_RECORDED",
            "该费用方向已确认",
        )
    now = utcnow()
    confirmation = SettlementStatementConfirmation(
        confirmation_id=f"confirmation-{uuid4().hex}",
        statement_id=statement.statement_id,
        fee_direction=CONFIRMATION_DIRECTION_TO_DB[direction],
        confirmation_status=1,
        confirmed_amount_cent=expected_amount,
        confirmed_by=current_user.username,
        confirmed_at=now,
        idempotency_key_hash=key_hash,
        request_payload_sha256=payload_hash,
    )
    session.add(confirmation)
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "CONFIRMATION_ALREADY_RECORDED",
            "该费用方向已确认或幂等键已被使用",
        )
    return _reporting_success(request, _confirmation_item(confirmation, statement))


@router.get("/store-settlements/{statement_id}/disputes")
def list_store_settlement_disputes(
    statement_id: str,
    request: Request,
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    statement = _get_billing_statement(session, statement_id, request)
    _require_billing_store_scope(current_user, statement.store_id, request)
    disputes = list(
        session.scalars(
            select(SettlementDispute)
            .where(SettlementDispute.statement_id == statement.statement_id)
            .order_by(SettlementDispute.submitted_at.desc(), SettlementDispute.dispute_id)
        )
    )
    return _reporting_success(
        request,
        {"list": [_dispute_item(session, dispute, request) for dispute in disputes]},
    )


@router.post("/store-settlements/{statement_id}/disputes")
def create_store_settlement_dispute(
    statement_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    statement = _get_billing_statement(session, statement_id, request)
    _require_billing_store_scope(current_user, statement.store_id, request)
    parsed = _parse_dispute_payload(payload, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(parsed)
    replay = session.scalar(
        select(SettlementDispute).where(
            SettlementDispute.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if replay.request_payload_sha256 != payload_hash:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        return _reporting_success(request, _dispute_item(session, replay, request))
    if not statement.is_current or statement.version_no != parsed["read_version"]:
        _raise_statement_version_conflict(request, statement)

    direction_code = CONFIRMATION_DIRECTION_TO_DB[parsed["fee_direction"]]
    for order in parsed["orders"]:
        conditions = [
            SettlementStatementEntry.statement_id == statement.statement_id,
            SettlementStatementEntry.fee_direction == direction_code,
            SettlementStatementEntry.order_id == order["order_id"],
        ]
        if order["coupon_id"] is not None:
            conditions.append(SettlementStatementEntry.coupon_id == order["coupon_id"])
        if session.scalar(select(SettlementStatementEntry).where(*conditions)) is None:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "DISPUTE_ORDER_NOT_IN_STATEMENT",
                "争议订单或券不属于当前账单版本和费用方向",
                field="orders",
            )

    dispute = SettlementDispute(
        dispute_id=f"dispute-{uuid4().hex}",
        statement_id=statement.statement_id,
        store_id=statement.store_id,
        statement_month=statement.statement_month,
        fee_direction=direction_code,
        dispute_type=DISPUTE_TYPE_TO_DB[parsed["dispute_type"]],
        status=1,
        disputed_amount_cent=parsed["disputed_amount_cent"],
        description=parsed["description"],
        contact_name=parsed["contact_name"],
        contact_phone_ciphertext=_encrypt_dispute_phone(
            parsed["contact_phone"], request
        ),
        evidence_json=parsed["evidence"],
        submitted_by=current_user.username,
        submitted_at=utcnow(),
        idempotency_key_hash=key_hash,
        request_payload_sha256=payload_hash,
    )
    session.add(dispute)
    for order in parsed["orders"]:
        session.add(
            SettlementDisputeOrder(
                dispute_id=dispute.dispute_id,
                order_id=order["order_id"],
                coupon_id=order["coupon_id"],
                disputed_amount_cent=order["disputed_amount_cent"],
            )
        )
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="DISPUTE_SUBMIT",
            target_type="SETTLEMENT_DISPUTE",
            target_id=dispute.dispute_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot=None,
            after_snapshot={
                "status": "PENDING",
                "statementId": statement.statement_id,
                "versionNo": statement.version_no,
            },
            result_status=1,
            request_id=request_id(request),
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
    )
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        replay = session.scalar(
            select(SettlementDispute).where(
                SettlementDispute.idempotency_key_hash == key_hash
            )
        )
        if replay is not None and replay.request_payload_sha256 == payload_hash:
            return _reporting_success(request, _dispute_item(session, replay, request))
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "DISPUTE_SUBMIT_CONFLICT",
            "异议提交冲突，请刷新后重试",
        )
    return _reporting_success(request, _dispute_item(session, dispute, request))


@router.post("/disputes/{dispute_id}/withdrawals")
def withdraw_store_settlement_dispute(
    dispute_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    dispute = session.scalar(
        select(SettlementDispute).where(SettlementDispute.dispute_id == dispute_id)
    )
    if dispute is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "异议不存在",
            field="disputeId",
        )
    _require_billing_store_scope(current_user, dispute.store_id, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(payload)
    replay = session.scalar(
        select(FinanceOperationAudit).where(
            FinanceOperationAudit.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if replay.request_payload_sha256 != payload_hash:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        return _reporting_success(
            request,
            {**_dispute_item(session, dispute, request), "adjustment_reversed": False},
        )
    reason = payload.get("reason")
    read_version = payload.get("readVersion")
    if not isinstance(reason, str) or not reason.strip():
        _raise_dispute_validation(request, "reason", "reason 为必填项")
    if (
        isinstance(read_version, bool)
        or not isinstance(read_version, int)
        or read_version < 1
    ):
        _raise_dispute_validation(request, "readVersion", "readVersion 必须为正整数")
    current_statement = session.scalar(
        select(SettlementStatement).where(
            SettlementStatement.store_id == dispute.store_id,
            SettlementStatement.statement_month == dispute.statement_month,
            SettlementStatement.is_current.is_(True),
        )
    )
    if current_statement is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "异议所属账期不存在当前账单",
            field="disputeId",
        )
    if current_statement.version_no != read_version:
        _raise_statement_version_conflict(request, current_statement)
    if dispute.status == 6:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "DISPUTE_ALREADY_WITHDRAWN",
            "异议已撤回",
        )

    before_status = DISPUTE_STATUS_NAMES[dispute.status]
    dispute.status = 6
    if dispute.result_statement_id is None:
        dispute.resolution_note = reason.strip()
    dispute.processed_by = current_user.username
    dispute.processed_at = utcnow()
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="DISPUTE_WITHDRAW",
            target_type="SETTLEMENT_DISPUTE",
            target_id=dispute.dispute_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot={"status": before_status},
            after_snapshot={"status": "WITHDRAWN", "adjustmentReversed": False},
            result_status=1,
            request_id=request_id(request),
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
    )
    session.commit()
    return _reporting_success(
        request,
        {**_dispute_item(session, dispute, request), "adjustment_reversed": False},
    )


@router.get("/admin/disputes")
def list_admin_disputes(
    request: Request,
    store_id: str | None = Query(default=None, alias="storeId"),
    month: str | None = Query(default=None),
    fee_direction: str | None = Query(default=None, alias="feeDirection"),
    dispute_status: str | None = Query(default=None, alias="status"),
    dispute_type: str | None = Query(default=None, alias="disputeType"),
    submitted_from: datetime | None = Query(default=None, alias="submittedFrom"),
    submitted_to: datetime | None = Query(default=None, alias="submittedTo"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    _require_finance_admin(current_user, request)
    session = _billing_session(store, request)
    conditions = []
    if store_id:
        conditions.append(SettlementDispute.store_id == store_id.strip())
    if month:
        _validate_month(month, "month", request)
        conditions.append(SettlementDispute.statement_month == month)
    if fee_direction:
        direction = _normalize_billing_direction(fee_direction, request)
        conditions.append(
            SettlementDispute.fee_direction == CONFIRMATION_DIRECTION_TO_DB[direction]
        )
    if dispute_status:
        normalized_status = dispute_status.strip().upper()
        _validate_enum(
            normalized_status, set(DISPUTE_STATUS_TO_DB), "status", request
        )
        conditions.append(SettlementDispute.status == DISPUTE_STATUS_TO_DB[normalized_status])
    if dispute_type:
        normalized_type = dispute_type.strip().upper()
        _validate_enum(normalized_type, set(DISPUTE_TYPE_TO_DB), "disputeType", request)
        conditions.append(SettlementDispute.dispute_type == DISPUTE_TYPE_TO_DB[normalized_type])
    if submitted_from:
        conditions.append(SettlementDispute.submitted_at >= submitted_from)
    if submitted_to:
        conditions.append(SettlementDispute.submitted_at <= submitted_to)

    total = int(
        session.scalar(select(func.count()).select_from(SettlementDispute).where(*conditions))
        or 0
    )
    disputes = list(
        session.scalars(
            select(SettlementDispute)
            .where(*conditions)
            .order_by(SettlementDispute.submitted_at.desc(), SettlementDispute.dispute_id)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    return _reporting_success(
        request,
        {
            "list": [_dispute_item(session, dispute, request) for dispute in disputes],
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    )


@router.post("/admin/disputes/{dispute_id}/transitions")
def transition_admin_dispute(
    dispute_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    _require_finance_admin(current_user, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    session = _billing_session(store, request)
    dispute = session.scalar(
        select(SettlementDispute).where(SettlementDispute.dispute_id == dispute_id)
    )
    if dispute is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "异议不存在",
            field="disputeId",
        )

    target_status = payload.get("targetStatus")
    resolution_note = payload.get("resolutionNote")
    read_version = payload.get("readVersion")
    adjustment_amount = payload.get("adjustmentAmountCent")
    payload_hash = _canonical_billing_sha256(payload)
    if not isinstance(target_status, str):
        _raise_dispute_validation(request, "targetStatus", "targetStatus 为必填项")
    normalized_target = target_status.strip().upper()
    allowed_targets = {
        "IN_REVIEW",
        "PENDING_ADMIN_APPROVAL",
        "ACCEPTED_WITH_ADJUSTMENT",
        "REJECTED",
    }
    _validate_enum(normalized_target, allowed_targets, "targetStatus", request)
    target_code = DISPUTE_STATUS_TO_DB[normalized_target]
    replay = session.scalar(
        select(FinanceOperationAudit).where(
            FinanceOperationAudit.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if replay.request_payload_sha256 != payload_hash:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        replay_data = _dispute_item(session, dispute, request)
        if replay.after_snapshot and "currentStatementId" in replay.after_snapshot:
            replay_data.update(
                {
                    "previous_statement_id": replay.before_snapshot.get("statementId"),
                    "previous_version": replay.before_snapshot.get("versionNo"),
                    "current_statement_id": replay.after_snapshot["currentStatementId"],
                    "current_version": replay.after_snapshot["currentVersion"],
                }
            )
        return _reporting_success(request, replay_data)
    if target_code not in DISPUTE_TRANSITIONS.get(dispute.status, set()):
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "DISPUTE_STATUS_CONFLICT",
            "异议状态已变化或不允许执行该流转",
            field="targetStatus",
        )
    if not isinstance(resolution_note, str) or not resolution_note.strip():
        _raise_dispute_validation(request, "resolutionNote", "resolutionNote 为必填项")
    if isinstance(read_version, bool) or not isinstance(read_version, int) or read_version < 1:
        _raise_dispute_validation(request, "readVersion", "readVersion 必须为正整数")
    if normalized_target == "ACCEPTED_WITH_ADJUSTMENT" and (
        isinstance(adjustment_amount, bool)
        or not isinstance(adjustment_amount, int)
        or adjustment_amount == 0
    ):
        _raise_dispute_validation(
            request, "adjustmentAmountCent", "成立并调整时必须提交非零整数调整金额"
        )

    current_statement = session.scalar(
        select(SettlementStatement).where(
            SettlementStatement.store_id == dispute.store_id,
            SettlementStatement.statement_month == dispute.statement_month,
            SettlementStatement.is_current.is_(True),
        )
    )
    if current_statement is None:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "STATEMENT_VERSION_CONFLICT",
            "当前账单版本不存在",
        )
    if current_statement.version_no != read_version:
        _raise_statement_version_conflict(request, current_statement)

    before_snapshot = {
        "status": DISPUTE_STATUS_NAMES[dispute.status],
        "statementId": current_statement.statement_id,
        "versionNo": current_statement.version_no,
    }
    previous_statement_id = current_statement.statement_id
    previous_version = current_statement.version_no
    current_result_statement = current_statement
    if normalized_target == "ACCEPTED_WITH_ADJUSTMENT":
        current_result_statement = _create_dispute_statement_version(
            session=session,
            dispute=dispute,
            current_statement=current_statement,
            adjustment_amount_cent=adjustment_amount,
            resolution_note=resolution_note.strip(),
            operator=current_user.username,
            request=request,
        )
        dispute.result_statement_id = current_result_statement.statement_id

    dispute.status = target_code
    dispute.resolution_note = resolution_note.strip()
    dispute.processed_by = current_user.username
    dispute.processed_at = utcnow()
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="DISPUTE_TRANSITION",
            target_type="SETTLEMENT_DISPUTE",
            target_id=dispute.dispute_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot=before_snapshot,
            after_snapshot={
                "status": normalized_target,
                "statementId": current_result_statement.statement_id,
                "versionNo": current_result_statement.version_no,
                "currentStatementId": current_result_statement.statement_id,
                "currentVersion": current_result_statement.version_no,
            },
            result_status=1,
            request_id=request_id(request),
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
    )
    try:
        session.commit()
    except Exception:
        session.rollback()
        raise

    data = _dispute_item(session, dispute, request)
    if normalized_target == "ACCEPTED_WITH_ADJUSTMENT":
        data.update(
            {
                "previous_statement_id": previous_statement_id,
                "previous_version": previous_version,
                "current_statement_id": current_result_statement.statement_id,
                "current_version": current_result_statement.version_no,
            }
        )
    return _reporting_success(request, data)


@router.post("/admin/finance-imports")
def create_finance_import(
    request: Request,
    import_type: str = Form(alias="importType"),
    statement_month: str = Form(alias="statementMonth"),
    file: UploadFile = File(),
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    _require_finance_admin(current_user, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    normalized_type = import_type.strip().upper()
    _validate_enum(normalized_type, set(FINANCE_IMPORT_TYPE_TO_DB), "importType", request)
    _validate_month(statement_month, "statementMonth", request)
    filename = (file.filename or "").strip()
    if not filename.lower().endswith((".csv", ".xlsx")):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "财务导入仅接受 UTF-8 CSV 或 XLSX 文件",
            field="file",
        )

    file_hasher = sha256()
    file_size = 0
    while chunk := file.file.read(1024 * 1024):
        file_size += len(chunk)
        if file_size > MAX_FINANCE_IMPORT_BYTES:
            _raise_reporting_error(
                request,
                status.HTTP_413_CONTENT_TOO_LARGE,
                "IMPORT_FILE_TOO_LARGE",
                "财务导入文件不得超过 10 MiB",
                field="file",
            )
        file_hasher.update(chunk)
    file.file.seek(0)
    session = _billing_session(store, request)
    upload_payload_hash = _canonical_billing_sha256(
        {
            "importType": normalized_type,
            "statementMonth": statement_month,
            "fileSha256": file_hasher.hexdigest(),
        }
    )
    replay = session.scalar(
        select(FinanceImportBatch).where(
            FinanceImportBatch.upload_idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if replay.upload_request_payload_sha256 != upload_payload_hash:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同的导入文件或参数",
            )
        return _reporting_success(
            request, _finance_import_upload_item(session, replay)
        )
    current_version = int(
        session.scalar(
            select(func.coalesce(func.max(FinanceImportBatch.current_version), 0)).where(
                FinanceImportBatch.import_type == FINANCE_IMPORT_TYPE_TO_DB[normalized_type],
                FinanceImportBatch.statement_month == statement_month,
                FinanceImportBatch.batch_status.in_((5, 8, 9)),
            )
        )
        or 0
    )
    batch = FinanceImportBatch(
        batch_id=f"finance-import-{uuid4().hex}",
        import_type=FINANCE_IMPORT_TYPE_TO_DB[normalized_type],
        statement_month=statement_month,
        file_name=filename,
        file_sha256=file_hasher.hexdigest(),
        normalized_sha256="",
        read_version=current_version,
        current_version=current_version,
        batch_status=1,
        upload_idempotency_key_hash=key_hash,
        upload_request_payload_sha256=upload_payload_hash,
        submitted_by=current_user.username,
        submitted_at=utcnow(),
    )
    session.add(batch)
    session.flush()

    normalized_hasher = sha256()
    total_rows = 0
    error_rows = 0
    first_errors: list[dict] = []
    seen_business_keys: set[str] = set()
    for row_number, raw_row in _iter_finance_import_rows(file, filename, request):
        if total_rows >= MAX_FINANCE_IMPORT_ROWS:
            session.rollback()
            _raise_reporting_error(
                request,
                status.HTTP_413_CONTENT_TOO_LARGE,
                "IMPORT_FILE_TOO_LARGE",
                "财务导入文件不得超过 5000 行",
                field="file",
            )
        total_rows += 1
        business_key, normalized_payload, errors = _validate_final_finance_import_row(
            session,
            import_type=normalized_type,
            statement_month=statement_month,
            row_number=row_number,
            raw_row=raw_row,
        )
        if business_key in seen_business_keys:
            errors.append(
                _finance_import_error(
                    row_number,
                    business_key,
                    "businessKey",
                    business_key,
                    "批次内业务唯一键重复",
                    "删除重复行",
                )
            )
        seen_business_keys.add(business_key)
        normalized_hasher.update(
            json.dumps(
                normalized_payload,
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            ).encode("utf-8")
        )
        if errors:
            error_rows += 1
            first_errors.extend(errors[: max(0, 100 - len(first_errors))])
        session.add(
            FinanceImportRow(
                batch_id=batch.batch_id,
                row_number=row_number,
                business_key=business_key,
                normalized_payload=normalized_payload,
                row_status=4 if errors else 1,
                validation_errors=errors,
            )
        )
        if total_rows % 500 == 0:
            session.flush()

    batch.normalized_sha256 = normalized_hasher.hexdigest()
    batch.total_rows = total_rows
    batch.error_rows = error_rows
    batch.success_rows = total_rows - error_rows
    if error_rows:
        batch.content_changed = False
        batch.batch_status = 6
        scenario = "BATCH_VALIDATION_FAILED"
    else:
        unchanged_batch_id = session.scalar(
            select(FinanceImportBatch.batch_id).where(
                FinanceImportBatch.batch_id != batch.batch_id,
                FinanceImportBatch.import_type == batch.import_type,
                FinanceImportBatch.statement_month == batch.statement_month,
                FinanceImportBatch.normalized_sha256 == batch.normalized_sha256,
                FinanceImportBatch.batch_status.in_((5, 8)),
            )
        )
        if unchanged_batch_id is not None:
            batch.content_changed = False
            batch.batch_status = 2
            scenario = "NO_CHANGE"
            session.execute(
                update(FinanceImportRow)
                .where(FinanceImportRow.batch_id == batch.batch_id)
                .values(row_status=2)
            )
        elif current_version > 0:
            batch.content_changed = True
            batch.batch_status = 3
            scenario = "DIFF_CONFIRMATION_REQUIRED"
            session.execute(
                update(FinanceImportRow)
                .where(FinanceImportRow.batch_id == batch.batch_id)
                .values(row_status=3)
            )
        else:
            batch.content_changed = True
            batch.batch_status = 4
            scenario = "FIRST_IMPORT_READY"
    session.commit()
    return _reporting_success(request, _finance_import_upload_item(session, batch))


@router.get("/admin/finance-imports")
def list_finance_imports(
    request: Request,
    import_type: str | None = Query(default=None, alias="importType"),
    statement_month: str | None = Query(default=None, alias="statementMonth"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=50, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    """List finance-import batches for administrators."""
    _require_finance_admin(current_user, request)
    session = _billing_session(store, request)
    conditions = []
    if import_type is not None:
        normalized_type = import_type.strip().upper()
        _validate_enum(
            normalized_type, set(FINANCE_IMPORT_TYPE_TO_DB), "importType", request
        )
        conditions.append(
            FinanceImportBatch.import_type
            == FINANCE_IMPORT_TYPE_TO_DB[normalized_type]
        )
    if statement_month is not None:
        _validate_month(statement_month, "statementMonth", request)
        conditions.append(FinanceImportBatch.statement_month == statement_month)
    total = int(
        session.scalar(
            select(func.count()).select_from(FinanceImportBatch).where(*conditions)
        )
        or 0
    )
    batches = list(
        session.scalars(
            select(FinanceImportBatch)
            .where(*conditions)
            .order_by(FinanceImportBatch.submitted_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    batch_ids = [batch.batch_id for batch in batches]
    reversed_by_batch_id = (
        {
            reverses_batch_id: reversal_batch_id
            for reversal_batch_id, reverses_batch_id in session.execute(
                select(
                    FinanceImportBatch.batch_id,
                    FinanceImportBatch.reverses_batch_id,
                ).where(FinanceImportBatch.reverses_batch_id.in_(batch_ids))
            )
        }
        if batch_ids
        else {}
    )
    return _reporting_success(
        request,
        {
            "list": [
                _finance_import_batch_item(
                    session,
                    batch,
                    reversed_by_batch_id=reversed_by_batch_id.get(batch.batch_id),
                )
                for batch in batches
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
        },
    )


@router.get("/admin/finance-imports/{batch_id}/error-file")
def download_finance_import_errors(
    batch_id: str,
    request: Request,
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    """Stream every validation error in a finance-import batch as CSV."""
    _require_finance_admin(current_user, request)
    session = _billing_session(store, request)
    batch = session.scalar(
        select(FinanceImportBatch).where(FinanceImportBatch.batch_id == batch_id)
    )
    if batch is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "财务导入批次不存在",
            field="batchId",
        )
    filename = quote(f"finance-import-errors-{batch_id}.csv")
    return StreamingResponse(
        _finance_import_error_csv(session, batch_id),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.get("/admin/finance-imports/{batch_id}")
def get_finance_import(
    batch_id: str,
    request: Request,
    error_page: int = Query(default=1, ge=1, alias="errorPage"),
    error_page_size: int = Query(
        default=20, ge=1, le=50, alias="errorPageSize"
    ),
    reversal_page: int = Query(default=1, ge=1, alias="reversalPage"),
    reversal_page_size: int = Query(
        default=20, ge=1, le=50, alias="reversalPageSize"
    ),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    """Return one import batch and a page of complete row-level errors."""
    _require_finance_admin(current_user, request)
    session = _billing_session(store, request)
    batch = session.scalar(
        select(FinanceImportBatch).where(FinanceImportBatch.batch_id == batch_id)
    )
    if batch is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "财务导入批次不存在",
            field="batchId",
        )
    error_rows = list(
        session.scalars(
            select(FinanceImportRow)
            .where(
                FinanceImportRow.batch_id == batch_id,
                FinanceImportRow.row_status == 4,
            )
            .order_by(FinanceImportRow.row_number)
            .offset((error_page - 1) * error_page_size)
            .limit(error_page_size)
        )
    )
    errors = [
        error
        for row in error_rows
        for error in (row.validation_errors or [])
    ]
    reversed_by_batch_id = session.scalar(
        select(FinanceImportBatch.batch_id).where(
            FinanceImportBatch.reverses_batch_id == batch_id
        )
    )
    data = _finance_import_batch_item(
        session,
        batch,
        reversed_by_batch_id=reversed_by_batch_id,
    )
    data.update(
        {
            "diff_summary": {
                "changedRows": batch.success_rows if batch.content_changed else 0
            },
            "errors": {
                "list": errors,
                "total": _count_finance_import_errors(session, batch_id),
                "page": error_page,
                "page_size": error_page_size,
            },
            "reversal_rows": _finance_import_reversal_rows(
                session,
                batch,
                reversed_by_batch_id=reversed_by_batch_id,
                page=reversal_page,
                page_size=reversal_page_size,
            ),
        }
    )
    return _reporting_success(request, data)


@router.post("/admin/finance-imports/{batch_id}/commits")
def commit_finance_import(
    batch_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    """Atomically commit a validated finance-import preview."""
    return _execute_finance_import_commit(
        batch_id=batch_id,
        payload=payload,
        request=request,
        idempotency_key=idempotency_key,
        current_user=current_user,
        store=store,
        is_correction=False,
    )


@router.post("/admin/finance-imports/{batch_id}/corrections")
def correct_finance_import(
    batch_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    """Commit a validated correction as a new immutable import version."""
    return _execute_finance_import_commit(
        batch_id=batch_id,
        payload=payload,
        request=request,
        idempotency_key=idempotency_key,
        current_user=current_user,
        store=store,
        is_correction=True,
    )


def _lock_finance_import_version_slot(
    session,
    *,
    import_type: int,
    statement_month: str,
) -> None:
    """Serialize finalized version allocation on PostgreSQL; DB uniqueness is fallback."""
    if session.get_bind().dialect.name != "postgresql":
        return
    session.execute(
        text("SELECT pg_advisory_xact_lock(hashtextextended(:slot_key, 0))"),
        {
            "slot_key": (
                f"finance-import-version:{import_type}:{statement_month}"
            )
        },
    )


def _finance_import_reversal_plan(session, batch: FinanceImportBatch, request: Request) -> list[dict]:
    """Lock and validate every business target before a reversal writes anything."""
    rows = list(
        session.scalars(
            select(FinanceImportRow)
            .where(
                FinanceImportRow.batch_id == batch.batch_id,
                FinanceImportRow.row_status == 5,
            )
            .order_by(FinanceImportRow.row_number)
        )
    )
    if not rows or len(rows) != batch.success_rows:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "REVERSAL_TARGET_INCOMPLETE",
            "原批次业务目标不完整，无法安全撤销",
        )
    plan: list[dict] = []
    for row in rows:
        if not row.target_record_id:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "REVERSAL_TARGET_MISSING",
                "原批次存在未关联业务版本的行",
            )
        if batch.import_type in (1, 4):
            target = session.scalar(
                select(StoreFinanceProfile)
                .where(StoreFinanceProfile.profile_id == row.target_record_id)
                .with_for_update()
            )
            if target is None or not target.is_current:
                _raise_reporting_error(
                    request,
                    status.HTTP_409_CONFLICT,
                    "REVERSAL_BUSINESS_VERSION_CONFLICT",
                    "门店财务资料已被后续导入或页面操作覆盖",
                    data={"business_key": row.business_key},
                )
            previous = session.scalar(
                select(StoreFinanceProfile)
                .where(
                    StoreFinanceProfile.store_id == target.store_id,
                    StoreFinanceProfile.profile_type == target.profile_type,
                    StoreFinanceProfile.version_no < target.version_no,
                )
                .order_by(StoreFinanceProfile.version_no.desc())
                .limit(1)
            )
            plan.append({"row": row, "kind": "PROFILE", "target": target, "previous": previous})
            continue
        if batch.import_type == 3:
            target = session.scalar(
                select(InvoiceRecord)
                .where(InvoiceRecord.invoice_id == row.target_record_id)
                .with_for_update()
            )
            if target is None or not target.is_current:
                _raise_reporting_error(
                    request,
                    status.HTTP_409_CONFLICT,
                    "REVERSAL_BUSINESS_VERSION_CONFLICT",
                    "管理服务费记录已被后续导入或页面操作覆盖",
                    data={"business_key": row.business_key},
                )
            previous = session.scalar(
                select(InvoiceRecord)
                .where(
                    InvoiceRecord.store_id == target.store_id,
                    InvoiceRecord.statement_month == target.statement_month,
                    InvoiceRecord.fee_direction == target.fee_direction,
                    InvoiceRecord.version_no < target.version_no,
                )
                .order_by(InvoiceRecord.version_no.desc())
                .limit(1)
            )
            plan.append({"row": row, "kind": "MANAGEMENT", "target": target, "previous": previous})
            continue
        target = session.scalar(
            select(PromotionInvoice)
            .where(PromotionInvoice.invoice_id == row.target_record_id)
            .with_for_update()
        )
        if target is None or not target.is_current:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "REVERSAL_BUSINESS_VERSION_CONFLICT",
                "推广服务费厂家结果已被后续导入或页面操作覆盖",
                data={"business_key": row.business_key},
            )
        previous = (
            session.scalar(
                select(PromotionInvoice).where(
                    PromotionInvoice.invoice_id == target.supersedes_invoice_id
                )
            )
            if target.supersedes_invoice_id
            else None
        )
        plan.append({"row": row, "kind": "PROMOTION", "target": target, "previous": previous})
    return plan


def _append_profile_reversal_version(
    session, *, target: StoreFinanceProfile, previous: StoreFinanceProfile | None,
    batch: FinanceImportBatch, now: datetime,
) -> StoreFinanceProfile:
    target.is_current = False
    session.flush()
    source = previous or target
    restored = StoreFinanceProfile(
        profile_id=f"store-finance-profile-{uuid4().hex}",
        store_id=target.store_id,
        profile_type=target.profile_type,
        source_type=3,
        version_no=target.version_no + 1,
        is_current=True,
        is_tombstone=previous is None or previous.is_tombstone,
        store_name_snapshot=source.store_name_snapshot,
        sap_code=source.sap_code,
        initial_sap_code=source.initial_sap_code,
        service_store_code=source.service_store_code,
        factory_confirmed=source.factory_confirmed,
        confirmed_at=source.confirmed_at,
        import_batch_id=batch.batch_id,
        created_at=now,
        updated_at=now,
    )
    session.add(restored)
    session.flush()
    return restored


def _append_management_reversal_version(
    session, *, target: InvoiceRecord, previous: InvoiceRecord | None,
    batch: FinanceImportBatch, operator_id: str, now: datetime,
) -> InvoiceRecord:
    target.is_current = False
    session.flush()
    source = previous or target
    restored = InvoiceRecord(
        invoice_id=f"management-invoice-{uuid4().hex}",
        store_id=target.store_id,
        statement_month=target.statement_month,
        statement_id=target.statement_id,
        fee_direction=target.fee_direction,
        version_no=target.version_no + 1,
        is_current=True,
        is_tombstone=previous is None or previous.is_tombstone,
        invoice_number=source.invoice_number,
        invoice_date=source.invoice_date,
        invoice_amount_cent=source.invoice_amount_cent,
        invoice_status=source.invoice_status,
        source_type=3,
        import_batch_id=batch.batch_id,
        factory_deduction_date=source.factory_deduction_date,
        factory_deduction_amount_cent=source.factory_deduction_amount_cent,
        registered_by=operator_id,
        registered_at=now,
    )
    session.add(restored)
    session.flush()
    return restored


def _append_promotion_reversal_version(
    session, *, target: PromotionInvoice, previous: PromotionInvoice | None,
    batch: FinanceImportBatch, operator_id: str, now: datetime,
) -> PromotionInvoice:
    current_allocations = list(
        session.scalars(
            select(PromotionInvoiceAllocation).where(
                PromotionInvoiceAllocation.invoice_id == target.invoice_id,
                PromotionInvoiceAllocation.is_current.is_(True),
            )
        )
    )
    for allocation in current_allocations:
        allocation.is_current = False
    target.is_current = False
    session.flush()
    source = previous or target
    restored = PromotionInvoice(
        invoice_id=f"promotion-invoice-{uuid4().hex}",
        physical_invoice_id=target.physical_invoice_id,
        store_id=target.store_id,
        version_no=target.version_no + 1,
        version_kind=2,
        is_current=True,
        is_tombstone=previous is None or previous.is_tombstone,
        supersedes_invoice_id=target.invoice_id,
        replaces_invoice_id=source.replaces_invoice_id,
        invoice_number=source.invoice_number,
        invoice_date=source.invoice_date,
        invoice_amount_cent=source.invoice_amount_cent,
        buyer_name=source.buyer_name,
        tax_rate_percent=source.tax_rate_percent,
        invoice_status=source.invoice_status,
        registered_by=operator_id,
        registered_at=now,
    )
    session.add(restored)
    session.flush()
    if previous is not None:
        previous_allocations = list(
            session.scalars(
                select(PromotionInvoiceAllocation).where(
                    PromotionInvoiceAllocation.invoice_id == previous.invoice_id
                )
            )
        )
        for allocation in previous_allocations:
            session.add(
                PromotionInvoiceAllocation(
                    allocation_id=f"promotion-allocation-{uuid4().hex}",
                    invoice_id=restored.invoice_id,
                    store_id=allocation.store_id,
                    statement_id=allocation.statement_id,
                    statement_month=allocation.statement_month,
                    settlement_batch_month=allocation.settlement_batch_month,
                    allocated_amount_cent=allocation.allocated_amount_cent,
                    is_current=True,
                )
            )
    session.add(
        InvoiceStatusEvent(
            event_id=f"invoice-event-{uuid4().hex}",
            invoice_id=restored.invoice_id,
            event_type=2,
            from_status=target.invoice_status,
            to_status=restored.invoice_status,
            operator_id=operator_id,
            import_batch_id=batch.batch_id,
            result_reason="FINANCE_IMPORT_REVERSAL",
            occurred_at=now,
        )
    )
    return restored


@router.post("/admin/finance-imports/{batch_id}/reversals")
def reverse_finance_import(
    batch_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    """Append one atomic reversal batch and one business version per original row."""
    _require_finance_admin(current_user, request)
    session = _billing_session(store, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(payload)
    replay = session.scalar(
        select(FinanceOperationAudit).where(
            FinanceOperationAudit.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if (
            replay.operation_type != "FINANCE_IMPORT_REVERSAL"
            or replay.request_payload_sha256 != payload_hash
            or replay.before_snapshot.get("reversesBatchId") != batch_id
        ):
            _persist_finance_conflict_audit(
                session,
                request=request,
                current_user=current_user,
                operation_type="FINANCE_IMPORT_REVERSAL",
                target_type="FINANCE_IMPORT_BATCH",
                target_id=batch_id,
                conflict_code="IDEMPOTENCY_KEY_REUSED",
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
            )
            _raise_reporting_error(
                request, status.HTTP_409_CONFLICT, "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        replay_batch = session.scalar(
            select(FinanceImportBatch).where(FinanceImportBatch.batch_id == replay.target_id)
        )
        if replay_batch is None:
            _raise_reporting_error(
                request, status.HTTP_409_CONFLICT, "REPLAY_TARGET_MISSING", "撤销批次不存在"
            )
        return _reporting_success(
            request, _finance_import_batch_item(session, replay_batch)
        )

    read_version = payload.get("readVersion")
    change_reason = str(payload.get("changeReason") or "").strip()
    if isinstance(read_version, bool) or not isinstance(read_version, int) or read_version < 1:
        _raise_reporting_error(
            request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED",
            "readVersion 必须为正整数", field="readVersion",
        )
    if not change_reason or len(change_reason) > 1000:
        _raise_reporting_error(
            request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED",
            "changeReason 必填且不得超过 1000 字", field="changeReason",
        )
    target_batch = session.scalar(
        select(FinanceImportBatch)
        .where(FinanceImportBatch.batch_id == batch_id)
        .with_for_update()
    )
    if target_batch is None:
        _raise_reporting_error(
            request, status.HTTP_404_NOT_FOUND, "RESOURCE_NOT_FOUND", "财务导入批次不存在"
        )
    if target_batch.batch_status not in (5, 8) or target_batch.current_version != read_version:
        conflict_data = {
            "current_version": target_batch.current_version,
            "scenario": FINANCE_IMPORT_SCENARIO_FROM_STATUS[
                target_batch.batch_status
            ],
        }
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="FINANCE_IMPORT_REVERSAL",
            target_type="FINANCE_IMPORT_BATCH",
            target_id=batch_id,
            conflict_code="REVERSAL_VERSION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            details=conflict_data,
        )
        _raise_reporting_error(
            request, status.HTTP_409_CONFLICT, "REVERSAL_VERSION_CONFLICT",
            "批次状态或读取版本已变化，无法撤销",
            data=conflict_data,
        )
    if session.scalar(
        select(FinanceImportBatch.batch_id).where(
            FinanceImportBatch.reverses_batch_id == target_batch.batch_id
        )
    ) is not None:
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="FINANCE_IMPORT_REVERSAL",
            target_type="FINANCE_IMPORT_BATCH",
            target_id=batch_id,
            conflict_code="IMPORT_ALREADY_REVERSED",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
        _raise_reporting_error(
            request, status.HTTP_409_CONFLICT, "IMPORT_ALREADY_REVERSED", "该批次已被撤销"
        )

    _lock_finance_import_version_slot(
        session,
        import_type=target_batch.import_type,
        statement_month=target_batch.statement_month,
    )
    try:
        plan = _finance_import_reversal_plan(session, target_batch, request)
    except HTTPException as exc:
        if exc.status_code == status.HTTP_409_CONFLICT:
            detail = exc.detail if isinstance(exc.detail, dict) else {}
            _persist_finance_conflict_audit(
                session,
                request=request,
                current_user=current_user,
                operation_type="FINANCE_IMPORT_REVERSAL",
                target_type="FINANCE_IMPORT_BATCH",
                target_id=batch_id,
                conflict_code=str(detail.get("code") or "REVERSAL_CONFLICT"),
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
                details=detail.get("data"),
            )
        raise
    management_applications_by_store: dict[str, list[dict]] = {}
    for item in plan:
        if item["kind"] == "MANAGEMENT":
            target = item["target"]
            _, projected_applications = _management_invoiceable_projection(
                session,
                store_id=target.store_id,
                through_month=target.statement_month,
            )
            management_applications_by_store[target.store_id] = projected_applications
    next_version = int(
        session.scalar(
            select(func.coalesce(func.max(FinanceImportBatch.current_version), 0)).where(
                FinanceImportBatch.import_type == target_batch.import_type,
                FinanceImportBatch.statement_month == target_batch.statement_month,
                FinanceImportBatch.batch_status.in_((5, 8, 9)),
            )
        )
        or 0
    ) + 1
    now = utcnow()
    reversal_batch = FinanceImportBatch(
        batch_id=f"finance-import-reversal-{uuid4().hex}",
        import_type=target_batch.import_type,
        statement_month=target_batch.statement_month,
        file_name=f"reversal-{target_batch.file_name}",
        file_sha256=sha256(f"reversal:{target_batch.batch_id}".encode()).hexdigest(),
        normalized_sha256=sha256(f"reversal:{target_batch.normalized_sha256}".encode()).hexdigest(),
        read_version=target_batch.current_version,
        current_version=next_version,
        batch_status=9,
        total_rows=len(plan),
        success_rows=len(plan),
        error_rows=0,
        content_changed=True,
        reverses_batch_id=target_batch.batch_id,
        submitted_by=current_user.username,
        committed_by=current_user.username,
        submitted_at=now,
        committed_at=now,
    )
    session.add(reversal_batch)
    session.flush()
    for index, item in enumerate(plan, start=2):
        target = item["target"]
        previous = item["previous"]
        if item["kind"] == "PROFILE":
            restored = _append_profile_reversal_version(
                session, target=target, previous=previous, batch=reversal_batch, now=now
            )
            restored_id = restored.profile_id
        elif item["kind"] == "MANAGEMENT":
            restored = _append_management_reversal_version(
                session, target=target, previous=previous, batch=reversal_batch,
                operator_id=current_user.username, now=now,
            )
            restored_id = restored.invoice_id
            if previous is None or previous.is_tombstone:
                _, projected_applications = _management_invoiceable_projection(
                    session,
                    store_id=target.store_id,
                    through_month=target.statement_month,
                )
                invoice_map: dict[str, str] = {}
            else:
                projected_applications = management_applications_by_store.get(
                    target.store_id, []
                )
                invoice_map = {target.statement_id: restored.invoice_id}
            _synchronize_management_carryforward_applications(
                session,
                applications=projected_applications,
                invoice_id_by_statement=invoice_map,
                scope_store_ids={target.store_id},
                scope_through_month=target.statement_month,
            )
        else:
            restored = _append_promotion_reversal_version(
                session, target=target, previous=previous, batch=reversal_batch,
                operator_id=current_user.username, now=now,
            )
            restored_id = restored.invoice_id
        original_row = item["row"]
        session.add(
            FinanceImportRow(
                batch_id=reversal_batch.batch_id,
                row_number=index,
                business_key=original_row.business_key,
                normalized_payload={
                    "changeReason": change_reason,
                    "reversesBatchId": target_batch.batch_id,
                    "reversesRowNumber": original_row.row_number,
                },
                row_status=5,
                validation_errors=[],
                target_record_id=restored_id,
                reversal_effect_type=(
                    1
                    if previous is not None and not previous.is_tombstone
                    else 2
                ),
                reverses_target_record_id=original_row.target_record_id,
                previous_target_record_id=(
                    previous.profile_id
                    if item["kind"] == "PROFILE" and previous is not None
                    else previous.invoice_id if previous is not None else None
                ),
            )
        )
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="FINANCE_IMPORT_REVERSAL",
            target_type="FINANCE_IMPORT_BATCH",
            target_id=reversal_batch.batch_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot={
                "reversesBatchId": target_batch.batch_id,
                "scenario": FINANCE_IMPORT_SCENARIO_FROM_STATUS[target_batch.batch_status],
                "currentVersion": target_batch.current_version,
            },
            after_snapshot={
                "batchId": reversal_batch.batch_id,
                "scenario": "REVERSED",
                "currentVersion": next_version,
                "changeReason": change_reason,
                "rowCount": len(plan),
            },
            result_status=1,
            request_id=request_id(request),
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            occurred_at=now,
        )
    )
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        replay = session.scalar(
            select(FinanceOperationAudit).where(
                FinanceOperationAudit.idempotency_key_hash == key_hash
            )
        )
        if replay is not None and replay.request_payload_sha256 == payload_hash:
            replay_batch = session.scalar(
                select(FinanceImportBatch).where(FinanceImportBatch.batch_id == replay.target_id)
            )
            if replay_batch is not None:
                return _reporting_success(
                    request, _finance_import_batch_item(session, replay_batch)
                )
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="FINANCE_IMPORT_REVERSAL",
            target_type="FINANCE_IMPORT_BATCH",
            target_id=batch_id,
            conflict_code="REVERSAL_COMMIT_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
        _raise_reporting_error(
            request, status.HTTP_409_CONFLICT, "FINANCE_IMPORT_REVERSAL_CONFLICT",
            "撤销期间业务版本已变化，请刷新后重试",
        )
    return _reporting_success(
        request, _finance_import_batch_item(session, reversal_batch)
    )


def _execute_finance_import_commit(
    *,
    batch_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None,
    current_user: AuthContext,
    store,
    is_correction: bool,
):
    """Execute the shared atomic commit path for imports and corrections."""
    _require_finance_admin(current_user, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(payload)
    session = _billing_session(store, request)
    batch = session.scalar(
        select(FinanceImportBatch).where(FinanceImportBatch.batch_id == batch_id)
    )
    if batch is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "财务导入批次不存在",
            field="batchId",
        )

    _lock_finance_import_version_slot(
        session,
        import_type=batch.import_type,
        statement_month=batch.statement_month,
    )
    session.refresh(batch)

    replay = session.scalar(
        select(FinanceOperationAudit).where(
            FinanceOperationAudit.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if (
            replay.target_id != batch_id
            or replay.request_payload_sha256 != payload_hash
        ):
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        return _reporting_success(request, _finance_import_commit_item(batch))

    read_version = payload.get("readVersion")
    change_reason = payload.get("changeReason")
    if (
        isinstance(read_version, bool)
        or not isinstance(read_version, int)
        or read_version < 0
    ):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "readVersion 必须为非负整数",
            field="readVersion",
        )
    if not isinstance(change_reason, str) or not change_reason.strip():
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "changeReason 为必填项",
            field="changeReason",
        )
    if batch.batch_status not in (3, 4):
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "IMPORT_BATCH_NOT_READY",
            "财务导入批次不是可提交状态",
        )

    current_version = int(
        session.scalar(
            select(func.coalesce(func.max(FinanceImportBatch.current_version), 0)).where(
                FinanceImportBatch.import_type == batch.import_type,
                FinanceImportBatch.statement_month == batch.statement_month,
                FinanceImportBatch.batch_status.in_((5, 8, 9)),
            )
        )
        or 0
    )
    if read_version != batch.read_version or current_version != batch.read_version:
        latest_batch = session.scalar(
            select(FinanceImportBatch)
            .where(
                FinanceImportBatch.import_type == batch.import_type,
                FinanceImportBatch.statement_month == batch.statement_month,
                FinanceImportBatch.batch_status.in_((5, 8, 9)),
            )
            .order_by(
                FinanceImportBatch.current_version.desc(),
                FinanceImportBatch.committed_at.desc(),
            )
        )
        batch.batch_status = 7
        batch.current_version = current_version
        session.commit()
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "VERSION_CONFLICT",
            "导入预览版本已变化，请刷新后重新校验",
            field="readVersion",
            data={
                "read_version": read_version,
                "current_version": current_version,
                "latest_operator": (
                    latest_batch.committed_by if latest_batch is not None else None
                ),
                "latest_operated_at": (
                    latest_batch.committed_at.isoformat()
                    if latest_batch is not None
                    and latest_batch.committed_at is not None
                    else None
                ),
            },
        )

    rows = session.scalars(
        select(FinanceImportRow)
        .where(FinanceImportRow.batch_id == batch.batch_id)
        .order_by(FinanceImportRow.row_number)
    ).yield_per(500)

    before_snapshot = {
        "batchStatus": batch.batch_status,
        "readVersion": batch.read_version,
        "currentVersion": current_version,
    }
    committed_at = utcnow()
    try:
        for row in rows:
            _commit_final_finance_import_row(
                session,
                batch=batch,
                row=row,
                operator_id=current_user.username,
                committed_at=committed_at,
                is_correction=is_correction,
                request=request,
            )
        batch.batch_status = 8 if is_correction else 5
        batch.current_version = current_version + 1
        batch.committed_by = current_user.username
        batch.committed_at = committed_at
        session.add(
            FinanceOperationAudit(
                audit_id=f"audit-{uuid4().hex}",
                operation_type=(
                    "FINANCE_IMPORT_CORRECTION"
                    if is_correction
                    else "FINANCE_IMPORT_COMMIT"
                ),
                target_type="FINANCE_IMPORT_BATCH",
                target_id=batch.batch_id,
                operator_id=current_user.username,
                operator_role=_finance_operator_role(current_user),
                before_snapshot=before_snapshot,
                after_snapshot={
                    "batchStatus": batch.batch_status,
                    "currentVersion": batch.current_version,
                    "changeReason": change_reason.strip(),
                },
                result_status=1,
                request_id=request_id(request),
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
                occurred_at=committed_at,
            )
        )
        session.commit()
    except IntegrityError:
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type=(
                "FINANCE_IMPORT_CORRECTION"
                if is_correction
                else "FINANCE_IMPORT_COMMIT"
            ),
            target_type="FINANCE_IMPORT_BATCH",
            target_id=batch_id,
            conflict_code="FINANCE_IMPORT_VERSION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "FINANCE_IMPORT_VERSION_CONFLICT",
            "同类型和账期的导入版本已被其他事务占用，请刷新后重试",
        )
    except Exception:
        session.rollback()
        raise
    return _reporting_success(request, _finance_import_commit_item(batch))


@router.get("/promotion-invoices")
def list_promotion_invoices(
    request: Request,
    store_id: str = Query(alias="storeId"),
    month: str = Query(),
    status_filter: str | None = Query(default=None, alias="status"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=50, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_billing_store_scope(current_user, store_id, request)
    _require_billing_store(session, store_id, request)
    _validate_month(month, "month", request)
    conditions = [
        PromotionInvoiceAllocation.store_id == store_id,
        PromotionInvoiceAllocation.statement_month == month,
        PromotionInvoiceAllocation.is_current.is_(True),
        PromotionInvoice.is_current.is_(True),
        PromotionInvoice.is_tombstone.is_(False),
    ]
    if status_filter is not None:
        normalized_status = status_filter.upper()
        if normalized_status not in set(INVOICE_STATUS_NAMES.values()):
            _validate_enum(normalized_status, set(INVOICE_STATUS_NAMES.values()), "status", request)
        if normalized_status == "PENDING_INVOICE":
            return _reporting_success(request, {"list": [], "total": 0, "page": page, "page_size": page_size})
        conditions.append(PromotionInvoice.invoice_status == _invoice_status_db(normalized_status))
    query = (
        select(PromotionInvoice, PromotionInvoiceAllocation)
        .join(PromotionInvoiceAllocation, PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id)
        .where(*conditions)
        .order_by(PromotionInvoice.registered_at.desc())
    )
    total = session.scalar(select(func.count()).select_from(query.subquery())) or 0
    rows = session.execute(query.offset((page - 1) * page_size).limit(page_size)).all()
    return _reporting_success(request, {
        "list": [_promotion_invoice_item(invoice, allocation) for invoice, allocation in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    })


@router.get("/promotion-invoices/replacement-candidates")
def list_promotion_invoice_replacement_candidates(
    request: Request,
    store_id: str = Query(alias="storeId"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_promotion_invoice_store_operator(current_user, request)
    _require_billing_store_scope(current_user, store_id, request)
    _require_billing_store(session, store_id, request)
    events = list(
        session.scalars(
            select(PromotionInvoiceLifecycleEvent)
            .join(
                PromotionInvoice,
                PromotionInvoice.invoice_id
                == PromotionInvoiceLifecycleEvent.invoice_id,
            )
            .where(
                PromotionInvoice.store_id == store_id,
                PromotionInvoiceLifecycleEvent.is_current.is_(True),
            )
            .order_by(
                PromotionInvoiceLifecycleEvent.occurred_at,
                PromotionInvoiceLifecycleEvent.lifecycle_event_id,
            )
        )
    )
    candidates = []
    for event in events:
        successor = session.scalar(
            select(PromotionInvoice.invoice_id).where(
                PromotionInvoice.replaces_invoice_id == event.invoice_id,
                PromotionInvoice.version_kind == 1,
            )
        )
        if successor is not None:
            continue
        invoice = session.scalar(
            select(PromotionInvoice).where(
                PromotionInvoice.invoice_id == event.invoice_id
            )
        )
        released_months = sorted(
            set(
                session.scalars(
                    select(PromotionInvoiceAllocation.statement_month).where(
                        PromotionInvoiceAllocation.invoice_id == event.invoice_id
                    )
                )
            )
        )
        candidates.append(
            {
                "invoice": _promotion_invoice_header_item_from_invoice(invoice),
                "lifecycle_event": _promotion_invoice_lifecycle_event_item(event),
                "released_statement_months": released_months,
            }
        )
    return _reporting_success(
        request, {"list": candidates, "total": len(candidates)}
    )


@router.get("/promotion-invoices/{invoice_id}")
def get_promotion_invoice_detail(
    invoice_id: str,
    request: Request,
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    invoice = session.scalar(
        select(PromotionInvoice).where(PromotionInvoice.invoice_id == invoice_id)
    )
    if invoice is None:
        _raise_reporting_error(
            request, status.HTTP_404_NOT_FOUND, "RESOURCE_NOT_FOUND", "推广费发票不存在"
        )
    _require_billing_store_scope(current_user, invoice.store_id, request)
    versions = list(
        session.scalars(
            select(PromotionInvoice)
            .where(
                PromotionInvoice.physical_invoice_id == invoice.physical_invoice_id
            )
            .order_by(PromotionInvoice.version_no)
        )
    )
    version_ids = [row.invoice_id for row in versions]
    allocations = list(
        session.scalars(
            select(PromotionInvoiceAllocation)
            .where(PromotionInvoiceAllocation.invoice_id.in_(version_ids))
            .order_by(
                PromotionInvoiceAllocation.statement_month,
                PromotionInvoiceAllocation.allocation_id,
            )
        )
    )
    status_events = list(
        session.scalars(
            select(InvoiceStatusEvent)
            .where(InvoiceStatusEvent.invoice_id.in_(version_ids))
            .order_by(InvoiceStatusEvent.occurred_at, InvoiceStatusEvent.event_id)
        )
    )
    lifecycle_events = list(
        session.scalars(
            select(PromotionInvoiceLifecycleEvent)
            .where(
                PromotionInvoiceLifecycleEvent.physical_invoice_id
                == invoice.physical_invoice_id
            )
            .order_by(
                PromotionInvoiceLifecycleEvent.occurred_at,
                PromotionInvoiceLifecycleEvent.lifecycle_event_id,
            )
        )
    )
    replacements = list(
        session.scalars(
            select(PromotionInvoice)
            .where(
                PromotionInvoice.replaces_invoice_id.in_(version_ids),
                PromotionInvoice.version_kind == 1,
            )
            .order_by(PromotionInvoice.registered_at, PromotionInvoice.invoice_id)
        )
    )
    replacement_chain = _promotion_invoice_replacement_chain(session, invoice)
    return _reporting_success(
        request,
        {
            **_promotion_invoice_header_item_from_invoice(invoice),
            "versions": [
                _promotion_invoice_header_item_from_invoice(row) for row in versions
            ],
            "allocations": [
                {
                    "allocation_id": row.allocation_id,
                    "invoice_id": row.invoice_id,
                    "statement_id": row.statement_id,
                    "statement_month": row.statement_month,
                    "settlement_batch_month": row.settlement_batch_month,
                    "allocated_amount_cent": row.allocated_amount_cent,
                    "is_current": row.is_current,
                }
                for row in allocations
            ],
            "status_events": [
                {
                    "event_id": row.event_id,
                    "invoice_id": row.invoice_id,
                    "from_status": (
                        INVOICE_STATUS_NAMES[row.from_status]
                        if row.from_status is not None
                        else None
                    ),
                    "to_status": INVOICE_STATUS_NAMES[row.to_status],
                    "operator_id": row.operator_id,
                    "occurred_at": row.occurred_at,
                }
                for row in status_events
            ],
            "lifecycle_events": [
                _promotion_invoice_lifecycle_event_item(row)
                for row in lifecycle_events
            ],
            "replacements": [
                _promotion_invoice_header_item_from_invoice(row)
                for row in replacements
            ],
            "replacement_chain": [
                _promotion_invoice_header_item_from_invoice(row)
                for row in replacement_chain
            ],
        },
    )


@router.post("/promotion-invoices/{invoice_id}/lifecycle-events")
def create_promotion_invoice_lifecycle_event(
    invoice_id: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_promotion_invoice_store_operator(current_user, request)
    event_type = payload.get("eventType")
    reason = payload.get("reason")
    read_version = payload.get("readVersion")
    if not isinstance(event_type, str) or event_type.upper() not in PROMOTION_LIFECYCLE_EVENT_TO_DB:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "eventType 只允许 RED_FLUSHED 或 VOIDED",
            field="eventType",
        )
    if not isinstance(reason, str) or not reason.strip():
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "reason 必填",
            field="reason",
        )
    if len(reason.strip()) > 1000:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "reason 最多 1000 个字符",
            field="reason",
        )
    if isinstance(read_version, bool) or not isinstance(read_version, int) or read_version < 1:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "readVersion 必须为正整数",
            field="readVersion",
        )
    normalized = {
        "invoice_id": invoice_id,
        "event_type": event_type.upper(),
        "reason": reason.strip(),
        "read_version": read_version,
    }
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(normalized)
    replay = session.scalar(
        select(PromotionInvoiceLifecycleEvent).where(
            PromotionInvoiceLifecycleEvent.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        return _promotion_invoice_lifecycle_replay_response(
            session, replay, payload_hash, request, current_user
        )
    invoice = session.scalar(
        select(PromotionInvoice)
        .where(PromotionInvoice.invoice_id == invoice_id)
        .with_for_update()
    )
    if invoice is None:
        _raise_reporting_error(
            request, status.HTTP_404_NOT_FOUND, "RESOURCE_NOT_FOUND", "推广费发票不存在"
        )
    _require_billing_store_scope(current_user, invoice.store_id, request)
    _lock_promotion_invoice_physical(session, invoice.physical_invoice_id, request)
    replay = session.scalar(
        select(PromotionInvoiceLifecycleEvent).where(
            PromotionInvoiceLifecycleEvent.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        return _promotion_invoice_lifecycle_replay_response(
            session, replay, payload_hash, request, current_user
        )
    if not invoice.is_current or invoice.version_no != read_version:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "PROMOTION_INVOICE_VERSION_CONFLICT",
            "发票版本已变化，请刷新后重试",
            data={"readVersion": read_version, "currentVersion": invoice.version_no},
        )
    terminated = session.scalar(
        select(PromotionInvoiceLifecycleEvent).where(
            PromotionInvoiceLifecycleEvent.physical_invoice_id
            == invoice.physical_invoice_id,
            PromotionInvoiceLifecycleEvent.is_current.is_(True),
        )
    )
    if terminated is not None:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "PROMOTION_INVOICE_ALREADY_TERMINATED",
            "同一物理发票已经登记红冲或作废",
        )
    allocations = list(
        session.scalars(
            select(PromotionInvoiceAllocation)
            .where(
                PromotionInvoiceAllocation.invoice_id == invoice.invoice_id,
                PromotionInvoiceAllocation.is_current.is_(True),
            )
            .with_for_update()
        )
    )
    now = utcnow()
    lifecycle_event = PromotionInvoiceLifecycleEvent(
        lifecycle_event_id=f"promotion-lifecycle-{uuid4().hex}",
        physical_invoice_id=invoice.physical_invoice_id,
        invoice_id=invoice.invoice_id,
        invoice_version=invoice.version_no,
        event_type=PROMOTION_LIFECYCLE_EVENT_TO_DB[event_type.upper()],
        reason=reason.strip(),
        read_version=read_version,
        is_current=True,
        operator_id=current_user.username,
        idempotency_key_hash=key_hash,
        request_payload_sha256=payload_hash,
        occurred_at=now,
    )
    invoice.is_current = False
    for allocation in allocations:
        allocation.is_current = False
    session.add(lifecycle_event)
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="PROMOTION_INVOICE_LIFECYCLE",
            target_type="PROMOTION_INVOICE",
            target_id=invoice.invoice_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot={
                "versionNo": invoice.version_no,
                "invoiceStatus": INVOICE_STATUS_NAMES[invoice.invoice_status],
                "isCurrent": True,
            },
            after_snapshot={
                "eventType": event_type.upper(),
                "reason": reason.strip(),
                "releasedStatementMonths": sorted(
                    row.statement_month for row in allocations
                ),
            },
            result_status=1,
            request_id=request_id(request),
            occurred_at=now,
        )
    )
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        replay = session.scalar(
            select(PromotionInvoiceLifecycleEvent).where(
                PromotionInvoiceLifecycleEvent.idempotency_key_hash == key_hash
            )
        )
        if replay is not None:
            return _promotion_invoice_lifecycle_replay_response(
                session, replay, payload_hash, request, current_user
            )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "PROMOTION_INVOICE_LIFECYCLE_CONFLICT",
            "发票版本或生命周期事件已变化，请刷新后重试",
        )
    return _reporting_success(
        request,
        {
            "invoice": _promotion_invoice_header_item_from_invoice(invoice),
            "lifecycle_event": _promotion_invoice_lifecycle_event_item(
                lifecycle_event
            ),
            "released_statement_months": sorted(
                row.statement_month for row in allocations
            ),
        },
    )


@router.post("/promotion-invoices")
def register_promotion_invoice(
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_promotion_invoice_store_operator(current_user, request)
    parsed = _parse_promotion_invoice_payload(payload, request)
    _require_billing_store_scope(current_user, parsed["store_id"], request)
    _require_billing_store(session, parsed["store_id"], request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(parsed)
    replay = session.scalar(select(PromotionInvoice).where(PromotionInvoice.idempotency_key_hash == key_hash))
    if replay is not None:
        return _promotion_invoice_registration_replay_response(
            session, replay, payload_hash, request
        )
    if session.scalar(
        select(PromotionInvoiceNumberRegistry).where(
            PromotionInvoiceNumberRegistry.invoice_number == parsed["invoice_number"]
        )
    ) is not None or session.scalar(
        select(PromotionInvoice.invoice_id).where(
            PromotionInvoice.invoice_number == parsed["invoice_number"]
        )
    ) is not None:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "PROMOTION_INVOICE_NUMBER_REUSED",
            "该发票号码已被使用，新的物理发票必须使用新号码",
            field="invoiceNumber",
        )
    promotion_projection, promotion_groups = (
        _promotion_invoice_carryforward_projection(
            session,
            store_id=parsed["store_id"],
        )
    )
    selected_statement_ids_for_lock = [
        item["statement_id"] for item in parsed["allocations"]
    ]
    selected_months_for_lock = [
        item["statement_month"] for item in parsed["allocations"]
    ]
    if selected_statement_ids_for_lock:
        list(
            session.scalars(
                select(SettlementStatement)
                .where(
                    SettlementStatement.statement_id.in_(
                        selected_statement_ids_for_lock
                    )
                )
                .with_for_update()
            )
        )
        list(
            session.scalars(
                select(SettlementStatementConfirmation)
                .where(
                    SettlementStatementConfirmation.statement_id.in_(
                        selected_statement_ids_for_lock
                    ),
                    SettlementStatementConfirmation.fee_direction == 1,
                )
                .with_for_update()
            )
        )
        list(
            session.scalars(
                select(PromotionInvoiceAllocation)
                .where(
                    PromotionInvoiceAllocation.store_id == parsed["store_id"],
                    PromotionInvoiceAllocation.statement_month.in_(
                        selected_months_for_lock
                    ),
                    PromotionInvoiceAllocation.is_current.is_(True),
                )
                .with_for_update()
            )
        )
    promotion_projection, promotion_groups = (
        _promotion_invoice_carryforward_projection(
            session,
            store_id=parsed["store_id"],
        )
    )
    groups_by_id = {group["group_id"]: group for group in promotion_groups}
    selected_statement_ids_by_group: dict[str, set[str]] = {}
    allocations = []
    for item in parsed["allocations"]:
        projected_group = promotion_projection.get(item["statement_id"])
        submitted_group_id = item["promotion_invoice_group_id"]
        if (
            projected_group is None
            or submitted_group_id is None
            or projected_group["group_id"] != submitted_group_id
        ):
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "PROMOTION_INVOICE_GROUP_CHANGED",
                "推广费抵扣组已变化，请刷新账单后重新选择",
                field="allocations",
            )
        if projected_group["invoiceable_amount_cent"] <= 0:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "PROMOTION_INVOICE_GROUP_NOT_CLOSED",
                "推广费抵扣组尚未形成正数净额，不能登记发票",
                field="allocations",
            )
        selected_statement_ids_by_group.setdefault(
            submitted_group_id,
            set(),
        ).add(item["statement_id"])
        statement = _get_billing_statement(session, item["statement_id"], request)
        if (not statement.is_current or statement.store_id != parsed["store_id"]
                or statement.statement_month != item["statement_month"]
                or statement.version_no != item["read_version"]):
            _raise_statement_version_conflict(request, statement)
        confirmed = session.scalar(select(SettlementStatementConfirmation).where(
            SettlementStatementConfirmation.statement_id == statement.statement_id,
            SettlementStatementConfirmation.fee_direction == 1,
            SettlementStatementConfirmation.confirmation_status == 1,
        ))
        if confirmed is None or confirmed.confirmed_amount_cent != item["allocated_amount_cent"]:
            _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "每个账期分配金额必须等于当前有效推广费确认金额", field="allocations")
        occupied = session.scalar(select(PromotionInvoiceAllocation).where(
            PromotionInvoiceAllocation.store_id == statement.store_id,
            PromotionInvoiceAllocation.statement_month == statement.statement_month,
            PromotionInvoiceAllocation.is_current.is_(True),
        ))
        if occupied is not None:
            _raise_reporting_error(request, status.HTTP_409_CONFLICT, "PROMOTION_INVOICE_PERIOD_OCCUPIED", "账期已存在当前有效推广费发票分配", field="allocations")
        allocations.append((statement, item))
    for group_id, selected_statement_ids in selected_statement_ids_by_group.items():
        required_statement_ids = set(
            groups_by_id[group_id]["required_statement_ids"]
        )
        if selected_statement_ids != required_statement_ids:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "PROMOTION_INVOICE_GROUP_INCOMPLETE",
                "登记发票必须一次包含抵扣组的全部正负账期",
                field="allocations",
                data={
                    "promotionInvoiceGroupId": group_id,
                    "requiredStatementIds": sorted(required_statement_ids),
                },
            )
    selected_statement_ids = [statement.statement_id for statement, _ in allocations]
    unmatched_terminations: list[PromotionInvoiceLifecycleEvent] = []
    if selected_statement_ids:
        termination_candidates = list(
            session.scalars(
                select(PromotionInvoiceLifecycleEvent)
                .join(
                    PromotionInvoiceAllocation,
                    PromotionInvoiceAllocation.invoice_id
                    == PromotionInvoiceLifecycleEvent.invoice_id,
                )
                .where(
                    PromotionInvoiceAllocation.statement_id.in_(
                        selected_statement_ids
                    ),
                    PromotionInvoiceLifecycleEvent.is_current.is_(True),
                )
                .distinct()
            )
        )
        for candidate in termination_candidates:
            successor = session.scalar(
                select(PromotionInvoiceReplacementSource.replacement_invoice_id).where(
                    PromotionInvoiceReplacementSource.source_invoice_id
                    == candidate.invoice_id
                )
            ) or session.scalar(
                select(PromotionInvoice.invoice_id).where(
                    PromotionInvoice.replaces_invoice_id == candidate.invoice_id,
                    PromotionInvoice.version_kind == 1,
                )
            )
            if successor is None:
                unmatched_terminations.append(candidate)
    required_replacement_ids = sorted(
        termination.invoice_id for termination in unmatched_terminations
    )
    if required_replacement_ids and (
        parsed["replaces_invoice_id"] is None
        or parsed["replaces_invoice_id"] not in required_replacement_ids
    ):
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "PROMOTION_INVOICE_REPLACEMENT_REQUIRED",
            "释放账期必须关联其中一张被红冲或作废的原发票；系统会审计关联同组全部来源",
            field="replacesInvoiceId",
            data={"requiredReplacesInvoiceIds": required_replacement_ids},
        )

    source_invoice_ids = required_replacement_ids or (
        [parsed["replaces_invoice_id"]]
        if parsed["replaces_invoice_id"] is not None
        else []
    )
    replaced_invoices: list[PromotionInvoice] = []
    replacement_terminations: list[PromotionInvoiceLifecycleEvent] = []
    selected_months = {item["statement_month"] for _, item in allocations}
    for source_invoice_id in source_invoice_ids:
        source_invoice = session.scalar(
            select(PromotionInvoice).where(
                PromotionInvoice.invoice_id == source_invoice_id
            )
        )
        if source_invoice is None:
            _raise_reporting_error(
                request,
                status.HTTP_404_NOT_FOUND,
                "RESOURCE_NOT_FOUND",
                "被替换发票不存在",
                field="replacesInvoiceId",
            )
        _lock_promotion_invoice_physical(
            session, source_invoice.physical_invoice_id, request
        )
        source_invoice = session.scalar(
            select(PromotionInvoice).where(
                PromotionInvoice.invoice_id == source_invoice_id
            )
        )
        if source_invoice.store_id != parsed["store_id"]:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "PROMOTION_INVOICE_REPLACEMENT_SCOPE_CONFLICT",
                "不得跨门店替换发票",
                field="replacesInvoiceId",
            )
        termination = session.scalar(
            select(PromotionInvoiceLifecycleEvent).where(
                PromotionInvoiceLifecycleEvent.physical_invoice_id
                == source_invoice.physical_invoice_id,
                PromotionInvoiceLifecycleEvent.is_current.is_(True),
            )
        )
        if termination is None or termination.invoice_id != source_invoice.invoice_id:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "PROMOTION_INVOICE_REPLACEMENT_VERSION_CONFLICT",
                "被替换发票必须指向当前待替换的红冲或作废版本",
                field="replacesInvoiceId",
            )
        existing_successor = session.scalar(
            select(PromotionInvoiceReplacementSource.replacement_invoice_id).where(
                PromotionInvoiceReplacementSource.source_invoice_id
                == source_invoice.invoice_id
            )
        ) or session.scalar(
            select(PromotionInvoice.invoice_id).where(
                PromotionInvoice.replaces_invoice_id == source_invoice.invoice_id,
                PromotionInvoice.version_kind == 1,
            )
        )
        if existing_successor is not None:
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "PROMOTION_INVOICE_REPLACEMENT_ALREADY_EXISTS",
                "被替换发票已经登记后继发票，不能形成分叉链",
                field="replacesInvoiceId",
                data={"successorInvoiceId": existing_successor},
            )
        released_months = set(
            session.scalars(
                select(PromotionInvoiceAllocation.statement_month).where(
                    PromotionInvoiceAllocation.invoice_id == source_invoice.invoice_id
                )
            )
        )
        if not released_months.issubset(selected_months):
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "PROMOTION_INVOICE_REPLACEMENT_PERIOD_MISMATCH",
                "替换发票必须覆盖全部来源发票释放的完整账期及其当前抵扣组",
                field="allocations",
            )
        if source_invoice.invoice_number == parsed["invoice_number"]:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "PROMOTION_INVOICE_NUMBER_REUSED",
                "替换发票必须使用新的 20 位发票号码",
                field="invoiceNumber",
            )
        replaced_invoices.append(source_invoice)
        replacement_terminations.append(termination)

    replaced_invoice = next(
        (
            invoice
            for invoice in replaced_invoices
            if invoice.invoice_id == parsed["replaces_invoice_id"]
        ),
        replaced_invoices[0] if replaced_invoices else None,
    )
    now = utcnow()
    settlement_batch_month = _promotion_invoice_settlement_batch_month(now)
    physical_invoice_id = f"physical-invoice-{uuid4().hex}"
    invoice_id = f"promotion-invoice-{uuid4().hex}"
    invoice = PromotionInvoice(
        invoice_id=invoice_id,
        physical_invoice_id=physical_invoice_id,
        store_id=parsed["store_id"],
        version_no=1,
        version_kind=1,
        supersedes_invoice_id=None,
        replaces_invoice_id=(
            replaced_invoice.invoice_id if replaced_invoice is not None else None
        ),
        invoice_number=parsed["invoice_number"], invoice_date=parsed["invoice_date"],
        invoice_amount_cent=parsed["invoice_amount_cent"], invoice_status=2,
        buyer_name=parsed["buyer_name"], tax_rate_percent=parsed["tax_rate_percent"],
        registered_by=current_user.username, registered_at=now,
        idempotency_key_hash=key_hash, request_payload_sha256=payload_hash,
    )
    session.add(invoice)
    session.add(
        PromotionInvoiceNumberRegistry(
            invoice_number=parsed["invoice_number"],
            physical_invoice_id=physical_invoice_id,
            first_invoice_id=invoice_id,
            store_id=parsed["store_id"],
            registered_at=now,
        )
    )
    for statement, item in allocations:
        session.add(PromotionInvoiceAllocation(
            allocation_id=f"promotion-allocation-{uuid4().hex}", invoice_id=invoice.invoice_id,
            store_id=statement.store_id, statement_id=statement.statement_id,
            statement_month=statement.statement_month, allocated_amount_cent=item["allocated_amount_cent"],
            settlement_batch_month=settlement_batch_month,
        ))
    session.add(InvoiceStatusEvent(
        event_id=f"invoice-event-{uuid4().hex}", invoice_id=invoice.invoice_id,
        event_type=1, from_status=None, to_status=2, operator_id=current_user.username,
        occurred_at=now,
    ))
    for source_invoice, termination in zip(
        replaced_invoices, replacement_terminations, strict=True
    ):
        session.add(
            PromotionInvoiceReplacementSource(
                replacement_invoice_id=invoice.invoice_id,
                source_invoice_id=source_invoice.invoice_id,
                source_physical_invoice_id=source_invoice.physical_invoice_id,
                linked_at=now,
            )
        )
        termination.is_current = False
    if replaced_invoices:
        session.add(
            FinanceOperationAudit(
                audit_id=f"audit-{uuid4().hex}",
                operation_type="PROMOTION_INVOICE_REPLACEMENT_REGISTER",
                target_type="PROMOTION_INVOICE",
                target_id=invoice.invoice_id,
                operator_id=current_user.username,
                operator_role=_finance_operator_role(current_user),
                before_snapshot={
                    "replacedInvoices": [
                        {
                            "invoiceId": source.invoice_id,
                            "physicalInvoiceId": source.physical_invoice_id,
                            "invoiceNumber": source.invoice_number,
                        }
                        for source in replaced_invoices
                    ],
                },
                after_snapshot={
                    "replacementInvoiceId": invoice.invoice_id,
                    "replacementPhysicalInvoiceId": invoice.physical_invoice_id,
                    "invoiceNumber": invoice.invoice_number,
                },
                result_status=1,
                request_id=request_id(request),
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
                occurred_at=now,
            )
        )
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        replay = session.scalar(
            select(PromotionInvoice).where(
                PromotionInvoice.idempotency_key_hash == key_hash
            )
        )
        if replay is not None:
            return _promotion_invoice_registration_replay_response(
                session, replay, payload_hash, request
            )
        _raise_reporting_error(request, status.HTTP_409_CONFLICT, "PROMOTION_INVOICE_CONFLICT", "发票号码或账期分配已存在")
    return _reporting_success(request, _promotion_invoice_header_item(session, invoice))


@router.get("/admin/finance/summary")
def get_admin_finance_summary(
    request: Request,
    month: str = Query(),
    fee_direction: str = Query(alias="feeDirection"),
    metric_scope: str = Query(alias="metricScope"),
    store_id: str | None = Query(default=None, alias="storeId"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_finance_admin(current_user, request)
    _validate_month(month, "month", request)
    direction = _normalize_billing_direction(fee_direction, request)
    normalized_scope = metric_scope.upper()
    _validate_enum(normalized_scope, METRIC_SCOPES, "metricScope", request)
    if store_id is not None:
        _require_billing_store(session, store_id, request)

    return _reporting_success(
        request,
        {
            "month": month,
            "store_id": store_id,
            "fee_direction": direction,
            "metric_scope": normalized_scope,
            "metrics": _finance_summary_metrics(
                session,
                month=month,
                fee_direction=direction,
                metric_scope=normalized_scope,
                store_id=store_id,
            ),
        },
    )


@router.get("/admin/finance/invoices")
def list_admin_finance_invoices(
    request: Request,
    month: str = Query(),
    fee_direction: str = Query(alias="feeDirection"),
    store_id: str | None = Query(default=None, alias="storeId"),
    invoice_status: str | None = Query(default=None, alias="invoiceStatus"),
    include_history: bool = Query(default=False, alias="includeHistory"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=50, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_finance_admin(current_user, request)
    _validate_month(month, "month", request)
    direction = _normalize_billing_direction(fee_direction, request)
    if store_id is not None:
        _require_billing_store(session, store_id, request)
    status_code = _normalize_invoice_status(invoice_status, request)

    if direction == "PROMOTION":
        conditions = [
            PromotionInvoiceAllocation.statement_month == month,
            PromotionInvoiceAllocation.is_current.is_(True),
            PromotionInvoice.is_current.is_(True),
            PromotionInvoice.is_tombstone.is_(False),
            SettlementStatement.is_current.is_(True),
        ]
        if store_id is not None:
            conditions.append(PromotionInvoiceAllocation.store_id == store_id)
        if status_code is not None:
            conditions.append(PromotionInvoice.invoice_status == status_code)
        query = (
            select(PromotionInvoice, PromotionInvoiceAllocation)
            .join(
                PromotionInvoiceAllocation,
                PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id,
            )
            .join(
                SettlementStatement,
                SettlementStatement.statement_id == PromotionInvoiceAllocation.statement_id,
            )
            .where(*conditions)
            .order_by(PromotionInvoice.registered_at.desc(), PromotionInvoice.invoice_id)
        )
        total = session.scalar(select(func.count()).select_from(query.subquery())) or 0
        rows = session.execute(query.offset((page - 1) * page_size).limit(page_size)).all()
        items = [_promotion_invoice_item(invoice, allocation) for invoice, allocation in rows]
    else:
        conditions = [
            InvoiceRecord.statement_month == month,
            InvoiceRecord.fee_direction == CONFIRMATION_DIRECTION_TO_DB[direction],
            InvoiceRecord.is_tombstone.is_(False),
            SettlementStatement.is_current.is_(True),
        ]
        if not include_history:
            conditions.append(InvoiceRecord.is_current.is_(True))
        if store_id is not None:
            conditions.append(InvoiceRecord.store_id == store_id)
        if status_code is not None:
            conditions.append(InvoiceRecord.invoice_status == status_code)
        query = (
            select(InvoiceRecord)
            .join(
                SettlementStatement,
                SettlementStatement.statement_id == InvoiceRecord.statement_id,
            )
            .where(*conditions)
            .order_by(InvoiceRecord.registered_at.desc(), InvoiceRecord.invoice_id)
        )
        total = session.scalar(select(func.count()).select_from(query.subquery())) or 0
        invoices = list(session.scalars(query.offset((page - 1) * page_size).limit(page_size)))
        items = [_management_invoice_item(invoice) for invoice in invoices]

    return _reporting_success(
        request,
        {"list": items, "total": total, "page": page, "page_size": page_size},
    )


@router.post(
    "/admin/finance/management-invoices/{store_id}/{statement_month}/corrections"
)
def correct_management_invoice(
    store_id: str,
    statement_month: str,
    payload: dict,
    request: Request,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    """Append one administrator correction to a management invoice slot."""

    _require_finance_admin(current_user, request)
    _validate_month(statement_month, "statementMonth", request)
    session = _billing_session(store, request)
    _require_billing_store(session, store_id, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(
        {
            "storeId": store_id,
            "statementMonth": statement_month,
            **payload,
        }
    )
    replay = session.scalar(
        select(FinanceOperationAudit).where(
            FinanceOperationAudit.idempotency_key_hash == key_hash
        )
    )
    if replay is not None:
        if (
            replay.operation_type != "MANAGEMENT_INVOICE_CORRECTION"
            or replay.request_payload_sha256 != payload_hash
        ):
            _persist_finance_conflict_audit(
                session,
                request=request,
                current_user=current_user,
                operation_type="MANAGEMENT_INVOICE_CORRECTION",
                target_type="INVOICE_RECORD",
                target_id=f"{store_id}|{statement_month}",
                conflict_code="IDEMPOTENCY_KEY_REUSED",
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
            )
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        replay_invoice = session.scalar(
            select(InvoiceRecord).where(
                InvoiceRecord.invoice_id == replay.target_id
            )
        )
        if replay_invoice is None:
            _persist_finance_conflict_audit(
                session,
                request=request,
                current_user=current_user,
                operation_type="MANAGEMENT_INVOICE_CORRECTION",
                target_type="INVOICE_RECORD",
                target_id=replay.target_id,
                conflict_code="IDEMPOTENCY_REPLAY_TARGET_MISSING",
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
            )
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_REPLAY_TARGET_MISSING",
                "幂等重放目标不存在，请联系管理员核对审计记录",
            )
        return _reporting_success(request, _management_invoice_item(replay_invoice))

    invoice_number = str(payload.get("invoiceNumber") or "").strip()
    change_reason = str(payload.get("changeReason") or "").strip()
    read_version = payload.get("readVersion")
    invoice_amount = payload.get("invoiceAmountCent")
    deduction_amount = payload.get("deductionAmountCent")
    if len(invoice_number) != 20 or not invoice_number.isdigit():
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "发票号码必须为 20 位数字",
            field="invoiceNumber",
        )
    if not change_reason or len(change_reason) > 1000:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "更正原因必填且不得超过 1000 字",
            field="changeReason",
        )
    if isinstance(read_version, bool) or not isinstance(read_version, int) or read_version <= 0:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "readVersion 必须为正整数",
            field="readVersion",
        )
    for field, value in (
        ("invoiceAmountCent", invoice_amount),
        ("deductionAmountCent", deduction_amount),
    ):
        if isinstance(value, bool) or not isinstance(value, int) or value <= 0:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "VALIDATION_FAILED",
                f"{field} 必须为正整数分",
                field=field,
            )
    if invoice_amount != deduction_amount:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "MANAGEMENT_AMOUNT_MISMATCH",
            "发票金额与厂家扣款金额必须全额一致",
            field="deductionAmountCent",
        )
    try:
        invoice_date_value = date.fromisoformat(str(payload.get("invoiceDate") or ""))
        deduction_date_value = date.fromisoformat(
            str(payload.get("deductionDate") or "")
        )
    except ValueError:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "发票日期和厂家扣款日期必须使用 YYYY-MM-DD",
            field="invoiceDate",
        )

    current = session.scalar(
        select(InvoiceRecord)
        .where(
            InvoiceRecord.store_id == store_id,
            InvoiceRecord.statement_month == statement_month,
            InvoiceRecord.fee_direction == 2,
            InvoiceRecord.is_current.is_(True),
            InvoiceRecord.is_tombstone.is_(False),
        )
        .with_for_update()
    )
    if current is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "当前管理服务费发票记录不存在",
        )
    if current.version_no != read_version:
        conflict_data = {
            "read_version": read_version,
            "current_version": current.version_no,
            "current_invoice_id": current.invoice_id,
        }
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="MANAGEMENT_INVOICE_CORRECTION",
            target_type="INVOICE_RECORD",
            target_id=current.invoice_id,
            conflict_code="VERSION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            details=conflict_data,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "VERSION_CONFLICT",
            "管理服务费发票版本已变化，请刷新后重试",
            field="readVersion",
            data=conflict_data,
        )
    projection_periods, carryforward_applications = (
        _management_invoiceable_projection(
            session,
            store_id=store_id,
            through_month=statement_month,
            excluded_invoice_ids={current.invoice_id},
        )
    )
    projected_period = next(
        (
            period
            for period in projection_periods
            if period["statement_id"] == current.statement_id
        ),
        None,
    )
    projected_amount = (
        int(projected_period["invoiceable_amount_cent"])
        if projected_period is not None
        else 0
    )
    if invoice_amount != projected_amount or deduction_amount != projected_amount:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "MANAGEMENT_AMOUNT_MISMATCH",
            "页面更正不得改变当前有效管理服务费结转后全额",
            field="invoiceAmountCent",
            data={"projected_amount_cent": projected_amount},
        )
    before_snapshot = _management_invoice_item(current)
    corrected_at = utcnow()
    current.is_current = False
    session.flush()
    corrected = InvoiceRecord(
        invoice_id=f"management-invoice-{uuid4().hex}",
        store_id=current.store_id,
        statement_month=current.statement_month,
        statement_id=current.statement_id,
        fee_direction=2,
        version_no=current.version_no + 1,
        is_current=True,
        invoice_number=invoice_number,
        invoice_date=invoice_date_value,
        invoice_amount_cent=invoice_amount,
        invoice_status=3,
        source_type=3,
        import_batch_id=None,
        factory_deduction_date=deduction_date_value,
        factory_deduction_amount_cent=deduction_amount,
        registered_by=current_user.username,
        registered_at=corrected_at,
    )
    session.add(corrected)
    session.flush()
    _synchronize_management_carryforward_applications(
        session,
        applications=carryforward_applications,
        invoice_id_by_statement={corrected.statement_id: corrected.invoice_id},
        scope_store_ids={store_id},
        scope_through_month=statement_month,
    )
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="MANAGEMENT_INVOICE_CORRECTION",
            target_type="INVOICE_RECORD",
            target_id=corrected.invoice_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot=before_snapshot,
            after_snapshot={
                **_management_invoice_item(corrected),
                "changeReason": change_reason,
            },
            result_status=1,
            request_id=request_id(request),
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            occurred_at=corrected_at,
        )
    )
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        replay = session.scalar(
            select(FinanceOperationAudit).where(
                FinanceOperationAudit.idempotency_key_hash == key_hash
            )
        )
        if replay is not None and replay.request_payload_sha256 == payload_hash:
            replay_invoice = session.scalar(
                select(InvoiceRecord).where(
                    InvoiceRecord.invoice_id == replay.target_id
                )
            )
            if replay_invoice is not None:
                return _reporting_success(
                    request, _management_invoice_item(replay_invoice)
                )
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="MANAGEMENT_INVOICE_CORRECTION",
            target_type="INVOICE_RECORD",
            target_id=f"{store_id}|{statement_month}",
            conflict_code="MANAGEMENT_INVOICE_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "MANAGEMENT_INVOICE_CONFLICT",
            "管理服务费发票已被其他操作更新，请刷新后重试",
        )
    return _reporting_success(request, _management_invoice_item(corrected))


@router.get("/admin/finance/order-details")
def list_admin_finance_order_details(
    request: Request,
    month: str = Query(),
    fee_direction: str = Query(alias="feeDirection"),
    store_id: str | None = Query(default=None, alias="storeId"),
    store_name: str | None = Query(default=None, alias="storeName"),
    sap_code: str | None = Query(default=None, alias="sapCode"),
    invoice_number: str | None = Query(default=None, alias="invoiceNumber"),
    order_id: str | None = Query(default=None, alias="orderId"),
    sku_id: str | None = Query(default=None, alias="skuId"),
    sale_channel: str | None = Query(default=None, alias="saleChannel"),
    invoice_status: str | None = Query(default=None, alias="invoiceStatus"),
    submitted_from: datetime | None = Query(default=None, alias="submittedFrom"),
    submitted_to: datetime | None = Query(default=None, alias="submittedTo"),
    verify_from: datetime | None = Query(default=None, alias="verifyFrom"),
    verify_to: datetime | None = Query(default=None, alias="verifyTo"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_finance_admin(current_user, request)
    _validate_month(month, "month", request)
    direction = _normalize_billing_direction(fee_direction, request)
    if store_id is not None:
        _require_billing_store(session, store_id, request)
    filters = _normalize_finance_order_detail_filters(
        request=request, month=month, fee_direction=direction, store_id=store_id,
        store_name=store_name,
        sap_code=sap_code, invoice_number=invoice_number, order_id=order_id,
        sku_id=sku_id, sale_channel=sale_channel, invoice_status=invoice_status,
        submitted_from=submitted_from, submitted_to=submitted_to,
        verify_from=verify_from, verify_to=verify_to,
    )
    query = _finance_order_details_query(filters)
    total = session.scalar(select(func.count()).select_from(query.order_by(None).subquery())) or 0
    rows = session.execute(query.offset((page - 1) * page_size).limit(page_size)).all()
    return _reporting_success(
        request,
        {
            "list": [_finance_order_detail_item(row, direction) for row in rows],
            "total": total,
            "page": page,
            "page_size": page_size,
            "definitions": FINANCE_ORDER_DETAIL_DEFINITIONS,
        },
    )


@router.get("/admin/finance/order-details/export")
def export_admin_finance_order_details(
    request: Request,
    month: str = Query(),
    fee_direction: str = Query(alias="feeDirection"),
    store_id: str | None = Query(default=None, alias="storeId"),
    store_name: str | None = Query(default=None, alias="storeName"),
    sap_code: str | None = Query(default=None, alias="sapCode"),
    invoice_number: str | None = Query(default=None, alias="invoiceNumber"),
    order_id: str | None = Query(default=None, alias="orderId"),
    sku_id: str | None = Query(default=None, alias="skuId"),
    sale_channel: str | None = Query(default=None, alias="saleChannel"),
    invoice_status: str | None = Query(default=None, alias="invoiceStatus"),
    submitted_from: datetime | None = Query(default=None, alias="submittedFrom"),
    submitted_to: datetime | None = Query(default=None, alias="submittedTo"),
    verify_from: datetime | None = Query(default=None, alias="verifyFrom"),
    verify_to: datetime | None = Query(default=None, alias="verifyTo"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_finance_admin(current_user, request)
    _validate_month(month, "month", request)
    direction = _normalize_billing_direction(fee_direction, request)
    if store_id is not None:
        _require_billing_store(session, store_id, request)
    filters = _normalize_finance_order_detail_filters(
        request=request, month=month, fee_direction=direction, store_id=store_id,
        store_name=store_name,
        sap_code=sap_code, invoice_number=invoice_number, order_id=order_id,
        sku_id=sku_id, sale_channel=sale_channel, invoice_status=invoice_status,
        submitted_from=submitted_from, submitted_to=submitted_to,
        verify_from=verify_from, verify_to=verify_to,
    )
    audit_filters = _finance_order_detail_audit_filters(filters)
    try:
        rows = session.execute(
            _finance_order_details_query(filters).limit(
                MAX_FINANCE_ORDER_EXPORT_ROWS + 1
            )
        ).all()
    except Exception:
        session.rollback()
        _persist_finance_order_export_audit(
            session,
            request=request,
            current_user=current_user,
            filters=audit_filters,
            row_count=0,
            result="QUERY_OR_PROJECTION_FAILED",
            result_status=3,
        )
        raise
    if len(rows) > MAX_FINANCE_ORDER_EXPORT_ROWS:
        _persist_finance_order_export_audit(
            session, request=request, current_user=current_user,
            filters=audit_filters, row_count=len(rows),
            result="LIMIT_EXCEEDED", result_status=3,
        )
        _raise_reporting_error(
            request,
            status.HTTP_413_CONTENT_TOO_LARGE,
            "EXPORT_LIMIT_EXCEEDED",
            f"命中记录超过 {MAX_FINANCE_ORDER_EXPORT_ROWS} 行，请缩小筛选范围后重试",
        )
    try:
        items = [_finance_order_detail_item(row, direction) for row in rows]
    except Exception:
        session.rollback()
        _persist_finance_order_export_audit(
            session,
            request=request,
            current_user=current_user,
            filters=audit_filters,
            row_count=len(rows),
            result="QUERY_OR_PROJECTION_FAILED",
            result_status=3,
        )
        raise
    export_result = "SUCCESS" if items else "EMPTY"
    try:
        csv_text = _finance_order_detail_csv(items)
    except Exception:
        session.rollback()
        _persist_finance_order_export_audit(
            session,
            request=request,
            current_user=current_user,
            filters=audit_filters,
            row_count=len(items),
            result="CSV_GENERATION_FAILED",
            result_status=3,
        )
        raise
    _persist_finance_order_export_audit(
        session,
        request=request,
        current_user=current_user,
        filters=audit_filters,
        row_count=len(items),
        result=export_result,
        result_status=1 if items else 2,
    )
    generated = generated_at().isoformat()
    filename = quote(f"finance-order-details-{generated[:10]}.csv")
    return StreamingResponse(
        iter((with_utf8_bom(csv_text),)),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "X-Export-Generated-At": generated,
            "X-Export-Filters": json.dumps(audit_filters, ensure_ascii=False, separators=(",", ":"), sort_keys=True),
            "X-Request-ID": request_id(request),
            "X-Export-Result": export_result,
        },
    )


def _sap_suggestion_item(
    suggestion: SapSuggestion,
    confirmed_profile: StoreFinanceProfile | None = None,
) -> dict:
    return {
        "suggestion_id": suggestion.suggestion_id,
        "store_id": suggestion.store_id,
        "version_no": suggestion.version_no,
        "is_current": suggestion.is_current,
        "suggested_sap_code": suggestion.suggested_sap_code,
        "suggestion_note": suggestion.suggestion_note,
        "status": SAP_SUGGESTION_STATUS_NAMES[suggestion.suggestion_status],
        "submitted_by": suggestion.submitted_by,
        "submitted_at": (
            _finance_datetime(suggestion.submitted_at).isoformat()
            if suggestion.submitted_at is not None
            else None
        ),
        "handled_by": suggestion.handled_by,
        "handled_at": (
            _finance_datetime(suggestion.handled_at).isoformat()
            if suggestion.handled_at is not None
            else None
        ),
        "handling_reason": suggestion.handling_reason,
        "confirmed_profile_id": suggestion.confirmed_profile_id,
        "confirmed_sap_code": (
            confirmed_profile.sap_code
            if confirmed_profile is not None and not confirmed_profile.is_tombstone
            else None
        ),
        "confirmed_version": (
            confirmed_profile.version_no if confirmed_profile is not None else 0
        ),
    }


@router.get("/stores/{store_id}/sap-suggestions")
def list_store_sap_suggestions(
    store_id: str,
    request: Request,
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    _require_store_actor(current_user, store_id)
    session = _billing_session(store, request)
    suggestions = list(
        session.scalars(
            select(SapSuggestion)
            .where(SapSuggestion.store_id == store_id)
            .order_by(SapSuggestion.version_no.desc())
        )
    )
    confirmed_profile = session.scalar(
        select(StoreFinanceProfile).where(
            StoreFinanceProfile.store_id == store_id,
            StoreFinanceProfile.profile_type == 2,
            StoreFinanceProfile.is_current.is_(True),
        )
    )
    return _reporting_success(
        request,
        {
            "list": [
                _sap_suggestion_item(item, confirmed_profile) for item in suggestions
            ],
            "total": len(suggestions),
            "current_version": suggestions[0].version_no if suggestions else 0,
            "confirmed_version": (
                confirmed_profile.version_no if confirmed_profile is not None else 0
            ),
        },
    )


@router.post("/stores/{store_id}/sap-suggestions")
def submit_store_sap_suggestion(
    store_id: str,
    request: Request,
    payload: dict,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    _require_store_actor(current_user, store_id)
    session = _billing_session(store, request)
    if session.scalar(select(DimStore.store_id).where(DimStore.store_id == store_id)) is None:
        _raise_reporting_error(
            request, status.HTTP_404_NOT_FOUND, "RESOURCE_NOT_FOUND", "门店不存在"
        )
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(payload)
    replay = session.scalar(
        select(SapSuggestion).where(SapSuggestion.idempotency_key_hash == key_hash)
    )
    if replay is not None:
        if replay.request_payload_sha256 != payload_hash or replay.store_id != store_id:
            _persist_finance_conflict_audit(
                session,
                request=request,
                current_user=current_user,
                operation_type="SAP_SUGGESTION_SUBMIT",
                target_type="SAP_SUGGESTION",
                target_id=store_id,
                conflict_code="IDEMPOTENCY_KEY_REUSED",
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
            )
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        confirmed_profile = session.scalar(
            select(StoreFinanceProfile).where(
                StoreFinanceProfile.store_id == store_id,
                StoreFinanceProfile.profile_type == 2,
                StoreFinanceProfile.is_current.is_(True),
            )
        )
        return _reporting_success(
            request, _sap_suggestion_item(replay, confirmed_profile)
        )

    suggested_sap_code = str(payload.get("suggestedSapCode") or "").strip()
    suggestion_note = str(payload.get("suggestionNote") or "").strip()
    read_version = payload.get("readVersion")
    if not suggested_sap_code or len(suggested_sap_code) > 128:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "建议 SAP 编码必填且不得超过 128 字",
            field="suggestedSapCode",
        )
    if not suggestion_note or len(suggestion_note) > 1000:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "建议说明必填且不得超过 1000 字",
            field="suggestionNote",
        )
    if isinstance(read_version, bool) or not isinstance(read_version, int) or read_version < 0:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "readVersion 必须为非负整数",
            field="readVersion",
        )
    current = session.scalar(
        select(SapSuggestion)
        .where(
            SapSuggestion.store_id == store_id,
            SapSuggestion.is_current.is_(True),
        )
        .with_for_update()
    )
    current_version = current.version_no if current is not None else 0
    if read_version != current_version:
        conflict_data = {
            "read_version": read_version,
            "current_version": current_version,
        }
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="SAP_SUGGESTION_SUBMIT",
            target_type="SAP_SUGGESTION",
            target_id=store_id,
            conflict_code="VERSION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            details=conflict_data,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "VERSION_CONFLICT",
            "SAP 建议版本已变化，请刷新后重试",
            field="readVersion",
            data=conflict_data,
        )
    now = utcnow()
    if current is not None:
        current.is_current = False
        session.flush()
    suggestion = SapSuggestion(
        suggestion_id=f"sap-suggestion-{uuid4().hex}",
        store_id=store_id,
        version_no=current_version + 1,
        is_current=True,
        supersedes_suggestion_id=(current.suggestion_id if current is not None else None),
        suggested_sap_code=suggested_sap_code,
        suggestion_note=suggestion_note,
        suggestion_status=1,
        submitted_by=current_user.username,
        submitted_at=now,
        idempotency_key_hash=key_hash,
        request_payload_sha256=payload_hash,
        created_at=now,
    )
    session.add(suggestion)
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="SAP_SUGGESTION_SUBMIT",
            target_type="SAP_SUGGESTION",
            target_id=suggestion.suggestion_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot=(
                _sap_suggestion_item(current) if current is not None else None
            ),
            after_snapshot={
                "storeId": store_id,
                "versionNo": suggestion.version_no,
                "suggestedSapCode": suggested_sap_code,
                "suggestionNote": suggestion_note,
                "status": "PENDING",
            },
            result_status=1,
            request_id=request_id(request),
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            occurred_at=now,
        )
    )
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        replay = session.scalar(
            select(SapSuggestion).where(SapSuggestion.idempotency_key_hash == key_hash)
        )
        if replay is not None and replay.request_payload_sha256 == payload_hash:
            return _reporting_success(request, _sap_suggestion_item(replay))
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="SAP_SUGGESTION_SUBMIT",
            target_type="SAP_SUGGESTION",
            target_id=store_id,
            conflict_code="SAP_SUGGESTION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "SAP_SUGGESTION_CONFLICT",
            "SAP 建议已被其他操作更新，请刷新后重试",
        )
    confirmed_profile = session.scalar(
        select(StoreFinanceProfile).where(
            StoreFinanceProfile.store_id == store_id,
            StoreFinanceProfile.profile_type == 2,
            StoreFinanceProfile.is_current.is_(True),
        )
    )
    return _reporting_success(
        request, _sap_suggestion_item(suggestion, confirmed_profile)
    )


@router.post("/admin/finance/sap-suggestions/{suggestion_id}/decisions")
def decide_sap_suggestion(
    suggestion_id: str,
    request: Request,
    payload: dict,
    idempotency_key: str | None = Header(default=None, alias="Idempotency-Key"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_finance_admin(current_user, request)
    key_hash = _billing_idempotency_key_hash(idempotency_key, request)
    payload_hash = _canonical_billing_sha256(payload)
    replay_suggestion = session.scalar(
        select(SapSuggestion).where(SapSuggestion.idempotency_key_hash == key_hash)
    )
    if replay_suggestion is not None:
        if (
            replay_suggestion.request_payload_sha256 != payload_hash
            or replay_suggestion.supersedes_suggestion_id != suggestion_id
            or replay_suggestion.suggestion_status == 1
        ):
            _persist_finance_conflict_audit(
                session,
                request=request,
                current_user=current_user,
                operation_type="SAP_SUGGESTION_DECISION",
                target_type="SAP_SUGGESTION",
                target_id=suggestion_id,
                conflict_code="IDEMPOTENCY_KEY_REUSED",
                idempotency_key_hash=key_hash,
                request_payload_sha256=payload_hash,
            )
            _raise_reporting_error(
                request,
                status.HTTP_409_CONFLICT,
                "IDEMPOTENCY_KEY_REUSED",
                "Idempotency-Key 已用于不同请求",
            )
        replay_suggestion = session.scalar(
            select(SapSuggestion).where(SapSuggestion.idempotency_key_hash == key_hash)
        )
        if replay_suggestion is None:
            _raise_reporting_error(
                request, status.HTTP_409_CONFLICT, "REPLAY_TARGET_MISSING", "审计目标不存在"
            )
        profile = session.scalar(
            select(StoreFinanceProfile).where(
                StoreFinanceProfile.profile_id == replay_suggestion.confirmed_profile_id
            )
        ) if replay_suggestion.confirmed_profile_id else None
        return _reporting_success(
            request, _sap_suggestion_item(replay_suggestion, profile)
        )

    action = str(payload.get("action") or "").upper()
    if action not in {"CONFIRM", "CORRECT", "REJECT"}:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "action 必须为 CONFIRM、CORRECT 或 REJECT",
            field="action",
        )
    suggestion_version = payload.get("suggestionVersion")
    expected_confirmed_version = payload.get("expectedConfirmedVersion")
    for field, value in (
        ("suggestionVersion", suggestion_version),
        ("expectedConfirmedVersion", expected_confirmed_version),
    ):
        if isinstance(value, bool) or not isinstance(value, int) or value < 0:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "VALIDATION_FAILED",
                f"{field} 必须为非负整数",
                field=field,
            )
    handling_reason = str(payload.get("handlingReason") or "").strip()
    if not handling_reason or len(handling_reason) > 1000:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "处理原因必填且不得超过 1000 字",
            field="handlingReason",
        )
    suggestion = session.scalar(
        select(SapSuggestion)
        .where(SapSuggestion.suggestion_id == suggestion_id)
        .with_for_update()
    )
    if suggestion is None:
        _raise_reporting_error(
            request, status.HTTP_404_NOT_FOUND, "RESOURCE_NOT_FOUND", "SAP 建议不存在"
        )
    if (
        not suggestion.is_current
        or suggestion.version_no != suggestion_version
        or suggestion.suggestion_status != 1
    ):
        conflict_data = {
            "current_version": suggestion.version_no,
            "current_status": SAP_SUGGESTION_STATUS_NAMES[
                suggestion.suggestion_status
            ],
        }
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="SAP_SUGGESTION_DECISION",
            target_type="SAP_SUGGESTION",
            target_id=suggestion_id,
            conflict_code="SUGGESTION_VERSION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            details=conflict_data,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "SUGGESTION_VERSION_CONFLICT",
            "SAP 建议已被更新或处理，请刷新后重试",
            data=conflict_data,
        )
    profile = session.scalar(
        select(StoreFinanceProfile)
        .where(
            StoreFinanceProfile.store_id == suggestion.store_id,
            StoreFinanceProfile.profile_type == 2,
            StoreFinanceProfile.is_current.is_(True),
        )
        .with_for_update()
    )
    confirmed_version = profile.version_no if profile is not None else 0
    if expected_confirmed_version != confirmed_version:
        conflict_data = {"current_confirmed_version": confirmed_version}
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="SAP_SUGGESTION_DECISION",
            target_type="SAP_SUGGESTION",
            target_id=suggestion_id,
            conflict_code="CONFIRMED_VERSION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            details=conflict_data,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "CONFIRMED_VERSION_CONFLICT",
            "当前有效 SAP 已变化，请刷新后重试",
            data=conflict_data,
        )
    confirmed_sap_code = str(payload.get("confirmedSapCode") or "").strip()
    if action == "CONFIRM" and not confirmed_sap_code:
        confirmed_sap_code = suggestion.suggested_sap_code
    if action != "REJECT" and (
        not confirmed_sap_code or len(confirmed_sap_code) > 128
    ):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "确认 SAP 编码必填且不得超过 128 字",
            field="confirmedSapCode",
        )

    before_snapshot = _sap_suggestion_item(suggestion, profile)
    now = utcnow()
    next_profile = None
    if action != "REJECT":
        if profile is not None:
            profile.is_current = False
            session.flush()
        store_name = session.scalar(
            select(DimStore.store_name).where(DimStore.store_id == suggestion.store_id)
        ) or suggestion.store_id
        next_profile = StoreFinanceProfile(
            profile_id=f"store-finance-profile-{uuid4().hex}",
            store_id=suggestion.store_id,
            profile_type=2,
            source_type=2,
            version_no=confirmed_version + 1,
            is_current=True,
            store_name_snapshot=store_name,
            sap_code=confirmed_sap_code,
            initial_sap_code=(profile.initial_sap_code if profile is not None else None),
            service_store_code=(profile.service_store_code if profile is not None else None),
            factory_confirmed=True,
            confirmed_at=now,
            import_batch_id=None,
            created_at=now,
            updated_at=now,
        )
        session.add(next_profile)
        session.flush()
    decided_status = {
        "CONFIRM": 2,
        "CORRECT": 3,
        "REJECT": 4,
    }[action]
    suggestion.is_current = False
    session.flush()
    decided_suggestion = SapSuggestion(
        suggestion_id=f"sap-suggestion-{uuid4().hex}",
        store_id=suggestion.store_id,
        version_no=suggestion.version_no + 1,
        is_current=True,
        supersedes_suggestion_id=suggestion.suggestion_id,
        suggested_sap_code=suggestion.suggested_sap_code,
        suggestion_note=suggestion.suggestion_note,
        suggestion_status=decided_status,
        submitted_by=suggestion.submitted_by,
        submitted_at=suggestion.submitted_at,
        handled_by=current_user.username,
        handled_at=now,
        handling_reason=handling_reason,
        confirmed_profile_id=(
            next_profile.profile_id if next_profile is not None else None
        ),
        idempotency_key_hash=key_hash,
        request_payload_sha256=payload_hash,
        created_at=now,
    )
    session.add(decided_suggestion)
    after_snapshot = _sap_suggestion_item(
        decided_suggestion, next_profile or profile
    )
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="SAP_SUGGESTION_DECISION",
            target_type="SAP_SUGGESTION",
            target_id=decided_suggestion.suggestion_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot=before_snapshot,
            after_snapshot={**after_snapshot, "action": action},
            result_status=1,
            request_id=request_id(request),
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
            occurred_at=now,
        )
    )
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        replay_suggestion = session.scalar(
            select(SapSuggestion).where(SapSuggestion.idempotency_key_hash == key_hash)
        )
        if (
            replay_suggestion is not None
            and replay_suggestion.request_payload_sha256 == payload_hash
            and replay_suggestion.supersedes_suggestion_id == suggestion_id
        ):
            replay_profile = (
                session.scalar(
                    select(StoreFinanceProfile).where(
                        StoreFinanceProfile.profile_id
                        == replay_suggestion.confirmed_profile_id
                    )
                )
                if replay_suggestion.confirmed_profile_id
                else None
            )
            return _reporting_success(
                request, _sap_suggestion_item(replay_suggestion, replay_profile)
            )
        _persist_finance_conflict_audit(
            session,
            request=request,
            current_user=current_user,
            operation_type="SAP_SUGGESTION_DECISION",
            target_type="SAP_SUGGESTION",
            target_id=suggestion_id,
            conflict_code="SAP_DECISION_CONFLICT",
            idempotency_key_hash=key_hash,
            request_payload_sha256=payload_hash,
        )
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "SAP_DECISION_CONFLICT",
            "SAP 建议或确认值已被其他操作更新，请刷新后重试",
        )
    return _reporting_success(
        request, _sap_suggestion_item(decided_suggestion, next_profile or profile)
    )


@router.get("/admin/finance/stores")
def list_admin_finance_stores(
    request: Request,
    month: str = Query(),
    fee_direction: str = Query(alias="feeDirection"),
    metric_scope: str = Query(alias="metricScope"),
    q: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=50, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    session = _billing_session(store, request)
    _require_finance_admin(current_user, request)
    _validate_month(month, "month", request)
    direction = _normalize_billing_direction(fee_direction, request)
    normalized_scope = metric_scope.upper()
    _validate_enum(normalized_scope, METRIC_SCOPES, "metricScope", request)
    conditions = [DimStore.is_active.is_(True)]
    normalized_query = (q or "").strip()
    if normalized_query:
        matching_profile_store_ids = select(StoreFinanceProfile.store_id).where(
            StoreFinanceProfile.profile_type == 2,
            StoreFinanceProfile.is_current.is_(True),
            StoreFinanceProfile.is_tombstone.is_(False),
            StoreFinanceProfile.sap_code.contains(normalized_query),
        )
        matching_suggestion_store_ids = select(SapSuggestion.store_id).where(
            SapSuggestion.is_current.is_(True),
            SapSuggestion.suggested_sap_code.contains(normalized_query),
        )
        conditions.append(
            (DimStore.store_id.contains(normalized_query))
            | (DimStore.store_name.contains(normalized_query))
            | (DimStore.store_id.in_(matching_profile_store_ids))
            | (DimStore.store_id.in_(matching_suggestion_store_ids))
        )
    store_ids_query = (
        select(DimStore.store_id)
        .where(*conditions)
        .distinct()
        .order_by(DimStore.store_id)
    )
    total = session.scalar(select(func.count()).select_from(store_ids_query.subquery())) or 0
    store_ids = list(
        session.scalars(store_ids_query.offset((page - 1) * page_size).limit(page_size))
    )
    stores = {
        store_row.store_id: store_row
        for store_row in session.scalars(select(DimStore).where(DimStore.store_id.in_(store_ids)))
    }
    current_profiles = {
        profile.store_id: profile
        for profile in session.scalars(
            select(StoreFinanceProfile).where(
                StoreFinanceProfile.store_id.in_(store_ids),
                StoreFinanceProfile.profile_type == 2,
                StoreFinanceProfile.is_current.is_(True),
            )
        )
    }
    current_suggestions = {
        suggestion.store_id: suggestion
        for suggestion in session.scalars(
            select(SapSuggestion).where(
                SapSuggestion.store_id.in_(store_ids),
                SapSuggestion.is_current.is_(True),
            )
        )
    }
    items = []
    for current_store_id in store_ids:
        store_row = stores[current_store_id]
        confirmed_profile = current_profiles.get(current_store_id)
        current_suggestion = current_suggestions.get(current_store_id)
        metrics = _finance_summary_metrics(
            session,
            month=month,
            fee_direction=direction,
            metric_scope=normalized_scope,
            store_id=current_store_id,
        )
        items.append(
            {
                "store_id": store_row.store_id,
                "store_name": store_row.store_name,
                "sap_code": (
                    confirmed_profile.sap_code
                    if confirmed_profile is not None and not confirmed_profile.is_tombstone
                    else None
                ),
                "confirmed_version": (
                    confirmed_profile.version_no if confirmed_profile is not None else 0
                ),
                "confirmed_source_type": (
                    confirmed_profile.source_type if confirmed_profile is not None else None
                ),
                "suggestion_id": (
                    current_suggestion.suggestion_id if current_suggestion is not None else None
                ),
                "suggested_sap_code": (
                    current_suggestion.suggested_sap_code if current_suggestion is not None else None
                ),
                "suggestion_note": (
                    current_suggestion.suggestion_note if current_suggestion is not None else None
                ),
                "suggestion_status": (
                    SAP_SUGGESTION_STATUS_NAMES[current_suggestion.suggestion_status]
                    if current_suggestion is not None
                    else None
                ),
                "suggestion_version": (
                    current_suggestion.version_no if current_suggestion is not None else 0
                ),
                "suggestion_updated_at": (
                    current_suggestion.handled_at or current_suggestion.submitted_at
                    if current_suggestion is not None
                    else None
                ),
                "updated_at": (
                    confirmed_profile.updated_at
                    if confirmed_profile is not None
                    else store_row.updated_at
                ),
                **metrics,
            }
        )
    return _reporting_success(
        request,
        {"list": items, "total": total, "page": page, "page_size": page_size},
    )


@router.get("/stores/{store_id}/monthly-settlement")
def monthly_settlement(
    request: Request,
    store_id: str,
    month: str,
    product_scope: str = Query(default="all", alias="productScope"),
    product_type: str = Query(default="all", alias="productType"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    _require_store_scope(current_user, store_id)
    _validate_month(month, "month", request)
    _validate_product_selection(store, product_scope, product_type, request)
    _validate_monthly_context(store, store_id, month, request)
    data = _call_reporting_store(
        request,
        store.monthly_settlement_report,
        {
            "store_id": store_id,
            "month": month,
            "product_scope": product_scope,
            "product_type": product_type,
        },
    )
    return _reporting_success(
        request, data, definitions=MONTHLY_SETTLEMENT_DEFINITIONS
    )


@router.get("/order-fee-details")
def order_fee_details(
    request: Request,
    statement_id: str | None = Query(default=None, alias="statementId"),
    statement_line_id: str | None = Query(default=None, alias="statementLineId"),
    store_id: str | None = Query(default=None, alias="storeId"),
    month: str | None = None,
    sale_month: str | None = Query(default=None, alias="saleMonth"),
    verify_month: str | None = Query(default=None, alias="verifyMonth"),
    fee_direction: str = Query(alias="feeDirection"),
    product_scope: str = Query(default="all", alias="productScope"),
    product_type: str = Query(default="all", alias="productType"),
    fee_rates: list[str] | None = Query(default=None, alias="feeRates"),
    rule_versions: list[str] | None = Query(default=None, alias="ruleVersions"),
    data_status: str | None = Query(default=None, alias="dataStatus"),
    q: str | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=100, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    filters = _order_fee_filters(
        request=request,
        store=store,
        current_user=current_user,
        statement_id=statement_id,
        statement_line_id=statement_line_id,
        store_id=store_id,
        month=month,
        sale_month=sale_month,
        verify_month=verify_month,
        fee_direction=fee_direction,
        product_scope=product_scope,
        product_type=product_type,
        fee_rates=fee_rates,
        rule_versions=rule_versions,
        data_status=data_status,
        q=q,
        page=page,
        page_size=page_size,
    )
    data = _call_reporting_store(request, store.order_fee_details, filters)
    return _reporting_success(request, data)


@router.get("/order-fee-details/export")
def order_fee_details_export(
    request: Request,
    statement_id: str | None = Query(default=None, alias="statementId"),
    statement_line_id: str | None = Query(default=None, alias="statementLineId"),
    store_id: str | None = Query(default=None, alias="storeId"),
    month: str | None = None,
    sale_month: str | None = Query(default=None, alias="saleMonth"),
    verify_month: str | None = Query(default=None, alias="verifyMonth"),
    fee_direction: str = Query(alias="feeDirection"),
    product_scope: str = Query(default="all", alias="productScope"),
    product_type: str = Query(default="all", alias="productType"),
    fee_rates: list[str] | None = Query(default=None, alias="feeRates"),
    rule_versions: list[str] | None = Query(default=None, alias="ruleVersions"),
    data_status: str | None = Query(default=None, alias="dataStatus"),
    q: str | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=100, alias="pageSize"),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    filters = _order_fee_filters(
        request=request,
        store=store,
        current_user=current_user,
        statement_id=statement_id,
        statement_line_id=statement_line_id,
        store_id=store_id,
        month=month,
        sale_month=sale_month,
        verify_month=verify_month,
        fee_direction=fee_direction,
        product_scope=product_scope,
        product_type=product_type,
        fee_rates=fee_rates,
        rule_versions=rule_versions,
        data_status=data_status,
        q=q,
        page=page,
        page_size=page_size,
    )
    csv_text = _call_reporting_store(
        request, store.order_fee_details_export_csv, filters
    )
    if not csv_text:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "EXPORT_EMPTY",
            "当前筛选无可导出明细",
        )
    generated = generated_at().isoformat()
    filename = quote(f"order-fee-details-{generated[:10]}.csv")
    current_request_id = request_id(request)
    return Response(
        content=with_utf8_bom(csv_text),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "X-Export-Generated-At": generated,
            "X-Export-Filters": store.export_filter_header(filters),
            "X-Request-ID": current_request_id,
        },
    )


@router.get("/dashboard/sales")
def sales_dashboard(
    store_id: str | None = None,
    month: str = "all",
    product_scope: str = "all",
    product_type: str = "all",
    trend_months: list[str] | None = Query(default=None),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    scoped_store_id = _resolve_sales_dashboard_store_id(current_user, store_id)
    data = SalesDashboardData(
        **store.sales_dashboard(
            store_id=scoped_store_id,
            month=month,
            product_scope=product_scope,
            product_type=product_type,
            trend_months=trend_months or [],
        )
    )
    return {
        "data": dump_model(data),
        "definitions": SALES_DASHBOARD_DEFINITIONS,
        "meta": {"generated_at": generated_at(), "source": "postgres"},
    }


@router.get("/order-details")
def order_details(
    product_scope: str = "all",
    product_type: str = "all",
    sale_store_id: str | None = None,
    exclude_sale_store_id: str | None = None,
    sale_month: str | None = None,
    is_verified: bool | None = None,
    verify_store_id: str | None = None,
    exclude_verify_store_id: str | None = None,
    verify_month: str | None = None,
    relation_type: str | None = None,
    is_commissionable: bool | None = None,
    q: str | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    filters = _filters_from_query(
        product_scope=product_scope,
        product_type=product_type,
        sale_store_id=sale_store_id,
        exclude_sale_store_id=exclude_sale_store_id,
        sale_month=sale_month,
        is_verified=is_verified,
        verify_store_id=verify_store_id,
        exclude_verify_store_id=exclude_verify_store_id,
        verify_month=verify_month,
        relation_type=relation_type,
        is_commissionable=is_commissionable,
        q=q,
        page=page,
        page_size=page_size,
    )
    if not current_user.has_global_data_access:
        filters["scope_store_ids"] = current_user.store_ids
    data = OrderDetailsData(**store.order_details(filters))
    return {
        "data": dump_model(data),
        "meta": {"generated_at": generated_at(), "source": "postgres"},
    }


@router.get("/order-details/export")
def order_details_export(
    product_scope: str = "all",
    product_type: str = "all",
    sale_store_id: str | None = None,
    exclude_sale_store_id: str | None = None,
    sale_month: str | None = None,
    is_verified: bool | None = None,
    verify_store_id: str | None = None,
    exclude_verify_store_id: str | None = None,
    verify_month: str | None = None,
    relation_type: str | None = None,
    is_commissionable: bool | None = None,
    q: str | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    current_user: AuthContext = Depends(get_current_user),
    store=Depends(get_data_store),
):
    filters = _filters_from_query(
        product_scope=product_scope,
        product_type=product_type,
        sale_store_id=sale_store_id,
        exclude_sale_store_id=exclude_sale_store_id,
        sale_month=sale_month,
        is_verified=is_verified,
        verify_store_id=verify_store_id,
        exclude_verify_store_id=exclude_verify_store_id,
        verify_month=verify_month,
        relation_type=relation_type,
        is_commissionable=is_commissionable,
        q=q,
        page=page,
        page_size=page_size,
    )
    if not current_user.has_global_data_access:
        filters["scope_store_ids"] = current_user.store_ids
    generated = generated_at().isoformat()
    filename = quote(f"order-details-{generated[:10]}.csv")
    return Response(
        content=with_utf8_bom(store.order_details_export_csv(filters)),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename}",
            "X-Export-Generated-At": generated,
            "X-Export-Filters": store.export_filter_header(filters),
        },
    )


def _reporting_success(
    request: Request, data: dict, *, definitions: list[dict] | None = None
) -> dict:
    payload = {
        "data": camelize(data),
        "meta": {
            "generatedAt": generated_at(),
            "source": "postgres",
            "requestId": request_id(request),
        },
    }
    if definitions:
        payload["definitions"] = [camelize(item) for item in definitions]
    return payload


def _billing_session(store, request: Request):
    session = getattr(store, "session", None)
    if session is None:
        _raise_reporting_error(
            request,
            status.HTTP_503_SERVICE_UNAVAILABLE,
            "DATABASE_UNAVAILABLE",
            "数据库暂不可用",
        )
    return session


def _require_billing_store_scope(
    current_user: AuthContext, store_id: str, request: Request
) -> None:
    if not current_user.has_global_data_access and store_id not in current_user.store_ids:
        _raise_reporting_error(
            request,
            status.HTTP_403_FORBIDDEN,
            "DATA_SCOPE_FORBIDDEN",
            "门店不在当前账号数据范围内",
            field="storeId",
        )


def _require_promotion_invoice_store_operator(
    current_user: AuthContext, request: Request
) -> None:
    if current_user.role != "store":
        _raise_reporting_error(
            request,
            status.HTTP_403_FORBIDDEN,
            "DATA_SCOPE_FORBIDDEN",
            "仅门店账号可以登记推广费发票、替换发票或生命周期事件",
        )


def _lock_promotion_invoice_physical(
    session, physical_invoice_id: str, request: Request
) -> PromotionInvoiceNumberRegistry:
    registry = session.scalar(
        select(PromotionInvoiceNumberRegistry)
        .where(
            PromotionInvoiceNumberRegistry.physical_invoice_id
            == physical_invoice_id
        )
        .with_for_update()
    )
    if registry is None:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "PROMOTION_INVOICE_PHYSICAL_LOCK_MISSING",
            "发票物理版本锁缺失，请先处理迁移异常",
        )
    return registry


def _require_billing_store(session, store_id: str, request: Request) -> None:
    if session.get(DimStore, store_id) is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "门店不存在",
            field="storeId",
        )


def _normalize_billing_direction(value: str | None, request: Request) -> str | None:
    if value is None:
        return None
    direction = value.upper()
    _validate_enum(direction, FEE_DIRECTIONS, "feeDirection", request)
    return direction


def _get_billing_statement(session, statement_id: str, request: Request) -> SettlementStatement:
    statement = session.scalar(
        select(SettlementStatement).where(SettlementStatement.statement_id == statement_id)
    )
    if statement is None:
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "账单不存在",
            field="statementId",
        )
    return statement


def _promotion_invoice_carryforward_projection(
    session,
    *,
    store_id: str | None = None,
) -> tuple[dict[str, dict], list[dict]]:
    """Project deterministic promotion-invoice groups from current facts."""

    effective_period_allocation_exists = (
        select(PromotionInvoiceAllocation.allocation_id)
        .join(
            PromotionInvoice,
            PromotionInvoice.invoice_id == PromotionInvoiceAllocation.invoice_id,
        )
        .where(
            PromotionInvoiceAllocation.store_id == SettlementStatement.store_id,
            PromotionInvoiceAllocation.statement_month
            == SettlementStatement.statement_month,
            PromotionInvoiceAllocation.is_current.is_(True),
            PromotionInvoice.is_current.is_(True),
            PromotionInvoice.is_tombstone.is_(False),
        )
        .exists()
    )
    conditions = [
        SettlementStatement.is_current.is_(True),
        SettlementStatement.statement_month >= FORMAL_PERIOD_START_MONTH,
        SettlementStatementConfirmation.fee_direction == 1,
        SettlementStatementConfirmation.confirmation_status == 1,
        ~effective_period_allocation_exists,
    ]
    if store_id is not None:
        conditions.append(SettlementStatement.store_id == store_id)
    rows = list(
        session.execute(
            select(SettlementStatement, SettlementStatementConfirmation)
            .join(
                SettlementStatementConfirmation,
                SettlementStatementConfirmation.statement_id
                == SettlementStatement.statement_id,
            )
            .where(*conditions)
            .order_by(
                SettlementStatement.store_id,
                SettlementStatement.statement_month,
                SettlementStatement.statement_id,
            )
        ).all()
    )
    rows_by_store: dict[
        str,
        list[tuple[SettlementStatement, SettlementStatementConfirmation]],
    ] = {}
    for statement, confirmation in rows:
        rows_by_store.setdefault(statement.store_id, []).append(
            (statement, confirmation)
        )

    projection: dict[str, dict] = {}
    groups: list[dict] = []

    def add_group(
        member_rows: list[
            tuple[SettlementStatement, SettlementStatementConfirmation]
        ],
        *,
        invoiceable_amount_cent: int,
    ) -> None:
        ordered_members = sorted(
            member_rows,
            key=lambda row: (
                row[0].statement_month,
                row[0].statement_id,
            ),
        )
        identity = [
            {
                "statementId": statement.statement_id,
                "statementVersion": statement.version_no,
                "confirmationId": confirmation.confirmation_id,
                "confirmationStatus": confirmation.confirmation_status,
            }
            for statement, confirmation in ordered_members
        ]
        digest = sha256(
            json.dumps(
                identity,
                ensure_ascii=False,
                separators=(",", ":"),
                sort_keys=True,
            ).encode("utf-8")
        ).hexdigest()
        group_id = f"promotion-invoice-group-{digest[:40]}"
        required_statement_ids = [
            statement.statement_id for statement, _ in ordered_members
        ]
        closing_month = (
            max(statement.statement_month for statement, _ in ordered_members)
            if invoiceable_amount_cent > 0
            else None
        )
        group = {
            "group_id": group_id,
            "store_id": ordered_members[0][0].store_id,
            "required_statement_ids": required_statement_ids,
            "invoiceable_amount_cent": invoiceable_amount_cent,
            "positive_amount_cent": sum(
                max(confirmation.confirmed_amount_cent, 0)
                for _, confirmation in ordered_members
            ),
            "negative_amount_cent": sum(
                min(confirmation.confirmed_amount_cent, 0)
                for _, confirmation in ordered_members
            ),
            "closing_month": closing_month,
        }
        groups.append(group)
        chronological_total = 0
        for statement, confirmation in ordered_members:
            chronological_total += confirmation.confirmed_amount_cent
            projection[statement.statement_id] = {
                **group,
                "statement_id": statement.statement_id,
                "statement_month": statement.statement_month,
                "statement_version": statement.version_no,
                "confirmation_id": confirmation.confirmation_id,
                "confirmed_amount_cent": confirmation.confirmed_amount_cent,
                "carryforward_balance_cent": min(chronological_total, 0),
            }

    for store_rows in rows_by_store.values():
        negative_rows = [
            row for row in store_rows if row[1].confirmed_amount_cent < 0
        ]
        positive_rows = [
            row for row in store_rows if row[1].confirmed_amount_cent > 0
        ]
        pending_rows: list[
            tuple[SettlementStatement, SettlementStatementConfirmation]
        ] = []
        running_balance = 0
        for statement, confirmation in negative_rows:
            pending_rows.append((statement, confirmation))
            running_balance += confirmation.confirmed_amount_cent

        positive_index = 0
        if pending_rows:
            while positive_index < len(positive_rows):
                statement, confirmation = positive_rows[positive_index]
                positive_index += 1
                pending_rows.append((statement, confirmation))
                running_balance += confirmation.confirmed_amount_cent
                if running_balance > 0:
                    add_group(
                        pending_rows,
                        invoiceable_amount_cent=running_balance,
                    )
                    pending_rows = []
                    running_balance = 0
                    break
            if pending_rows:
                add_group(
                    pending_rows,
                    invoiceable_amount_cent=0,
                )

        for statement, confirmation in positive_rows[positive_index:]:
            add_group(
                [(statement, confirmation)],
                invoiceable_amount_cent=confirmation.confirmed_amount_cent,
            )

    return projection, groups


def _statement_list_item(
    session,
    statement: SettlementStatement,
    *,
    direction: str | None,
    promotion_projection: dict[str, dict] | None = None,
) -> dict:
    item = _statement_header_item(
        session,
        statement,
        promotion_projection=promotion_projection,
    )
    if direction is not None:
        item["fee_direction"] = direction
    return item


def _statement_metrics(session, store_id: str, month: str, metric_scope: str) -> dict:
    monthly_conditions = [
        SettlementStatement.store_id == store_id,
        SettlementStatement.statement_month == month,
        SettlementStatement.is_current.is_(True),
    ]
    monthly = session.execute(
        select(
            func.coalesce(func.sum(SettlementStatement.promotion_net_fee_cent), 0),
            func.coalesce(func.sum(SettlementStatement.management_net_fee_cent), 0),
        ).where(*monthly_conditions)
    ).one()
    _, promotion_groups = _promotion_invoice_carryforward_projection(
        session,
        store_id=store_id,
    )
    monthly_invoiceable = sum(
        group["invoiceable_amount_cent"]
        for group in promotion_groups
        if group["closing_month"] == month
    )
    data = {
        "month": {
            "promotion_amount_cent": int(monthly[0]),
            "management_amount_cent": int(monthly[1]),
            "promotion_invoiceable_amount_cent": monthly_invoiceable,
        }
    }
    if metric_scope == "CUMULATIVE":
        cumulative = session.execute(
            select(
                func.coalesce(func.sum(SettlementStatement.promotion_net_fee_cent), 0),
                func.coalesce(func.sum(SettlementStatement.management_net_fee_cent), 0),
            ).where(
                SettlementStatement.store_id == store_id,
                SettlementStatement.statement_month >= FORMAL_PERIOD_START_MONTH,
                SettlementStatement.statement_month <= month,
                SettlementStatement.is_current.is_(True),
            )
        ).one()
        data["cumulative"] = {
            "promotion_amount_cent": int(cumulative[0]),
            "management_amount_cent": int(cumulative[1]),
            "promotion_invoiceable_amount_cent": sum(
                group["invoiceable_amount_cent"]
                for group in promotion_groups
                if group["closing_month"] is not None
                and group["closing_month"] <= month
            ),
        }
    return data


def _require_finance_admin(current_user: AuthContext, request: Request) -> None:
    if not current_user.is_admin:
        _raise_reporting_error(
            request,
            status.HTTP_403_FORBIDDEN,
            "DATA_SCOPE_FORBIDDEN",
            "仅管理员和最高管理员可以查询财务汇总",
        )


def _finance_operator_role(current_user: AuthContext) -> int:
    if current_user.role == "highest_admin":
        return 3
    if current_user.is_admin:
        return 2
    return 1


def _persist_finance_conflict_audit(
    session,
    *,
    request: Request,
    current_user: AuthContext,
    operation_type: str,
    target_type: str,
    target_id: str,
    conflict_code: str,
    idempotency_key_hash: str | None,
    request_payload_sha256: str | None,
    details: dict | None = None,
) -> None:
    """Rollback business writes, then persist a conflict without claiming replay ownership."""
    session.rollback()
    now = utcnow()
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type=operation_type,
            target_type=target_type,
            target_id=target_id,
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot=None,
            after_snapshot={
                "conflictCode": conflict_code,
                "attemptedIdempotencyKeyHash": idempotency_key_hash,
                "details": details or {},
            },
            result_status=2,
            request_id=request_id(request),
            # Only successful operations own the replay key. Keeping failure rows
            # NULL avoids the existing unique constraint hiding conflict evidence.
            idempotency_key_hash=None,
            request_payload_sha256=request_payload_sha256,
            occurred_at=now,
        )
    )
    session.commit()


def _raise_finance_conflict(
    session,
    *,
    request: Request,
    current_user: AuthContext,
    operation_type: str,
    target_type: str,
    target_id: str,
    conflict_code: str,
    message: str,
    idempotency_key_hash: str | None,
    request_payload_sha256: str | None,
    field: str | None = None,
    data: dict | None = None,
) -> None:
    _persist_finance_conflict_audit(
        session,
        request=request,
        current_user=current_user,
        operation_type=operation_type,
        target_type=target_type,
        target_id=target_id,
        conflict_code=conflict_code,
        idempotency_key_hash=idempotency_key_hash,
        request_payload_sha256=request_payload_sha256,
        details=data,
    )
    _raise_reporting_error(
        request,
        status.HTTP_409_CONFLICT,
        conflict_code,
        message,
        field=field,
        data=data,
    )


def _finance_import_error(
    row_number: int,
    business_key: str,
    field: str,
    original_value: str,
    reason: str,
    suggestion: str,
) -> dict:
    return {
        "row_number": row_number,
        "business_key": business_key,
        "field": field,
        "original_value": original_value,
        "reason": reason,
        "suggestion": suggestion,
    }


def _finance_import_cell_text(value) -> str:
    """Normalize a CSV or XLSX cell to stable text."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _iter_finance_import_rows(
    file: UploadFile, filename: str, request: Request
):
    """Yield finance-import rows from CSV or XLSX without loading the file fully."""
    if filename.lower().endswith(".csv"):
        text_stream = TextIOWrapper(file.file, encoding="utf-8-sig", newline="")
        try:
            for row_number, raw_row in enumerate(DictReader(text_stream), start=2):
                yield row_number, raw_row
        except UnicodeDecodeError:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "VALIDATION_FAILED",
                "CSV 文件必须使用 UTF-8 编码",
                field="file",
            )
        finally:
            text_stream.detach()
        return

    try:
        workbook = load_workbook(file.file, read_only=True, data_only=True)
    except Exception:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "XLSX 文件无法读取",
            field="file",
        )
    try:
        values = workbook.active.iter_rows(values_only=True)
        header_row = next(values, None)
        if header_row is None:
            return
        headers = [_finance_import_cell_text(value) for value in header_row]
        for row_number, row_values in enumerate(values, start=2):
            if all(value is None for value in row_values):
                continue
            yield row_number, {
                header: value for header, value in zip(headers, row_values)
            }
    finally:
        workbook.close()


def _validate_final_finance_import_row(
    session,
    *,
    import_type: str,
    statement_month: str,
    row_number: int,
    raw_row: dict,
) -> tuple[str, dict, list[dict]]:
    """Validate the four finance templates frozen in the DYDATA-19 issue body."""
    errors: list[dict] = []
    store_id = _finance_import_cell_text(raw_row.get("storeId"))
    store_name = _finance_import_cell_text(raw_row.get("storeName"))
    invoice_number = _finance_import_cell_text(raw_row.get("invoiceNumber"))
    business_key = store_id
    normalized_payload: dict = {"statementMonth": statement_month}

    def add_error(field: str, value: str, reason: str, suggestion: str) -> None:
        errors.append(
            _finance_import_error(
                row_number, business_key, field, value, reason, suggestion
            )
        )

    def exact_store() -> DimStore | None:
        if not store_id:
            add_error("storeId", store_id, "门店 ID 必填", "填写系统门店 ID")
            return None
        item = session.scalar(select(DimStore).where(DimStore.store_id == store_id))
        if item is None:
            add_error("storeId", store_id, "门店 ID 不存在", "使用系统中存在的门店 ID")
            return None
        if not store_name:
            add_error("storeName", store_name, "门店名称必填", "填写门店当前名称")
        elif item.store_name != store_name:
            add_error(
                "storeName",
                store_name,
                "门店名称与门店 ID 不一致",
                "以门店 ID 对应的系统门店名称为准",
            )
        return item

    def parse_date(field: str, required: bool = True) -> str:
        text = _finance_import_cell_text(raw_row.get(field))
        if not text and not required:
            return ""
        try:
            return date.fromisoformat(text).isoformat()
        except ValueError:
            add_error(field, text, "日期必须使用 YYYY-MM-DD", "填写有效日期")
            return ""

    def parse_datetime(field: str) -> str:
        text = _finance_import_cell_text(raw_row.get(field))
        try:
            parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
            if parsed.tzinfo is None:
                raise ValueError
            return parsed.isoformat()
        except ValueError:
            add_error(
                field,
                text,
                "时间必须是包含时区的 ISO 8601 时间",
                "例如 2026-08-21T10:00:00+08:00",
            )
            return ""

    def parse_positive_cent(field: str, required: bool = True) -> tuple[int | None, bool]:
        text = _finance_import_cell_text(raw_row.get(field))
        if not text and not required:
            return None, True
        try:
            value = int(text)
            if value <= 0:
                raise ValueError
            return value, True
        except ValueError:
            add_error(field, text, "金额必须为正整数分", "填写大于 0 的整数")
            return None, False

    if import_type == "BASIC_INFO":
        exact_store()
        sap_code = _finance_import_cell_text(raw_row.get("sapCode"))
        if not sap_code:
            add_error("sapCode", sap_code, "SAP 编码必填", "填写财务 SAP 编码")
        imported_at = parse_datetime("importedAt")
        normalized_payload.update(
            {
                "storeId": store_id,
                "storeName": store_name,
                "sapCode": sap_code,
                "importedAt": imported_at,
            }
        )
        return business_key, normalized_payload, errors

    if import_type == "SAP_CONFIRMATION":
        exact_store()
        initial_sap = _finance_import_cell_text(raw_row.get("financeInitialSap"))
        service_store_code = _finance_import_cell_text(raw_row.get("serviceStoreCode"))
        result = _finance_import_cell_text(raw_row.get("factoryConfirmationResult")).upper()
        if not initial_sap:
            add_error(
                "financeInitialSap", initial_sap, "财务初始 SAP 必填", "填写财务初始 SAP"
            )
        if not service_store_code:
            add_error(
                "serviceStoreCode", service_store_code, "服务门店编码必填", "填写服务门店编码"
            )
        if result not in {"CONFIRMED", "REJECTED"}:
            add_error(
                "factoryConfirmationResult",
                result,
                "厂家确认结果枚举无效",
                "填写 CONFIRMED 或 REJECTED",
            )
        confirmed_at = parse_datetime("confirmedAt")
        normalized_payload.update(
            {
                "storeId": store_id,
                "storeName": store_name,
                "financeInitialSap": initial_sap,
                "serviceStoreCode": service_store_code,
                "factoryConfirmationResult": result,
                "confirmedAt": confirmed_at,
            }
        )
        return business_key, normalized_payload, errors

    if import_type == "PROMOTION_FACTORY_RESULT":
        business_key = invoice_number
        if len(invoice_number) != 20 or not invoice_number.isdigit():
            add_error(
                "invoiceNumber", invoice_number, "发票号码必须为 20 位数字", "填写有效数电专票号码"
            )
        result = _finance_import_cell_text(raw_row.get("reviewResult")).upper()
        if result not in {"APPROVED_SETTLED", "REJECTED_REUPLOAD"}:
            add_error(
                "reviewResult",
                result,
                "厂家结果枚举无效",
                "填写 APPROVED_SETTLED 或 REJECTED_REUPLOAD",
            )
        rejection_reason = _finance_import_cell_text(raw_row.get("rejectionReason"))
        settlement_date = parse_date(
            "settlementDate", required=result == "APPROVED_SETTLED"
        )
        settlement_amount, amount_valid = parse_positive_cent(
            "settlementAmountCent", required=result == "APPROVED_SETTLED"
        )
        if result == "REJECTED_REUPLOAD" and not rejection_reason:
            add_error(
                "rejectionReason", rejection_reason, "审核不通过原因必填", "填写厂家退回原因"
            )
        current_invoice = session.scalar(
            select(PromotionInvoice)
            .join(
                PromotionInvoiceAllocation,
                PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id,
            )
            .where(
                PromotionInvoice.invoice_number == invoice_number,
                PromotionInvoice.is_current.is_(True),
                PromotionInvoiceAllocation.statement_month == statement_month,
                PromotionInvoiceAllocation.is_current.is_(True),
            )
        )
        if invoice_number and current_invoice is None:
            add_error(
                "invoiceNumber",
                invoice_number,
                "未找到该账期精确匹配的当前推广费发票",
                "核对发票号码和导入账期，禁止模糊匹配",
            )
        elif (
            current_invoice is not None
            and amount_valid
            and settlement_amount is not None
            and current_invoice.invoice_amount_cent != settlement_amount
        ):
            add_error(
                "settlementAmountCent",
                str(settlement_amount),
                "结算金额必须等于当前推广费发票全额",
                "填写当前有效发票全额，不允许部分结算",
            )
        normalized_payload.update(
            {
                "storeId": current_invoice.store_id if current_invoice is not None else "",
                "invoiceNumber": invoice_number,
                "reviewResult": result,
                "rejectionReason": rejection_reason,
                "settlementDate": settlement_date,
                "settlementAmountCent": settlement_amount,
            }
        )
        return business_key, normalized_payload, errors

    business_key = f"{store_id}|{statement_month}"
    exact_store()
    row_month = _finance_import_cell_text(raw_row.get("statementMonth"))
    if row_month != statement_month:
        add_error(
            "statementMonth",
            row_month,
            "行账期必须与批次账期一致",
            f"填写 {statement_month}",
        )
    if len(invoice_number) != 20 or not invoice_number.isdigit():
        add_error(
            "invoiceNumber", invoice_number, "发票号码必须为 20 位数字", "填写有效数电专票号码"
        )
    invoice_date = parse_date("invoiceDate")
    deduction_date = parse_date("deductionDate")
    deduction_amount, amount_valid = parse_positive_cent("deductionAmountCent")
    _append_management_amount_errors(
        session,
        errors=errors,
        row_number=row_number,
        business_key=business_key,
        store_id=store_id,
        statement_month=statement_month,
        amount_text=str(deduction_amount or ""),
        amount_cent=deduction_amount or 0,
        amount_is_valid=amount_valid,
    )
    normalized_payload.update(
        {
            "storeId": store_id,
            "storeName": store_name,
            "invoiceNumber": invoice_number,
            "invoiceDate": invoice_date,
            "deductionDate": deduction_date,
            "deductionAmountCent": deduction_amount,
        }
    )
    return business_key, normalized_payload, errors


def _validate_finance_import_row(
    session,
    *,
    import_type: str,
    statement_month: str,
    row_number: int,
    raw_row: dict,
) -> tuple[str, dict, list[dict]]:
    """Validate and normalize one of the four finance-import row types."""
    store_id = _finance_import_cell_text(raw_row.get("storeId"))
    invoice_number = _finance_import_cell_text(raw_row.get("invoiceNumber"))
    amount_text = _finance_import_cell_text(raw_row.get("amountCent"))
    business_key = f"{store_id}|{statement_month}|{invoice_number}"
    errors: list[dict] = []
    if not store_id:
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "storeId",
                store_id,
                "门店 ID 必填",
                "填写系统门店 ID",
            )
        )
    elif session.scalar(
        select(DimStore.store_id).where(DimStore.store_id == store_id)
    ) is None:
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "storeId",
                store_id,
                "门店 ID 不存在",
                "使用系统中存在的门店 ID",
            )
        )
    if len(invoice_number) != 20 or not invoice_number.isdigit():
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "invoiceNumber",
                invoice_number,
                "发票号码必须为 20 位数字",
                "按模板填写数电专票号码",
            )
        )
    try:
        amount_cent = int(amount_text)
        if amount_cent <= 0:
            raise ValueError
        amount_is_valid = True
    except ValueError:
        amount_cent = 0
        amount_is_valid = False
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "amountCent",
                amount_text,
                "金额必须为正整数分",
                "填写大于 0 的整数",
            )
        )

    normalized_payload = {
        "storeId": store_id,
        "statementMonth": statement_month,
        "invoiceNumber": invoice_number,
        "amountCent": amount_cent,
    }
    if import_type in {
        "PROMOTION_REVIEW_RESULT",
        "PROMOTION_SETTLEMENT_RESULT",
    }:
        imported_status = (
            "APPROVED_SETTLED"
            if import_type == "PROMOTION_SETTLEMENT_RESULT"
            else _finance_import_cell_text(raw_row.get("status")).upper()
        )
        allowed_statuses = (
            {"APPROVED_SETTLED", "REJECTED_REUPLOAD"}
            if import_type == "PROMOTION_REVIEW_RESULT"
            else {"APPROVED_SETTLED"}
        )
        if imported_status not in allowed_statuses:
            errors.append(
                _finance_import_error(
                    row_number,
                    business_key,
                    "status",
                    imported_status,
                    "推广费结果枚举无效",
                    f"填写 {' 或 '.join(sorted(allowed_statuses))}",
                )
            )
        normalized_payload["status"] = imported_status
        current_invoice = _find_current_promotion_invoice(
            session,
            store_id=store_id,
            statement_month=statement_month,
            invoice_number=invoice_number,
        )
        if store_id and invoice_number and current_invoice is None:
            errors.append(
                _finance_import_error(
                    row_number,
                    business_key,
                    "businessKey",
                    business_key,
                    "未找到精确匹配的当前推广费发票",
                    "核对门店 ID、账期和发票号码",
                )
            )
        elif (
            current_invoice is not None
            and amount_is_valid
            and current_invoice.invoice_amount_cent != amount_cent
        ):
            errors.append(
                _finance_import_error(
                    row_number,
                    business_key,
                    "amountCent",
                    amount_text,
                    "金额与当前推广费发票不一致",
                    "填写当前有效发票全额",
                )
            )
    elif import_type == "MANAGEMENT_INVOICE_DETAIL":
        invoice_date_text = _finance_import_cell_text(raw_row.get("invoiceDate"))
        try:
            invoice_date_value = date.fromisoformat(invoice_date_text).isoformat()
        except ValueError:
            invoice_date_value = ""
            errors.append(
                _finance_import_error(
                    row_number,
                    business_key,
                    "invoiceDate",
                    invoice_date_text,
                    "开票日期必须使用 YYYY-MM-DD",
                    "填写有效开票日期",
                )
            )
        normalized_payload["invoiceDate"] = invoice_date_value
        _append_management_amount_errors(
            session,
            errors=errors,
            row_number=row_number,
            business_key=business_key,
            store_id=store_id,
            statement_month=statement_month,
            amount_text=amount_text,
            amount_cent=amount_cent,
            amount_is_valid=amount_is_valid,
        )
    else:
        current_invoice = _find_current_management_invoice(
            session,
            store_id=store_id,
            statement_month=statement_month,
            invoice_number=invoice_number,
        )
        if store_id and invoice_number and current_invoice is None:
            errors.append(
                _finance_import_error(
                    row_number,
                    business_key,
                    "businessKey",
                    business_key,
                    "未找到精确匹配的当前管理服务费发票",
                    "核对门店 ID、账期和发票号码",
                )
            )
        elif (
            current_invoice is not None
            and amount_is_valid
            and current_invoice.invoice_amount_cent != amount_cent
        ):
            errors.append(
                _finance_import_error(
                    row_number,
                    business_key,
                    "amountCent",
                    amount_text,
                    "厂家扣款金额必须等于当前发票全额",
                    "填写当前有效发票全额，不允许部分扣款",
                )
            )
        _append_management_amount_errors(
            session,
            errors=errors,
            row_number=row_number,
            business_key=business_key,
            store_id=store_id,
            statement_month=statement_month,
            amount_text=amount_text,
            amount_cent=amount_cent,
            amount_is_valid=amount_is_valid,
        )
    return business_key, normalized_payload, errors


def _find_current_promotion_invoice(
    session, *, store_id: str, statement_month: str, invoice_number: str
) -> PromotionInvoice | None:
    """Find a promotion invoice using only its exact business key."""
    return session.scalar(
        select(PromotionInvoice)
        .join(
            PromotionInvoiceAllocation,
            PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id,
        )
        .where(
            PromotionInvoice.store_id == store_id,
            PromotionInvoice.invoice_number == invoice_number,
            PromotionInvoice.is_current.is_(True),
            PromotionInvoiceAllocation.statement_month == statement_month,
            PromotionInvoiceAllocation.is_current.is_(True),
        )
    )


def _find_current_management_invoice(
    session, *, store_id: str, statement_month: str, invoice_number: str
) -> InvoiceRecord | None:
    """Find a management invoice using only its exact business key."""
    return session.scalar(
        select(InvoiceRecord).where(
            InvoiceRecord.store_id == store_id,
            InvoiceRecord.statement_month == statement_month,
            InvoiceRecord.fee_direction == 2,
            InvoiceRecord.invoice_number == invoice_number,
            InvoiceRecord.is_current.is_(True),
            InvoiceRecord.is_tombstone.is_(False),
        )
    )


def _append_management_amount_errors(
    session,
    *,
    errors: list[dict],
    row_number: int,
    business_key: str,
    store_id: str,
    statement_month: str,
    amount_text: str,
    amount_cent: int,
    amount_is_valid: bool,
) -> None:
    """Validate one management row against the shared invoiceable projection."""
    projected_periods, _ = _management_invoiceable_projection(
        session,
        store_id=store_id,
        through_month=statement_month,
    )
    projected_period = next(
        (
            period
            for period in projected_periods
            if period["statement_month"] == statement_month
        ),
        None,
    )
    if store_id and projected_period is None:
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "businessKey",
                business_key,
                "未找到当前账期有效的管理服务费确认",
                "先完成对应账期管理服务费确认",
            )
        )
    elif projected_period is not None and projected_period["invoice_id"] is not None:
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "businessKey",
                business_key,
                "当前账期已存在有效管理服务费发票",
                "更正当前版本或选择仍有可开票金额的账期",
            )
        )
    elif projected_period is not None and projected_period["invoiceable_amount_cent"] <= 0:
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "deductionAmountCent",
                amount_text,
                "当前账期经负数结转后无可开票金额",
                "等待后续正数账期或核对当前有效账单",
            )
        )
    elif (
        projected_period is not None
        and amount_is_valid
        and projected_period["invoiceable_amount_cent"] != amount_cent
    ):
        errors.append(
            _finance_import_error(
                row_number,
                business_key,
                "deductionAmountCent",
                amount_text,
                "金额必须等于当前管理服务费可开票金额",
                "填写负数结转后的可开票全额，不允许部分扣款",
            )
        )


def _finance_import_upload_item(session, batch: FinanceImportBatch) -> dict:
    """Build a stable upload response for first execution and idempotent replay."""
    error_rows = session.scalars(
        select(FinanceImportRow)
        .where(
            FinanceImportRow.batch_id == batch.batch_id,
            FinanceImportRow.row_status == 4,
        )
        .order_by(FinanceImportRow.row_number)
        .limit(100)
    )
    errors = [
        error
        for row in error_rows
        for error in (row.validation_errors or [])
    ][:100]
    return {
        "batch_id": batch.batch_id,
        "scenario": FINANCE_IMPORT_SCENARIO_FROM_STATUS[batch.batch_status],
        "read_version": batch.read_version,
        "current_version": batch.current_version,
        "content_changed": batch.content_changed,
        "total_rows": batch.total_rows,
        "success_rows": batch.success_rows,
        "error_rows": batch.error_rows,
        "errors": {
            "list": errors,
            "total": _count_finance_import_errors(session, batch.batch_id),
        },
    }


def _finance_import_reversal_eligibility(
    session,
    batch: FinanceImportBatch,
    *,
    reversed_by_batch_id: str | None,
) -> tuple[bool, str | None, str | None]:
    """Evaluate every imported business key without mutating persisted state."""
    if reversed_by_batch_id is not None:
        return False, "IMPORT_ALREADY_REVERSED", "该批次已被撤销"
    if batch.batch_status not in (5, 8):
        return False, "BATCH_STATUS_NOT_REVERSIBLE", "当前批次状态不允许撤销"
    rows = list(
        session.scalars(
            select(FinanceImportRow).where(
                FinanceImportRow.batch_id == batch.batch_id,
                FinanceImportRow.row_status == 5,
            )
        )
    )
    if not rows or len(rows) != batch.success_rows:
        return False, "REVERSAL_TARGET_INCOMPLETE", "原批次业务目标不完整"
    for row in rows:
        if not row.target_record_id:
            return (
                False,
                "REVERSAL_TARGET_MISSING",
                f"业务键 {row.business_key} 未关联业务版本",
            )
        if not _finance_import_target_is_current(
            session, batch.import_type, row.target_record_id
        ):
            return (
                False,
                "REVERSAL_BUSINESS_VERSION_CONFLICT",
                f"业务键 {row.business_key} 已被后续版本覆盖",
            )
    return True, None, None


def _finance_import_target_is_current(
    session,
    import_type: int,
    target_record_id: str,
) -> bool:
    """Return whether one import target is the current immutable version."""
    if import_type in (1, 4):
        is_current = session.scalar(
            select(StoreFinanceProfile.is_current).where(
                StoreFinanceProfile.profile_id == target_record_id
            )
        )
    elif import_type == 3:
        is_current = session.scalar(
            select(InvoiceRecord.is_current).where(
                InvoiceRecord.invoice_id == target_record_id
            )
        )
    else:
        is_current = session.scalar(
            select(PromotionInvoice.is_current).where(
                PromotionInvoice.invoice_id == target_record_id
            )
        )
    return is_current is True


def _finance_import_reversal_rows(
    session,
    batch: FinanceImportBatch,
    *,
    reversed_by_batch_id: str | None,
    page: int,
    page_size: int,
) -> dict:
    """Return a paged per-business-key reversal lineage for one batch."""
    lineage_batch_id = (
        batch.batch_id
        if batch.reverses_batch_id is not None
        else reversed_by_batch_id or batch.batch_id
    )
    conditions = [
        FinanceImportRow.batch_id == lineage_batch_id,
        FinanceImportRow.row_status == 5,
    ]
    total = int(
        session.scalar(
            select(func.count()).select_from(FinanceImportRow).where(*conditions)
        )
        or 0
    )
    rows = list(
        session.scalars(
            select(FinanceImportRow)
            .where(*conditions)
            .order_by(FinanceImportRow.row_number)
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    )
    items = []
    for row in rows:
        is_reversal_row = row.reversal_effect_type is not None
        current_target_id = row.target_record_id
        items.append(
            {
                "business_key": row.business_key,
                "original_target_record_id": (
                    row.reverses_target_record_id
                    if is_reversal_row
                    else row.target_record_id
                ),
                "previous_target_record_id": row.previous_target_record_id,
                "reversal_target_record_id": (
                    row.target_record_id if is_reversal_row else None
                ),
                "effect_type": (
                    {1: "VALUE", 2: "TOMBSTONE"}.get(row.reversal_effect_type)
                ),
                "is_current": (
                    _finance_import_target_is_current(
                        session, batch.import_type, current_target_id
                    )
                    if current_target_id
                    else False
                ),
            }
        )
    return {
        "list": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }


def _finance_import_batch_item(
    session,
    batch: FinanceImportBatch,
    *,
    reversed_by_batch_id: str | None = None,
) -> dict:
    """Build the shared finance-import batch representation."""
    can_reverse, reverse_code, reverse_reason = _finance_import_reversal_eligibility(
        session,
        batch,
        reversed_by_batch_id=reversed_by_batch_id,
    )
    if batch.reverses_batch_id is not None:
        reversal_chain = [batch.reverses_batch_id, batch.batch_id]
    elif reversed_by_batch_id is not None:
        reversal_chain = [batch.batch_id, reversed_by_batch_id]
    else:
        reversal_chain = [batch.batch_id]
    return {
        "batch_id": batch.batch_id,
        "import_type": FINANCE_IMPORT_TYPE_FROM_DB[batch.import_type],
        "statement_month": batch.statement_month,
        "file_name": batch.file_name,
        "scenario": FINANCE_IMPORT_SCENARIO_FROM_STATUS[batch.batch_status],
        "read_version": batch.read_version,
        "current_version": batch.current_version,
        "content_changed": batch.content_changed,
        "reverses_batch_id": batch.reverses_batch_id,
        "reversed_by_batch_id": reversed_by_batch_id,
        "reversal_chain": reversal_chain,
        "can_reverse": can_reverse,
        "reverse_not_allowed_code": reverse_code,
        "reverse_not_allowed_reason": reverse_reason,
        "total_rows": batch.total_rows,
        "success_rows": batch.success_rows,
        "error_rows": batch.error_rows,
        "submitted_by": batch.submitted_by,
        "submitted_at": batch.submitted_at,
        "committed_by": batch.committed_by,
        "committed_at": batch.committed_at,
    }


def _count_finance_import_errors(session, batch_id: str) -> int:
    """Count nested validation errors without retaining all row objects."""
    count = 0
    rows = session.scalars(
        select(FinanceImportRow)
        .where(
            FinanceImportRow.batch_id == batch_id,
            FinanceImportRow.row_status == 4,
        )
        .order_by(FinanceImportRow.row_number)
    ).yield_per(500)
    for row in rows:
        count += len(row.validation_errors or [])
    return count


def _finance_import_error_csv(session, batch_id: str):
    """Yield a UTF-8 BOM CSV containing every stored validation error."""
    fieldnames = [
        "rowNumber",
        "businessKey",
        "field",
        "originalValue",
        "reason",
        "suggestion",
    ]
    buffer = StringIO()
    buffer.write("\ufeff")
    writer = DictWriter(buffer, fieldnames=fieldnames, lineterminator="\r\n")
    writer.writeheader()
    yield buffer.getvalue()
    buffer.seek(0)
    buffer.truncate(0)
    rows = session.scalars(
        select(FinanceImportRow)
        .where(
            FinanceImportRow.batch_id == batch_id,
            FinanceImportRow.row_status == 4,
        )
        .order_by(FinanceImportRow.row_number)
    ).yield_per(500)
    for row in rows:
        for error in row.validation_errors or []:
            writer.writerow(
                {
                    "rowNumber": error["row_number"],
                    "businessKey": error["business_key"],
                    "field": error["field"],
                    "originalValue": error["original_value"],
                    "reason": error["reason"],
                    "suggestion": error["suggestion"],
                }
            )
            yield buffer.getvalue()
            buffer.seek(0)
            buffer.truncate(0)


def _finance_import_commit_item(batch: FinanceImportBatch) -> dict:
    """Build the stable response for a committed import batch."""
    return {
        "batch_id": batch.batch_id,
        "status": "CORRECTED" if batch.batch_status == 8 else "COMMITTED",
        "read_version": batch.read_version,
        "current_version": batch.current_version,
        "total_rows": batch.total_rows,
        "success_rows": batch.success_rows,
        "error_rows": batch.error_rows,
        "committed_by": batch.committed_by,
        "committed_at": batch.committed_at,
    }


def _commit_final_finance_import_row(
    session,
    *,
    batch: FinanceImportBatch,
    row: FinanceImportRow,
    operator_id: str,
    committed_at: datetime,
    is_correction: bool,
    request: Request,
) -> None:
    """Route one valid row to the exact frozen template target."""
    if batch.import_type in (1, 4):
        _commit_store_finance_profile_row(
            session, batch=batch, row=row, committed_at=committed_at
        )
    elif batch.import_type == 2:
        _commit_promotion_factory_result_row(
            session,
            batch=batch,
            row=row,
            operator_id=operator_id,
            committed_at=committed_at,
            request=request,
        )
    else:
        _commit_management_factory_result_row(
            session,
            batch=batch,
            row=row,
            operator_id=operator_id,
            committed_at=committed_at,
            is_correction=is_correction,
            request=request,
        )


def _commit_store_finance_profile_row(
    session,
    *,
    batch: FinanceImportBatch,
    row: FinanceImportRow,
    committed_at: datetime,
) -> None:
    """Append a basic-information or SAP-confirmation profile version."""
    payload = row.normalized_payload
    profile_type = 1 if batch.import_type == 1 else 2
    current = session.scalar(
        select(StoreFinanceProfile).where(
            StoreFinanceProfile.store_id == payload["storeId"],
            StoreFinanceProfile.profile_type == profile_type,
            StoreFinanceProfile.is_current.is_(True),
        )
    )
    next_version = 1
    if current is not None:
        next_version = current.version_no + 1
        current.is_current = False
        session.flush()
    profile_id = f"store-finance-profile-{uuid4().hex}"
    session.add(
        StoreFinanceProfile(
            profile_id=profile_id,
            store_id=payload["storeId"],
            profile_type=profile_type,
            version_no=next_version,
            is_current=True,
            store_name_snapshot=payload["storeName"],
            sap_code=payload.get("sapCode"),
            initial_sap_code=payload.get("financeInitialSap"),
            service_store_code=payload.get("serviceStoreCode"),
            factory_confirmed=(
                payload.get("factoryConfirmationResult") == "CONFIRMED"
                if profile_type == 2
                else None
            ),
            confirmed_at=(
                datetime.fromisoformat(payload["confirmedAt"])
                if profile_type == 2
                else None
            ),
            import_batch_id=batch.batch_id,
            created_at=committed_at,
            updated_at=committed_at,
        )
    )
    row.row_status = 5
    row.target_record_id = profile_id


def _commit_promotion_factory_result_row(
    session,
    *,
    batch: FinanceImportBatch,
    row: FinanceImportRow,
    operator_id: str,
    committed_at: datetime,
    request: Request,
) -> None:
    """Append the externally reviewed promotion invoice result."""
    payload = row.normalized_payload
    current_invoice = session.scalar(
        select(PromotionInvoice)
        .join(
            PromotionInvoiceAllocation,
            PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id,
        )
        .where(
            PromotionInvoice.invoice_number == payload["invoiceNumber"],
            PromotionInvoice.is_current.is_(True),
            PromotionInvoiceAllocation.statement_month == batch.statement_month,
            PromotionInvoiceAllocation.is_current.is_(True),
        )
        .with_for_update()
    )
    if current_invoice is None:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "IMPORT_TARGET_NOT_FOUND",
            "推广费发票当前版本不存在",
            field="invoiceNumber",
        )
    _lock_promotion_invoice_physical(
        session, current_invoice.physical_invoice_id, request
    )
    current_invoice = session.scalar(
        select(PromotionInvoice)
        .join(
            PromotionInvoiceAllocation,
            PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id,
        )
        .where(
            PromotionInvoice.invoice_number == payload["invoiceNumber"],
            PromotionInvoice.is_current.is_(True),
            PromotionInvoiceAllocation.statement_month == batch.statement_month,
            PromotionInvoiceAllocation.is_current.is_(True),
        )
        .with_for_update()
    )
    if current_invoice is None:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "PROMOTION_INVOICE_VERSION_CONFLICT",
            "推广费发票版本已变化，请重新预检导入文件",
            field="invoiceNumber",
        )
    settlement_amount = payload.get("settlementAmountCent")
    if (
        settlement_amount is not None
        and current_invoice.invoice_amount_cent != settlement_amount
    ):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "IMPORT_AMOUNT_MISMATCH",
            "导入结算金额与当前推广费发票不一致",
            field="settlementAmountCent",
        )
    allocations = list(
        session.scalars(
            select(PromotionInvoiceAllocation).where(
                PromotionInvoiceAllocation.invoice_id == current_invoice.invoice_id,
                PromotionInvoiceAllocation.is_current.is_(True),
            )
        )
    )
    current_invoice.is_current = False
    for allocation in allocations:
        allocation.is_current = False
    session.flush()
    next_status = 3 if payload["reviewResult"] == "APPROVED_SETTLED" else 4
    next_invoice_id = f"promotion-invoice-{uuid4().hex}"
    session.add(
        PromotionInvoice(
            invoice_id=next_invoice_id,
            physical_invoice_id=current_invoice.physical_invoice_id,
            store_id=current_invoice.store_id,
            version_no=current_invoice.version_no + 1,
            version_kind=2,
            is_current=True,
            supersedes_invoice_id=current_invoice.invoice_id,
            replaces_invoice_id=current_invoice.replaces_invoice_id,
            invoice_number=current_invoice.invoice_number,
            invoice_date=current_invoice.invoice_date,
            invoice_amount_cent=current_invoice.invoice_amount_cent,
            buyer_name=current_invoice.buyer_name,
            tax_rate_percent=current_invoice.tax_rate_percent,
            invoice_status=next_status,
            registered_by=current_invoice.registered_by,
            registered_at=current_invoice.registered_at,
        )
    )
    for allocation in allocations:
        session.add(
            PromotionInvoiceAllocation(
                allocation_id=f"promotion-allocation-{uuid4().hex}",
                invoice_id=next_invoice_id,
                store_id=allocation.store_id,
                statement_id=allocation.statement_id,
                statement_month=allocation.statement_month,
                settlement_batch_month=allocation.settlement_batch_month,
                allocated_amount_cent=allocation.allocated_amount_cent,
                is_current=True,
            )
        )
    session.add(
        InvoiceStatusEvent(
            event_id=f"invoice-event-{uuid4().hex}",
            invoice_id=next_invoice_id,
            event_type=2,
            from_status=current_invoice.invoice_status,
            to_status=next_status,
            operator_id=operator_id,
            import_batch_id=batch.batch_id,
            result_reason=payload.get("rejectionReason") or None,
            business_date=(
                date.fromisoformat(payload["settlementDate"])
                if payload.get("settlementDate")
                else None
            ),
            business_amount_cent=settlement_amount,
            occurred_at=committed_at,
        )
    )
    row.row_status = 5
    row.target_record_id = next_invoice_id


def _commit_management_factory_result_row(
    session,
    *,
    batch: FinanceImportBatch,
    row: FinanceImportRow,
    operator_id: str,
    committed_at: datetime,
    is_correction: bool,
    request: Request,
) -> None:
    """Append the combined management invoice and factory-deduction version."""
    payload = row.normalized_payload
    statement = session.scalar(
        select(SettlementStatement).where(
            SettlementStatement.store_id == payload["storeId"],
            SettlementStatement.statement_month == batch.statement_month,
            SettlementStatement.is_current.is_(True),
        )
    )
    if statement is None:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "IMPORT_TARGET_NOT_FOUND",
            "管理服务费账期不存在当前账单",
            field="businessKey",
        )
    _, carryforward_applications = _management_invoiceable_projection(
        session,
        store_id=payload["storeId"],
        through_month=batch.statement_month,
    )
    current = session.scalar(
        select(InvoiceRecord).where(
            InvoiceRecord.store_id == payload["storeId"],
            InvoiceRecord.statement_month == batch.statement_month,
            InvoiceRecord.fee_direction == 2,
            InvoiceRecord.is_current.is_(True),
        )
    )
    next_version = 1
    if current is not None:
        next_version = current.version_no + 1
        current.is_current = False
        session.flush()
    next_invoice_id = f"management-invoice-{uuid4().hex}"
    next_invoice = InvoiceRecord(
            invoice_id=next_invoice_id,
            store_id=payload["storeId"],
            statement_month=batch.statement_month,
            statement_id=statement.statement_id,
            fee_direction=2,
            version_no=next_version,
            is_current=True,
            invoice_number=payload["invoiceNumber"],
            invoice_date=date.fromisoformat(payload["invoiceDate"]),
            invoice_amount_cent=payload["deductionAmountCent"],
            invoice_status=3,
            source_type=3 if is_correction else 2,
            import_batch_id=batch.batch_id,
            factory_deduction_date=date.fromisoformat(payload["deductionDate"]),
            factory_deduction_amount_cent=payload["deductionAmountCent"],
            registered_by=operator_id,
            registered_at=committed_at,
        )
    session.add(next_invoice)
    session.flush()
    _synchronize_management_carryforward_applications(
        session,
        applications=carryforward_applications,
        invoice_id_by_statement={statement.statement_id: next_invoice_id},
        scope_store_ids={payload["storeId"]},
        scope_through_month=batch.statement_month,
    )
    row.row_status = 5
    row.target_record_id = next_invoice_id


def _commit_promotion_review_row(
    session,
    *,
    batch: FinanceImportBatch,
    row: FinanceImportRow,
    operator_id: str,
    committed_at: datetime,
    request: Request,
) -> None:
    """Create the next immutable promotion-invoice version for one review row."""
    payload = row.normalized_payload
    current_invoice = session.scalar(
        select(PromotionInvoice)
        .join(
            PromotionInvoiceAllocation,
            PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id,
        )
        .where(
            PromotionInvoice.store_id == payload["storeId"],
            PromotionInvoice.invoice_number == payload["invoiceNumber"],
            PromotionInvoice.is_current.is_(True),
            PromotionInvoiceAllocation.statement_month == batch.statement_month,
            PromotionInvoiceAllocation.is_current.is_(True),
        )
        .with_for_update()
    )
    if current_invoice is None:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "IMPORT_TARGET_NOT_FOUND",
            "推广费发票当前版本不存在",
            field="businessKey",
        )
    if current_invoice.invoice_amount_cent != payload["amountCent"]:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "IMPORT_AMOUNT_MISMATCH",
            "导入金额与当前推广费发票金额不一致",
            field="amountCent",
        )

    allocations = list(
        session.scalars(
            select(PromotionInvoiceAllocation).where(
                PromotionInvoiceAllocation.invoice_id == current_invoice.invoice_id,
                PromotionInvoiceAllocation.is_current.is_(True),
            )
        )
    )
    current_invoice.is_current = False
    for allocation in allocations:
        allocation.is_current = False
    session.flush()

    next_status = 3 if payload["status"] == "APPROVED_SETTLED" else 4
    next_invoice_id = f"promotion-invoice-{uuid4().hex}"
    next_invoice = PromotionInvoice(
        invoice_id=next_invoice_id,
        physical_invoice_id=current_invoice.physical_invoice_id,
        store_id=current_invoice.store_id,
        version_no=current_invoice.version_no + 1,
        version_kind=2,
        is_current=True,
        supersedes_invoice_id=current_invoice.invoice_id,
        replaces_invoice_id=current_invoice.replaces_invoice_id,
        invoice_number=current_invoice.invoice_number,
        invoice_date=current_invoice.invoice_date,
        invoice_amount_cent=current_invoice.invoice_amount_cent,
        buyer_name=current_invoice.buyer_name,
        tax_rate_percent=current_invoice.tax_rate_percent,
        invoice_status=next_status,
        registered_by=current_invoice.registered_by,
        registered_at=current_invoice.registered_at,
    )
    session.add(next_invoice)
    for allocation in allocations:
        session.add(
            PromotionInvoiceAllocation(
                allocation_id=f"promotion-allocation-{uuid4().hex}",
                invoice_id=next_invoice_id,
                store_id=allocation.store_id,
                statement_id=allocation.statement_id,
                statement_month=allocation.statement_month,
                settlement_batch_month=allocation.settlement_batch_month,
                allocated_amount_cent=allocation.allocated_amount_cent,
                is_current=True,
            )
        )
    session.add(
        InvoiceStatusEvent(
            event_id=f"invoice-event-{uuid4().hex}",
            invoice_id=next_invoice_id,
            event_type=2,
            from_status=current_invoice.invoice_status,
            to_status=next_status,
            operator_id=operator_id,
            import_batch_id=batch.batch_id,
            occurred_at=committed_at,
        )
    )
    row.row_status = 5
    row.target_record_id = next_invoice_id


def _commit_management_import_row(
    session,
    *,
    batch: FinanceImportBatch,
    row: FinanceImportRow,
    operator_id: str,
    committed_at: datetime,
    is_correction: bool,
    request: Request,
) -> None:
    """Create the next immutable management-invoice import version."""
    payload = row.normalized_payload
    statement = session.scalar(
        select(SettlementStatement).where(
            SettlementStatement.store_id == payload["storeId"],
            SettlementStatement.statement_month == batch.statement_month,
            SettlementStatement.is_current.is_(True),
        )
    )
    if statement is None:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "IMPORT_TARGET_NOT_FOUND",
            "管理服务费账期不存在当前账单",
            field="businessKey",
        )
    if batch.import_type == FINANCE_IMPORT_TYPE_TO_DB["FACTORY_DEDUCTION_RESULT"]:
        current_invoice = _find_current_management_invoice(
            session,
            store_id=payload["storeId"],
            statement_month=batch.statement_month,
            invoice_number=payload["invoiceNumber"],
        )
        if current_invoice is None:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "IMPORT_TARGET_NOT_FOUND",
                "厂家扣款未找到当前管理服务费发票",
                field="businessKey",
            )
        invoice_date_value = current_invoice.invoice_date
    else:
        current_invoice = session.scalar(
            select(InvoiceRecord).where(
                InvoiceRecord.store_id == payload["storeId"],
                InvoiceRecord.statement_month == batch.statement_month,
                InvoiceRecord.fee_direction == 2,
                InvoiceRecord.is_current.is_(True),
            )
        )
        invoice_date_value = date.fromisoformat(payload["invoiceDate"])

    next_version = 1
    if current_invoice is not None:
        next_version = current_invoice.version_no + 1
        current_invoice.is_current = False
        session.flush()
    next_invoice_id = f"management-invoice-{uuid4().hex}"
    session.add(
        InvoiceRecord(
            invoice_id=next_invoice_id,
            store_id=payload["storeId"],
            statement_month=batch.statement_month,
            statement_id=statement.statement_id,
            fee_direction=2,
            version_no=next_version,
            is_current=True,
            invoice_number=payload["invoiceNumber"],
            invoice_date=invoice_date_value,
            invoice_amount_cent=payload["amountCent"],
            invoice_status=3,
            source_type=3 if is_correction else 2,
            import_batch_id=batch.batch_id,
            registered_by=operator_id,
            registered_at=committed_at,
        )
    )
    row.row_status = 5
    row.target_record_id = next_invoice_id


def _finance_statement_conditions(
    *, month: str, metric_scope: str, store_id: str | None
) -> list:
    conditions = [SettlementStatement.is_current.is_(True)]
    if metric_scope == "CUMULATIVE":
        conditions.extend(
            [
                SettlementStatement.statement_month >= FORMAL_PERIOD_START_MONTH,
                SettlementStatement.statement_month <= month,
            ]
        )
    else:
        conditions.append(SettlementStatement.statement_month == month)
    if store_id is not None:
        conditions.append(SettlementStatement.store_id == store_id)
    return conditions


def _management_invoiceable_projection(
    session,
    *,
    store_id: str | None = None,
    through_month: str | None = None,
    excluded_invoice_ids: set[str] | None = None,
) -> tuple[list[dict], list[dict]]:
    """Project management invoiceable periods and signed carry-forward applications.

    Current confirmed management amounts are evaluated per store in business-month
    order. Current completed management invoices lock their own period out of the
    candidate pool; negative periods then offset the earliest still-uninvoiced
    positive balance, while any remainder is carried to future positive periods.
    """

    conditions = [
        SettlementStatement.is_current.is_(True),
        SettlementStatement.statement_month >= FORMAL_PERIOD_START_MONTH,
        SettlementStatementConfirmation.fee_direction == 2,
        SettlementStatementConfirmation.confirmation_status == 1,
    ]
    if store_id is not None:
        conditions.append(SettlementStatement.store_id == store_id)
    if through_month is not None:
        conditions.append(SettlementStatement.statement_month <= through_month)
    statement_rows = list(
        session.execute(
            select(SettlementStatement, SettlementStatementConfirmation)
            .join(
                SettlementStatementConfirmation,
                SettlementStatementConfirmation.statement_id
                == SettlementStatement.statement_id,
            )
            .where(*conditions)
            .order_by(
                SettlementStatement.store_id,
                SettlementStatement.statement_month,
                SettlementStatement.statement_id,
            )
        ).all()
    )
    if not statement_rows:
        return [], []

    statement_ids = [statement.statement_id for statement, _ in statement_rows]
    invoice_conditions = [
        InvoiceRecord.statement_id.in_(statement_ids),
        InvoiceRecord.fee_direction == 2,
        InvoiceRecord.is_current.is_(True),
        InvoiceRecord.is_tombstone.is_(False),
        InvoiceRecord.invoice_status == 3,
        InvoiceRecord.invoice_amount_cent > 0,
    ]
    if excluded_invoice_ids:
        invoice_conditions.append(
            InvoiceRecord.invoice_id.not_in(excluded_invoice_ids)
        )
    current_invoices = list(
        session.scalars(select(InvoiceRecord).where(*invoice_conditions))
    )
    invoices_by_statement = {
        invoice.statement_id: invoice for invoice in current_invoices
    }

    current_invoice_ids = {invoice.invoice_id for invoice in current_invoices}
    locked_application_rows = (
        list(
            session.scalars(
                select(ManagementCarryforwardApplication).where(
                    ManagementCarryforwardApplication.is_current.is_(True),
                    ManagementCarryforwardApplication.invoice_id.in_(
                        current_invoice_ids
                    ),
                )
            )
        )
        if current_invoice_ids
        else []
    )
    reserved_negative_by_statement: dict[str, int] = {}
    for application in locked_application_rows:
        reserved_negative_by_statement[application.source_statement_id] = (
            reserved_negative_by_statement.get(application.source_statement_id, 0)
            + application.applied_amount_cent
        )

    periods: list[dict] = []
    period_by_statement: dict[str, dict] = {}
    applications: list[dict] = [
        {
            "store_id": application.store_id,
            "source_statement_id": application.source_statement_id,
            "source_statement_month": application.source_statement_month,
            "target_statement_id": application.target_statement_id,
            "target_statement_month": application.target_statement_month,
            "invoice_id": application.invoice_id,
            "applied_amount_cent": application.applied_amount_cent,
        }
        for application in locked_application_rows
    ]
    current_store_id: str | None = None
    pending_positives: list[dict] = []
    pending_negatives: list[dict] = []

    for statement, confirmation in statement_rows:
        if statement.store_id != current_store_id:
            current_store_id = statement.store_id
            pending_positives = []
            pending_negatives = []

        confirmed_amount = int(confirmation.confirmed_amount_cent or 0)
        current_invoice = invoices_by_statement.get(statement.statement_id)
        period = {
            "store_id": statement.store_id,
            "statement_id": statement.statement_id,
            "statement_month": statement.statement_month,
            "confirmed_amount_cent": confirmed_amount,
            "invoice_id": current_invoice.invoice_id if current_invoice else None,
            "invoice_version": current_invoice.version_no if current_invoice else None,
            "invoiceable_amount_cent": 0,
            "carryforward_balance_cent": 0,
        }
        periods.append(period)
        period_by_statement[statement.statement_id] = period

        if confirmed_amount > 0:
            if current_invoice is not None:
                continue
            remaining_positive = confirmed_amount
            while remaining_positive > 0 and pending_negatives:
                negative = pending_negatives[0]
                applied_amount = min(remaining_positive, negative["remaining_cent"])
                applications.append(
                    {
                        "store_id": statement.store_id,
                        "source_statement_id": negative["statement_id"],
                        "source_statement_month": negative["statement_month"],
                        "target_statement_id": statement.statement_id,
                        "target_statement_month": statement.statement_month,
                        "invoice_id": None,
                        "applied_amount_cent": applied_amount,
                    }
                )
                remaining_positive -= applied_amount
                negative["remaining_cent"] -= applied_amount
                if negative["remaining_cent"] == 0:
                    pending_negatives.pop(0)
            period["invoiceable_amount_cent"] = remaining_positive
            if remaining_positive > 0:
                pending_positives.append(
                    {
                        "statement_id": statement.statement_id,
                        "statement_month": statement.statement_month,
                        "remaining_cent": remaining_positive,
                    }
                )
            continue

        if confirmed_amount < 0:
            remaining_negative = max(
                -confirmed_amount
                - reserved_negative_by_statement.get(statement.statement_id, 0),
                0,
            )
            while remaining_negative > 0 and pending_positives:
                positive = pending_positives[0]
                applied_amount = min(remaining_negative, positive["remaining_cent"])
                applications.append(
                    {
                        "store_id": statement.store_id,
                        "source_statement_id": statement.statement_id,
                        "source_statement_month": statement.statement_month,
                        "target_statement_id": positive["statement_id"],
                        "target_statement_month": positive["statement_month"],
                        "invoice_id": None,
                        "applied_amount_cent": applied_amount,
                    }
                )
                remaining_negative -= applied_amount
                positive["remaining_cent"] -= applied_amount
                period_by_statement[positive["statement_id"]][
                    "invoiceable_amount_cent"
                ] = positive["remaining_cent"]
                if positive["remaining_cent"] == 0:
                    pending_positives.pop(0)
            if remaining_negative > 0:
                pending_negatives.append(
                    {
                        "statement_id": statement.statement_id,
                        "statement_month": statement.statement_month,
                        "remaining_cent": remaining_negative,
                    }
                )
            period["carryforward_balance_cent"] = -remaining_negative

    return periods, applications


def _synchronize_management_carryforward_applications(
    session,
    *,
    applications: list[dict],
    invoice_id_by_statement: dict[str, str],
    scope_store_ids: set[str] | None = None,
    scope_through_month: str | None = None,
) -> list[ManagementCarryforwardApplication]:
    """Append immutable carry-forward application versions for one projection."""
    desired_by_pair: dict[tuple[str, str], dict] = {}
    for application in applications:
        normalized = {
            **application,
            "invoice_id": invoice_id_by_statement.get(
                application["target_statement_id"], application.get("invoice_id")
            ),
        }
        normalized["projection_sha256"] = _canonical_billing_sha256(
            {
                "storeId": normalized["store_id"],
                "sourceStatementId": normalized["source_statement_id"],
                "targetStatementId": normalized["target_statement_id"],
                "invoiceId": normalized.get("invoice_id"),
                "appliedAmountCent": normalized["applied_amount_cent"],
            }
        )
        desired_by_pair[
            (normalized["source_statement_id"], normalized["target_statement_id"])
        ] = normalized
    effective_store_ids = scope_store_ids or {
        application["store_id"] for application in applications
    }
    if not effective_store_ids:
        return []
    scope_conditions = [
        ManagementCarryforwardApplication.is_current.is_(True),
        ManagementCarryforwardApplication.store_id.in_(effective_store_ids),
    ]
    if scope_through_month is not None:
        scope_conditions.extend(
            [
                ManagementCarryforwardApplication.source_statement_month
                <= scope_through_month,
                ManagementCarryforwardApplication.target_statement_month
                <= scope_through_month,
            ]
        )
    current_rows = list(
        session.scalars(
            select(ManagementCarryforwardApplication).where(*scope_conditions)
        )
    )
    current_by_pair = {
        (row.source_statement_id, row.target_statement_id): row
        for row in current_rows
    }
    materialized: list[ManagementCarryforwardApplication] = []
    for pair, desired in desired_by_pair.items():
        current = current_by_pair.get(pair)
        if (
            current is not None
            and current.projection_sha256 == desired["projection_sha256"]
        ):
            materialized.append(current)
            continue
        if current is not None:
            current.is_current = False
            session.flush()
        next_row = ManagementCarryforwardApplication(
            application_id=f"management-carryforward-{uuid4().hex}",
            store_id=desired["store_id"],
            source_statement_id=desired["source_statement_id"],
            source_statement_month=desired["source_statement_month"],
            target_statement_id=desired["target_statement_id"],
            target_statement_month=desired["target_statement_month"],
            invoice_id=desired.get("invoice_id"),
            applied_amount_cent=desired["applied_amount_cent"],
            version_no=current.version_no + 1 if current is not None else 1,
            is_current=True,
            supersedes_application_id=(
                current.application_id if current is not None else None
            ),
            projection_sha256=desired["projection_sha256"],
            created_at=utcnow(),
        )
        session.add(next_row)
        session.flush()
        materialized.append(next_row)
    for pair, current in current_by_pair.items():
        if pair not in desired_by_pair:
            current.is_current = False
    return materialized


def _finance_summary_metrics(
    session,
    *,
    month: str,
    fee_direction: str,
    metric_scope: str,
    store_id: str | None,
) -> dict:
    statement_conditions = _finance_statement_conditions(
        month=month, metric_scope=metric_scope, store_id=store_id
    )
    direction_code = CONFIRMATION_DIRECTION_TO_DB[fee_direction]
    fee_column = (
        SettlementStatement.promotion_net_fee_cent
        if fee_direction == "PROMOTION"
        else SettlementStatement.management_net_fee_cent
    )
    statement_total = session.scalar(
        select(func.coalesce(func.sum(fee_column), 0)).where(*statement_conditions)
    )
    monthly_confirmation_total = _finance_confirmation_total(
        session,
        statement_conditions=_finance_statement_conditions(
            month=month, metric_scope="MONTH", store_id=store_id
        ),
        direction_code=direction_code,
    )
    pending_confirmation_total = (
        monthly_confirmation_total
        if metric_scope == "MONTH"
        else _finance_confirmation_total(
            session,
            statement_conditions=statement_conditions,
            direction_code=direction_code,
        )
    )
    if fee_direction == "PROMOTION":
        _, promotion_groups = _promotion_invoice_carryforward_projection(
            session,
            store_id=store_id,
        )
        if metric_scope == "MONTH":
            pending_invoice_amount = sum(
                group["invoiceable_amount_cent"]
                for group in promotion_groups
                if group["closing_month"] == month
            )
        else:
            pending_invoice_amount = sum(
                group["invoiceable_amount_cent"]
                for group in promotion_groups
                if group["closing_month"] is not None
                and group["closing_month"] <= month
            )
        allocation_conditions = [
            PromotionInvoiceAllocation.is_current.is_(True),
            PromotionInvoice.is_current.is_(True),
            PromotionInvoice.is_tombstone.is_(False),
        ]
        if metric_scope == "CUMULATIVE":
            allocation_conditions.extend(
                [
                    PromotionInvoiceAllocation.statement_month
                    >= FORMAL_PERIOD_START_MONTH,
                    PromotionInvoiceAllocation.statement_month <= month,
                ]
            )
        else:
            allocation_conditions.append(
                PromotionInvoiceAllocation.statement_month == month
            )
        if store_id is not None:
            allocation_conditions.append(
                PromotionInvoiceAllocation.store_id == store_id
            )
        issued_total = session.scalar(
            select(func.coalesce(func.sum(PromotionInvoiceAllocation.allocated_amount_cent), 0))
            .join(
                PromotionInvoice,
                PromotionInvoice.invoice_id == PromotionInvoiceAllocation.invoice_id,
            )
            .where(
                *allocation_conditions,
                PromotionInvoice.invoice_status.in_((2, 3)),
            )
        )
        settled_total = session.scalar(
            select(func.coalesce(func.sum(PromotionInvoiceAllocation.allocated_amount_cent), 0))
            .join(
                PromotionInvoice,
                PromotionInvoice.invoice_id == PromotionInvoiceAllocation.invoice_id,
            )
            .where(
                *allocation_conditions,
                PromotionInvoice.invoice_status == 3,
            )
        )
    else:
        management_periods, _ = _management_invoiceable_projection(
            session,
            store_id=store_id,
            through_month=month,
        )
        if metric_scope == "MONTH":
            pending_invoice_amount = sum(
                period["invoiceable_amount_cent"]
                for period in management_periods
                if period["statement_month"] == month
            )
        else:
            pending_invoice_amount = sum(
                period["invoiceable_amount_cent"]
                for period in management_periods
                if period["statement_month"] <= month
            )
        issued_total = session.scalar(
            select(func.coalesce(func.sum(InvoiceRecord.invoice_amount_cent), 0))
            .join(
                SettlementStatement,
                SettlementStatement.statement_id == InvoiceRecord.statement_id,
            )
            .where(
                *statement_conditions,
                InvoiceRecord.is_current.is_(True),
                InvoiceRecord.is_tombstone.is_(False),
                InvoiceRecord.fee_direction == direction_code,
                InvoiceRecord.invoice_status == 3,
            )
        )
        settled_total = session.scalar(
            select(func.coalesce(func.sum(InvoiceRecord.invoice_amount_cent), 0))
            .join(
                SettlementStatement,
                SettlementStatement.statement_id == InvoiceRecord.statement_id,
            )
            .where(
                *statement_conditions,
                InvoiceRecord.is_current.is_(True),
                InvoiceRecord.is_tombstone.is_(False),
                InvoiceRecord.fee_direction == direction_code,
                InvoiceRecord.invoice_status == 3,
            )
        )

    confirmed_amount = int(monthly_confirmation_total or 0)
    pending_confirmation_amount = int(pending_confirmation_total or 0)
    issued_amount = int(issued_total or 0)
    return {
        "statement_total_cent": int(statement_total or 0),
        "confirmed_amount_cent": confirmed_amount,
        "pending_invoice_amount_cent": (
            pending_invoice_amount
            if fee_direction == "PROMOTION"
            else max(pending_invoice_amount, 0)
        ),
        "issued_amount_cent": issued_amount,
        "settled_or_deducted_amount_cent": int(settled_total or 0),
    }


def _finance_confirmation_total(session, *, statement_conditions: list, direction_code: int) -> int:
    return int(
        session.scalar(
            select(
                func.coalesce(
                    func.sum(SettlementStatementConfirmation.confirmed_amount_cent), 0
                )
            )
            .join(
                SettlementStatement,
                SettlementStatement.statement_id
                == SettlementStatementConfirmation.statement_id,
            )
            .where(
                *statement_conditions,
                SettlementStatementConfirmation.fee_direction == direction_code,
                SettlementStatementConfirmation.confirmation_status == 1,
            )
        )
        or 0
    )


def _statement_header_item(
    session,
    statement: SettlementStatement,
    *,
    promotion_projection: dict[str, dict] | None = None,
) -> dict:
    store = session.get(DimStore, statement.store_id)
    confirmations = {
        row.fee_direction: row
        for row in session.scalars(
            select(SettlementStatementConfirmation).where(
                SettlementStatementConfirmation.statement_id == statement.statement_id
            )
        )
    }
    invoices = {
        row.fee_direction: row
        for row in session.scalars(
            select(InvoiceRecord).where(
                InvoiceRecord.statement_id == statement.statement_id,
                InvoiceRecord.is_current.is_(True),
                InvoiceRecord.is_tombstone.is_(False),
            )
        )
    }
    promotion_invoice = session.scalar(
        select(PromotionInvoice)
        .join(
            PromotionInvoiceAllocation,
            PromotionInvoiceAllocation.invoice_id == PromotionInvoice.invoice_id,
        )
        .where(
            PromotionInvoiceAllocation.store_id == statement.store_id,
            PromotionInvoiceAllocation.statement_month
            == statement.statement_month,
            PromotionInvoiceAllocation.is_current.is_(True),
            PromotionInvoice.is_current.is_(True),
            PromotionInvoice.is_tombstone.is_(False),
        )
    )
    if promotion_projection is None:
        promotion_projection, _ = _promotion_invoice_carryforward_projection(
            session,
            store_id=statement.store_id,
        )
    projected_group = promotion_projection.get(statement.statement_id)
    return {
        "statement_id": statement.statement_id,
        "store_id": statement.store_id,
        "store_name": store.store_name if store is not None else None,
        "month": statement.statement_month,
        "version_no": statement.version_no,
        "is_current": statement.is_current,
        "supersedes_statement_id": statement.supersedes_statement_id,
        "status": STATEMENT_STATUS_NAMES[statement.statement_status],
        "promotion_amount_cent": statement.promotion_net_fee_cent,
        "management_amount_cent": statement.management_net_fee_cent,
        "promotion_confirmation": _confirmation_summary(confirmations.get(1)),
        "management_confirmation": _confirmation_summary(confirmations.get(2)),
        "promotion_invoice_status": (
            INVOICE_STATUS_NAMES[promotion_invoice.invoice_status]
            if promotion_invoice is not None
            else "PENDING_INVOICE"
        ),
        "promotion_invoiceable_amount_cent": (
            projected_group["invoiceable_amount_cent"]
            if projected_group is not None
            else 0
        ),
        "promotion_carryforward_balance_cent": (
            projected_group["carryforward_balance_cent"]
            if projected_group is not None
            else 0
        ),
        "promotion_invoice_group_id": (
            projected_group["group_id"]
            if projected_group is not None
            else None
        ),
        "promotion_required_statement_ids": (
            projected_group["required_statement_ids"]
            if projected_group is not None
            else []
        ),
        "promotion_positive_amount_cent": (
            projected_group["positive_amount_cent"]
            if projected_group is not None
            else 0
        ),
        "promotion_negative_amount_cent": (
            projected_group["negative_amount_cent"]
            if projected_group is not None
            else 0
        ),
        "management_invoice_status": _invoice_status(invoices.get(2)),
    }


def _confirmation_summary(
    confirmation: SettlementStatementConfirmation | None,
) -> dict | None:
    if confirmation is None:
        return None
    return {
        "confirmation_id": confirmation.confirmation_id,
        "status": CONFIRMATION_STATUS_NAMES[confirmation.confirmation_status],
        "confirmed_amount_cent": confirmation.confirmed_amount_cent,
        "confirmed_at": confirmation.confirmed_at,
    }


def _invoice_status(invoice: InvoiceRecord | None) -> str:
    return INVOICE_STATUS_NAMES[invoice.invoice_status] if invoice else "PENDING_INVOICE"


def _statement_version_item(statement: SettlementStatement) -> dict:
    return {
        "statement_id": statement.statement_id,
        "version_no": statement.version_no,
        "is_current": statement.is_current,
        "supersedes_statement_id": statement.supersedes_statement_id,
        "status": STATEMENT_STATUS_NAMES[statement.statement_status],
        "created_at": statement.created_at,
    }


def _statement_line_item(line: SettlementStatementLine) -> dict:
    return {
        "statement_line_id": line.statement_line_id,
        "fee_direction": CONFIRMATION_DIRECTION_FROM_DB[line.fee_direction],
        "product_scope": line.product_scope,
        "product_type": line.product_type,
        "original_entry_count": line.original_entry_count,
        "adjustment_entry_count": line.adjustment_entry_count,
        "original_base_cent": line.original_base_cent,
        "adjustment_base_cent": line.adjustment_base_cent,
        "net_base_cent": line.net_base_cent,
        "original_fee_cent": line.original_fee_cent,
        "adjustment_fee_cent": line.adjustment_fee_cent,
        "net_fee_cent": line.net_fee_cent,
    }


def _parse_confirmation_payload(payload: dict, request: Request) -> dict:
    fee_direction = payload.get("feeDirection")
    confirmed_amount = payload.get("confirmedAmountCent")
    read_version = payload.get("readVersion")
    if not isinstance(fee_direction, str):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "feeDirection 为必填项",
            field="feeDirection",
        )
    direction = _normalize_billing_direction(fee_direction, request)
    if isinstance(confirmed_amount, bool) or not isinstance(confirmed_amount, int):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "confirmedAmountCent 必须为整数",
            field="confirmedAmountCent",
        )
    if isinstance(read_version, bool) or not isinstance(read_version, int) or read_version < 1:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "readVersion 必须为正整数",
            field="readVersion",
        )
    return {
        "fee_direction": direction,
        "confirmed_amount_cent": confirmed_amount,
        "read_version": read_version,
    }


def _billing_idempotency_key_hash(value: str | None, request: Request) -> str:
    normalized = (value or "").strip()
    if not 16 <= len(normalized) <= 128:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "Idempotency-Key 长度必须为 16～128",
            field="Idempotency-Key",
        )
    return sha256(normalized.encode("utf-8")).hexdigest()


def _canonical_billing_sha256(value: dict) -> str:
    return sha256(
        json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode(
            "utf-8"
        )
    ).hexdigest()


def _confirmation_item(
    confirmation: SettlementStatementConfirmation, statement: SettlementStatement
) -> dict:
    return {
        "confirmation_id": confirmation.confirmation_id,
        "status": CONFIRMATION_STATUS_NAMES[confirmation.confirmation_status],
        "confirmed_amount_cent": confirmation.confirmed_amount_cent,
        "confirmed_at": confirmation.confirmed_at,
        "statement_id": statement.statement_id,
        "version_no": statement.version_no,
        "is_current": statement.is_current,
    }


def _parse_dispute_payload(payload: dict, request: Request) -> dict:
    fee_direction = payload.get("feeDirection")
    dispute_type = payload.get("disputeType")
    description = payload.get("description")
    contact_name = payload.get("contactName")
    contact_phone = payload.get("contactPhone")
    disputed_amount = payload.get("disputedAmountCent")
    orders = payload.get("orders")
    evidence = payload.get("evidence")
    read_version = payload.get("readVersion")

    if not isinstance(fee_direction, str):
        _raise_dispute_validation(request, "feeDirection", "feeDirection 为必填项")
    normalized_direction = _normalize_billing_direction(fee_direction, request)
    if not isinstance(dispute_type, str):
        _raise_dispute_validation(request, "disputeType", "disputeType 为必填项")
    normalized_type = dispute_type.upper()
    _validate_enum(normalized_type, set(DISPUTE_TYPE_TO_DB), "disputeType", request)
    if not isinstance(description, str) or not description.strip():
        _raise_dispute_validation(request, "description", "description 为必填项")
    if not isinstance(contact_name, str) or not contact_name.strip():
        _raise_dispute_validation(request, "contactName", "contactName 为必填项")
    if (
        not isinstance(contact_phone, str)
        or not contact_phone.isdigit()
        or len(contact_phone) != 11
    ):
        _raise_dispute_validation(request, "contactPhone", "contactPhone 必须为 11 位手机号")
    if (
        isinstance(disputed_amount, bool)
        or not isinstance(disputed_amount, int)
        or disputed_amount <= 0
    ):
        _raise_dispute_validation(
            request, "disputedAmountCent", "disputedAmountCent 必须为正整数"
        )
    if (
        isinstance(read_version, bool)
        or not isinstance(read_version, int)
        or read_version < 1
    ):
        _raise_dispute_validation(request, "readVersion", "readVersion 必须为正整数")
    if not isinstance(orders, list) or not orders:
        _raise_dispute_validation(request, "orders", "orders 至少包含一条争议订单")
    if not isinstance(evidence, list) or not evidence:
        _raise_dispute_validation(request, "evidence", "evidence 至少包含一个受控对象键")

    parsed_orders = []
    seen_order_keys = set()
    for order in orders:
        if not isinstance(order, dict):
            _raise_dispute_validation(request, "orders", "orders 元素必须为对象")
        order_id = order.get("orderId")
        coupon_id = order.get("couponId")
        order_amount = order.get("disputedAmountCent")
        if not isinstance(order_id, str) or not order_id.strip():
            _raise_dispute_validation(request, "orders", "orderId 为必填项")
        if coupon_id is not None and (
            not isinstance(coupon_id, str) or not coupon_id.strip()
        ):
            _raise_dispute_validation(request, "orders", "couponId 必须为非空字符串")
        if (
            isinstance(order_amount, bool)
            or not isinstance(order_amount, int)
            or order_amount <= 0
        ):
            _raise_dispute_validation(
                request, "orders", "订单 disputedAmountCent 必须为正整数"
            )
        order_key = (order_id.strip(), coupon_id.strip() if coupon_id else None)
        if order_key in seen_order_keys:
            _raise_dispute_validation(request, "orders", "争议订单和券不得重复")
        seen_order_keys.add(order_key)
        parsed_orders.append(
            {
                "order_id": order_key[0],
                "coupon_id": order_key[1],
                "disputed_amount_cent": order_amount,
            }
        )
    if sum(order["disputed_amount_cent"] for order in parsed_orders) != disputed_amount:
        _raise_dispute_validation(
            request, "orders", "订单争议金额合计必须等于异议总金额"
        )

    parsed_evidence = []
    for item in evidence:
        object_key = item.get("objectKey") if isinstance(item, dict) else None
        if (
            not isinstance(object_key, str)
            or not object_key.strip()
            or object_key.startswith(("http://", "https://"))
            or ".." in object_key
        ):
            _raise_dispute_validation(
                request, "evidence", "evidence 仅接受受控 objectKey"
            )
        parsed_evidence.append({"objectKey": object_key.strip()})

    return {
        "fee_direction": normalized_direction,
        "dispute_type": normalized_type,
        "description": description.strip(),
        "contact_name": contact_name.strip(),
        "contact_phone": contact_phone,
        "disputed_amount_cent": disputed_amount,
        "orders": parsed_orders,
        "evidence": parsed_evidence,
        "read_version": read_version,
    }


def _raise_dispute_validation(request: Request, field: str, message: str) -> None:
    _raise_reporting_error(
        request,
        status.HTTP_422_UNPROCESSABLE_ENTITY,
        "VALIDATION_FAILED",
        message,
        field=field,
    )


def _dispute_fernet(request: Request) -> Fernet:
    secret = (
        os.getenv("DY_FINANCE_PII_SECRET", "").strip()
        or os.getenv("DY_SESSION_SECRET", "").strip()
    )
    if not secret and os.getenv("DY_API_TEST_MODE", "").strip().lower() in {
        "1",
        "true",
        "yes",
        "on",
    }:
        secret = "dydata-test-finance-pii-secret"
    if not secret:
        _raise_reporting_error(
            request,
            status.HTTP_503_SERVICE_UNAVAILABLE,
            "PII_ENCRYPTION_UNAVAILABLE",
            "敏感信息加密密钥未配置",
        )
    key = urlsafe_b64encode(sha256(secret.encode("utf-8")).digest())
    return Fernet(key)


def _encrypt_dispute_phone(phone: str, request: Request) -> str:
    return _dispute_fernet(request).encrypt(phone.encode("utf-8")).decode("ascii")


def _decrypt_dispute_phone(ciphertext: str, request: Request) -> str:
    try:
        return _dispute_fernet(request).decrypt(ciphertext.encode("ascii")).decode("utf-8")
    except (InvalidToken, UnicodeDecodeError, ValueError):
        return ""


def _mask_dispute_phone(phone: str) -> str:
    if len(phone) != 11 or not phone.isdigit():
        return ""
    return f"{phone[:3]}****{phone[-4:]}"


def _dispute_item(session, dispute: SettlementDispute, request: Request) -> dict:
    orders = list(
        session.scalars(
            select(SettlementDisputeOrder)
            .where(SettlementDisputeOrder.dispute_id == dispute.dispute_id)
            .order_by(SettlementDisputeOrder.order_id, SettlementDisputeOrder.coupon_id)
        )
    )
    return {
        "dispute_id": dispute.dispute_id,
        "statement_id": dispute.statement_id,
        "store_id": dispute.store_id,
        "statement_month": dispute.statement_month,
        "fee_direction": CONFIRMATION_DIRECTION_FROM_DB[dispute.fee_direction],
        "dispute_type": DISPUTE_TYPE_FROM_DB[dispute.dispute_type],
        "status": DISPUTE_STATUS_NAMES[dispute.status],
        "disputed_amount_cent": dispute.disputed_amount_cent,
        "description": dispute.description,
        "contact_name": dispute.contact_name,
        "contact_phone_masked": _mask_dispute_phone(
            _decrypt_dispute_phone(dispute.contact_phone_ciphertext, request)
        ),
        "evidence": dispute.evidence_json,
        "orders": [
            {
                "order_id": order.order_id,
                "coupon_id": order.coupon_id,
                "disputed_amount_cent": order.disputed_amount_cent,
            }
            for order in orders
        ],
        "submitted_at": dispute.submitted_at,
        "resolution_note": dispute.resolution_note,
        "result_statement_id": dispute.result_statement_id,
    }


def _validate_billing_month(month: str, request: Request) -> None:
    try:
        datetime.strptime(month, "%Y-%m")
    except (TypeError, ValueError):
        _raise_dispute_validation(request, "month", "month 必须使用 YYYY-MM 格式")


def _create_dispute_statement_version(
    *,
    session,
    dispute: SettlementDispute,
    current_statement: SettlementStatement,
    adjustment_amount_cent: int,
    resolution_note: str,
    operator: str,
    request: Request,
) -> SettlementStatement:
    current_lines = list(
        session.scalars(
            select(SettlementStatementLine).where(
                SettlementStatementLine.statement_id == current_statement.statement_id
            )
        )
    )
    current_entries = list(
        session.scalars(
            select(SettlementStatementEntry).where(
                SettlementStatementEntry.statement_id == current_statement.statement_id
            )
        )
    )
    if not current_entries:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "STATEMENT_SNAPSHOT_INCOMPLETE",
            "当前账单缺少完整来源项，不能生成调整版本",
        )

    known_line_ids = {line.statement_line_id for line in current_lines}
    missing_line_ids = {
        entry.statement_line_id
        for entry in current_entries
        if entry.statement_line_id not in known_line_ids
    }
    for missing_line_id in missing_line_ids:
        grouped_entries = [
            entry for entry in current_entries if entry.statement_line_id == missing_line_id
        ]
        first_entry = grouped_entries[0]
        original_entries = [entry for entry in grouped_entries if entry.source_type == 1]
        adjustment_entries = [entry for entry in grouped_entries if entry.source_type == 2]
        current_lines.append(
            SettlementStatementLine(
                statement_line_id=missing_line_id,
                statement_id=current_statement.statement_id,
                fee_direction=first_entry.fee_direction,
                product_scope=first_entry.product_scope,
                product_type=first_entry.product_type,
                original_entry_count=len(original_entries),
                adjustment_entry_count=len(adjustment_entries),
                original_base_cent=sum(entry.base_amount_cent for entry in original_entries),
                adjustment_base_cent=sum(
                    entry.base_amount_cent for entry in adjustment_entries
                ),
                net_base_cent=sum(entry.base_amount_cent for entry in grouped_entries),
                original_fee_cent=sum(entry.fee_amount_cent for entry in original_entries),
                adjustment_fee_cent=sum(
                    entry.fee_amount_cent for entry in adjustment_entries
                ),
                net_fee_cent=sum(entry.fee_amount_cent for entry in grouped_entries),
            )
        )

    entries_by_order = {
        (entry.order_id, entry.coupon_id): entry for entry in current_entries
    }
    dispute_orders = list(
        session.scalars(
            select(SettlementDisputeOrder)
            .where(SettlementDisputeOrder.dispute_id == dispute.dispute_id)
            .order_by(SettlementDisputeOrder.order_id, SettlementDisputeOrder.coupon_id)
        )
    )
    selected_entries = []
    for order in dispute_orders:
        selected = entries_by_order.get((order.order_id, order.coupon_id))
        if selected is None or selected.fee_direction != dispute.fee_direction:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "DISPUTE_ORDER_SCOPE_INVALID",
                "异议订单无法精确映射到当前账单费用方向",
                field="orders",
            )
        selected_entries.append((order, selected))

    new_statement_id = f"statement-{uuid4().hex}"
    operated_at = utcnow()
    should_auto_confirm = operated_at.day > 6
    current_statement.is_current = False
    session.flush()
    new_statement = SettlementStatement(
        statement_id=new_statement_id,
        store_id=current_statement.store_id,
        statement_month=current_statement.statement_month,
        version_no=current_statement.version_no + 1,
        is_current=True,
        supersedes_statement_id=current_statement.statement_id,
        statement_status=3 if should_auto_confirm else 2,
        promotion_original_fee_cent=current_statement.promotion_original_fee_cent,
        promotion_adjustment_fee_cent=(
            current_statement.promotion_adjustment_fee_cent
            + (adjustment_amount_cent if dispute.fee_direction == 1 else 0)
        ),
        promotion_net_fee_cent=(
            current_statement.promotion_net_fee_cent
            + (adjustment_amount_cent if dispute.fee_direction == 1 else 0)
        ),
        management_original_fee_cent=current_statement.management_original_fee_cent,
        management_adjustment_fee_cent=(
            current_statement.management_adjustment_fee_cent
            + (adjustment_amount_cent if dispute.fee_direction == 2 else 0)
        ),
        management_net_fee_cent=(
            current_statement.management_net_fee_cent
            + (adjustment_amount_cent if dispute.fee_direction == 2 else 0)
        ),
        confirmed_by="system:auto-confirmation" if should_auto_confirm else None,
        confirmed_at=operated_at if should_auto_confirm else None,
        locked_by=None,
        locked_at=None,
        lock_version=uuid4().hex,
        store_name_snapshot=current_statement.store_name_snapshot,
        sap_code_snapshot=current_statement.sap_code_snapshot,
        store_snapshot_status=current_statement.store_snapshot_status,
        store_snapshot_profile_id=current_statement.store_snapshot_profile_id,
    )
    session.add(new_statement)
    session.flush()

    carryforward_adjustment_id_map: dict[str, str] = {}
    current_carryforward_applications = list(
        session.scalars(
            select(SettlementCarryforwardApplication)
            .where(
                SettlementCarryforwardApplication.target_statement_id
                == current_statement.statement_id,
                SettlementCarryforwardApplication.is_current.is_(True),
            )
            .order_by(SettlementCarryforwardApplication.carryforward_source_id)
            .with_for_update()
        )
    )
    copied_adjustments: list[
        tuple[SettlementCarryforwardApplication, SettlementFeeAdjustment, str]
    ] = []
    for application in current_carryforward_applications:
        original_adjustment = session.scalar(
            select(SettlementFeeAdjustment).where(
                SettlementFeeAdjustment.adjustment_id
                == application.target_adjustment_id
            )
        )
        if original_adjustment is None:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "CARRYFORWARD_LINEAGE_INCOMPLETE",
                "当前账单的顺延应用缺少对应调整记录，不能生成新版本",
            )
        copied_adjustment_id = f"carryforward-adjustment-{uuid4().hex}"
        application.is_current = False
        carryforward_adjustment_id_map[
            original_adjustment.adjustment_id
        ] = copied_adjustment_id
        copied_adjustments.append(
            (application, original_adjustment, copied_adjustment_id)
        )
    if copied_adjustments:
        session.flush()
    for application, original_adjustment, copied_adjustment_id in copied_adjustments:
        session.add(
            SettlementFeeAdjustment(
                adjustment_id=copied_adjustment_id,
                original_fee_result_id=original_adjustment.original_fee_result_id,
                refund_event_id=original_adjustment.refund_event_id,
                coupon_id=original_adjustment.coupon_id,
                order_id=original_adjustment.order_id,
                fee_direction=original_adjustment.fee_direction,
                original_business_month=original_adjustment.original_business_month,
                adjustment_posting_month=original_adjustment.adjustment_posting_month,
                adjustment_type=original_adjustment.adjustment_type,
                adjustment_base_cent=original_adjustment.adjustment_base_cent,
                adjustment_fee_cent=original_adjustment.adjustment_fee_cent,
                rule_version=original_adjustment.rule_version,
                adjustment_reason=original_adjustment.adjustment_reason,
                occurred_at=original_adjustment.occurred_at,
                created_by=f"dispute-version:{operator}",
            )
        )
        session.add(
            SettlementCarryforwardApplication(
                carryforward_application_id=f"carryforward-application-{uuid4().hex}",
                carryforward_source_id=application.carryforward_source_id,
                target_statement_id=new_statement.statement_id,
                target_statement_version=new_statement.version_no,
                target_adjustment_id=copied_adjustment_id,
                target_posting_month=application.target_posting_month,
                application_version=application.application_version + 1,
                is_current=True,
                applied_by=f"dispute-version:{operator}",
                applied_at=operated_at,
            )
        )
    if copied_adjustments:
        session.flush()

    line_id_map: dict[str, str] = {}
    cloned_lines: dict[str, SettlementStatementLine] = {}
    for line in current_lines:
        new_line_id = f"statement-line-{uuid4().hex}"
        line_id_map[line.statement_line_id] = new_line_id
        cloned_line = SettlementStatementLine(
            statement_line_id=new_line_id,
            statement_id=new_statement_id,
            fee_direction=line.fee_direction,
            product_scope=line.product_scope,
            product_type=line.product_type,
            original_entry_count=line.original_entry_count,
            adjustment_entry_count=line.adjustment_entry_count,
            original_base_cent=line.original_base_cent,
            adjustment_base_cent=line.adjustment_base_cent,
            net_base_cent=line.net_base_cent,
            original_fee_cent=line.original_fee_cent,
            adjustment_fee_cent=line.adjustment_fee_cent,
            net_fee_cent=line.net_fee_cent,
        )
        cloned_lines[line.statement_line_id] = cloned_line
        session.add(cloned_line)

    for entry in current_entries:
        new_line_id = line_id_map.get(entry.statement_line_id)
        if new_line_id is None:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "STATEMENT_SNAPSHOT_INCOMPLETE",
                "账单来源项无法映射到汇总行",
            )
        session.add(
            SettlementStatementEntry(
                statement_entry_id=f"statement-entry-{uuid4().hex}",
                statement_id=new_statement_id,
                statement_line_id=new_line_id,
                source_type=entry.source_type,
                source_record_id=carryforward_adjustment_id_map.get(
                    entry.source_record_id,
                    entry.source_record_id,
                ),
                original_fee_result_id=entry.original_fee_result_id,
                coupon_id=entry.coupon_id,
                order_id=entry.order_id,
                fee_direction=entry.fee_direction,
                original_business_month=entry.original_business_month,
                statement_posting_month=entry.statement_posting_month,
                product_scope=entry.product_scope,
                product_type=entry.product_type,
                base_amount_cent=entry.base_amount_cent,
                fee_amount_cent=entry.fee_amount_cent,
                rule_version=entry.rule_version,
                order_status_snapshot=entry.order_status_snapshot,
                coupon_status_snapshot=entry.coupon_status_snapshot,
                product_name_snapshot=entry.product_name_snapshot,
                sku_id_snapshot=entry.sku_id_snapshot,
                sku_name_snapshot=entry.sku_name_snapshot,
                sale_channel_snapshot=entry.sale_channel_snapshot,
                sale_store_id_snapshot=entry.sale_store_id_snapshot,
                sale_store_snapshot=entry.sale_store_snapshot,
                verify_store_id_snapshot=entry.verify_store_id_snapshot,
                verify_store_snapshot=entry.verify_store_snapshot,
                sale_time_snapshot=entry.sale_time_snapshot,
                verify_time_snapshot=entry.verify_time_snapshot,
                received_amount_cent_snapshot=entry.received_amount_cent_snapshot,
                fee_rate_snapshot=entry.fee_rate_snapshot,
                refund_at_snapshot=entry.refund_at_snapshot,
                adjustment_type_snapshot=entry.adjustment_type_snapshot,
            )
        )

    allocated_amounts = _allocate_dispute_adjustment(
        adjustment_amount_cent,
        [order.disputed_amount_cent for order, _ in selected_entries],
    )
    for index, ((_, original_entry), allocated_amount) in enumerate(
        zip(selected_entries, allocated_amounts, strict=True), start=1
    ):
        cloned_line = cloned_lines[original_entry.statement_line_id]
        cloned_line.adjustment_entry_count += 1
        cloned_line.adjustment_fee_cent += allocated_amount
        cloned_line.net_fee_cent += allocated_amount
        session.add(
            SettlementStatementEntry(
                statement_entry_id=f"statement-entry-{uuid4().hex}",
                statement_id=new_statement_id,
                statement_line_id=cloned_line.statement_line_id,
                source_type=2,
                source_record_id=f"dispute:{dispute.dispute_id}:{index}",
                original_fee_result_id=original_entry.original_fee_result_id,
                coupon_id=original_entry.coupon_id,
                order_id=original_entry.order_id,
                fee_direction=dispute.fee_direction,
                original_business_month=original_entry.original_business_month,
                statement_posting_month=current_statement.statement_month,
                product_scope=original_entry.product_scope,
                product_type=original_entry.product_type,
                base_amount_cent=0,
                fee_amount_cent=allocated_amount,
                rule_version="dispute-adjustment-v1",
                order_status_snapshot=original_entry.order_status_snapshot,
                coupon_status_snapshot=original_entry.coupon_status_snapshot,
                product_name_snapshot=original_entry.product_name_snapshot,
                sku_id_snapshot=original_entry.sku_id_snapshot,
                sku_name_snapshot=original_entry.sku_name_snapshot,
                sale_channel_snapshot=original_entry.sale_channel_snapshot,
                sale_store_id_snapshot=original_entry.sale_store_id_snapshot,
                sale_store_snapshot=original_entry.sale_store_snapshot,
                verify_store_id_snapshot=original_entry.verify_store_id_snapshot,
                verify_store_snapshot=original_entry.verify_store_snapshot,
                sale_time_snapshot=original_entry.sale_time_snapshot,
                verify_time_snapshot=original_entry.verify_time_snapshot,
                received_amount_cent_snapshot=original_entry.received_amount_cent_snapshot,
                fee_rate_snapshot=original_entry.fee_rate_snapshot,
                refund_at_snapshot=operated_at,
                adjustment_type_snapshot=3,
            )
        )

    adjusted_line_total = sum(
        line.net_fee_cent for line in cloned_lines.values() if line.fee_direction == dispute.fee_direction
    )
    expected_total = (
        new_statement.promotion_net_fee_cent
        if dispute.fee_direction == 1
        else new_statement.management_net_fee_cent
    )
    if adjusted_line_total != expected_total:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "STATEMENT_SNAPSHOT_INCONSISTENT",
            "调整后的账单头与汇总行金额不一致",
        )
    if should_auto_confirm:
        session.add(
            SettlementStatementConfirmation(
                confirmation_id=f"confirmation-dispute-{uuid4().hex}",
                statement_id=new_statement_id,
                fee_direction=dispute.fee_direction,
                confirmation_status=1,
                confirmed_amount_cent=expected_total,
                confirmed_by="system:auto-confirmation",
                confirmed_at=operated_at,
                idempotency_key_hash=sha256(
                    f"dispute-auto-confirmation-{dispute.dispute_id}".encode("utf-8")
                ).hexdigest(),
                request_payload_sha256=_canonical_billing_sha256(
                    {
                        "disputeId": dispute.dispute_id,
                        "resolutionNote": resolution_note,
                        "versionNo": new_statement.version_no,
                    }
                ),
            )
        )
    return new_statement


def _allocate_dispute_adjustment(total: int, weights: list[int]) -> list[int]:
    weight_total = sum(weights)
    absolute_total = abs(total)
    sign = 1 if total > 0 else -1
    allocated: list[int] = []
    remaining = absolute_total
    for index, weight in enumerate(weights):
        amount = (
            remaining
            if index == len(weights) - 1
            else absolute_total * weight // weight_total
        )
        allocated.append(sign * amount)
        remaining -= amount
    return allocated


def _raise_statement_version_conflict(
    request: Request, statement: SettlementStatement
) -> None:
    _raise_reporting_error(
        request,
        status.HTTP_409_CONFLICT,
        "STATEMENT_VERSION_CONFLICT",
        "账单版本已变化或不是当前版本",
        field="readVersion",
    )


def _invoice_status_db(value: str) -> int:
    return next(key for key, status_name in INVOICE_STATUS_NAMES.items() if status_name == value)


def _promotion_invoice_item(
    invoice: PromotionInvoice, allocation: PromotionInvoiceAllocation
) -> dict:
    return {
        **_promotion_invoice_header_item_from_invoice(invoice),
        "statement_id": allocation.statement_id,
        "statement_month": allocation.statement_month,
        "settlement_batch_month": allocation.settlement_batch_month,
        "allocated_amount_cent": allocation.allocated_amount_cent,
    }


def _promotion_invoice_header_item(session, invoice: PromotionInvoice) -> dict:
    allocations = list(session.scalars(select(PromotionInvoiceAllocation).where(
        PromotionInvoiceAllocation.invoice_id == invoice.invoice_id,
        PromotionInvoiceAllocation.is_current.is_(True),
    )))
    return {
        **_promotion_invoice_header_item_from_invoice(invoice),
        "allocations": [{
            "statement_id": allocation.statement_id,
            "statement_month": allocation.statement_month,
            "settlement_batch_month": allocation.settlement_batch_month,
            "allocated_amount_cent": allocation.allocated_amount_cent,
        } for allocation in allocations],
    }


def _promotion_invoice_header_item_from_invoice(invoice: PromotionInvoice) -> dict:
    return {
        "invoice_id": invoice.invoice_id,
        "physical_invoice_id": invoice.physical_invoice_id,
        "store_id": invoice.store_id,
        "version_no": invoice.version_no,
        "version_kind": "REGISTRATION" if invoice.version_kind == 1 else "FACTORY_RESULT",
        "is_current": invoice.is_current,
        "supersedes_invoice_id": invoice.supersedes_invoice_id,
        "replaces_invoice_id": invoice.replaces_invoice_id,
        "invoice_number": invoice.invoice_number,
        "invoice_date": invoice.invoice_date.isoformat(),
        "invoice_amount_cent": invoice.invoice_amount_cent,
        "buyer_name": invoice.buyer_name,
        "tax_rate_percent": invoice.tax_rate_percent,
        "status": INVOICE_STATUS_NAMES[invoice.invoice_status],
        "registered_at": _promotion_invoice_beijing_datetime(invoice.registered_at),
    }


def _promotion_invoice_replacement_chain(
    session,
    invoice: PromotionInvoice,
) -> list[PromotionInvoice]:
    def registration_for_physical(physical_invoice_id: str) -> PromotionInvoice | None:
        return session.scalar(
            select(PromotionInvoice).where(
                PromotionInvoice.physical_invoice_id == physical_invoice_id,
                PromotionInvoice.version_kind == 1,
            )
        )

    root = registration_for_physical(invoice.physical_invoice_id) or invoice
    connected: dict[str, PromotionInvoice] = {}
    pending = [root]
    while pending:
        current = pending.pop()
        if current.physical_invoice_id in connected:
            continue
        connected[current.physical_invoice_id] = current
        current_version_ids = list(
            session.scalars(
                select(PromotionInvoice.invoice_id).where(
                    PromotionInvoice.physical_invoice_id
                    == current.physical_invoice_id
                )
            )
        )
        source_ids = set(
            session.scalars(
                select(PromotionInvoiceReplacementSource.source_invoice_id).where(
                    PromotionInvoiceReplacementSource.replacement_invoice_id.in_(
                        current_version_ids
                    )
                )
            )
        )
        source_ids.update(
            source_id
            for source_id in session.scalars(
                select(PromotionInvoice.replaces_invoice_id).where(
                    PromotionInvoice.invoice_id.in_(current_version_ids),
                    PromotionInvoice.replaces_invoice_id.is_not(None),
                )
            )
            if source_id is not None
        )
        replacement_ids = set(
            session.scalars(
                select(
                    PromotionInvoiceReplacementSource.replacement_invoice_id
                ).where(
                    PromotionInvoiceReplacementSource.source_invoice_id.in_(
                        current_version_ids
                    )
                )
            )
        )
        replacement_ids.update(
            session.scalars(
                select(PromotionInvoice.invoice_id).where(
                    PromotionInvoice.replaces_invoice_id.in_(current_version_ids),
                    PromotionInvoice.version_kind == 1,
                )
            )
        )
        for related_id in source_ids | replacement_ids:
            related = session.scalar(
                select(PromotionInvoice).where(
                    PromotionInvoice.invoice_id == related_id
                )
            )
            if related is None:
                continue
            registration = (
                registration_for_physical(related.physical_invoice_id) or related
            )
            if registration.physical_invoice_id not in connected:
                pending.append(registration)
    return sorted(
        connected.values(),
        key=lambda item: (item.registered_at, item.invoice_id),
    )


def _promotion_invoice_lifecycle_event_item(
    event: PromotionInvoiceLifecycleEvent,
) -> dict:
    return {
        "lifecycle_event_id": event.lifecycle_event_id,
        "physical_invoice_id": event.physical_invoice_id,
        "invoice_id": event.invoice_id,
        "invoice_version": event.invoice_version,
        "event_type": PROMOTION_LIFECYCLE_EVENT_FROM_DB[event.event_type],
        "reason": event.reason,
        "read_version": event.read_version,
        "is_current": event.is_current,
        "operator_id": event.operator_id,
        "occurred_at": event.occurred_at,
    }


def _promotion_invoice_lifecycle_replay_response(
    session,
    event: PromotionInvoiceLifecycleEvent,
    payload_hash: str,
    request: Request,
    current_user: AuthContext,
):
    invoice = session.scalar(
        select(PromotionInvoice).where(PromotionInvoice.invoice_id == event.invoice_id)
    )
    _require_billing_store_scope(current_user, invoice.store_id, request)
    if event.request_payload_sha256 != payload_hash:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "IDEMPOTENCY_KEY_REUSED",
            "Idempotency-Key 已用于不同请求",
        )
    return _reporting_success(
        request,
        {
            "invoice": _promotion_invoice_header_item_from_invoice(invoice),
            "lifecycle_event": _promotion_invoice_lifecycle_event_item(event),
            "released_statement_months": sorted(
                session.scalars(
                    select(PromotionInvoiceAllocation.statement_month).where(
                        PromotionInvoiceAllocation.invoice_id == event.invoice_id
                    )
                )
            ),
        },
    )


def _promotion_invoice_registration_replay_response(
    session,
    invoice: PromotionInvoice,
    payload_hash: str,
    request: Request,
):
    if invoice.request_payload_sha256 != payload_hash:
        _raise_reporting_error(
            request,
            status.HTTP_409_CONFLICT,
            "IDEMPOTENCY_KEY_REUSED",
            "Idempotency-Key 已用于不同请求",
        )
    return _reporting_success(
        request,
        _promotion_invoice_header_item(session, invoice),
    )


def _management_invoice_item(invoice: InvoiceRecord) -> dict:
    return {
        "invoice_id": invoice.invoice_id,
        "store_id": invoice.store_id,
        "statement_id": invoice.statement_id,
        "statement_month": invoice.statement_month,
        "fee_direction": CONFIRMATION_DIRECTION_FROM_DB[invoice.fee_direction],
        "version_no": invoice.version_no,
        "is_current": invoice.is_current,
        "invoice_number": invoice.invoice_number,
        "invoice_date": invoice.invoice_date.isoformat(),
        "invoice_amount_cent": invoice.invoice_amount_cent,
        "status": INVOICE_STATUS_NAMES[invoice.invoice_status],
        "source_type": invoice.source_type,
        "import_batch_id": invoice.import_batch_id,
        "registered_at": (
            _finance_datetime(invoice.registered_at).isoformat()
            if invoice.registered_at is not None
            else None
        ),
        "factory_deduction_date": (
            invoice.factory_deduction_date.isoformat()
            if invoice.factory_deduction_date is not None
            else None
        ),
        "factory_deduction_amount_cent": invoice.factory_deduction_amount_cent,
        "settled_at": (
            _finance_datetime(invoice.registered_at).isoformat()
            if invoice.invoice_status == 3 and invoice.registered_at is not None
            else None
        ),
    }


def _finance_adjustment_type(value: int | None) -> str | None:
    """Map immutable adjustment facts to user-safe display values."""
    return {1: "REFUND", 2: "VERIFY_CANCEL", 3: "DISPUTE", 4: "OTHER"}.get(value)


def _finance_order_detail_definitions() -> list[dict]:
    """Describe immutable field sources for the finance order-detail UI."""
    return [
        {"group": "账单冻结", "fields": ["storeName", "sapCode", "statementMonth", "feeDirection"], "source": "current settlement statement version snapshot"},
        {"group": "订单冻结", "fields": ["orderStatus", "productName", "skuId", "skuName", "saleChannel", "saleTime", "verifyTime", "receivedAmountCent", "feeRate"], "source": "statement entry snapshot"},
        {"group": "发票事实", "fields": ["invoiceNumber", "submittedAt", "invoiceStatus", "settledAt"], "source": "current valid invoice fact"},
    ]


def _normalize_finance_order_detail_filters(
    *,
    request: Request,
    month: str,
    fee_direction: str,
    store_id: str | None,
    store_name: str | None,
    sap_code: str | None,
    invoice_number: str | None,
    order_id: str | None,
    sku_id: str | None,
    sale_channel: str | None,
    invoice_status: str | None,
    submitted_from: datetime | None,
    submitted_to: datetime | None,
    verify_from: datetime | None,
    verify_to: datetime | None,
) -> dict:
    """Normalize the one filter contract shared by list, count, export, and audit."""
    status_name = invoice_status.strip().upper() if invoice_status else None
    allowed_statuses = (
        set(INVOICE_STATUS_NAMES.values())
        if fee_direction == "PROMOTION"
        else {"SETTLED", "UNSETTLED"}
    )
    if status_name is not None:
        _validate_enum(status_name, allowed_statuses, "invoiceStatus", request)
    if submitted_from and submitted_to and submitted_from > submitted_to:
        _raise_reporting_error(
            request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED",
            "submittedFrom 不得晚于 submittedTo", field="submittedFrom",
        )
    if verify_from and verify_to and verify_from > verify_to:
        _raise_reporting_error(
            request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED",
            "verifyFrom 不得晚于 verifyTo", field="verifyFrom",
        )
    values = {
        "month": month,
        "feeDirection": fee_direction,
        "storeId": store_id.strip() if store_id else None,
        "storeName": store_name.strip() if store_name else None,
        "sapCode": sap_code.strip() if sap_code else None,
        "invoiceNumber": invoice_number.strip() if invoice_number else None,
        "orderId": order_id.strip() if order_id else None,
        "skuId": sku_id.strip() if sku_id else None,
        "saleChannel": sale_channel.strip().lower() if sale_channel else None,
        "invoiceStatus": status_name,
        "submittedFrom": submitted_from,
        "submittedTo": submitted_to,
        "verifyFrom": verify_from,
        "verifyTo": verify_to,
    }
    return {key: value for key, value in values.items() if value is not None}


def _finance_order_detail_audit_filters(filters: dict) -> dict:
    return {
        key: value.isoformat() if isinstance(value, datetime) else value
        for key, value in filters.items()
    }


def _finance_order_details_query(filters: dict):
    """Return the immutable projection used without variation by list and export."""
    direction = filters["feeDirection"]
    direction_code = CONFIRMATION_DIRECTION_TO_DB[direction]
    promotion_settled_date = (
        select(InvoiceStatusEvent.business_date)
        .where(
            InvoiceStatusEvent.invoice_id == PromotionInvoice.invoice_id,
            InvoiceStatusEvent.to_status == 3,
        )
        .order_by(InvoiceStatusEvent.occurred_at.desc(), InvoiceStatusEvent.event_id.desc())
        .limit(1)
        .correlate(PromotionInvoice)
        .scalar_subquery()
    )
    promotion_rejection_reason = (
        select(InvoiceStatusEvent.result_reason)
        .where(
            InvoiceStatusEvent.invoice_id == PromotionInvoice.invoice_id,
            InvoiceStatusEvent.to_status == 4,
        )
        .order_by(InvoiceStatusEvent.occurred_at.desc(), InvoiceStatusEvent.event_id.desc())
        .limit(1)
        .correlate(PromotionInvoice)
        .scalar_subquery()
    )
    sku_expr = SettlementStatementEntry.sku_id_snapshot
    sale_channel_expr = SettlementStatementEntry.sale_channel_snapshot
    verify_time_expr = SettlementStatementEntry.verify_time_snapshot
    conditions = [
        SettlementStatement.statement_month == filters["month"],
        SettlementStatement.is_current.is_(True),
        SettlementStatementEntry.fee_direction == direction_code,
    ]
    if filters.get("storeId"):
        conditions.append(SettlementStatement.store_id == filters["storeId"])
    if filters.get("storeName"):
        conditions.append(func.lower(SettlementStatement.store_name_snapshot).contains(filters["storeName"].lower()))
    if filters.get("sapCode"):
        conditions.append(SettlementStatement.sap_code_snapshot == filters["sapCode"])
    if filters.get("orderId"):
        conditions.append(SettlementStatementEntry.order_id == filters["orderId"])
    if filters.get("skuId"):
        conditions.append(sku_expr == filters["skuId"])
    if filters.get("saleChannel"):
        conditions.append(func.lower(sale_channel_expr) == filters["saleChannel"])
    if filters.get("verifyFrom"):
        conditions.append(verify_time_expr >= filters["verifyFrom"])
    if filters.get("verifyTo"):
        conditions.append(verify_time_expr <= filters["verifyTo"])
    active_invoice = PromotionInvoice if direction == "PROMOTION" else InvoiceRecord
    if filters.get("invoiceNumber"):
        conditions.append(active_invoice.invoice_number == filters["invoiceNumber"])
    if filters.get("submittedFrom"):
        conditions.append(active_invoice.registered_at >= filters["submittedFrom"])
    if filters.get("submittedTo"):
        conditions.append(active_invoice.registered_at <= filters["submittedTo"])
    if filters.get("invoiceStatus"):
        if direction == "PROMOTION":
            conditions.append(
                PromotionInvoice.invoice_status
                == next(key for key, value in INVOICE_STATUS_NAMES.items() if value == filters["invoiceStatus"])
            )
        elif filters["invoiceStatus"] == "SETTLED":
            conditions.append(InvoiceRecord.invoice_status == 3)
        else:
            conditions.append((InvoiceRecord.invoice_id.is_(None)) | (InvoiceRecord.invoice_status != 3))
    return (
        select(
            SettlementStatementEntry.statement_entry_id.label("statement_entry_id"),
            SettlementStatement.statement_id.label("statement_id"),
            SettlementStatement.store_id.label("store_id"),
            SettlementStatement.store_name_snapshot.label("store_name"),
            SettlementStatement.sap_code_snapshot.label("sap_code"),
            SettlementStatement.statement_month.label("statement_month"),
            SettlementStatementEntry.source_type.label("source_type"),
            SettlementStatementEntry.order_id.label("order_id"),
            SettlementStatementEntry.coupon_id.label("coupon_id"),
            SettlementStatementEntry.order_status_snapshot.label("order_status"),
            SettlementStatementEntry.coupon_status_snapshot.label("coupon_status"),
            SettlementStatementEntry.product_name_snapshot.label("product_name"),
            sku_expr.label("sku_id"),
            SettlementStatementEntry.sku_name_snapshot.label("sku_name"),
            sale_channel_expr.label("sale_channel"),
            SettlementStatementEntry.sale_store_id_snapshot.label("sale_store_id"),
            SettlementStatementEntry.sale_store_snapshot.label("sale_store_name"),
            SettlementStatementEntry.verify_store_id_snapshot.label("verify_store_id"),
            SettlementStatementEntry.verify_store_snapshot.label("verify_store_name"),
            SettlementStatementEntry.sale_time_snapshot.label("sale_time"),
            verify_time_expr.label("verify_time"),
            SettlementStatementEntry.received_amount_cent_snapshot.label("received_amount_cent"),
            SettlementStatementEntry.base_amount_cent.label("frozen_fee_base_cent"),
            SettlementStatementEntry.fee_rate_snapshot.label("actual_fee_rate"),
            SettlementStatementEntry.fee_amount_cent.label("frozen_fee_amount_cent"),
            SettlementStatementEntry.refund_at_snapshot.label("refund_time"),
            SettlementStatementEntry.adjustment_type_snapshot.label("adjustment_type"),
            PromotionInvoice.invoice_number.label("promotion_invoice_number"),
            PromotionInvoice.registered_at.label("promotion_submitted_at"),
            PromotionInvoice.invoice_status.label("promotion_invoice_status"),
            promotion_settled_date.label("promotion_settled_date"),
            promotion_rejection_reason.label("promotion_rejection_reason"),
            InvoiceRecord.invoice_number.label("management_invoice_number"),
            InvoiceRecord.registered_at.label("management_imported_at"),
            InvoiceRecord.invoice_status.label("management_invoice_status"),
            InvoiceRecord.factory_deduction_date.label("factory_deduction_date"),
            InvoiceRecord.factory_deduction_amount_cent.label("factory_deduction_amount_cent"),
        )
        .join(SettlementStatement, SettlementStatement.statement_id == SettlementStatementEntry.statement_id)
        .outerjoin(PromotionInvoiceAllocation, and_(PromotionInvoiceAllocation.statement_id == SettlementStatement.statement_id, PromotionInvoiceAllocation.is_current.is_(True)))
        .outerjoin(PromotionInvoice, and_(PromotionInvoice.invoice_id == PromotionInvoiceAllocation.invoice_id, PromotionInvoice.is_current.is_(True), PromotionInvoice.is_tombstone.is_(False)))
        .outerjoin(InvoiceRecord, and_(InvoiceRecord.statement_id == SettlementStatement.statement_id, InvoiceRecord.fee_direction == 2, InvoiceRecord.is_current.is_(True), InvoiceRecord.is_tombstone.is_(False)))
        .where(*conditions)
        .order_by(SettlementStatementEntry.order_id, SettlementStatementEntry.coupon_id, SettlementStatementEntry.source_type, SettlementStatementEntry.statement_entry_id)
    )


def _finance_order_detail_item(row, direction: str) -> dict:
    values = row._mapping
    as_datetime = lambda value: _finance_datetime(value).isoformat() if value is not None else None
    as_date = lambda value: value.isoformat() if value is not None else None
    rate = values["actual_fee_rate"]
    promotion_status = values["promotion_invoice_status"]
    management_status = values["management_invoice_status"]
    return {
        "statement_entry_id": values["statement_entry_id"],
        "statement_id": values["statement_id"],
        "store_id": values["store_id"],
        "store_name": values["store_name"],
        "sap_code": values["sap_code"],
        "statement_month": values["statement_month"],
        "fee_direction": direction,
        "order_id": values["order_id"],
        "coupon_id": values["coupon_id"],
        "order_status": values["order_status"],
        "coupon_status": values["coupon_status"],
        "product_name": values["product_name"],
        "sku_id": values["sku_id"],
        "sku_name": values["sku_name"],
        "sale_channel": values["sale_channel"],
        "sale_store_id": values["sale_store_id"],
        "sale_store_name": values["sale_store_name"],
        "verify_store_id": values["verify_store_id"],
        "verify_store_name": values["verify_store_name"],
        "sale_time": as_datetime(values["sale_time"]),
        "verify_time": as_datetime(values["verify_time"]),
        "received_amount_cent": values["received_amount_cent"],
        "frozen_fee_base_cent": values["frozen_fee_base_cent"],
        "actual_fee_rate": f"{rate:.6f}" if rate is not None else None,
        "frozen_fee_amount_cent": values["frozen_fee_amount_cent"],
        "refund_time": as_datetime(values["refund_time"]),
        "adjustment_type": _finance_adjustment_type(values["adjustment_type"]),
        "row_type": "ADJUSTMENT" if values["source_type"] == 2 else "ORIGINAL",
        "invoice_number": values["promotion_invoice_number"] if direction == "PROMOTION" else values["management_invoice_number"],
        "submitted_at": as_datetime(values["promotion_submitted_at"]) if direction == "PROMOTION" else None,
        "invoice_status": INVOICE_STATUS_NAMES.get(promotion_status) if direction == "PROMOTION" else None,
        "settled_at": as_date(values["promotion_settled_date"]) if direction == "PROMOTION" else as_date(values["factory_deduction_date"]),
        "rejection_reason": values["promotion_rejection_reason"] if direction == "PROMOTION" else None,
        "imported_at": as_datetime(values["management_imported_at"]) if direction == "MANAGEMENT" else None,
        "settlement_status": ("SETTLED" if management_status == 3 else "UNSETTLED") if direction == "MANAGEMENT" else None,
        "factory_deduction_date": as_date(values["factory_deduction_date"]) if direction == "MANAGEMENT" else None,
        "factory_deduction_amount_cent": values["factory_deduction_amount_cent"] if direction == "MANAGEMENT" else None,
    }


def _persist_finance_order_export_audit(
    session,
    *,
    request: Request,
    current_user: AuthContext,
    filters: dict,
    row_count: int,
    result: str,
    result_status: int,
) -> None:
    session.add(
        FinanceOperationAudit(
            audit_id=f"audit-{uuid4().hex}",
            operation_type="FINANCE_ORDER_DETAILS_EXPORT",
            target_type="FINANCE_ORDER_DETAILS",
            target_id=f"{filters['feeDirection']}|{filters['month']}",
            operator_id=current_user.username,
            operator_role=_finance_operator_role(current_user),
            before_snapshot=None,
            after_snapshot={"filters": filters, "rowCount": row_count, "result": result},
            result_status=result_status,
            request_id=request_id(request),
            occurred_at=utcnow(),
        )
    )
    session.commit()


def _finance_order_detail_csv(rows: list[dict]) -> str:
    fieldnames = [
        "statement_entry_id", "statement_id", "store_id", "store_name", "sap_code",
        "statement_month", "fee_direction", "order_id", "coupon_id", "order_status",
        "coupon_status", "product_name", "sku_id", "sku_name", "sale_channel",
        "sale_store_id", "sale_store_name", "verify_store_id", "verify_store_name",
        "sale_time", "verify_time", "received_amount_cent", "frozen_fee_base_cent",
        "actual_fee_rate", "frozen_fee_amount_cent", "refund_time", "adjustment_type",
        "row_type", "invoice_number", "submitted_at", "invoice_status", "settled_at",
        "rejection_reason", "imported_at", "settlement_status", "factory_deduction_date",
        "factory_deduction_amount_cent",
    ]
    buffer = StringIO(newline="")
    writer = DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()


def _normalize_invoice_status(value: str | None, request: Request) -> int | None:
    if value is None:
        return None
    normalized = value.upper()
    _validate_enum(normalized, set(INVOICE_STATUS_NAMES.values()), "invoiceStatus", request)
    return _invoice_status_db(normalized)


def _finance_datetime(value: datetime | None) -> datetime | None:
    if value is None or value.tzinfo is not None:
        return value
    return value.replace(tzinfo=timezone.utc)


def _promotion_invoice_beijing_datetime(value: datetime) -> datetime:
    normalized = value if value.tzinfo is not None else value.replace(tzinfo=timezone.utc)
    return normalized.astimezone(BEIJING_TIME_ZONE)


def _promotion_invoice_settlement_batch_month(value: datetime) -> str:
    beijing_time = _promotion_invoice_beijing_datetime(value)
    if beijing_time.day > 10:
        return f"{beijing_time.year:04d}-{beijing_time.month:02d}"
    if beijing_time.month == 1:
        return f"{beijing_time.year - 1:04d}-12"
    return f"{beijing_time.year:04d}-{beijing_time.month - 1:02d}"


def _parse_promotion_invoice_payload(payload: dict, request: Request) -> dict:
    store_id = payload.get("storeId")
    buyer_name = payload.get("buyerName")
    tax_rate_percent = payload.get("taxRatePercent")
    invoice_number = payload.get("invoiceNumber")
    invoice_date_value = payload.get("invoiceDate")
    invoice_amount = payload.get("invoiceAmountCent")
    rows = payload.get("allocations")
    replaces_invoice_id = payload.get("replacesInvoiceId")
    if not isinstance(store_id, str) or not store_id.strip():
        _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "storeId 为必填项", field="storeId")
    if buyer_name != PROMOTION_INVOICE_BUYER_NAME:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            f"buyerName 必须为 {PROMOTION_INVOICE_BUYER_NAME}",
            field="buyerName",
        )
    if (
        isinstance(tax_rate_percent, bool)
        or not isinstance(tax_rate_percent, int)
        or tax_rate_percent != PROMOTION_INVOICE_TAX_RATE_PERCENT
    ):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "taxRatePercent 必须为整数 6",
            field="taxRatePercent",
        )
    if not isinstance(invoice_number, str) or not invoice_number.isdigit() or len(invoice_number) != 20:
        _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "invoiceNumber 必须为 20 位数字", field="invoiceNumber")
    try:
        parsed_date = date.fromisoformat(invoice_date_value) if isinstance(invoice_date_value, str) else None
    except ValueError:
        parsed_date = None
    if parsed_date is None:
        _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "invoiceDate 必须为 YYYY-MM-DD", field="invoiceDate")
    if isinstance(invoice_amount, bool) or not isinstance(invoice_amount, int) or invoice_amount <= 0:
        _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "invoiceAmountCent 必须为正整数", field="invoiceAmountCent")
    if not isinstance(rows, list) or not rows:
        _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "allocations 至少包含一个完整账期", field="allocations")
    if replaces_invoice_id is not None and (
        not isinstance(replaces_invoice_id, str) or not replaces_invoice_id.strip()
    ):
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "replacesInvoiceId 必须为非空字符串",
            field="replacesInvoiceId",
        )
    allocations = []
    seen_months = set()
    for item in rows:
        if not isinstance(item, dict):
            _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "allocations 元素必须为对象", field="allocations")
        statement_id, statement_month = item.get("statementId"), item.get("statementMonth")
        allocated, read_version = item.get("allocatedAmountCent"), item.get("readVersion")
        promotion_invoice_group_id = item.get("promotionInvoiceGroupId")
        if not isinstance(statement_id, str) or not isinstance(statement_month, str):
            _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "分配必须包含 statementId 和 statementMonth", field="allocations")
        _validate_month(statement_month, "statementMonth", request)
        if statement_month in seen_months:
            _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "同一账期不得拆分多张发票", field="allocations")
        seen_months.add(statement_month)
        if isinstance(allocated, bool) or not isinstance(allocated, int) or isinstance(read_version, bool) or not isinstance(read_version, int) or read_version < 1:
            _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "分配金额和 readVersion 不合法", field="allocations")
        allocations.append({
            "statement_id": statement_id,
            "statement_month": statement_month,
            "allocated_amount_cent": allocated,
            "read_version": read_version,
            "promotion_invoice_group_id": (
                promotion_invoice_group_id.strip()
                if isinstance(promotion_invoice_group_id, str)
                and promotion_invoice_group_id.strip()
                else None
            ),
        })
    if sum(item["allocated_amount_cent"] for item in allocations) != invoice_amount:
        _raise_reporting_error(request, status.HTTP_422_UNPROCESSABLE_ENTITY, "VALIDATION_FAILED", "分配金额合计必须等于发票金额", field="invoiceAmountCent")
    return {
        "store_id": store_id.strip(),
        "buyer_name": buyer_name,
        "tax_rate_percent": tax_rate_percent,
        "invoice_number": invoice_number,
        "invoice_date": parsed_date,
        "invoice_amount_cent": invoice_amount,
        "replaces_invoice_id": (
            replaces_invoice_id.strip()
            if isinstance(replaces_invoice_id, str)
            else None
        ),
        "allocations": allocations,
    }


def _call_reporting_store(request: Request, operation, filters: dict):
    try:
        return operation(filters)
    except ReportingPermissionError as exc:
        _raise_reporting_error(
            request,
            status.HTTP_403_FORBIDDEN,
            "DATA_SCOPE_FORBIDDEN",
            str(exc),
        )
    except ReportingValidationError as exc:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            str(exc),
            field=exc.field,
        )


def _raise_reporting_error(
    request: Request,
    status_code: int,
    code: str,
    message: str,
    *,
    field: str | None = None,
    data: dict | None = None,
) -> None:
    errors = [] if field is None else [{"field": field, "reason": message}]
    detail = {
        "code": code,
        "message": message,
        "errors": errors,
        "requestId": request_id(request),
    }
    if data is not None:
        detail["data"] = camelize(data)
    raise HTTPException(
        status_code=status_code,
        detail=detail,
    )


def _validate_enum(
    value: str, allowed: set[str], field: str, request: Request
) -> None:
    if value not in allowed:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            f"{field} 必须是 {', '.join(sorted(allowed))} 之一",
            field=field,
        )


def _validate_month(value: str, field: str, request: Request) -> None:
    if len(value) != 7 or value[4:5] != "-" or not value[:4].isdigit() or not value[5:].isdigit():
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            f"{field} 必须使用 YYYY-MM 格式",
            field=field,
        )
    month_number = int(value[5:])
    if month_number < 1 or month_number > 12:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            f"{field} 月份必须在 01 到 12 之间",
            field=field,
        )


def _validate_product_selection(
    store, product_scope: str, product_type: str, request: Request
) -> None:
    scope_map = getattr(store, "product_scope_type_map", lambda: {})()
    available_types = set(getattr(store, "list_product_types", lambda: ["all"])())
    if product_scope != "all" and product_scope not in scope_map:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "productScope 不在可用产品范围内",
            field="productScope",
        )
    if product_type == "all":
        return
    allowed_types = (
        available_types if product_scope == "all" else set(scope_map.get(product_scope, []))
    )
    if product_type not in allowed_types:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "productType 不属于所选 productScope",
            field="productType",
        )


def _validate_monthly_context(
    store, store_id: str, month: str, request: Request
) -> None:
    store_exists = getattr(store, "store_exists", lambda _store_id: True)
    if not store_exists(store_id):
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "门店不存在",
            field="storeId",
        )
    context_exists = getattr(
        store,
        "monthly_settlement_context_exists",
        lambda _store_id, _month: True,
    )
    if not context_exists(store_id, month):
        _raise_reporting_error(
            request,
            status.HTTP_404_NOT_FOUND,
            "RESOURCE_NOT_FOUND",
            "门店账期不存在",
            field="month",
        )


def _order_fee_filters(
    *,
    request: Request,
    store,
    current_user: AuthContext,
    statement_id: str | None,
    statement_line_id: str | None,
    store_id: str | None,
    month: str | None,
    sale_month: str | None,
    verify_month: str | None,
    fee_direction: str,
    product_scope: str,
    product_type: str,
    fee_rates: list[str] | None,
    rule_versions: list[str] | None,
    data_status: str | None,
    q: str | None,
    page: int,
    page_size: int,
) -> dict:
    statement_id = (statement_id or "").strip() or None
    statement_line_id = (statement_line_id or "").strip() or None
    store_id = (store_id or "").strip() or None
    month = (month or "").strip() or None
    if statement_line_id and not statement_id:
        _raise_reporting_error(
            request,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            "VALIDATION_FAILED",
            "statementLineId 必须与 statementId 同时提供",
            field="statementLineId",
        )
    if statement_id:
        if not statement_line_id:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "VALIDATION_FAILED",
                "statementId 存在时 statementLineId 必填",
                field="statementLineId",
            )
        if store_id or month:
            _raise_reporting_error(
                request,
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                "VALIDATION_FAILED",
                "锁账明细上下文不能与门店月份上下文混用",
                field="storeId" if store_id else "month",
            )
    if store_id:
        _require_store_scope(current_user, store_id)
    if month:
        _validate_month(month, "month", request)
    if store_id and month:
        _validate_monthly_context(store, store_id, month, request)
    if sale_month:
        _validate_month(sale_month, "saleMonth", request)
    if verify_month:
        _validate_month(verify_month, "verifyMonth", request)
    fee_direction = fee_direction.upper()
    _validate_enum(fee_direction, FEE_DIRECTIONS, "feeDirection", request)
    normalized_data_status = data_status.upper() if data_status else None
    if normalized_data_status:
        _validate_enum(normalized_data_status, DATA_STATUSES, "dataStatus", request)
    _validate_product_selection(store, product_scope, product_type, request)
    has_source_context = bool(statement_id and statement_line_id) or bool(
        store_id and month
    )
    return {
        "statement_id": statement_id,
        "statement_line_id": statement_line_id,
        "store_id": store_id,
        "month": month,
        "sale_month": sale_month,
        "verify_month": verify_month,
        "fee_direction": fee_direction,
        "product_scope": product_scope,
        "product_type": product_type,
        "fee_rates": (fee_rates or []) if has_source_context else [],
        "rule_versions": (rule_versions or []) if has_source_context else [],
        "data_status": normalized_data_status,
        "q": (q or "").strip() or None,
        "page": page,
        "page_size": page_size,
        "scope_store_ids": (
            None if current_user.has_global_data_access else current_user.store_ids
        ),
    }


def _require_store_scope(current_user: AuthContext, store_id: str) -> None:
    if current_user.has_global_data_access:
        return
    if store_id not in current_user.store_ids:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Store is outside current account scope",
        )


def _require_store_actor(current_user: AuthContext, store_id: str) -> None:
    if current_user.role != "store" or store_id not in current_user.store_ids:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Store API requires the matching store account",
        )


def _resolve_sales_dashboard_store_id(
    current_user: AuthContext, store_id: str | None
) -> str | None:
    normalized_store_id = (store_id or "").strip()
    if normalized_store_id:
        _require_store_scope(current_user, normalized_store_id)
        return normalized_store_id
    if current_user.has_global_data_access:
        return None
    raise HTTPException(
        status_code=status.HTTP_403_FORBIDDEN,
        detail="Store is required for current account scope",
    )
