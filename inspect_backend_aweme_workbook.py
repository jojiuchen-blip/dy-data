import json
from pathlib import Path

import openpyxl


path = Path(r"D:\浏览器下载\抖音号明细-2026-06-09.xlsx")
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

result = []
for ws in wb.worksheets:
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = ["" if v is None else str(v).strip() for v in row]
        if any(values):
            rows.append({"row": i, "values": values})
        if len(rows) >= 20:
            break
    result.append({
        "sheet": ws.title,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "first_nonempty_rows": rows,
    })

print(json.dumps(result, ensure_ascii=False, indent=2))
